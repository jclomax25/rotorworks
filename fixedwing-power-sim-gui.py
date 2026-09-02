"""
fixedwing-power-sim-gui.py
--------------------------
Fixed-Wing UAV performance simulator.

Physics backend is fully rewritten for fixed-wing aerodynamics:
  - Lift / Drag polar model (CD0 + k·CL²)
  - Stall speed, pitch speed, tip speed
  - Angle of attack at cruise
  - Reynolds number
  - Available thrust and specific thrust
  - Rate of climb (instantaneous and max)
  - Max angle of climb
  - Best endurance speed (max L^1.5/D)
  - Best range speed     (max L/D)
  - Takeoff ground roll distance
  - Max propeller power check
  - Flight time and range vs speed curves

Battery / Motor / ESC / Avionics / Propeller models are identical to
the multicopter sim so the same hardware specs can be entered.

Modes:
  - Motor electrical model  (KV, Rm, I0, limits)
  - Motor/prop test table   (CSV with Thrust_g / Power_W columns)
  - Theoretical actuator    (ideal momentum + efficiency)

GUI: Tkinter notebook with tabs:
  Airframe | Battery | Motor | ESC | Avionics | Propeller | Environment
Output panels (right side):
  Plots | Status | Metrics | Mission Plots

CLI usage (all units SI unless noted):
  python fixedwing-power-sim-gui.py --gui
  python fixedwing-power-sim-gui.py --weight 2.5 --wing_area 0.45 --wing_span 1.6 \\
      --CD0 0.028 --CL_max 1.3 --oswald 0.82 \\
      --battery_operating_voltage_min 3.5 --battery_operating_voltage_max 4.2 \\
      --battery_operating_voltage_nominal 3.8 --battery_series_units 4 \\
      --battery_cell_capacity 5000 --battery_cell_weight_g 75 \\
      --battery_charge_current_max 5 --battery_discharge_c_cont 25 \\
      --battery_resistance_cell 5 \\
      --motor_kv 920 --motor_idle_current 0.8 --motor_idle_voltage 7.0 \\
      --motor_rated_voltage 16 --motor_resistance 0.06 \\
      --motor_max_current 40 --motor_max_power 500 \\
      --prop_diameter 10 --prop_pitch 4.5 \\
      --plot
"""

from __future__ import annotations

import math
import argparse
import json
import os
import re
from dataclasses import dataclass, field
from typing import Optional, List, Tuple, Dict
import sys

import pandas as pd
import numpy as np

# ------------------------------------------------------------------
# Shared core, extracted so the two simulators cannot drift apart.
# See rotorworks_core.py for what lives there and what deliberately
# does not. It must sit beside this file.
# ------------------------------------------------------------------
try:
    import rotorworks_core as core
except ImportError as _exc:      # pragma: no cover - install/deploy problem
    raise SystemExit(
        "rotorworks_core.py could not be imported. It must sit in the same "
        f"folder as this script.\nOriginal error: {_exc}"
    )

# Names re-exported under their historical spellings so existing call sites,
# saved scripts and the test suite keep working unchanged.
SOC_PRESETS = core.SOC_PRESETS
SOC_PRESET_ALIASES = core.SOC_PRESET_ALIASES
Tooltip = _Tooltip = core.Tooltip
parse_float_list = core.parse_float_list
parse_soc_breakpoints = core.parse_soc_breakpoints
wind_components_mps = core.wind_components_mps
groundspeed_along_track_mps = core.groundspeed_along_track_mps
thermal_step = core.thermal_step
_eval_poly = core.eval_poly
_interp_linear_clamped = core.interp_linear_clamped
_battery_preset_key = core.battery_preset_key
_normalize_soc_curves = core.normalize_soc_curves
_load_soc_curve_csv = core.load_soc_curve_csv
_configure_battery_soc_model = core.configure_battery_soc_model
_soc_model_short_label = core.soc_model_short_label
_fit_propeller_curve = core.fit_propeller_curve
kinetic_power_term_W = core.kinetic_power_term_W
ramp_speed = core.ramp_speed

# Build identifier. Shown in the title bar, the Output pane and Help > About
# so you can always tell which copy of the script you are running.
SIM_VERSION = "2.12.0"
SIM_BUILD_NOTE = "Airspeed-dependent thrust available; climb rates corrected"
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
# FigureCanvasTkAgg is imported lazily inside launch_gui() so that
# the physics engine can be used on headless / no-Tkinter systems.

# ============================================================
# CONSTANTS — ISA (International Standard Atmosphere)
# ============================================================
RHO0    = 1.225       # kg/m³  — sea-level air density
R_AIR   = 287.05      # J/kg/K — specific gas constant for dry air
T0      = 288.15      # K      — sea-level temperature (15 °C)
P0      = 101325.0    # Pa     — sea-level pressure
L_LAPSE = 0.0065      # K/m    — temperature lapse rate (troposphere)
G0      = 9.80665     # m/s²   — standard gravity
MU_AIR  = 1.789e-5    # Pa·s   — dynamic viscosity of air at 15 °C (≈ constant for subsonic UAV)


def isa_density(altitude_m: float,
                temperature_C: Optional[float] = None,
                pressure_Pa: Optional[float] = None) -> float:
    """
    Air density [kg/m^3]. Thin wrapper over the shared core implementation.

    Kept under this name because both simulators, their CLIs and saved
    scripts refer to it. The two used to carry separate implementations
    that disagreed about pressure overrides; there is now exactly one.
    """
    return core.air_density(altitude_m, temperature_C, pressure_Pa)


# ============================================================
# BATTERY MODEL  (identical to multicopter sim)
# ============================================================
# ============================================================
# BATTERY STATE-OF-CHARGE (SoC) MODEL
# ============================================================
# A real pack's open-circuit voltage sags non-linearly as it empties, and its
# internal resistance climbs steeply at low SoC.  Modelling that matters for
# endurance: the linear fallback anchors pack voltage at full charge, which
# flatters current draw late in a flight.
#
# These tables mirror the multicopter simulator so the two tools agree.
# They are deliberately conservative approximations, not cell datasheets.




class BatteryConfig:
    """
    Li-polymer / Li-Ion battery pack model.

    Pack voltage layout:
        V_pack  = V_cell × N_series_cells

    Under-load voltage (Thevenin model):
        V_load = V_max_pack - I_pack × R_pack
        where R_pack = R_cell_series × N_series / N_parallel  [Ω]

    Usable energy:
        E_usable = C_pack_Ah × V_nom_pack × (discharge_percent / 100)  [Wh]
    """
    def __init__(self,
                 chemistry: Optional[str],
                 operating_voltage_min: float,
                 operating_voltage_nominal: float,
                 operating_voltage_max: float,
                 unit_mode: str = "cell",
                 series_units: int = 1,
                 parallel_units: int = 1,
                 cells_series_per_unit: int = 1,
                 cells_parallel_per_unit: int = 1,
                 pack_weight_g: Optional[float] = None,
                 cell_weight_g: Optional[float] = None,
                 cell_capacity_mAh: Optional[float] = None,
                 pack_capacity_mAh: Optional[float] = None,
                 unit_energy_density: Optional[float] = None,
                 charge_current_max: Optional[float] = None,
                 discharge_cont_A: Optional[float] = None,
                 discharge_max_A: Optional[float] = None,
                 discharge_c_cont: Optional[float] = None,
                 discharge_c_max: Optional[float] = None,
                 discharge_percent: float = 100.0,
                 resistance_cell_mOhm: float = 0.0,
                 soc_model: str = "auto",
                 soc_curve_csv: Optional[str] = None,
                 soc_bp: Optional[List[float]] = None,
                 ocv_cell_bp: Optional[List[float]] = None,
                 r_scale_bp: Optional[List[float]] = None):
        self.chemistry = chemistry
        self.operating_voltage_min      = float(operating_voltage_min)
        self.operating_voltage_nominal  = float(operating_voltage_nominal)
        self.operating_voltage_max      = float(operating_voltage_max)
        self.unit_mode = str(unit_mode).strip().lower() if unit_mode is not None else "cell"

        if self.unit_mode not in ("cell", "pack"):
            self.unit_mode = "pack" if pack_weight_g is not None else "cell"

        if self.unit_mode == "cell":
            cells_series_per_unit   = 1
            cells_parallel_per_unit = 1

        # Counts must be at least 1 — 0 would give a zero-volt pack and
        # divide-by-zero downstream.
        self.cells_series_per_unit   = max(int(cells_series_per_unit), 1)
        self.cells_parallel_per_unit = max(int(cells_parallel_per_unit), 1)
        self.series_units   = max(int(series_units), 1)
        self.parallel_units = max(int(parallel_units), 1)

        self.series_cells   = self.series_units   * self.cells_series_per_unit
        self.parallel_cells = self.parallel_units * self.cells_parallel_per_unit
        self.total_cells    = self.series_cells   * self.parallel_cells

        self.vmin_pack  = self.operating_voltage_min     * self.series_cells
        self.vnom_pack  = self.operating_voltage_nominal * self.series_cells
        self.vmax_pack  = self.operating_voltage_max     * self.series_cells

        self.pack_weight_g    = float(pack_weight_g)    if pack_weight_g    is not None else None
        self.cell_weight_g    = float(cell_weight_g)    if cell_weight_g    is not None else None
        self.cell_capacity_mAh = float(cell_capacity_mAh) if cell_capacity_mAh is not None else None
        self.pack_capacity_mAh = float(pack_capacity_mAh) if pack_capacity_mAh is not None else None

        # ------------------------------------------------------------------
        # Effective pack capacity in mAh.
        #
        # PHYSICS: capacity (Ah) is set by the number of PARALLEL branches.
        # Wiring units in SERIES raises voltage, NOT capacity.  Total energy
        # still rises with series count because  E = C_Ah x V_pack.
        #
        #   4S1P of 5000 mAh packs ->  5000 mAh @ 14.8 V =  74 Wh
        #   4S2P of 5000 mAh packs -> 10000 mAh @ 14.8 V = 148 Wh
        #   8S1P of 5000 mAh packs ->  5000 mAh @ 29.6 V = 148 Wh  <- same energy, 2x voltage
        # ------------------------------------------------------------------
        if self.unit_mode == "cell":
            self.capacity_mAh = (self.cell_capacity_mAh or 0.0) * self.parallel_cells
        else:  # "pack" — only parallel packs add capacity
            self.capacity_mAh = (self.pack_capacity_mAh or 0.0) * self.parallel_units
        self.capacity_Ah = self.capacity_mAh / 1000.0

        # Weight
        if self.unit_mode == "cell":
            self.weight_g = (self.cell_weight_g or 0.0) * self.total_cells
        else:
            self.weight_g = (self.pack_weight_g or 0.0) * self.series_units * self.parallel_units

        # Derived energy density  [Wh/kg]
        if unit_energy_density is not None:
            self.energy_density_Wh_per_kg = float(unit_energy_density)
        else:
            wkg = self.weight_g / 1000.0
            self.energy_density_Wh_per_kg = (self.capacity_Wh / wkg) if wkg > 0 else 0.0

        self.charge_current_max = float(charge_current_max) if charge_current_max is not None else 0.0

        # Discharge limits — prefer explicit A, fall back to C-rate × capacity
        if discharge_cont_A is not None:
            self.discharge_cont_A = float(discharge_cont_A)
        elif discharge_c_cont is not None:
            self.discharge_cont_A = float(discharge_c_cont) * self.capacity_Ah
        else:
            self.discharge_cont_A = float("inf")

        if discharge_max_A is not None:
            self.discharge_max_A = float(discharge_max_A)
        elif discharge_c_max is not None:
            self.discharge_max_A = float(discharge_c_max) * self.capacity_Ah
        else:
            self.discharge_max_A = self.discharge_cont_A

        self.discharge_c_cont = (discharge_c_cont if discharge_c_cont is not None
                                 else (self.discharge_cont_A / self.capacity_Ah if self.capacity_Ah > 0 else None))
        self.discharge_c_max  = (discharge_c_max  if discharge_c_max  is not None
                                 else (self.discharge_max_A  / self.capacity_Ah if self.capacity_Ah > 0 else None))

        self.discharge_percent = min(max(float(discharge_percent), 0.0), 100.0)
        self.usable_fraction   = self.discharge_percent / 100.0

        # Internal resistance  [Ω per pack]
        self.resistance_cell = float(resistance_cell_mOhm) / 1000.0   # Ω per cell

        # ---- State-of-charge model ----------------------------------
        # "auto" picks a preset from the chemistry label; "linear" disables
        # the curve and anchors pack voltage at full charge (legacy behaviour).
        self.soc_model = str(soc_model or "auto").strip().lower()
        self.soc_curve_csv = (str(soc_curve_csv).strip() if soc_curve_csv else None)
        self.soc_nonlinear_enabled: bool = False
        self.soc_model_source: str = "linear-fallback"
        self.soc_bp: List[float] = []
        self.ocv_cell_bp: List[float] = []
        self.r_scale_bp: List[float] = []
        _configure_battery_soc_model(
            self,
            model=self.soc_model,
            curve_csv=self.soc_curve_csv,
            soc_bp=soc_bp,
            ocv_cell_bp=ocv_cell_bp,
            r_scale_bp=r_scale_bp,
        )

    # ---- SoC-dependent pack behaviour --------------------------------
    def ocv_at_soc(self, soc: float) -> float:
        """Pack open-circuit voltage at a given state of charge (0..1)."""
        if self.soc_nonlinear_enabled and self.soc_bp:
            ocv_cell = _interp_linear_clamped(
                min(max(float(soc), 0.0), 1.0),
                list(self.soc_bp), list(self.ocv_cell_bp))
            return max(float(ocv_cell) * float(self.series_cells), float(self.vmin_pack))
        # Linear fallback: anchored at full-charge voltage.
        return float(self.vmax_pack)

    def resistance_at_soc(self, soc: float) -> float:
        """Pack internal resistance at a given SoC — rises as the pack empties."""
        base_r = max(float(self.pack_resistance), 0.0)
        if self.soc_nonlinear_enabled and self.soc_bp:
            scale = _interp_linear_clamped(
                min(max(float(soc), 0.0), 1.0),
                list(self.soc_bp), list(self.r_scale_bp))
            return base_r * max(float(scale), 0.05)
        return base_r

    @property
    def pack_resistance(self) -> float:
        """R_pack = R_cell × N_series / N_parallel  [Ω]"""
        return (self.resistance_cell * self.series_cells
                / max(self.parallel_cells, 1))

    @property
    def capacity_Wh(self) -> float:
        return self.capacity_Ah * self.vnom_pack

    @property
    def usable_Wh(self) -> float:
        return self.capacity_Wh * self.usable_fraction

    def voltage_under_load(self, current_A: float,
                           soc: Optional[float] = None) -> float:
        """
        V_load = OCV(soc) - I x R_pack(soc), clamped at V_min.

        `soc` defaults to 1.0 (full charge), which reproduces the previous
        behaviour exactly when no SoC is supplied.
        """
        soc_eval = 1.0 if soc is None else min(max(float(soc), 0.0), 1.0)
        ocv = self.ocv_at_soc(soc_eval)
        r   = self.resistance_at_soc(soc_eval)
        v = ocv - float(current_A) * r
        return max(float(v), float(self.vmin_pack))

    def soc_after_energy_draw(self, soc_now: float, energy_draw_Wh: float) -> float:
        """Advance SoC after drawing a given amount of energy."""
        usable_wh = max(float(self.usable_Wh), 1e-9)
        return min(max(float(soc_now) - max(float(energy_draw_Wh), 0.0) / usable_wh, 0.0), 1.0)


# ============================================================
# MOTOR MODEL
# ============================================================
class MotorConfig:
    """
    Brushless motor electrical model.

    Back-EMF constant:
        K_e = 1 / KV  [V·s/rad] when KV is in RPM/V

    Shaft power (mechanical):
        P_shaft = (V_motor - I·Rm) × I - I0 × V0
               = η_motor × V_motor × I  (approximate)

    More precisely, using KV model:
        ω  [rad/s] = KV × 2π/60 × V_back_emf
        V_back_emf = V_motor - I × Rm
        I0 = no-load current at idle voltage V0

    We store the raw params and compute shaft power in the physics layer.
    """
    def __init__(self,
                 kv:           Optional[float],
                 idle_current:  float,
                 idle_voltage:  float,
                 rated_voltage: int,
                 resistance:    float,
                 max_current:   float,
                 max_power:     float,
                 pole_count:    Optional[int]   = None,
                 weight_g:      Optional[float] = None,
                 size_mm:       Optional[str]   = None):
        self.kv           = None if kv is None else float(kv)
        self.idle_current = float(idle_current)
        self.idle_voltage = float(idle_voltage)
        self.resistance   = float(resistance)
        self.max_current  = float(max_current)
        self.max_power    = float(max_power)
        self.rated_voltage= int(rated_voltage)
        self.pole_count   = pole_count
        self.weight_g     = weight_g
        self.size_mm      = size_mm


# ============================================================
# ESC MODEL
# ============================================================
class ESCConfig:
    """
    Electronic Speed Controller model.
    Losses:  P_loss = I² × R_esc + I_idle × V  [W per ESC]
    """
    def __init__(self,
                 voltage_rating:      int,
                 continuous_current_A: float,
                 max_current_A:        float,
                 idle_current_A:       float,
                 resistance:           float,
                 weight_g: Optional[float] = None):
        self.voltage_rating     = int(voltage_rating)
        self.continuous_rating_A = float(continuous_current_A)
        self.max_current_A      = float(max_current_A)
        self.idle_current_A     = float(idle_current_A)
        self.resistance         = float(resistance)
        self.weight_g           = weight_g


# ============================================================
# AVIONICS / PERIPHERALS
# ============================================================
class AvionicsConfig:
    """
    BEC-regulated avionics loads.
    voltage_tree: {V_rail: (I_rail [A], BEC_efficiency)} dict

    Input power from battery for each rail:
        P_in = (V_rail × I_rail) / η_bec   [W]
    """
    def __init__(self, voltage_tree: Optional[dict] = None):
        self.voltage_tree = voltage_tree or {}


def avionics_input_power_W(avionics: Optional[AvionicsConfig]) -> float:
    if avionics is None or not avionics.voltage_tree:
        return 0.0
    total = 0.0
    for v, (i, eff) in avionics.voltage_tree.items():
        total += (float(v) * float(i)) / max(float(eff), 1e-9)
    return total


def parse_voltage_tree(spec: Optional[str]) -> dict:
    """Parse 'V:(I,eff), V2:(I2,eff2)' or 'V:I:eff, ...' string.

    Splits on commas that are NOT inside parentheses so that the
    inner 'I,eff' tuple survives the tokenisation step.
    """
    if not spec:
        return {}
    if isinstance(spec, dict):
        return {float(k): (float(v[0]), float(v[1])) for k, v in spec.items()}

    # Split on commas outside parentheses
    parts, depth, buf = [], 0, ""
    for ch in str(spec):
        if ch == "(":
            depth += 1; buf += ch
        elif ch == ")":
            depth -= 1; buf += ch
        elif ch == "," and depth == 0:
            parts.append(buf.strip()); buf = ""
        else:
            buf += ch
    if buf.strip():
        parts.append(buf.strip())

    out: Dict[float, Tuple[float, float]] = {}
    for p in parts:
        if not p:
            continue
        m = re.match(r"^([0-9.]+)\s*:\s*\(\s*([0-9.]+)\s*,\s*([0-9.]+)\s*\)$", p)
        if m:
            v, i, e = map(float, m.groups())
        else:
            m2 = re.match(r"^([0-9.]+)\s*:\s*([0-9.]+)\s*:\s*([0-9.]+)$", p)
            if not m2:
                raise ValueError(f"Bad avionics spec token: {p!r}")
            v, i, e = map(float, m2.groups())
        out[v] = (i, e)
    return out


# ============================================================
# PROPELLER MODEL
# ============================================================
def load_prop_table(path: str) -> pd.DataFrame:
    """
    Load motor/prop test CSV. Accepts eCalc-style and simple headers.
    Required columns after renaming: Thrust_g, Power_W
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    try:
        df0 = pd.read_csv(path)
    except Exception:
        df0 = pd.DataFrame()

    def looks_ok(df):
        cols = {str(c).strip().lower() for c in df.columns}
        return "thrust_g" in cols or "thrust (g)" in cols

    if len(df0.columns) > 1 and looks_ok(df0):
        df = df0
    else:
        raw = pd.read_csv(path, header=None)
        header_row = None
        for i in range(min(len(raw), 25)):
            # Coerce every cell to a string explicitly. Series.astype(str)
            # leaves NaN as a real float under pandas' newer "str" dtype, so a
            # sparse title row ("Test Data,,,,,") puts floats in this list and
            # the membership test below raises
            # "argument of type 'float' is not iterable".
            row = ["" if pd.isna(v) else str(v).strip().lower()
                   for v in raw.iloc[i].tolist()]
            if any("thrust" in x for x in row) and any("power" in x for x in row):
                header_row = i
                break
        if header_row is None:
            raise ValueError(f"Cannot locate header row in {path}")
        df = pd.read_csv(path, header=header_row)

    rmap = {}
    for c in df.columns:
        cl = str(c).strip().lower()
        if cl in ("thrust_g", "thrust (g)"): rmap[c] = "Thrust_g"
        elif cl in ("power_w",  "power (w)"): rmap[c] = "Power_W"
        elif cl in ("current_a","current (a)"): rmap[c] = "Current_A"
        elif cl in ("voltage_v","voltage (v)"): rmap[c] = "Voltage_V"
        elif cl == "rpm":                       rmap[c] = "RPM"
        elif cl.startswith("throttle"):         rmap[c] = "Throttle_pct"
        elif "efficiency" in cl:                rmap[c] = "Efficiency_gW"
    df = df.rename(columns=rmap)

    for col in ("Thrust_g","Power_W","Current_A","Voltage_V","RPM","Efficiency_gW","Throttle_pct"):
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace("%","",regex=False).str.strip(), errors="coerce")

    if "Thrust_g" not in df.columns or "Power_W" not in df.columns:
        raise ValueError("CSV must have Thrust and Power columns.")

    df = (df.dropna(subset=["Thrust_g","Power_W"])
            .sort_values("Thrust_g")
            .drop_duplicates(subset=["Thrust_g"], keep="last")
            .reset_index(drop=True))
    return df


class PropellerConfig:
    """
    Propeller geometry used for:
      - Pitch speed  = pitch_m × n  [m/s]  (speed at advance ratio J = 1)
      - Tip speed    = π × D × n   [m/s]  where n [rev/s]
      - Thrust / power from KV-based or table-based motor model
    """
    def __init__(self,
                 diameter_in:  float,
                 pitch_in:     float,
                 blades:       int   = 2,
                 max_rpm:      float = 0.0,
                 max_thrust_g: float = 0.0,
                 table_csv:    Optional[str]   = None,
                 TConst:       Optional[float] = None,
                 PConst:       Optional[float] = None,
                 weight_g:     Optional[float] = None):
        self.diameter_in  = float(diameter_in)
        self.pitch_in     = float(pitch_in)
        self.blades       = int(blades)
        self.max_rpm      = float(max_rpm)
        self.max_thrust_g = float(max_thrust_g)
        self.TConst       = TConst
        self.PConst       = PConst
        self.weight_g     = float(weight_g) if weight_g is not None else 20.0
        self.table: Optional[pd.DataFrame] = None
        if table_csv:
            self.table = load_prop_table(table_csv)
            self._cache_table_arrays()

    def _cache_table_arrays(self) -> None:
        """
        Precompute the scalars and arrays the hot paths need.

        Every one of these was previously recomputed from the DataFrame on
        each call. The climb-rate and best-speed searches evaluate thrust
        hundreds of times per run, so a pandas reduction per evaluation turned
        a 5 ms calculation into 50 ms and made the window stop responding.
        """
        self._thrust_g_arr = None
        self._power_w_arr = None
        self._thrust_g_min = None
        self._thrust_g_max = None
        if self.table is None or "Thrust_g" not in self.table:
            return
        self._thrust_g_arr = self.table["Thrust_g"].to_numpy(dtype=float)
        if "Power_W" in self.table:
            self._power_w_arr = self.table["Power_W"].to_numpy(dtype=float)
        if self._thrust_g_arr.size:
            self._thrust_g_min = float(self._thrust_g_arr.min())
            self._thrust_g_max = float(self._thrust_g_arr.max())
        self._fit_static_power_model()

    def _fit_static_power_model(self) -> None:
        """
        Fit  P = a * T^1.5 + b  to the measured static data.

        Momentum theory gives ideal static power proportional to T^1.5, and
        `b` captures the losses that do NOT vanish with thrust: motor no-load
        current, iron losses, ESC quiescent draw.

        That second term matters enormously when extrapolating below the
        table. Scaling the measured power purely as T^1.5 implicitly assumes
        efficiency is CONSTANT, but the sample table shows efficiency falling
        from 0.508 mid-range to 0.441 at its lowest point — it is already
        dropping, and it keeps dropping as fixed losses take over. A pure
        power law therefore predicts efficiency rising without bound
        (40 g/W at 50 g of thrust, against a best measured 7.6 g/W), which is
        not physical: near zero thrust a motor still draws its idle power, so
        g/W must collapse toward zero, not diverge.

        On the sample table this fits to 2.7% with b = 20.5 W.
        """
        self._static_power_a = None
        self._static_power_b = None
        if (self._thrust_g_arr is None or self._power_w_arr is None
                or self._thrust_g_arr.size < 3):
            return
        thrust_N = self._thrust_g_arr * 9.80665 / 1000.0
        design = np.vstack([thrust_N ** 1.5, np.ones_like(thrust_N)]).T
        try:
            (a, b), *_ = np.linalg.lstsq(design, self._power_w_arr, rcond=None)
        except Exception:
            return
        # A negative fixed term is unphysical; fall back to pure scaling.
        if not (np.isfinite(a) and np.isfinite(b)) or a <= 0 or b < 0:
            return
        self._static_power_a = float(a)
        self._static_power_b = float(b)

    @property
    def diameter_m(self) -> float:
        return self.diameter_in * 0.0254

    @property
    def pitch_m(self) -> float:
        return self.pitch_in * 0.0254

    def disk_area(self) -> float:
        """Actuator disk area  A = π/4 · D²  [m²]"""
        return math.pi / 4.0 * self.diameter_m ** 2


# ============================================================
# AIRFRAME CONFIG  (fixed-wing specific)
# ============================================================
class AirframeConfig:
    """
    Fixed-wing aerodynamic model using a simple drag polar:

        CD = CD0 + k · CL²

    where:
        k = 1 / (π · AR · e)      (induced drag factor)
        AR = b² / S                (aspect ratio)
        e  = Oswald efficiency     (typically 0.75 – 0.90)

    Lift and drag:
        L = ½ ρ V² S CL
        D = ½ ρ V² S CD

    In straight-level flight L = W, so:
        CL = 2W / (ρ V² S)

    Stall occurs when CL = CL_max:
        V_stall = sqrt(2W / (ρ S CL_max))

    Angle of attack at cruise speed V (relative to zero-lift line):
        AoA = CL / a0  where a0 = lift-curve slope ≈ 2π per radian
        (corrected for AR via Prandtl lifting-line theory)
    """
    def __init__(self,
                 wing_span_m:    float,
                 wing_area_m2:   float,
                 CD0:            float  = 0.030,
                 CL_max:         float  = 1.30,
                 oswald:         float  = 0.80,
                 mu_roll:        float  = 0.04,
                 mu_brake:       Optional[float] = None,
                 CL_takeoff:     float  = 0.80,
                 prop_efficiency: float = 0.75,
                 prop_eff_model: str = "curve",
                 num_motors:     int    = 1):
        self.wing_span_m   = float(wing_span_m)
        self.wing_area_m2  = float(wing_area_m2)
        self.CD0           = float(CD0)
        self.CL_max        = float(CL_max)
        self.oswald        = float(oswald)
        self.mu_roll       = float(mu_roll)       # rolling friction coefficient
        # Landing rollout braking friction can be tuned independently from takeoff roll.
        self.mu_brake      = float(mu_brake) if mu_brake is not None else max(self.mu_roll * 1.8, 0.08)
        self.CL_takeoff    = float(CL_takeoff)    # CL at take-off rotation
        # PEAK combined motor+prop efficiency (see propeller_efficiency_at_speed).
        self.prop_efficiency = float(prop_efficiency)
        # "curve"    -> efficiency varies with advance ratio (default, realistic)
        # "constant" -> flat value at every speed (pre-2.4.0 behaviour)
        self.prop_eff_model = str(prop_eff_model or "curve").strip().lower()
        self.num_motors    = int(num_motors)

    @property
    def chord_m(self) -> float:
        """Mean aerodynamic chord  c = S / b  [m]"""
        return self.wing_area_m2 / max(self.wing_span_m, 1e-6)

    @property
    def aspect_ratio(self) -> float:
        """AR = b² / S"""
        return self.wing_span_m ** 2 / max(self.wing_area_m2, 1e-6)

    @property
    def k(self) -> float:
        """Induced drag factor  k = 1/(π·AR·e)"""
        return 1.0 / (math.pi * self.aspect_ratio * max(self.oswald, 1e-6))

    def lift_curve_slope(self) -> float:
        """
        Finite-wing lift-curve slope using Prandtl lifting-line theory:
            a = a0 / (1 + a0 / (π · AR))
        where a0 = 2π [1/rad] (thin-aerofoil theory).
        Returns slope in [1/rad].
        """
        a0 = 2.0 * math.pi
        return a0 / (1.0 + a0 / (math.pi * self.aspect_ratio))

    def reynolds_number(self, speed_mps: float, rho: float) -> float:
        """
        Re = ρ · V · c / μ
        Uses chord as characteristic length.
        """
        return rho * speed_mps * self.chord_m / MU_AIR

    def cl_at_speed(self, weight_N: float, speed_mps: float, rho: float) -> float:
        """
        Lift coefficient in straight-level flight (L = W):
            CL = 2W / (ρ V² S)
        """
        q = 0.5 * rho * speed_mps ** 2
        return weight_N / (q * self.wing_area_m2) if (q * self.wing_area_m2) > 0 else 0.0

    def cd_at_cl(self, CL: float) -> float:
        """Drag polar:  CD = CD0 + k · CL²"""
        return self.CD0 + self.k * CL ** 2

    def ld_ratio(self, CL: float) -> float:
        """Lift-to-drag ratio = CL / CD"""
        cd = self.cd_at_cl(CL)
        return CL / cd if cd > 0 else 0.0

    def aoa_deg(self, CL: float) -> float:
        """
        Angle of attack above zero-lift line [degrees]:
            α = CL / a   where a = finite-wing lift-curve slope [1/rad]
        """
        a = self.lift_curve_slope()
        return math.degrees(CL / a) if a > 0 else 0.0


# ============================================================
# FIXED-WING AIRCRAFT CONFIG
# ============================================================
class FixedWingConfig:
    """
    Top-level configuration object passed to all physics functions.
    Combines airframe, propulsion, and battery specs.
    """
    def __init__(self,
                 airframe:        AirframeConfig,
                 battery:         BatteryConfig,
                 motor:           MotorConfig,
                 propeller:       PropellerConfig,
                 aircraft_weight_g: float,
                 cruise_speed_mps:  float,
                 payload_mass_g:   float = 0.0,
                 periph_current_A:  float   = 0.0,
                 esc:             Optional[ESCConfig]      = None,
                 avionics:        Optional[AvionicsConfig] = None,
                 air_density:     float     = RHO0,
                 reference_altitude_m: float = 0.0,
                 cruise_altitude_m: Optional[float] = None):
        self.airframe          = airframe
        self.battery           = battery
        self.motor             = motor
        self.propeller         = propeller
        self.base_aircraft_weight_g = float(aircraft_weight_g)
        self.payload_mass_g    = max(float(payload_mass_g), 0.0)
        self.aircraft_weight_g = self.base_aircraft_weight_g + self.payload_mass_g
        self.cruise_speed_mps  = float(cruise_speed_mps)
        self.periph_current_A  = float(periph_current_A)
        self.esc               = esc
        self.avionics          = avionics
        self.air_density       = float(air_density)
        # Two DIFFERENT altitudes, previously conflated into one field:
        #   reference_altitude_m -> the airfield elevation. Used to express the
        #                           service ceiling as height above the ground.
        #   cruise_altitude_m    -> the height actually flown at. Used for the
        #                           glide-distance estimate.
        # Leaving cruise altitude unset falls back to the field elevation,
        # which is the old behaviour; it made glide distance read 0 m whenever
        # the field elevation was 0, which looked like a broken output.
        self.reference_altitude_m = max(float(reference_altitude_m), 0.0)
        self.cruise_altitude_m = (None if cruise_altitude_m is None
                                  else max(float(cruise_altitude_m), 0.0))
        # Alias for shared code
        self.num_motors        = airframe.num_motors

    @property
    def glide_reference_altitude_m(self) -> float:
        """Height used for glide distance: cruise altitude if set, else field elevation."""
        if self.cruise_altitude_m is not None:
            return float(self.cruise_altitude_m)
        return float(self.reference_altitude_m)

    @property
    def weight_N(self) -> float:
        """Convert total aircraft weight from grams to Newtons."""
        return (self.aircraft_weight_g / 1000.0) * G0


# ============================================================
# PROPULSION PHYSICS
# ============================================================
def rpm_from_kv_and_throttle(motor: MotorConfig, v_pack: float, throttle: float = 1.0) -> float:
    """
    Approximate no-load RPM at a given throttle fraction:
        RPM_no_load = KV × V_pack × throttle
    """
    if motor.kv is None:
        return 0.0
    return motor.kv * v_pack * max(0.0, min(1.0, throttle))


def prop_thrust_from_rpm(propeller: PropellerConfig, rpm: float, rho: float) -> float:
    """
    Simple momentum-disk thrust estimate from RPM:

        T = CT · ρ · n² · D⁴

    where n = RPM/60 [rev/s] and CT is estimated from pitch ratio.

    Approximation for a standard fixed-pitch prop:
        CT ≈ 0.12 × (pitch/diameter) × (1 – advance_ratio)
    clamped above zero.

    If TConst is supplied, CT is taken directly from it.
    """
    if rpm <= 0:
        return 0.0
    n  = rpm / 60.0          # rev/s
    D  = propeller.diameter_m
    D4 = D ** 4

    if propeller.TConst is not None:
        CT = float(propeller.TConst)
    else:
        # Empirical estimate from pitch/diameter ratio
        ratio = propeller.pitch_m / max(D, 1e-6)
        CT = 0.10 * ratio            # rough approximation

    T = CT * rho * (n ** 2) * D4
    return max(T, 0.0)


def prop_power_from_rpm(propeller: PropellerConfig, rpm: float, rho: float) -> float:
    """
    Propeller shaft power from RPM using power coefficient:

        P_shaft = CP · ρ · n³ · D⁵

    CP is either user-supplied (PConst) or estimated from CT via the
    figure-of-merit relation:  CP ≈ CT × π × pitch / D  (dimensional)
    """
    if rpm <= 0:
        return 0.0
    n  = rpm / 60.0
    D  = propeller.diameter_m

    if propeller.PConst is not None:
        CP = float(propeller.PConst)
    else:
        ratio = propeller.pitch_m / max(D, 1e-6)
        # Approximate: CP ≈ 0.045 × (pitch/diameter) roughly ties with CT model
        CP = 0.045 * ratio

    P = CP * rho * (n ** 3) * D ** 5
    return max(P, 0.0)


def _table_power_for_thrust(df, thrust_N: float, prop_cache=None) -> Optional[float]:
    """
    Electrical power for ONE motor at a given thrust, from a bench table.

    Returns None when the table cannot answer. Values below the measured range
    are extrapolated from a fitted curve; values above it are clamped to the
    last measured point rather than extrapolated, because a motor past its
    tested limit is not something to guess about.
    """
    if df is None or "Thrust_g" not in df or "Power_W" not in df:
        return None
    thrust_g = max(float(thrust_N), 0.0) * 1000.0 / G0
    lo = float(df["Thrust_g"].iloc[0])
    hi = float(df["Thrust_g"].iloc[-1])

    if thrust_g < lo:
        # BELOW the measured range. A polynomial fitted to the measured band
        # and extrapolated downward is unsafe: on a table spanning
        # 1426-6733 g it crosses zero near 250 g and goes negative, which
        # showed up as an operating point at -12.7 W and an implied 53 g/W.
        #
        # Momentum theory gives the right shape instead. Static ideal power is
        #     P = T * sqrt(T / (2*rho*A))    ->    P proportional to T^1.5
        # so anchor on the lowest MEASURED point and scale from there. That is
        # monotonic, always positive, and exactly reproduces the measurement
        # at the edge of the table.
        a = getattr(prop_cache, "_static_power_a", None) if prop_cache else None
        b = getattr(prop_cache, "_static_power_b", None) if prop_cache else None
        if a is not None and b is not None:
            # P = a*T^1.5 + b, so power tends to the idle draw b rather than to
            # zero, and efficiency correctly collapses near zero thrust.
            return max(a * (max(float(thrust_N), 0.0) ** 1.5) + b, 1e-6)

        # No usable fit (too few rows, or a negative fixed term): fall back to
        # pure momentum scaling anchored on the lowest measured point.
        p_lo = float(df["Power_W"].iloc[0])
        if lo > 0 and p_lo > 0:
            return max(p_lo * (thrust_g / lo) ** 1.5, 1e-6)
        return p_lo
    if thrust_g >= hi:
        return float(df["Power_W"].iloc[-1])
    return float(_interp1d(df["Thrust_g"].values, df["Power_W"].values, thrust_g))


def motor_shaft_power_from_thrust(config: FixedWingConfig, thrust_N: float,
                                  airspeed_mps: float = 0.0) -> float:
    """
    Electrical input power required to produce `thrust_N` at `airspeed_mps`,
    using FORWARD-FLIGHT actuator-disk (momentum) theory.

    Why airspeed matters
    --------------------
    A propeller moving through the air at speed V does useful work at the
    rate  T·V.  The momentum-theory induced velocity in forward flight is
    the positive root of  vi² + V·vi − T/(2ρA) = 0 :

        vi = −V/2 + sqrt( (V/2)² + T / (2·ρ·A) )

    and the ideal shaft power is

        P_shaft = T · (V + vi)

    At V = 0 this reduces to the familiar static/hover form
    P = T·sqrt(T/(2ρA)), so hover and take-off still behave correctly.

    Using the STATIC form at cruise is a serious error for a fixed-wing
    aircraft: at V = 18 m/s it returns roughly T·3.7 instead of T·18.7,
    understating cruise power by about 5x and overstating endurance by
    the same factor.

    Step 3: divide by the combined motor+prop efficiency to get the
    electrical input:

        P_electrical = P_shaft / (η_motor × η_prop)

    If a measured prop table is available we interpolate it instead, since
    the table already embeds the real efficiency at each operating point.
    """
    prop = config.propeller
    rho  = config.air_density
    V    = max(float(airspeed_mps), 0.0)

    # `thrust_N` is the TOTAL thrust the aircraft needs. Split it across the
    # motors before touching any single-propeller quantity: momentum theory
    # and a measured table both describe ONE propeller.
    #
    # Getting this wrong is not a rounding error. Feeding total thrust into a
    # single disc makes the induced velocity far too high, and it made motor
    # count have no effect whatsoever — a twin reported exactly the same power
    # as a single, when spreading the same thrust over two discs should
    # measurably reduce induced power.
    n_motors = max(int(getattr(config, "num_motors", 1) or 1), 1)
    thrust_per_motor_N = max(float(thrust_N), 0.0) / n_motors

    if prop.table is not None:
        # ------------------------------------------------------------------
        # A bench table is STATIC data (V = 0). Reading it directly is right
        # for a static case, and badly wrong in cruise: at V = 0 the ideal
        # power is T*sqrt(T/2rhoA), while at speed it is T*(V + vi) — for a
        # 6.4 N thrust at 22 m/s on an 18 in prop that is 5.7x more power.
        # Using the static number in cruise overstated endurance ~3.8x.
        #
        # What the table DOES give reliably is a measured combined
        # motor+propeller efficiency at each thrust:
        #
        #     eta(T) = P_ideal_static(T) / P_table(T)
        #
        # So in forward flight we take the efficiency from the measurement and
        # apply it to the correct forward-flight ideal power. That keeps the
        # value of the test data — a real efficiency instead of a guessed
        # one — without pretending a static test describes cruise.
        #
        # The multicopter can read its table directly because a multirotor
        # hovers (genuinely static) and in forward flight the freestream is
        # nearly edgewise to the disc. A fixed-wing propeller meets the
        # freestream head-on, so the correction is unavoidable here.
        # ------------------------------------------------------------------
        if V > 0.5:
            A = prop.disk_area()
            static_power_W = _table_power_for_thrust(prop.table, thrust_per_motor_N, prop)
            if static_power_W is not None and static_power_W > 0.0:
                vh = math.sqrt(max(thrust_per_motor_N, 0.0) /
                               max(2.0 * rho * A, 1e-9))
                ideal_static = thrust_per_motor_N * vh
                eta_measured = ideal_static / static_power_W
                # Clamp to a physically sensible band: a real motor+prop
                # combination is not below 15% or above 90% efficient.
                eta_measured = min(max(eta_measured, 0.15), 0.90)
                vi = _induced_velocity_forward(thrust_per_motor_N, V, rho, A)
                return thrust_per_motor_N * (V + vi) / eta_measured * n_motors

        # --- Static / very low speed: the table applies directly ------------
        thrust_g = thrust_per_motor_N * 1000.0 / G0   # per motor, as the table is
        df       = prop.table
        min_thrust = float(df["Thrust_g"].iloc[0])
        max_thrust = float(df["Thrust_g"].iloc[-1])
        
        # Below minimum: try to extrapolate
        if thrust_g < min_thrust:
            extrapolated = _extrapolate_motor_value_fw(df, thrust_g, "Power_W")
            if extrapolated is not None:
                return float(extrapolated) * n_motors
            else:
                return float(df["Power_W"].iloc[0]) * n_motors
        
        # Above maximum: use last value
        if thrust_g >= max_thrust:
            return float(df["Power_W"].iloc[-1]) * n_motors
        
        # Within range: linear interpolation
        return float(_interp1d(df["Thrust_g"].values,
                              df["Power_W"].values, thrust_g)) * n_motors

    if config.motor.kv is not None:
        # --- KV-based electrical model (forward-flight momentum theory) ---
        A      = prop.disk_area()                 # ONE propeller
        vi     = _induced_velocity_forward(thrust_per_motor_N, V, rho, A)
        P_shaft = thrust_per_motor_N * (V + vi) * n_motors   # all motors [W]
        # Combined motor + propeller efficiency, evaluated at THIS airspeed:
        # a prop far from its design advance ratio is markedly less efficient.
        eta_combined = max(propeller_efficiency_at_speed(config, V), 0.10)
        P_elec = P_shaft / eta_combined            # electrical input [W]
        return max(P_elec, 0.0)

    # --- Theoretical fallback (no table, no KV) ---
    A       = prop.disk_area()                    # ONE propeller
    vi      = _induced_velocity_forward(thrust_per_motor_N, V, rho, A)
    P_shaft = thrust_per_motor_N * (V + vi) * n_motors
    return P_shaft / max(propeller_efficiency_at_speed(config, V), 0.10)


def propeller_efficiency_at_speed(config: "FixedWingConfig",
                                  airspeed_mps: float) -> float:
    """
    Propeller (plus motor) efficiency at a given airspeed.

    A propeller is only efficient near its design operating point.  Efficiency
    is governed by the advance ratio  J = V / (n*D), which is most usefully
    normalised by the pitch/diameter ratio:

        x = J / (pitch/D) = V / V_pitch

    where V_pitch is the speed at which the blade's geometric pitch equals the
    distance travelled per revolution.  Physically:

        x -> 0   the blade is stalled, thrust is high but efficiency is poor
                 (this is the static / take-off condition)
        x ~ 0.6  near the design point, efficiency peaks
        x -> 1   the blade reaches zero angle of attack, thrust and efficiency
                 both collapse

    A parabola through zero at x = 0 and x = 1, peaking at x_peak, captures
    that shape well enough for sizing work:

        eta(x) = eta_peak * [ x(1-x) ] / [ x_peak(1-x_peak) ]

    `AirframeConfig.prop_efficiency` is treated as the PEAK efficiency, so a
    correctly matched prop at its design speed gets exactly the value the user
    entered, and mismatched operating points are penalised.

    V_pitch is evaluated at the full-throttle RPM implied by Kv and pack
    voltage, which is the same reference the "pitch speed vs cruise" status
    check uses.  This avoids a circular dependency (efficiency -> power ->
    throttle -> RPM -> efficiency).

    Set airframe.prop_eff_model = "constant" to disable the curve and use the
    flat value, which is how versions before 2.4.0 behaved.
    """
    eta_peak = max(float(config.airframe.prop_efficiency), 0.10)
    model = str(getattr(config.airframe, "prop_eff_model", "curve")).strip().lower()
    if model in ("constant", "flat", "fixed"):
        return eta_peak

    kv = getattr(config.motor, "kv", None)
    pitch_m = float(getattr(config.propeller, "pitch_m", 0.0) or 0.0)
    if not kv or kv <= 0 or pitch_m <= 0:
        return eta_peak                      # not enough data to place the curve

    # Full-throttle pitch speed [m/s]
    n_max = (float(kv) * float(config.battery.vmax_pack)) / 60.0     # rev/s
    v_pitch = pitch_m * n_max
    if v_pitch <= 1e-6:
        return eta_peak

    x = max(float(airspeed_mps), 0.0) / v_pitch

    X_PEAK = 0.60                            # efficiency peaks near 60% of pitch speed
    shape = (x * (1.0 - x)) / (X_PEAK * (1.0 - X_PEAK))

    # Floor: a real prop still does useful work when static (take-off) and
    # slightly past pitch speed (windmilling edge), so never return ~0, which
    # would make required power blow up.
    eta = eta_peak * max(shape, 0.0)
    return min(max(eta, 0.25 * eta_peak), eta_peak)


def _induced_velocity_forward(thrust_N: float, airspeed_mps: float,
                              rho: float, disk_area_m2: float) -> float:
    """
    Propeller induced velocity in forward flight, from momentum theory.

    Solves  vi² + V·vi − T/(2·ρ·A) = 0  for the positive root:

        vi = −V/2 + sqrt( (V/2)² + T/(2·ρ·A) )

    At V = 0 this collapses to the static case  vi = sqrt(T/(2ρA)).
    As V grows, vi falls toward T/(2ρAV) — the propeller works against a
    faster-moving stream, so less velocity has to be added to it.
    """
    T = max(float(thrust_N), 0.0)
    V = max(float(airspeed_mps), 0.0)
    A = max(float(disk_area_m2), 1e-9)
    vh_sq = T / (2.0 * max(rho, 1e-9) * A)     # = vi² at hover
    return -0.5 * V + math.sqrt((0.5 * V) ** 2 + vh_sq)


def _interp1d(x: "np.ndarray", y: "np.ndarray", xq: float) -> float:
    """Simple numpy-free linear interpolation on sorted x."""
    import bisect
    idx = bisect.bisect_left(x.tolist(), xq)
    idx = max(1, min(idx, len(x) - 1))
    x0, x1 = x[idx-1], x[idx]
    y0, y1 = y[idx-1], y[idx]
    t = (xq - x0) / (x1 - x0) if x1 != x0 else 0.0
    return y0 + t * (y1 - y0)


def max_thrust_N(config: FixedWingConfig) -> float:
    """
    Maximum static thrust available from the motor/prop system.

    For KV-based model at full throttle (advance ratio = 0, static):
        T_max = prop_thrust_from_rpm(RPM_max, rho)
    where RPM_max = KV × V_nom_pack (conservative — full throttle at nominal V).

    Clamped by motor max power:
        T_limited = min(T_max, P_max / V_pitch_speed)
    """
    motor = config.motor
    prop  = config.propeller
    rho   = config.air_density
    batt  = config.battery

    if prop.table is not None:
        # Read peak thrust from table
        cached = getattr(prop, "_thrust_g_max", None)
        max_g = cached if cached is not None else float(prop.table["Thrust_g"].max())
        return float(max_g) * G0 / 1000.0 * config.num_motors

    if motor.kv is not None:
        rpm_max = motor.kv * batt.vnom_pack
        T_ideal = prop_thrust_from_rpm(prop, rpm_max, rho)

        # Limit by motor max power:
        #   T = P_shaft / v_pitch_speed  (at full throttle pitch speed)
        n_max        = rpm_max / 60.0
        v_pitch      = prop.pitch_m * n_max      # pitch speed [m/s]
        if v_pitch > 1.0:
            T_pmax = motor.max_power / v_pitch
            T_ideal = min(T_ideal, T_pmax)

        return T_ideal * config.num_motors

    # Fallback: actuator-disk at max power
    A     = prop.disk_area()
    # T_max from P_max = T·vi with vi = sqrt(T/2ρA):  T^(3/2) = P·sqrt(2ρA)
    # → T_max = (P_max · sqrt(2ρA))^(2/3)
    factor = (motor.max_power * math.sqrt(2 * rho * A))
    return (factor ** (2.0/3.0)) * config.num_motors


def tip_speed_mps(config: FixedWingConfig, rpm: Optional[float] = None) -> float:
    """
    Propeller tip speed:
        V_tip = π · D · n   where n = RPM/60 [rev/s]
    """
    if rpm is None:
        rpm = config.motor.kv * config.battery.vnom_pack if config.motor.kv else 0.0
    n = rpm / 60.0
    return math.pi * config.propeller.diameter_m * n


def pitch_speed_mps(config: FixedWingConfig, rpm: Optional[float] = None) -> float:
    """
    Pitch speed (speed at advance ratio J = 1):
        V_pitch = pitch_m × n   where n = RPM/60
    This is the airspeed at which the blade sees zero angle of attack
    (theoretical no-thrust speed for an ideal propeller).
    """
    if rpm is None:
        rpm = config.motor.kv * config.battery.vnom_pack if config.motor.kv else 0.0
    n = rpm / 60.0
    return config.propeller.pitch_m * n


def esc_losses_W(config: FixedWingConfig, v_pack: float, motor_P_total_W: float) -> Tuple[float, str]:
    """
    ESC conduction and idle losses:
        P_esc_cond = I_motor² · R_esc
        P_esc_idle = I_idle  · V_pack
    Returns (total_esc_loss_W, status_note).
    """
    esc = config.esc
    if esc is None:
        return 0.0, ""
    v        = max(v_pack, 1.0)
    p_pm     = motor_P_total_W / max(config.num_motors, 1)
    i_motor  = p_pm / v
    p_cond   = i_motor ** 2 * max(esc.resistance, 0.0)
    p_idle   = max(esc.idle_current_A, 0.0) * v
    total    = (p_cond + p_idle) * config.num_motors
    note     = ""
    if i_motor > esc.continuous_rating_A:
        note = f"ESC over cont current ({i_motor:.1f} A > {esc.continuous_rating_A:.1f} A)"
    return total, note


# ============================================================
# AERODYNAMIC PERFORMANCE FUNCTIONS
# ============================================================
def stall_speed(config: FixedWingConfig) -> float:
    """
    Minimum flying speed (stall speed) in straight-level flight:

        V_stall = sqrt(2 · W / (ρ · S · CL_max))

    At V < V_stall the wing cannot generate enough lift.
    """
    rho = config.air_density
    W   = config.weight_N
    S   = config.airframe.wing_area_m2
    CL_max = config.airframe.CL_max
    denom = rho * S * CL_max
    return math.sqrt(2.0 * W / denom) if denom > 0 else 0.0


def drag_N(config: FixedWingConfig, speed_mps: float) -> float:
    """
    Total aerodynamic drag in straight-level flight:

        L = W  →  CL = 2W / (ρ V² S)
        CD = CD0 + k · CL²
        D  = ½ ρ V² S · CD

    The drag equals the thrust required for level flight.
    """
    _, _, total_drag = drag_components_N(config, speed_mps)
    return total_drag


def wing_loading_N_m2(config: FixedWingConfig) -> float:
    """Wing loading W/S in N/m²."""
    S = config.airframe.wing_area_m2
    return config.weight_N / S if S > 0 else 0.0


def wing_loading_kg_m2(config: FixedWingConfig) -> float:
    """Wing loading in kg/m² (mass per wing area)."""
    S = config.airframe.wing_area_m2
    return (config.aircraft_weight_g / 1000.0) / S if S > 0 else 0.0


def drag_components_N(config: FixedWingConfig,
                      speed_mps: float,
                      load_factor: float = 1.0) -> Tuple[float, float, float]:
    """
    Split level-flight drag into induced and parasitic components.
    Returns (induced_drag_N, parasitic_drag_N, total_drag_N).
    """
    af  = config.airframe
    rho = config.air_density
    W   = config.weight_N
    V   = max(speed_mps, 0.1)
    q   = 0.5 * rho * V ** 2
    S   = af.wing_area_m2
    n   = max(float(load_factor), 1.0)
    CL  = af.cl_at_speed(W * n, V, rho)

    cd_parasitic = af.CD0
    cd_induced   = af.k * CL ** 2

    d_parasitic = q * S * cd_parasitic
    d_induced   = q * S * cd_induced
    return d_induced, d_parasitic, d_induced + d_parasitic


def bank_load_factor(bank_deg: float) -> float:
    """Load factor in a coordinated turn: n = 1 / cos(phi)."""
    phi = math.radians(max(min(float(bank_deg), 85.0), -85.0))
    c = math.cos(phi)
    if abs(c) < 1e-6:
        return float("inf")
    return 1.0 / c


def turn_radius_m(speed_mps: float, bank_deg: float) -> float:
    """Coordinated turn radius: R = V² / (g * tan(phi))."""
    phi = math.radians(abs(float(bank_deg)))
    t = math.tan(phi)
    if abs(t) < 1e-6:
        return float("inf")
    return max(speed_mps, 0.0) ** 2 / (G0 * t)


def turn_rate_deg_s(speed_mps: float, bank_deg: float) -> float:
    """Coordinated turn rate in deg/s: omega = g*tan(phi)/V."""
    V = max(float(speed_mps), 0.1)
    phi = math.radians(float(bank_deg))
    return math.degrees(G0 * math.tan(phi) / V)


def glide_sink_rate_mps(config: FixedWingConfig, speed_mps: float) -> float:
    """
    Unpowered sink rate approximation:
        sink = V * D/L = V / (L/D)
    """
    af  = config.airframe
    rho = config.air_density
    W   = config.weight_N
    V   = max(speed_mps, 0.1)
    CL  = af.cl_at_speed(W, V, rho)
    LD  = af.ld_ratio(CL)
    return V / LD if LD > 0 else float("inf")


def minimum_sink_speed(config: FixedWingConfig,
                       v_stall: float,
                       v_max: float = 80.0,
                       steps: int = 500) -> Tuple[float, float]:
    """
    Minimum sink speed for glide, numerically searched.
    Returns (V_min_sink [m/s], sink_rate_min [m/s]).
    """
    v_lo = max(v_stall * 1.05, 1.0)
    v_hi = min(v_max, v_lo + 40.0)
    best_v = v_lo
    best_sink = float("inf")
    for i in range(steps + 1):
        V = v_lo + (v_hi - v_lo) * i / steps
        sink = glide_sink_rate_mps(config, V)
        if sink < best_sink:
            best_sink = sink
            best_v = V
    if not math.isfinite(best_sink):
        return best_v, 0.0
    return best_v, best_sink


def _max_rate_of_climb_at_altitude_m(config: FixedWingConfig,
                                     altitude_m: float,
                                     v_max: float = 80.0) -> float:
    """Maximum ROC at a given ISA altitude (helper for service ceiling)."""
    rho_prev = config.air_density
    try:
        config.air_density = isa_density(altitude_m)
        v_min = max(stall_speed(config) * 1.05, 1.0)
        _, rc = max_rate_of_climb_mps(config, v_min=v_min, v_max=v_max, steps=180)
        return rc
    finally:
        config.air_density = rho_prev


def service_ceiling_m(config: FixedWingConfig,
                      rc_threshold_mps: float = 0.508,
                      max_alt_m: float = 12000.0,
                      coarse_step_m: float = 250.0) -> float:
    """
    Service ceiling (ISA altitude ASL) where max rate of climb falls to
    rc_threshold_mps (default 100 ft/min ≈ 0.508 m/s).
    Returns +inf if threshold is still exceeded at max_alt_m.
    """
    # Cache by configuration state to avoid recomputing in speed sweeps/GUI refreshes.
    cache = getattr(config, "_service_ceiling_cache", {})
    cache_key = (
        round(rc_threshold_mps, 4),
        round(max_alt_m, 1),
        round(coarse_step_m, 1),
        round(config.reference_altitude_m, 1),
        round(config.aircraft_weight_g, 2),
        round(config.airframe.wing_area_m2, 5),
        round(config.airframe.CD0, 5),
        round(config.airframe.CL_max, 5),
        round(config.airframe.oswald, 5),
        round(config.airframe.prop_efficiency, 5),
        round(float(config.motor.kv or 0.0), 3),
        round(config.motor.max_power, 2),
        round(config.propeller.diameter_in, 3),
        round(config.propeller.pitch_in, 3),
        int(config.num_motors),
        bool(config.propeller.table is not None),
    )
    if cache_key in cache:
        return cache[cache_key]

    base_alt = max(config.reference_altitude_m, 0.0)
    rc_base = _max_rate_of_climb_at_altitude_m(config, base_alt)
    if rc_base < rc_threshold_mps:
        cache[cache_key] = base_alt
        setattr(config, "_service_ceiling_cache", cache)
        return base_alt

    lo = base_alt
    hi = None
    alt = base_alt + coarse_step_m
    while alt <= max_alt_m:
        rc_here = _max_rate_of_climb_at_altitude_m(config, alt)
        if rc_here < rc_threshold_mps:
            hi = alt
            break
        lo = alt
        alt += coarse_step_m

    if hi is None:
        cache[cache_key] = float("inf")
        setattr(config, "_service_ceiling_cache", cache)
        return float("inf")

    # Refine the "RC crosses threshold" altitude with binary search.
    for _ in range(12):
        mid = 0.5 * (lo + hi)
        rc_mid = _max_rate_of_climb_at_altitude_m(config, mid)
        if rc_mid >= rc_threshold_mps:
            lo = mid
        else:
            hi = mid
    cache[cache_key] = lo
    setattr(config, "_service_ceiling_cache", cache)
    return lo


def landing_distance_m(config: FixedWingConfig,
                       obstacle_height_m: float = 15.0) -> float:
    """
    Simplified landing distance from 15 m obstacle:
      total = approach + flare + rollout
    """
    af  = config.airframe
    rho = config.air_density
    W   = config.weight_N
    S   = af.wing_area_m2
    if S <= 0:
        return float("inf")

    V_stall = stall_speed(config)
    V_app   = 1.30 * V_stall
    V_td    = 1.15 * V_stall

    CL_app  = af.cl_at_speed(W, V_app, rho)
    LD_app  = max(af.ld_ratio(CL_app), 1e-3)
    approach_dist = obstacle_height_m * LD_app

    q_td    = 0.5 * rho * V_td ** 2

    # After touchdown the pilot (or the flight controller) dumps lift —
    # stick forward, spoilers, flaps retracted — so the wing is NOT still
    # flying at its approach CL.  Modelling the ground roll at CL_takeoff
    # leaves almost no weight on the wheels (L can even exceed W), which
    # drives the braking force to ~0 and produces absurdly long rollouts.
    # Use a low ground-roll CL so the brakes actually see the aircraft weight.
    CL_ground = min(0.25, af.CL_max)
    CD_land   = af.cd_at_cl(CL_ground)
    D_land    = q_td * S * CD_land
    L_land    = q_td * S * CL_ground

    # Use the dedicated braking-friction input for landing rollout.
    mu_brake = max(float(getattr(af, "mu_brake", af.mu_roll * 1.8)), 0.08)
    F_brake  = mu_brake * max(W - L_land, 0.0)
    mass_kg  = max(W / G0, 1e-6)
    decel    = (F_brake + D_land) / mass_kg
    if decel <= 0:
        return float("inf")

    rollout = V_td ** 2 / (2.0 * decel)
    flare   = 0.5 * V_td
    return approach_dist + flare + rollout


def cruise_efficiency(speed_mps: float, total_power_W: float) -> Tuple[float, float]:
    """
    Cruise efficiency at the operating point.
    Returns (specific_range_m_per_Wh, specific_endurance_min_per_Wh).
    """
    P = max(total_power_W, 0.0)
    if P <= 0:
        return 0.0, 0.0
    specific_range_m_per_Wh = max(speed_mps, 0.0) * 3600.0 / P
    specific_endurance_min_per_Wh = 60.0 / P
    return specific_range_m_per_Wh, specific_endurance_min_per_Wh


def _extrapolate_motor_value_fw(df: pd.DataFrame, thrust_g: float, column: str) -> Optional[float]:
    """Extrapolate a motor property value using fitted curve if thrust is below minimum.
    
    Args:
        df: Propeller table dataframe
        thrust_g: Target thrust in grams
        column: Column name to extrapolate (Power_W, Current_A, etc.)
    
    Returns:
        Extrapolated value or None if not available
    """
    if column not in df.columns:
        return None
    
    thrust_g_data = df["Thrust_g"].values
    try:
        y_data = pd.to_numeric(df[column], errors='coerce').values
    except:
        return None
    
    min_thrust = thrust_g_data.min()
    
    # If we're within data range, don't extrapolate
    if thrust_g >= min_thrust:
        return None
    
    # Fit a 2nd degree polynomial to the data
    fit_result = _fit_propeller_curve(thrust_g_data, y_data, degree=2)
    if fit_result is None:
        return None
    
    coeffs, _, _ = fit_result
    extrapolated = _eval_poly(coeffs, thrust_g)
    
    # For some properties, clamp to minimum reasonable values
    if column == "Current_A" and extrapolated is not None:
        extrapolated = max(extrapolated, 0.0)
    elif column == "Power_W" and extrapolated is not None:
        extrapolated = max(extrapolated, 0.0)
    elif column == "Efficiency_gW" and extrapolated is not None:
        extrapolated = max(extrapolated, 0.0)
    elif column == "RPM" and extrapolated is not None:
        extrapolated = max(extrapolated, 0.0)
    
    return extrapolated


def power_required_W(config: FixedWingConfig, speed_mps: float) -> float:
    """
    Shaft power required for level flight = D × V  [W].
    This is the propulsive power that must be delivered to the air.
    Electrical input power is higher by (η_motor × η_prop).
    """
    return drag_N(config, speed_mps) * max(speed_mps, 0.1)


def electrical_power_required_W(config: FixedWingConfig, speed_mps: float) -> float:
    """
    Total electrical input power from battery [W]:
        P_elec = P_prop_required / η_combined + P_avionics
    where P_prop_required is the motor shaft + prop power chain input.
    """
    T_req  = drag_N(config, speed_mps)     # thrust = drag for level flight
    P_mech = motor_shaft_power_from_thrust(config, T_req, speed_mps)
    return P_mech


def max_shaft_power_W(config: FixedWingConfig) -> float:
    """Best available shaft power across all motors [W]."""
    prop = config.propeller
    n_motors = max(int(getattr(config, "num_motors", 1) or 1), 1)

    p_elec_per_motor = None
    if prop.table is not None and "Power_W" in prop.table:
        p_elec_per_motor = float(prop.table["Power_W"].max())
    elif getattr(config.motor, "max_power", None):
        p_elec_per_motor = float(config.motor.max_power)
    if not p_elec_per_motor or p_elec_per_motor <= 0:
        return 0.0

    eta = max(float(config.airframe.prop_efficiency), 0.10)
    return p_elec_per_motor * n_motors * eta


def thrust_available_N(config: FixedWingConfig,
                       airspeed_mps: float = 0.0) -> float:
    """
    Thrust available at a given airspeed [N].

    A propeller's thrust falls sharply with forward speed. Returning the
    STATIC bench figure at every airspeed produced impossible results: on a
    3 m glider the model reported a best climb of 3597 m/min at 56 m/s, using
    66 N of thrust at a speed where the propeller — pitch speed 25 m/s at its
    maximum tested RPM — would actually be windmilling. The implied climb
    power was 2469 W against a measured maximum of 1680 W, so it violated
    energy conservation outright.

    Momentum theory bounds it properly. For the available shaft power P:

        P = T * (V + vi),   vi = -V/2 + sqrt((V/2)^2 + T/(2*rho*A))

    which is solved here for T by bisection. At V = 0 this returns the static
    result; as V rises, T falls roughly as P/V. The answer is additionally
    capped by the static thrust, since forward flight cannot beat it.
    """
    static_T = max_thrust_N(config)
    V = max(float(airspeed_mps), 0.0)
    if V < 0.1:
        return static_T

    # Cache by speed. The searches evaluate the same handful of speeds
    # repeatedly, and the signature guards against reusing a result after the
    # aircraft or the air has changed.
    signature = (round(static_T, 6), round(float(config.air_density), 6),
                 int(getattr(config, "num_motors", 1) or 1),
                 round(float(config.propeller.diameter_in), 4),
                 round(float(config.airframe.prop_efficiency), 4))
    cache = getattr(config, "_thrust_avail_cache", None)
    if cache is None or cache.get("signature") != signature:
        cache = {"signature": signature}
        config._thrust_avail_cache = cache
    key = round(V, 3)
    if key in cache:
        return cache[key]

    P_shaft = max_shaft_power_W(config)
    if P_shaft <= 0.0:
        cache[key] = static_T
        return static_T

    rho = max(float(config.air_density), 1e-9)
    n_motors = max(int(getattr(config, "num_motors", 1) or 1), 1)
    A = config.propeller.disk_area()
    P_per_motor = P_shaft / n_motors

    def power_needed(thrust_per_motor: float) -> float:
        vi = _induced_velocity_forward(thrust_per_motor, V, rho, A)
        return thrust_per_motor * (V + vi)

    # Bisection: power_needed is monotonic in thrust. This runs inside the
    # climb-rate and best-speed searches, so it is called on the order of a
    # thousand times per evaluation — stop as soon as the bracket is tight
    # rather than burning a fixed 60 iterations.
    lo, hi = 0.0, max(static_T / n_motors, 1e-6)
    if power_needed(hi) <= P_per_motor:
        cache[key] = min(static_T, hi * n_motors)
        return cache[key]
    for _ in range(40):
        mid = 0.5 * (lo + hi)
        if power_needed(mid) > P_per_motor:
            hi = mid
        else:
            lo = mid
        if hi - lo < 1e-4 * hi:          # 0.01% is far finer than the inputs
            break
    cache[key] = min(static_T, lo * n_motors)
    return cache[key]


def specific_thrust(config: FixedWingConfig) -> float:
    """
    Specific thrust = T_available / W  [dimensionless]
    Also called thrust-to-weight ratio.
    """
    return thrust_available_N(config) / max(config.weight_N, 1e-6)


def available_excess_power_W(config: FixedWingConfig, speed_mps: float) -> float:
    """
    Excess propulsive power available for climbing or acceleration:
        P_excess = (T_available - D) × V   [W]
    """
    T_avail = thrust_available_N(config, speed_mps)
    D       = drag_N(config, speed_mps)
    V       = max(speed_mps, 0.1)
    return max((T_avail - D) * V, 0.0)


def rate_of_climb_mps(config: FixedWingConfig, speed_mps: float) -> float:
    """
    Rate of climb at a given airspeed using all available thrust:

        RC = (T_available - D) · V / W   [m/s]

    This is the excess thrust times velocity divided by weight (i.e., excess
    power divided by weight). Clamped to zero if T < D (no climb possible).
    """
    T_avail = thrust_available_N(config, speed_mps)
    D       = drag_N(config, speed_mps)
    W       = config.weight_N
    V       = max(speed_mps, 0.1)
    return max((T_avail - D) * V / W, 0.0)


def max_rate_of_climb_mps(config: FixedWingConfig,
                           v_min: float = 1.0,
                           v_max: float = 80.0,
                           steps: int   = 500) -> Tuple[float, float]:
    """
    Scan airspeed range to find maximum rate of climb.
    Returns (V_at_max_RC [m/s], max_RC [m/s]).
    """
    best_v  = v_min
    best_rc = 0.0
    for i in range(steps + 1):
        V  = v_min + (v_max - v_min) * i / steps
        rc = rate_of_climb_mps(config, V)
        if rc > best_rc:
            best_rc = rc
            best_v  = V
    return best_v, best_rc


def max_angle_of_climb_deg(config: FixedWingConfig, speed_mps: float) -> float:
    """
    Maximum climb angle at a given speed:

        γ_max = arcsin((T - D) / W)   [degrees]

    This is limited to 90° (vertical) and clamped at 0° if T ≤ D.
    """
    T = thrust_available_N(config, speed_mps)
    D = drag_N(config, speed_mps)
    W = max(config.weight_N, 1e-6)
    ratio = (T - D) / W
    ratio = max(-1.0, min(1.0, ratio))
    return math.degrees(math.asin(ratio)) if ratio > 0 else 0.0


def best_angle_of_climb_speed(config: FixedWingConfig,
                               v_min: float = 1.0,
                               v_max: float = 80.0,
                               steps: int   = 500) -> Tuple[float, float]:
    """
    Best angle of climb speed (Vx): maximizes γ = arcsin((T-D)/W).
    Returns (Vx [m/s], γ_max [deg]).
    """
    best_v = v_min
    best_g = 0.0
    for i in range(steps + 1):
        V = v_min + (v_max - v_min) * i / steps
        g = max_angle_of_climb_deg(config, V)
        if g > best_g:
            best_g = g
            best_v = V
    return best_v, best_g


def best_endurance_speed(config: FixedWingConfig,
                          v_stall: float,
                          v_max:   float = 80.0,
                          steps:   int   = 500) -> Tuple[float, float]:
    """
    Best endurance speed for a propeller aircraft:
        Minimise power required  →  Maximise CL^1.5 / CD

    The analytic optimum (from drag polar):
        V_be = (2W/ρS)^0.5 · (k/(3·CD0))^0.25

    Returns (V_be [m/s], P_min [W]).
    """
    af   = config.airframe
    rho  = config.air_density
    W    = config.weight_N
    S    = af.wing_area_m2

    # Analytic guess
    k       = af.k
    CD0     = af.CD0
    V_be_analytic = ((2 * W / (rho * S)) ** 0.5
                     * (k / (3 * CD0)) ** 0.25) if CD0 > 0 else v_stall * 1.3

    # Numerical scan to confirm (also respects stall speed floor)
    v_lo  = max(v_stall * 1.0, 1.0)
    v_hi  = min(v_max, v_lo + 40.0)
    best_v = V_be_analytic
    best_P = 1e12
    for i in range(steps + 1):
        V = v_lo + (v_hi - v_lo) * i / steps
        P = power_required_W(config, V)
        if P < best_P:
            best_P = P
            best_v = V
    return best_v, best_P


def best_range_speed(config: FixedWingConfig,
                     v_stall: float,
                     v_max:   float = 80.0,
                     steps:   int   = 500) -> Tuple[float, float]:
    """
    Best range speed for a propeller aircraft:
        Maximise L/D ratio  →  Minimise CD/CL

    Analytic optimum:
        V_br = (2W/ρS)^0.5 · (k/CD0)^0.25

    Returns (V_br [m/s], L/D at V_br).
    """
    af  = config.airframe
    rho = config.air_density
    W   = config.weight_N
    S   = af.wing_area_m2

    k   = af.k
    CD0 = af.CD0
    V_br_analytic = ((2 * W / (rho * S)) ** 0.5
                     * (k / CD0) ** 0.25) if CD0 > 0 else v_stall * 1.5

    v_lo = max(v_stall * 1.0, 1.0)
    v_hi = min(v_max, v_lo + 40.0)
    best_v = V_br_analytic
    best_ld = 0.0
    for i in range(steps + 1):
        V  = v_lo + (v_hi - v_lo) * i / steps
        CL = af.cl_at_speed(W, V, rho)
        ld = af.ld_ratio(CL)
        if ld > best_ld:
            best_ld = ld
            best_v  = V
    return best_v, best_ld


def takeoff_distance_m(config: FixedWingConfig) -> float:
    """
    Ground roll to lift-off (simplified Raymer method):

        s_g = 1.44 · W² / (g · ρ · S · CL_TO · (T_avg - μr · W))

    where:
        CL_TO = CL at takeoff rotation (user-supplied, typically 0.6–1.0)
        T_avg = average thrust during roll ≈ T_static × 0.75
                (thrust decreases as speed builds up)
        μr    = rolling friction coefficient (paved: 0.02–0.05)

    If the net force (T - μr·W) ≤ 0, no takeoff is possible (∞ distance).
    """
    af    = config.airframe
    rho   = config.air_density
    W     = config.weight_N
    # Thrust decays with speed during the roll. The standard treatment
    # evaluates thrust once at 0.707 x lift-off speed, which is the speed at
    # which the instantaneous thrust equals the mean value over a roll with
    # V^2-proportional drag. That is more defensible than either a flat 75%
    # of static (the old assumption) or averaging the two endpoints.
    T_static = thrust_available_N(config, 0.0)
    _v_liftoff = 1.2 * stall_speed(config)
    T_max = T_static
    T_avg = thrust_available_N(config, 0.707 * _v_liftoff)
    mu_r  = af.mu_roll
    CL_to = max(af.CL_takeoff, 1e-3)
    S     = af.wing_area_m2

    net_force = T_avg - mu_r * W
    if net_force <= 0:
        return float("inf")
    return 1.44 * W ** 2 / (G0 * rho * S * CL_to * net_force)


# ============================================================
# FULL OPERATING METRICS AT A GIVEN SPEED
# ============================================================
def compute_metrics(config: FixedWingConfig,
                    speed_mps: float,
                    bank_deg: float = 0.0,
                    ambient_temp_C: float = 25.0,
                    wind_head_mps: float = 0.0,
                    wind_cross_mps: float = 0.0,
                    glide_altitude_m: Optional[float] = None) -> dict:
    """
    Compute all performance metrics at the given cruise airspeed.
    Returns a dict matching the eCalc-style output columns.
    """
    rho  = config.air_density
    W    = config.weight_N
    af   = config.airframe
    batt = config.battery
    motor= config.motor

    # Coordinated turn load factor raises required lift and therefore CL/CD and stall speed.
    n_turn  = bank_load_factor(bank_deg)
    V_stall = stall_speed(config)
    V_stall_turn = V_stall * math.sqrt(max(n_turn, 1.0))
    V       = max(speed_mps, V_stall_turn + 0.01)
    CL      = af.cl_at_speed(W * n_turn, V, rho)
    CD      = af.cd_at_cl(CL)
    LD      = af.ld_ratio(CL)
    d_induced, d_parasitic, D = drag_components_N(config, V, load_factor=n_turn)
    P_prop  = D * V                                # shaft / propulsive power  [W]
    T_req   = D                                    # thrust required [N]
    T_avail = thrust_available_N(config, V)
    P_elec  = motor_shaft_power_from_thrust(config, T_req, V)  # motor electrical [W]

    # Avionics / peripheral power
    P_avionics = avionics_input_power_W(config.avionics)
    if P_avionics <= 0.0:
        P_avionics = batt.vnom_pack * max(config.periph_current_A, 0.0)

    # ESC losses
    esc_loss, esc_note = esc_losses_W(config, batt.vmax_pack, P_elec)

    P_total = P_elec + esc_loss + P_avionics
    pack_I  = P_total / max(batt.vnom_pack, 1.0)
    V_load  = batt.voltage_under_load(pack_I)

    # RPM estimate
    rpm_est = motor.kv * V_load if motor.kv else 0.0

    # Performance metrics
    V_be, P_be = best_endurance_speed(config, V_stall)
    V_br, LD_br = best_range_speed(config, V_stall)
    v_rc_max, rc_max = max_rate_of_climb_mps(config)
    rc_at_V         = rate_of_climb_mps(config, V)
    v_gamma, gamma  = best_angle_of_climb_speed(config)
    S_to            = takeoff_distance_m(config)
    S_ld            = landing_distance_m(config)
    aoa             = af.aoa_deg(CL)
    Re              = af.reynolds_number(V, rho)
    V_tip           = tip_speed_mps(config, rpm_est)
    V_pitch         = pitch_speed_mps(config, rpm_est)
    spec_thrust     = specific_thrust(config)
    max_prop_P      = motor.max_power * config.num_motors
    wing_load_N_m2  = wing_loading_N_m2(config)
    wing_load_kg_m2 = wing_loading_kg_m2(config)
    V_min_sink, sink_min = minimum_sink_speed(config, V_stall)

    # Flight time/range at this operating point.
    # Ground distance uses along-track groundspeed (airspeed corrected for wind vector).
    if P_total > 0 and pack_I <= batt.discharge_max_A and V_load >= batt.vmin_pack:
        t_min = (batt.usable_Wh / P_total) * 60.0
        gs_track = groundspeed_along_track_mps(V, wind_head_mps, wind_cross_mps)
        d_km  = gs_track * (t_min * 60.0) / 1000.0
    else:
        t_min = 0.0
        d_km  = 0.0
        gs_track = groundspeed_along_track_mps(V, wind_head_mps, wind_cross_mps)

    specific_range_m_Wh, specific_endurance_min_Wh = cruise_efficiency(V, P_total)
    service_ceiling_abs_m = service_ceiling_m(config)
    if math.isfinite(service_ceiling_abs_m):
        service_ceiling_agl_m = max(service_ceiling_abs_m - config.reference_altitude_m, 0.0)
    else:
        service_ceiling_agl_m = float("inf")
    glide_ratio = LD
    # Glide distance is measured from the CRUISE altitude, not the field elevation.
    glide_alt = max(float(config.glide_reference_altitude_m
                          if glide_altitude_m is None else glide_altitude_m), 0.0)
    glide_dist_m = glide_ratio * glide_alt

    # Turning-flight metrics
    turn_r_m = turn_radius_m(V, bank_deg)
    turn_rate_dps = turn_rate_deg_s(V, bank_deg)
    turn_period_s = (360.0 / abs(turn_rate_dps)) if abs(turn_rate_dps) > 1e-9 else float("inf")
    loiter_circles = ((t_min * 60.0) / turn_period_s) if (turn_period_s > 0 and math.isfinite(turn_period_s)) else 0.0

    # Thermal model (single-point steady estimate, not a transient RC network).
    motor_i_per_motor = P_elec / max(V_load, 1.0) / max(config.num_motors, 1)
    motor_copper_loss_W = (motor_i_per_motor ** 2) * max(config.motor.resistance, 0.0) * max(config.num_motors, 1)
    battery_loss_W = (pack_I ** 2) * max(config.battery.pack_resistance, 0.0)
    motor_temp_est_C = ambient_temp_C + motor_copper_loss_W * 0.35
    esc_temp_est_C = ambient_temp_C + esc_loss * 0.70
    battery_temp_est_C = ambient_temp_C + battery_loss_W * 0.25
    max_temp = max(motor_temp_est_C, esc_temp_est_C, battery_temp_est_C)
    thermal_status = "OK" if max_temp < 65.0 else ("WARN" if max_temp < 85.0 else "HOT")

    return dict(
        # Speed
        airspeed_mps       = V,
        groundspeed_mps    = gs_track,
        wind_head_mps      = float(wind_head_mps),
        wind_cross_mps     = float(wind_cross_mps),
        stall_speed_mps    = V_stall,
        stall_speed_turn_mps = V_stall_turn,
        # Aerodynamics
        CL                 = CL,
        CD                 = CD,
        LD_ratio           = LD,
        glide_ratio        = glide_ratio,
        glide_distance_m   = glide_dist_m,
        glide_distance_km  = glide_dist_m / 1000.0,
        induced_drag_N     = d_induced,
        parasitic_drag_N   = d_parasitic,
        drag_N             = D,
        wing_loading_N_m2  = wing_load_N_m2,
        wing_loading_kg_m2 = wing_load_kg_m2,
        aoa_deg            = aoa,
        reynolds_number    = Re,
        bank_deg           = float(bank_deg),
        load_factor        = n_turn,
        turn_radius_m      = turn_r_m,
        turn_rate_deg_s    = turn_rate_dps,
        turn_period_s      = turn_period_s,
        loiter_circles     = loiter_circles,
        # Thrust / power
        thrust_required_N  = T_req,
        thrust_available_N = T_avail,
        specific_thrust    = spec_thrust,
        power_required_W   = P_prop,
        motor_power_W      = P_elec,
        esc_loss_W         = esc_loss,
        avionics_power_W   = P_avionics,
        total_power_W      = P_total,
        max_prop_power_W   = max_prop_P,
        # Battery
        pack_current_A     = pack_I,
        v_load_V           = V_load,
        # Propeller
        rpm_est            = rpm_est,
        tip_speed_mps      = V_tip,
        pitch_speed_mps    = V_pitch,
        # Climb
        rate_of_climb_mps  = rc_at_V,
        max_rc_mps         = rc_max,
        v_max_rc_mps       = v_rc_max,
        max_aoc_deg        = gamma,
        v_max_aoc_mps      = v_gamma,
        # Optimal speeds
        best_endurance_speed_mps  = V_be,
        best_range_speed_mps      = V_br,
        best_ld_ratio             = LD_br,
        min_sink_speed_mps        = V_min_sink,
        min_sink_rate_mps         = sink_min,
        service_ceiling_m         = service_ceiling_abs_m,
        service_ceiling_agl_m     = service_ceiling_agl_m,
        # Ground
        takeoff_dist_m     = S_to,
        landing_dist_m     = S_ld,
        # Duration
        flight_time_min    = t_min,
        flight_range_km    = d_km,
        specific_range_m_per_Wh        = specific_range_m_Wh,
        specific_range_km_per_kWh      = specific_range_m_Wh,
        specific_endurance_min_per_Wh  = specific_endurance_min_Wh,
        specific_endurance_h_per_kWh   = specific_endurance_min_Wh * (1000.0 / 60.0),
        # Thermal
        motor_copper_loss_W  = motor_copper_loss_W,
        battery_loss_W       = battery_loss_W,
        motor_temp_est_C     = motor_temp_est_C,
        esc_temp_est_C       = esc_temp_est_C,
        battery_temp_est_C   = battery_temp_est_C,
        thermal_status       = thermal_status,
        # ESC note
        esc_note           = esc_note,
    )


def _format_distance_m(value_m: float) -> str:
    """Format distance in m or km for compact display."""
    if not math.isfinite(value_m):
        return "∞"
    if abs(value_m) >= 1000.0:
        return f"{value_m/1000.0:.2f} km"
    return f"{value_m:.1f} m"


# ============================================================
# FLIGHT TIME / RANGE SWEEP
# ============================================================
def flight_time_min(config: FixedWingConfig, speed_mps: float) -> float:
    """Flight time at constant airspeed [min], respecting all limits."""
    V_stall = stall_speed(config)
    if speed_mps < V_stall:
        return 0.0
    m = compute_metrics(config, speed_mps)
    return m["flight_time_min"]


def flight_range_km(config: FixedWingConfig, speed_mps: float) -> float:
    return compute_metrics(config, speed_mps)["flight_range_km"]


def find_optimal_speeds(config: FixedWingConfig,
                        v_max: float = 80.0) -> Tuple[float, float, float, float]:
    """
    Numerical scan for best endurance and best range speeds.
    Returns (V_be, t_max_min, V_br, d_max_km).
    """
    V_s      = stall_speed(config)
    v_lo     = max(V_s, 1.0)
    n_steps  = 500
    best_t   = -1.0;  best_vt = v_lo
    best_d   = -1.0;  best_vd = v_lo
    for i in range(n_steps + 1):
        V = v_lo + (v_max - v_lo) * i / n_steps
        t = flight_time_min(config, V)
        d = flight_range_km(config, V)
        if t > best_t: best_t = t; best_vt = V
        if d > best_d: best_d = d; best_vd = V
    return best_vt, best_t, best_vd, best_d


# ============================================================
# PERFORMANCE FIGURE (for GUI plot canvas)
# ============================================================
def make_performance_figure(config: FixedWingConfig,
                             max_speed: float = 40.0,
                             figsize: tuple = (15, 9)) -> Figure:
    """
    Six-panel performance figure matching eCalc fixed-wing output layout.

    Panels:
      1. Flight Time & Range vs Speed
      2. Thrust Required vs Available vs Speed
      3. Power Required vs Available vs Speed
      4. Rate of Climb vs Speed
      5. Drag Polar  (CL vs CD)
      6. Induced vs Parasitic Drag vs Speed
    """
    V_stall = stall_speed(config)
    v_lo    = max(V_stall, 1.0)
    v_hi    = max_speed
    speeds  = [v_lo + (v_hi - v_lo) * i / 300 for i in range(301)]

    times, ranges, drags, T_avail_v = [], [], [], []
    powers_req, powers_avail        = [], []
    rcs, CLs, CDs, LDs              = [], [], [], []
    induced_drags, parasitic_drags  = [], []

    # Evaluated per speed inside the loop below; kept here only as the
    # static reference for any caller that wants it.
    T_av = thrust_available_N(config, 0.0)
    af   = config.airframe
    rho  = config.air_density
    W    = config.weight_N

    for V in speeds:
        times.append(flight_time_min(config, V))
        ranges.append(flight_range_km(config, V))
        d_ind, d_par, D = drag_components_N(config, V)
        induced_drags.append(d_ind)
        parasitic_drags.append(d_par)
        drags.append(D)
        T_avail_v.append(T_av)
        powers_req.append(power_required_W(config, V))
        powers_avail.append(T_av * V)
        rcs.append(rate_of_climb_mps(config, V))
        CL = af.cl_at_speed(W, V, rho)
        CD = af.cd_at_cl(CL)
        CLs.append(CL)
        CDs.append(CD)
        LDs.append(af.ld_ratio(CL))

    fig, axes = plt.subplots(2, 3, figsize=figsize)
    fig.suptitle("Fixed-Wing UAV Performance", fontsize=13, fontweight="bold")

    # ---- 1. Time & Range ----
    ax = axes[0, 0]
    ax2 = ax.twinx()
    l1, = ax.plot(speeds, times,  color="royalblue", label="Flight Time (min)")
    l2, = ax2.plot(speeds, ranges, color="darkorange", label="Range (km)", linestyle="--")
    ax.set_xlabel("Airspeed (m/s)"); ax.set_ylabel("Time (min)"); ax2.set_ylabel("Range (km)")
    ax.set_title("Flight Time & Range vs Airspeed")
    ax.axvline(V_stall, color="red", linestyle=":", linewidth=1, label="V_stall")
    ax.legend(handles=[l1, l2], loc="upper right", fontsize=8)
    ax.grid(True, alpha=0.4)

    # ---- 2. Thrust ----
    ax = axes[0, 1]
    ax.plot(speeds, drags,     label="Thrust Required (N)", color="crimson")
    ax.plot(speeds, T_avail_v, label="Thrust Available (N)", color="green", linestyle="--")
    ax.set_xlabel("Airspeed (m/s)"); ax.set_ylabel("Thrust (N)")
    ax.set_title("Thrust Required vs Available")
    ax.legend(fontsize=8); ax.grid(True, alpha=0.4)
    ax.axvline(V_stall, color="red", linestyle=":", linewidth=1)

    # ---- 3. Power ----
    ax = axes[0, 2]
    ax.plot(speeds, [p/1000 for p in powers_req],   label="Power Required (kW)", color="crimson")
    ax.plot(speeds, [p/1000 for p in powers_avail], label="Power Available (kW)", color="green", linestyle="--")
    ax.set_xlabel("Airspeed (m/s)"); ax.set_ylabel("Power (kW)")
    ax.set_title("Power Required vs Available")
    ax.legend(fontsize=8); ax.grid(True, alpha=0.4)
    ax.axvline(V_stall, color="red", linestyle=":", linewidth=1)

    # ---- 4. Rate of Climb ----
    ax = axes[1, 0]
    ax.plot(speeds, [r*60 for r in rcs], color="teal", label="Rate of Climb (m/min)")
    ax.set_xlabel("Airspeed (m/s)"); ax.set_ylabel("RC (m/min)")
    ax.set_title("Rate of Climb vs Airspeed")
    ax.axhline(0, color="gray", linewidth=0.8); ax.grid(True, alpha=0.4)
    ax.axvline(V_stall, color="red", linestyle=":", linewidth=1)
    ax.legend(fontsize=8)

    # ---- 5. Drag Polar ----
    ax = axes[1, 1]
    ax.plot(CDs, CLs, color="purple")
    ax.set_xlabel("CD (Drag Coefficient)"); ax.set_ylabel("CL (Lift Coefficient)")
    ax.set_title("Drag Polar (CD vs CL)")
    # Mark CL_max
    ax.axhline(config.airframe.CL_max, color="red", linestyle=":", linewidth=1, label="CL_max")
    ax.grid(True, alpha=0.4); ax.legend(fontsize=8)

    # ---- 6. Induced vs Parasitic Drag ----
    ax = axes[1, 2]
    ax.plot(speeds, induced_drags,   color="navy",      label="Induced Drag")
    ax.plot(speeds, parasitic_drags, color="darkorange", label="Parasitic Drag")
    ax.plot(speeds, drags, color="gray", linestyle="--", linewidth=1.0, label="Total Drag")
    ax.set_xlabel("Airspeed (m/s)"); ax.set_ylabel("Drag (N)")
    ax.set_title("Induced vs Parasitic Drag")
    ax.axvline(V_stall, color="red", linestyle=":", linewidth=1)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.4)

    fig.tight_layout(rect=[0, 0, 1, 0.95])
    return fig


def make_airframe_diagram_figure(config: "FixedWingConfig", figsize: tuple = (9, 7)):
    """
    Plan-view stick diagram of the wing with its propellers, to scale.

    The wing is drawn as a rectangle of span x mean chord (chord = area/span),
    which is what the drag model already assumes. Propellers sit on the wing:
    a single motor on the centreline as a nose tractor, and multiple motors
    spread symmetrically along the span.

    Discs that intersect each other are drawn in red, and a disc reaching past
    the wing tip is flagged, since both are easy to miss on paper.
    """
    import matplotlib.pyplot as _plt
    from matplotlib.patches import Circle as _Circle, Rectangle as _Rectangle

    span = max(float(config.airframe.wing_span_m), 1e-6)
    area = max(float(config.airframe.wing_area_m2), 1e-9)
    chord = area / span
    n_motors = max(int(config.num_motors), 1)
    prop_d = float(config.propeller.diameter_in) * 0.0254

    layout = core.wing_rotor_positions(n_motors, span, prop_d)
    r_prop = layout["prop_radius_m"]
    overlap = layout["overlaps"]
    past_tip = layout["tip_overhang_m"] > 0
    disc_colour = "#C62828" if (overlap or past_tip) else "#2E7D32"

    fig, ax = _plt.subplots(figsize=figsize)

    # Conventional plan view: span across X, chord down Y, nose toward -Y.
    ax.add_patch(_Rectangle((-span / 2.0, -chord / 2.0), span, chord,
                            facecolor="#CFD8DC", edgecolor="#37474F",
                            linewidth=1.8, zorder=2))
    # Fuselage stripe, purely indicative
    ax.add_patch(_Rectangle((-chord * 0.18, -chord * 0.95), chord * 0.36, chord * 2.1,
                            facecolor="#B0BEC5", edgecolor="#37474F",
                            linewidth=1.2, zorder=1))

    # Propellers ahead of the leading edge, as a tractor installation
    y_prop = -chord / 2.0 - r_prop * 0.15
    for x in layout["positions_y_m"]:
        ax.add_patch(_Circle((x, y_prop), r_prop, fill=False,
                             edgecolor=disc_colour, linewidth=1.6, zorder=3))
        ax.plot([x], [y_prop], marker="o", markersize=5, color="#37474F", zorder=4)
        ax.plot([x, x], [y_prop, -chord / 2.0], color="#37474F",
                linewidth=2.0, zorder=2)

    # Span dimension, below the wing
    y_dim = chord / 2.0 + max(span * 0.10, chord * 0.6)
    ax.annotate("", xy=(-span / 2.0, y_dim), xytext=(span / 2.0, y_dim),
                arrowprops=dict(arrowstyle="<->", color="#1565C0", lw=1.2))
    ax.text(0, y_dim, f"span {span * 1000:.0f} mm", color="#1565C0",
            fontsize=8, ha="center", va="bottom")
    # Chord dimension, off the right tip
    x_dim = span / 2.0 + max(span * 0.06, chord * 0.5)
    ax.annotate("", xy=(x_dim, -chord / 2.0), xytext=(x_dim, chord / 2.0),
                arrowprops=dict(arrowstyle="<->", color="#6A1B9A", lw=1.2))
    ax.text(x_dim + span * 0.01, 0, f" chord\n {chord * 1000:.0f} mm",
            color="#6A1B9A", fontsize=8, va="center")

    ax.set_aspect("equal", adjustable="box")
    margin = max(span * 0.22, r_prop * 1.4)
    ax.set_xlim(-span / 2.0 - margin, span / 2.0 + margin * 1.6)
    ax.set_ylim(y_prop - r_prop - margin * 0.5, y_dim + margin * 0.6)
    ax.invert_yaxis()                      # nose up, as a plan view reads
    ax.grid(True, linestyle=":", alpha=0.4)
    ax.set_xlabel("metres (spanwise)")
    ax.set_ylabel("metres (aft +)")

    gap_mm = layout["adjacent_gap_m"] * 1000.0
    if n_motors == 1:
        verdict = "single tractor"
    elif overlap:
        verdict = f"OVERLAP {abs(gap_mm):.0f} mm"
    else:
        verdict = f"tip gap {gap_mm:.0f} mm"
    ax.set_title(f"{n_motors} motor(s)  |  {config.propeller.diameter_in:g} in props"
                 f"  |  {verdict}", fontsize=11,
                 color="#C62828" if (overlap or past_tip) else "#000000")

    notes = [
        f"Wing span       {span * 1000:.0f} mm",
        f"Wing area       {area:.4f} m2",
        f"Mean chord      {chord * 1000:.0f} mm",
        f"Aspect ratio    {config.airframe.aspect_ratio:.2f}",
        f"Prop diameter   {prop_d * 1000:.0f} mm",
    ]
    if n_motors > 1:
        notes.append(f"Prop spacing    {layout['motor_spacing_m'] * 1000:.0f} mm")
        notes.append(f"Tip-to-tip gap  {gap_mm:+.0f} mm")
    if past_tip:
        notes.append("")
        notes.append(f"Disc extends {layout['tip_overhang_m'] * 1000:.0f} mm past the tip")
    ax.text(0.02, 0.02, "\n".join(notes), transform=ax.transAxes,
            fontsize=8, family="monospace", va="bottom",
            bbox=dict(boxstyle="round", facecolor="#FFFDE7", edgecolor="#BDBDBD"))

    fig.tight_layout()
    return fig


def make_motor_operating_point_figure(config: FixedWingConfig, metrics: dict, figsize: tuple = (12, 8)):
    """
    Create a figure showing motor/propeller operating curves with the current operating point marked.
    Uses propeller table data if available.
    
    Two subplots:
      1. Thrust vs Power (left), Thrust vs Current (right) on same plot
      2. Thrust vs Efficiency g/W (left), Thrust vs RPM (right) on same plot
    """
    if config.propeller.table is None:
        # Return empty figure if no propeller table
        fig, ax = plt.subplots(1, 1, figsize=figsize)
        ax.text(0.5, 0.5, "Propeller table not available\nCannot plot operating curves",
                ha="center", va="center", transform=ax.transAxes, fontsize=12)
        ax.axis("off")
        return fig
    
    df = config.propeller.table
    thrust_N = float(metrics.get("thrust_required_N", 0.0))
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=figsize)
    
    # Get data from propeller table
    thrust_g = df["Thrust_g"].values if "Thrust_g" in df.columns else []
    
    # Subplot 1: Thrust vs Power & Thrust vs Current
    if "Power_W" in df.columns and len(thrust_g) > 0:
        power_W = df["Power_W"].values
        ax1_1 = ax1
        ax1_1.plot(thrust_g, power_W, "b-", linewidth=2, label="Power")
        ax1_1.set_xlabel("Thrust (gf)", fontsize=10)
        ax1_1.set_ylabel("Power (W)", fontsize=10, color="b")
        ax1_1.tick_params(axis="y", labelcolor="b")
        ax1_1.grid(True, alpha=0.3)
    
    if "Current_A" in df.columns and len(thrust_g) > 0:
        current_A = df["Current_A"].values
        ax1_2 = ax1.twinx()
        ax1_2.plot(thrust_g, current_A, "r--", linewidth=2, label="Current")
        ax1_2.set_ylabel("Current (A)", fontsize=10, color="r")
        ax1_2.tick_params(axis="y", labelcolor="r")
    
    # Mark operating point on subplot 1
    thrust_g_op = thrust_N * 1000.0 / 9.81
    if "Power_W" in df.columns and len(thrust_g) > 1:
        try:
            # Use the SAME lookup the model uses, rather than a second
            # hand-rolled interpolation. The local version here walked the
            # bracketing pair and extrapolated linearly off the end of the
            # table, which produced a negative operating power (-12.7 W) and
            # an impossible efficiency whenever the aircraft cruised below the
            # lowest measured thrust.
            power_op = _table_power_for_thrust(df, thrust_N, config.propeller)
            if power_op is not None and power_op > 0:
                ax1.plot(thrust_g_op, power_op, "go", markersize=10,
                         label=f"Operating ({power_op:.1f} W)",
                         markeredgewidth=2, markeredgecolor="darkgreen")
        except Exception:
            pass
    
    ax1.set_title("Thrust vs Power & Current", fontsize=11, fontweight="bold")
    ax1.legend(loc="upper left", fontsize=9)
    
    # Subplot 2: Thrust vs Efficiency & Thrust vs RPM
    if "Efficiency_gW" in df.columns and len(thrust_g) > 0:
        eff_gW = df["Efficiency_gW"].values
        ax2_1 = ax2
        ax2_1.plot(thrust_g, eff_gW, "g-", linewidth=2, label="Efficiency")
        ax2_1.set_xlabel("Thrust (gf)", fontsize=10)
        ax2_1.set_ylabel("Efficiency (g/W)", fontsize=10, color="g")
        ax2_1.tick_params(axis="y", labelcolor="g")
        ax2_1.grid(True, alpha=0.3)
    
    if "RPM" in df.columns and len(thrust_g) > 0:
        rpm = df["RPM"].values
        ax2_2 = ax2.twinx()
        ax2_2.plot(thrust_g, rpm, "m--", linewidth=2, label="RPM")
        ax2_2.set_ylabel("RPM", fontsize=10, color="m")
        ax2_2.tick_params(axis="y", labelcolor="m")
    
    # Mark operating point on subplot 2
    if "Efficiency_gW" in df.columns and len(thrust_g) > 1:
        try:
            eff_gW = df["Efficiency_gW"].values
            idx = 0
            for i in range(len(thrust_g) - 1):
                if thrust_g[i] <= thrust_g_op <= thrust_g[i + 1]:
                    idx = i
                    break
            frac = (thrust_g_op - thrust_g[idx]) / (thrust_g[idx + 1] - thrust_g[idx]) if thrust_g[idx + 1] != thrust_g[idx] else 0
            eff_op = eff_gW[idx] + frac * (eff_gW[idx + 1] - eff_gW[idx])
            # Efficiency is derived from the power above so the two markers
            # can never disagree.
            _p = _table_power_for_thrust(df, thrust_N, config.propeller)
            if _p and _p > 0:
                eff_op = thrust_g_op / _p
                ax2.plot(thrust_g_op, eff_op, "o", markersize=10, color="cyan",
                         label=f"Operating ({eff_op:.2f} g/W)",
                         markeredgewidth=2, markeredgecolor="teal")
        except:
            pass
    
    ax2.set_title("Thrust vs Efficiency & RPM", fontsize=11, fontweight="bold")
    ax2.legend(loc="upper left", fontsize=9)
    
    # Say plainly when the operating point sits outside the measured data.
    # Anything outside the table is an extrapolation, and a reader deserves to
    # know that before trusting the marker's position.
    _t_lo = float(df["Thrust_g"].min()) if "Thrust_g" in df else None
    _t_hi = float(df["Thrust_g"].max()) if "Thrust_g" in df else None
    _outside = (_t_lo is not None
                and (thrust_g_op < _t_lo or thrust_g_op > _t_hi))
    _range_note = ""
    if _outside:
        _where = "below" if thrust_g_op < _t_lo else "above"
        _range_note = (f"   —  EXTRAPOLATED, {_where} the measured "
                       f"{_t_lo:.0f}-{_t_hi:.0f} g range")
    fig.suptitle(f"Motor/Propeller Operating Curves (Thrust: {thrust_g_op:.0f}g)"
                 + _range_note,
                 color=("#C62828" if _outside else "black"), 
                 fontsize=12, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    return fig


# ============================================================
# MISSION PROFILE  (fixed-wing version)
# ============================================================
from dataclasses import dataclass, field as _field

@dataclass
class MissionPhase:
    """One leg of a fixed-wing mission.

    Either duration (seconds) or distance (metres) must be supplied.
    altitude_m sets the ISA layer for that leg — air density is recomputed
    per-phase so climb/descent effects are captured.
    """
    name:       str
    speed:      float               # airspeed [m/s]
    duration:   Optional[float] = None   # seconds (mutually exclusive with distance)
    distance:   Optional[float] = None   # metres
    altitude:   float = 0.0         # metres ASL
    bank_deg:   float = 0.0
    course_deg: float = 0.0
    climb_rate_mps: Optional[float] = None
    descent_rate_mps: Optional[float] = None


@dataclass
class MissionProfile:
    phases: list
    reserve_percent: float = 20.0
    rth_reserve_Wh: float = 0.0
    diversion_reserve_Wh: float = 0.0
    wind_direction_deg: float = 0.0
    # Transient (acceleration / deceleration) model. Same JSON keys the
    # multicopter simulator uses, so a mission file reads the same either way.
    transient_dt_s: float = 0.5
    max_accel_mps2: float = 1.5
    max_decel_mps2: float = 2.0
    decel_regen_eff: float = 0.0

    @staticmethod
    def from_json(path: str) -> "MissionProfile":
        """Load a mission from a JSON file.

        Expected format (same as multicopter sim):
        {
          "phases": [
            {"name": "Climb",   "speed": 18, "duration": 60,   "altitude": 150},
            {"name": "Cruise",  "speed": 22, "distance": 5000, "altitude": 150},
            {"name": "Return",  "speed": 20, "distance": 5000, "altitude": 80},
            {"name": "Land",    "speed": 16, "duration": 30,   "altitude": 0}
          ]
        }
        """
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        phases = []
        for p in data.get("phases", []):
            phases.append(MissionPhase(
                name     = str(p["name"]),
                speed    = float(p["speed"]),
                duration = float(p["duration"])  if "duration"  in p else None,
                distance = float(p["distance"])  if "distance"  in p else None,
                altitude = float(p.get("altitude", 0.0)),
                bank_deg = float(p.get("bank_deg", 0.0)),
                course_deg = float(p.get("course_deg", 0.0)),
                climb_rate_mps = (float(p["climb_rate_mps"]) if "climb_rate_mps" in p else None),
                descent_rate_mps = (float(p["descent_rate_mps"]) if "descent_rate_mps" in p else None),
            ))
        return MissionProfile(
            phases=phases,
            reserve_percent=float(data.get("reserve_percent", 20.0)),
            rth_reserve_Wh=float(data.get("rth_reserve_Wh", 0.0)),
            diversion_reserve_Wh=float(data.get("diversion_reserve_Wh", 0.0)),
            wind_direction_deg=float(data.get("wind_direction_deg", 0.0)),
            transient_dt_s=float(data.get("transient_dt_s", 0.5)),
            max_accel_mps2=float(data.get("max_accel_mps2", 1.5)),
            max_decel_mps2=float(data.get("max_decel_mps2", 2.0)),
            decel_regen_eff=float(data.get("decel_regen_eff", 0.0)),
        )


def simulate_fw_mission(
    cfg: "FixedWingConfig",
    mission: MissionProfile,
    wind_mps: float = 0.0,
    temperature_C: Optional[float] = None,
    pressure_Pa:   Optional[float] = None,
) -> tuple:
    """
    Simulate a fixed-wing mission phase-by-phase, draining battery energy.

    Wind sign convention: +wind = headwind (reduces groundspeed and range).

    Returns
    -------
    results      : list of (name, time_min, dist_km, status_str)
    worst_metrics: dict — worst-case single-point metrics across all phases
    series       : dict of lists — time-series data for the Mission Plots tab
    """
    usable_Wh = cfg.battery.usable_Wh
    reserve_target_Wh = max(
        usable_Wh * (float(mission.reserve_percent) / 100.0),
        float(mission.rth_reserve_Wh) + float(mission.diversion_reserve_Wh),
    )
    remaining_Wh = usable_Wh
    results: list  = []
    worst:   dict  = {}
    t_s    = 0.0
    dist_km= 0.0
    reserve_breached = False
    reserve_min_Wh = remaining_Wh
    ambient_c = float(temperature_C) if temperature_C is not None else 25.0
    motor_temp_C = ambient_c
    esc_temp_C = ambient_c
    battery_temp_C = ambient_c
    prev_alt_m = float(mission.phases[0].altitude) if mission.phases else 0.0

    series: dict = {
        "t_s":              [],
        "phase":            [],
        "airspeed_mps":     [],
        "groundspeed_mps":  [],
        "headwind_mps":     [],
        "crosswind_mps":    [],
        "bank_deg":         [],
        "load_factor":      [],
        "turn_radius_m":    [],
        "stall_turn_mps":   [],
        "climb_rate_cmd_mps": [],
        "descent_rate_cmd_mps": [],
        "climb_power_add_W": [],
        "potential_power_W": [],
        "altitude_m":       [],
        "battery_voltage_V":[],
        "battery_current_A":[],
        "battery_energy_Wh":[],
        "soc_percent":      [],
        "reserve_target_Wh":[],
        "reserve_margin_Wh":[],
        "reserve_breach":   [],
        "total_power_W":    [],
        "motor_power_W":    [],
        "drag_N":           [],
        "thrust_avail_N":   [],
        "rate_of_climb_mps":[],
        "lift_drag_ratio":  [],
        "cl_cruise":        [],
        "motor_temp_est_C": [],
        "esc_temp_est_C":   [],
        "battery_temp_est_C": [],
        "thermal_status":   [],
    }

    def _append(phase_name, alt_m, m, t_now, d_now, e_now, reserve_hit):
        series["t_s"].append(t_now)
        series["phase"].append(phase_name)
        series["airspeed_mps"].append(m.get("airspeed_mps", 0.0))
        series["groundspeed_mps"].append(m.get("groundspeed_mps", 0.0))
        series["headwind_mps"].append(m.get("wind_head_mps", 0.0))
        series["crosswind_mps"].append(m.get("wind_cross_mps", 0.0))
        series["bank_deg"].append(m.get("bank_deg", 0.0))
        series["load_factor"].append(m.get("load_factor", 1.0))
        series["turn_radius_m"].append(m.get("turn_radius_m", float("inf")))
        series["stall_turn_mps"].append(m.get("stall_speed_turn_mps", 0.0))
        series["climb_rate_cmd_mps"].append(m.get("climb_rate_cmd_mps", 0.0))
        series["descent_rate_cmd_mps"].append(m.get("descent_rate_cmd_mps", 0.0))
        series["climb_power_add_W"].append(m.get("climb_power_add_W", 0.0))
        series["potential_power_W"].append(m.get("potential_power_W", 0.0))
        series["altitude_m"].append(float(alt_m))
        series["battery_voltage_V"].append(m.get("v_load_V", 0.0))
        series["battery_current_A"].append(m.get("pack_current_A", 0.0))
        series["battery_energy_Wh"].append(e_now)
        series["soc_percent"].append(
            (max(e_now, 0.0) / max(usable_Wh, 1e-9)) * 100.0 if usable_Wh > 0 else 100.0)
        series["reserve_target_Wh"].append(reserve_target_Wh)
        series["reserve_margin_Wh"].append(e_now - reserve_target_Wh)
        series["reserve_breach"].append(1 if reserve_hit else 0)
        series["total_power_W"].append(m.get("total_power_W", 0.0))
        series["motor_power_W"].append(m.get("motor_power_W", 0.0))
        series["drag_N"].append(m.get("drag_N", 0.0))
        series["thrust_avail_N"].append(m.get("thrust_available_N", 0.0))
        series["rate_of_climb_mps"].append(m.get("rate_of_climb_mps", 0.0))
        series["lift_drag_ratio"].append(m.get("LD_ratio", 0.0))
        series["cl_cruise"].append(m.get("CL", 0.0))
        series["motor_temp_est_C"].append(motor_temp_C)
        series["esc_temp_est_C"].append(esc_temp_C)
        series["battery_temp_est_C"].append(battery_temp_C)
        series["thermal_status"].append(m.get("thermal_status", "OK"))

    def _merge_worst(w, m):
        if not w:
            return dict(m)
        for k in ("total_power_W","motor_power_W","pack_current_A",
                  "drag_N","thrust_required_N","esc_loss_W",
                  "load_factor","motor_temp_est_C","esc_temp_est_C","battery_temp_est_C"):
            w[k] = max(float(w.get(k,0)), float(m.get(k,0)))
        w["v_load_V"] = min(float(w.get("v_load_V",1e9)),
                            float(m.get("v_load_V",1e9)))
        w["reserve_margin_Wh"] = min(float(w.get("reserve_margin_Wh", 1e9)),
                                     float(m.get("reserve_margin_Wh", 1e9)))
        return w

    # Airspeed carried across phases so a speed change costs a real ramp.
    # Starts at the first phase's speed, so a mission that never changes speed
    # behaves exactly as it did before the transient model existed.
    current_speed_mps = float(mission.phases[0].speed) if mission.phases else 0.0

    for phase in mission.phases:
        # Recompute air density per-phase altitude
        rho = isa_density(float(phase.altitude), temperature_C, pressure_Pa)
        cfg.air_density = rho

        headwind_mps, crosswind_mps = wind_components_mps(
            wind_speed_mps=wind_mps,
            wind_direction_deg=mission.wind_direction_deg,
            course_deg=float(phase.course_deg),
        )
        n_turn = bank_load_factor(phase.bank_deg)
        V_air = max(float(phase.speed), stall_speed(cfg) * math.sqrt(max(n_turn, 1.0)) + 0.01)
        V_gs  = max(groundspeed_along_track_mps(V_air, headwind_mps, crosswind_mps), 0.1)
        climb_cmd = max(float(phase.climb_rate_mps or 0.0), 0.0)
        descent_cmd = max(float(phase.descent_rate_mps or 0.0), 0.0)
        if climb_cmd > 0 and descent_cmd > 0:
            descent_cmd = 0.0
        if phase.climb_rate_mps is None and phase.descent_rate_mps is None:
            dh = float(phase.altitude) - prev_alt_m
            if phase.duration is not None and float(phase.duration) > 0:
                est_vz = dh / float(phase.duration)
                if est_vz > 0:
                    climb_cmd = est_vz
                elif est_vz < 0:
                    descent_cmd = -est_vz
        potential_power_W = cfg.weight_N * (climb_cmd - descent_cmd)

        m = compute_metrics(
            cfg,
            V_air,
            bank_deg=float(phase.bank_deg),
            ambient_temp_C=ambient_c,
            wind_head_mps=headwind_mps,
            wind_cross_mps=crosswind_mps,
            glide_altitude_m=float(phase.altitude),
        )
        m["climb_rate_cmd_mps"] = climb_cmd
        m["descent_rate_cmd_mps"] = descent_cmd
        m["potential_power_W"] = potential_power_W
        m["climb_power_add_W"] = potential_power_W
        base_total_P = float(m.get("total_power_W", 0.0))
        total_P = max(base_total_P + potential_power_W, 0.0)
        if base_total_P > 0:
            scale = total_P / base_total_P
            m["motor_power_W"] = float(m.get("motor_power_W", 0.0)) * scale
            m["esc_loss_W"] = float(m.get("esc_loss_W", 0.0)) * scale
        m["total_power_W"] = total_P
        m["pack_current_A"] = total_P / max(cfg.battery.vnom_pack, 1.0)
        # Evaluate the pack at its CURRENT state of charge, so voltage sag and
        # the rise in internal resistance late in the flight are captured.
        soc_now = (max(remaining_Wh, 0.0) / max(usable_Wh, 1e-9)) if usable_Wh > 0 else 1.0
        soc_now = min(max(soc_now, 0.0), 1.0)
        m["soc_percent"] = soc_now * 100.0
        m["soc_model_source"] = _soc_model_short_label(
            getattr(cfg.battery, "soc_model_source", None))
        m["v_load_V"] = cfg.battery.voltage_under_load(m["pack_current_A"], soc=soc_now)
        motor_i_per_motor = m["motor_power_W"] / max(m["v_load_V"], 1.0) / max(cfg.num_motors, 1)
        motor_copper_loss_W = (motor_i_per_motor ** 2) * max(cfg.motor.resistance, 0.0) * max(cfg.num_motors, 1)
        battery_loss_W = (m["pack_current_A"] ** 2) * max(
            cfg.battery.resistance_at_soc(soc_now), 0.0)
        m["motor_copper_loss_W"] = motor_copper_loss_W
        m["battery_loss_W"] = battery_loss_W
        m["reserve_margin_Wh"] = remaining_Wh - reserve_target_Wh
        worst = _merge_worst(worst, m)

        total_P   = float(m.get("total_power_W", 0.0))
        pack_I    = float(m.get("pack_current_A", 0.0))
        V_load    = float(m.get("v_load_V", 0.0))

        # Hard limits
        if pack_I > cfg.battery.discharge_max_A:
            results.append((phase.name, 0.0, 0.0, "Battery limit exceeded (current)"))
            break
        if V_load < cfg.battery.vmin_pack:
            results.append((phase.name, 0.0, 0.0, "Battery limit exceeded (voltage)"))
            break
        if V_air <= stall_speed(cfg):
            results.append((phase.name, 0.0, 0.0, f"Below stall speed ({stall_speed(cfg):.1f} m/s)"))
            break

        _append(phase.name, phase.altitude, m, t_s, dist_km, remaining_Wh, reserve_breached)

        # ---- Transient lead-in: ramp airspeed to the commanded value -------
        # A fixed-wing cannot change speed instantly. Accelerating costs extra
        # power on top of steady drag, and the ramp takes time and covers
        # ground, both of which must come out of this phase's budget.
        #
        # Modelled as a short segment at the head of the phase, integrated in
        # sub-steps, followed by the steady remainder. When the commanded
        # speed already matches the current speed the ramp is zero-length and
        # this reduces exactly to the previous behaviour.
        ramp_s = 0.0
        ramp_km = 0.0
        ramp_Wh = 0.0
        dv_needed = V_air - current_speed_mps
        if abs(dv_needed) > 1e-6:
            a_lim = (float(mission.max_accel_mps2) if dv_needed > 0
                     else float(mission.max_decel_mps2))
            dt_step = max(float(mission.transient_dt_s), 0.05)
            v_now = current_speed_mps
            guard = 0
            while abs(V_air - v_now) > 1e-6 and guard < 10000:
                guard += 1
                v_next, accel = ramp_speed(v_now, V_air, dt_step, a_lim, a_lim)

                # Evaluate steady power at the midpoint of the sub-step, then
                # add the kinetic term for the speed change itself.
                v_eval = 0.5 * (v_now + v_next)
                if v_eval <= stall_speed(cfg):
                    # Too slow to fly during the ramp; skip the transient model
                    # rather than report a stall that the pilot would simply
                    # have avoided by starting the acceleration earlier.
                    ramp_s = ramp_km = ramp_Wh = 0.0
                    break
                m_ramp = compute_metrics(
                    cfg, v_eval,
                    bank_deg=float(phase.bank_deg),
                    ambient_temp_C=ambient_c,
                    wind_head_mps=headwind_mps,
                    wind_cross_mps=crosswind_mps,
                    glide_altitude_m=float(phase.altitude),
                )
                p_kin = kinetic_power_term_W(
                    cfg.aircraft_weight_g, v_now, v_next, dt_step,
                    float(mission.decel_regen_eff))
                p_step = max(float(m_ramp.get("total_power_W", 0.0))
                             + potential_power_W + p_kin, 0.0)

                gs_step = groundspeed_along_track_mps(v_eval, headwind_mps, crosswind_mps)
                ramp_s += dt_step
                ramp_km += gs_step * dt_step / 1000.0
                ramp_Wh += p_step * (dt_step / 3600.0)
                v_now = v_next

            current_speed_mps = v_now

            if ramp_Wh >= remaining_Wh:
                # The battery runs out during the acceleration itself.
                actual_s = (remaining_Wh / max(ramp_Wh / max(ramp_s, 1e-9), 1e-9)) if ramp_s > 0 else 0.0
                t_s += actual_s
                dist_km += ramp_km * (actual_s / max(ramp_s, 1e-9))
                remaining_Wh = 0.0
                _append(phase.name, phase.altitude, m, t_s, dist_km, remaining_Wh, True)
                results.append((phase.name, actual_s / 60.0,
                                dist_km, "Battery depleted (during acceleration)"))
                break

            remaining_Wh -= ramp_Wh
            t_s += ramp_s
            dist_km += ramp_km
            reserve_breached = reserve_breached or (remaining_Wh < reserve_target_Wh)
            reserve_min_Wh = min(reserve_min_Wh, remaining_Wh)
        else:
            current_speed_mps = V_air

        phase_status = "OK"
        if reserve_breached:
            phase_status = "RESERVE VIOLATION"
        if phase.duration is not None:
            # The ramp already consumed part of this phase's time.
            dur_s      = max(float(phase.duration) - ramp_s, 0.0)
            energy_Wh  = total_P * (dur_s / 3600.0)
            if energy_Wh > remaining_Wh:
                # Battery runs out mid-phase
                actual_s   = (remaining_Wh / total_P) * 3600.0 if total_P > 0 else 0.0
                actual_km  = V_gs * actual_s / 1000.0
                t_s       += actual_s;  dist_km += actual_km;  remaining_Wh = 0.0
                motor_temp_C = thermal_step(motor_temp_C, ambient_c, m.get("motor_copper_loss_W", 0.0), 0.35, 240.0, actual_s)
                esc_temp_C = thermal_step(esc_temp_C, ambient_c, m.get("esc_loss_W", 0.0), 0.70, 180.0, actual_s)
                battery_temp_C = thermal_step(battery_temp_C, ambient_c, m.get("battery_loss_W", 0.0), 0.25, 500.0, actual_s)
                _append(phase.name, phase.altitude, m, t_s, dist_km, remaining_Wh, True)
                results.append((phase.name, actual_s/60.0, actual_km, "Battery depleted"))
                break
            remaining_Wh -= energy_Wh
            t_s          += dur_s
            dist_km      += V_gs * dur_s / 1000.0
            reserve_breached = reserve_breached or (remaining_Wh < reserve_target_Wh)
            reserve_min_Wh = min(reserve_min_Wh, remaining_Wh)
            motor_temp_C = thermal_step(motor_temp_C, ambient_c, m.get("motor_copper_loss_W", 0.0), 0.35, 240.0, dur_s)
            esc_temp_C = thermal_step(esc_temp_C, ambient_c, m.get("esc_loss_W", 0.0), 0.70, 180.0, dur_s)
            battery_temp_C = thermal_step(battery_temp_C, ambient_c, m.get("battery_loss_W", 0.0), 0.25, 500.0, dur_s)
            _append(phase.name, phase.altitude, m, t_s, dist_km, remaining_Wh, reserve_breached)
            results.append((phase.name, (dur_s + ramp_s)/60.0,
                            V_gs*dur_s/1000.0 + ramp_km, phase_status))

        elif phase.distance is not None:
            # The ramp already covered part of this phase's distance.
            dist_m  = max(float(phase.distance) - ramp_km * 1000.0, 0.0)
            time_s  = dist_m / V_gs if V_gs > 1e-9 else 0.0
            energy_Wh = total_P * (time_s / 3600.0)
            if energy_Wh > remaining_Wh:
                actual_s  = (remaining_Wh / total_P) * 3600.0 if total_P > 0 else 0.0
                actual_km = V_gs * actual_s / 1000.0
                t_s      += actual_s;  dist_km += actual_km;  remaining_Wh = 0.0
                motor_temp_C = thermal_step(motor_temp_C, ambient_c, m.get("motor_copper_loss_W", 0.0), 0.35, 240.0, actual_s)
                esc_temp_C = thermal_step(esc_temp_C, ambient_c, m.get("esc_loss_W", 0.0), 0.70, 180.0, actual_s)
                battery_temp_C = thermal_step(battery_temp_C, ambient_c, m.get("battery_loss_W", 0.0), 0.25, 500.0, actual_s)
                _append(phase.name, phase.altitude, m, t_s, dist_km, remaining_Wh, True)
                results.append((phase.name, actual_s/60.0, actual_km, "Battery depleted"))
                break
            remaining_Wh -= energy_Wh
            t_s          += time_s
            dist_km      += dist_m / 1000.0
            reserve_breached = reserve_breached or (remaining_Wh < reserve_target_Wh)
            reserve_min_Wh = min(reserve_min_Wh, remaining_Wh)
            motor_temp_C = thermal_step(motor_temp_C, ambient_c, m.get("motor_copper_loss_W", 0.0), 0.35, 240.0, time_s)
            esc_temp_C = thermal_step(esc_temp_C, ambient_c, m.get("esc_loss_W", 0.0), 0.70, 180.0, time_s)
            battery_temp_C = thermal_step(battery_temp_C, ambient_c, m.get("battery_loss_W", 0.0), 0.25, 500.0, time_s)
            _append(phase.name, phase.altitude, m, t_s, dist_km, remaining_Wh, reserve_breached)
            results.append((phase.name, (time_s + ramp_s)/60.0,
                            dist_m/1000.0 + ramp_km, phase_status))

        else:
            results.append((phase.name, 0.0, 0.0, "Invalid: no duration or distance"))
            break
        prev_alt_m = float(phase.altitude)

    if worst:
        worst["reserve_target_Wh"] = reserve_target_Wh
        worst["reserve_min_Wh"] = reserve_min_Wh
        worst["reserve_margin_Wh"] = reserve_min_Wh - reserve_target_Wh
        worst["reserve_breached"] = bool(reserve_breached)
        worst["motor_temp_est_C"] = motor_temp_C
        worst["esc_temp_est_C"] = esc_temp_C
        worst["battery_temp_est_C"] = battery_temp_C
        mt = max(motor_temp_C, esc_temp_C, battery_temp_C)
        worst["thermal_status"] = "OK" if mt < 65.0 else ("WARN" if mt < 85.0 else "HOT")
    return results, worst, series


# ============================================================
# GUI
# ============================================================
# ============================================================
# SHARED REPORTING / EXPORT UTILITIES
# ============================================================
import csv as _csv_mod
import io  as _io_mod
import datetime as _dt_mod

def _extract_weight_budget(cfg) -> list:
    """Return [(label, unit_g, count, total_g), ...] finishing with a TOTAL row."""
    rows = []
    total_g = float(getattr(cfg, "drone_weight_g",
                    getattr(cfg, "aircraft_weight_g", 0.0)))
    payload_g = max(float(getattr(cfg, "payload_mass_g", 0.0) or 0.0), 0.0)
    num_motors = int(getattr(cfg, "num_motors",
                    getattr(getattr(cfg, "airframe", None), "num_motors", 1)))
    batt  = getattr(cfg, "battery",   None)
    motor = getattr(cfg, "motor",     None)
    esc   = getattr(cfg, "esc",       None)
    prop  = getattr(cfg, "propeller", None)
    accounted = 0.0
    # Battery: weight_g already includes all series/parallel units
    if batt:
        w = float(getattr(batt, "weight_g", 0.0) or 0.0)
        rows.append(("Battery", w, 1, w)); accounted += w
    # Motor: per-motor weight, multiply by num_motors
    if motor:
        w = float(getattr(motor, "weight_g", 0.0) or 0.0)
        rows.append(("Motor", w, num_motors, w * num_motors)); accounted += w * num_motors
    # ESC: per-ESC weight, multiply by num_motors
    if esc:
        w = float(getattr(esc, "weight_g", 0.0) or 0.0)
        rows.append(("ESC", w, num_motors, w * num_motors)); accounted += w * num_motors
    # Propeller: per-propeller weight, multiply by num_motors
    if prop:
        w = float(getattr(prop, "weight_g", 0.0) or 0.0)
        rows.append(("Propeller", w, num_motors, w * num_motors)); accounted += w * num_motors
    if payload_g > 0:
        rows.append(("Payload", payload_g, 1, payload_g)); accounted += payload_g
    airframe_g = max(0.0, total_g - accounted)
    rows.append(("Airframe / Structure", airframe_g, 1, airframe_g))
    rows.append(("TOTAL", total_g, 1, total_g))
    return rows

def _export_csv_file(path: str, sweep: dict, metrics: list) -> None:
    import csv
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["[Performance Sweep]"])
        if sweep:
            headers = list(sweep.keys())
            w.writerow(headers)
            n = max(len(v) for v in sweep.values())
            for i in range(n):
                w.writerow([sweep[h][i] if i < len(sweep[h]) else "" for h in headers])
        w.writerow([])
        w.writerow(["[Metrics]"])
        w.writerow(["Metric", "Value"])
        for label, value in metrics:
            w.writerow([label, value])

def _export_excel_file(path: str, sweep: dict, metrics: list, weight_budget: list) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active; ws.title = "Performance Sweep"
    if sweep:
        headers = list(sweep.keys())
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
        n = max(len(v) for v in sweep.values())
        for ri in range(n):
            for ci, h in enumerate(headers, 1):
                ws.cell(row=ri+2, column=ci,
                        value=sweep[h][ri] if ri < len(sweep[h]) else None)
    ws2 = wb.create_sheet("Metrics")
    ws2.append(["Metric", "Value"])
    for r in [ws2["A1"], ws2["B1"]]: r.font = Font(bold=True)
    for label, value in metrics: ws2.append([label, value])
    ws3 = wb.create_sheet("Weight Budget")
    ws3.append(["Component", "Unit Weight (g)", "Count", "Total Weight (g)", "% of Total"])
    for c in ws3[1]: c.font = Font(bold=True)
    total_g = weight_budget[-1][3] if weight_budget else 1.0
    for label, uw, cnt, tw in weight_budget[:-1]:
        pct = round(tw / total_g * 100, 1) if total_g > 0 else 0
        ws3.append([label, round(uw, 1), cnt, round(tw, 1), pct])
    if weight_budget:
        label, uw, cnt, tw = weight_budget[-1]
        ws3.append([label, "", "", round(tw, 1), 100.0])
        for c in list(ws3.rows)[-1]: c.font = Font(bold=True)
    wb.save(path)

def _generate_pdf_report(path: str, report_title: str,
                          inputs_rows: list, metrics_rows: list,
                          status_sections: list, log_text: str,
                          figures: list, weight_budget: list) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, PageBreak, Image,
                                    HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units  import cm
    from reportlab.lib        import colors
    from reportlab.lib.enums  import TA_CENTER
    import io
    PAGE_W, PAGE_H = A4
    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2.0*cm,  bottomMargin=2.0*cm)
    styles = getSampleStyleSheet()
    story  = []
    NAVY  = colors.HexColor("#1F3864")
    TEAL  = colors.HexColor("#2E75B6")
    LGREY = colors.HexColor("#F2F2F2")
    DGREY = colors.HexColor("#595959")
    sTitle = ParagraphStyle("rTitle", fontSize=22, textColor=NAVY,
                             spaceAfter=6, alignment=TA_CENTER, fontName="Helvetica-Bold")
    sSub   = ParagraphStyle("rSub", fontSize=11, textColor=DGREY,
                             spaceAfter=20, alignment=TA_CENTER, fontName="Helvetica")
    sH1    = ParagraphStyle("rH1", fontSize=13, textColor=NAVY,
                             spaceBefore=12, spaceAfter=4, fontName="Helvetica-Bold")
    def _ts(hbg=TEAL):
        return TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), hbg),
            ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LGREY]),
            ("GRID",          (0,0),(-1,-1), 0.3, colors.lightgrey),
            ("LEFTPADDING",   (0,0),(-1,-1), 4),
            ("RIGHTPADDING",  (0,0),(-1,-1), 4),
            ("TOPPADDING",    (0,0),(-1,-1), 2),
            ("BOTTOMPADDING", (0,0),(-1,-1), 2),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ])
    usable_w = PAGE_W - 3.6*cm
    story.append(Spacer(1, 3*cm))
    story.append(Paragraph(report_title, sTitle))
    story.append(Paragraph(
        _dt_mod.datetime.now().strftime("Generated %d %B %Y  %H:%M"), sSub))
    story.append(HRFlowable(width="100%", thickness=1.5, color=TEAL, spaceAfter=12))
    if weight_budget:
        story.append(Paragraph("Component Weight Summary", sH1))
        total_g = weight_budget[-1][3] if weight_budget else 1.0
        wb_data = [["Component", "Unit (g)", "Qty", "Total (g)", "%"]]
        for label, uw, cnt, tw in weight_budget[:-1]:
            pct = f"{tw/total_g*100:.1f}" if total_g > 0 else "0"
            wb_data.append([label, f"{uw:.1f}", str(cnt), f"{tw:.1f}", pct])
        last = weight_budget[-1]
        wb_data.append(["TOTAL", "", "", f"{last[3]:.1f}", "100.0"])
        ts = _ts()
        ts.add("FONTNAME",(0,len(wb_data)-1),(-1,len(wb_data)-1),"Helvetica-Bold")
        t = Table(wb_data, colWidths=[usable_w*0.38, usable_w*0.15,
                                       usable_w*0.1, usable_w*0.17, usable_w*0.1])
        t.setStyle(ts); story.append(t)
    story.append(PageBreak())
    if inputs_rows:
        story.append(Paragraph("Design Inputs", sH1))
        mid = (len(inputs_rows)+1)//2
        left = inputs_rows[:mid]; right = inputs_rows[mid:]
        while len(right) < len(left): right.append(("",""))
        rows_data = [["Parameter","Value","Parameter","Value"]]
        for (la,va),(lb,vb) in zip(left,right): rows_data.append([la,va,lb,vb])
        col_w = usable_w/4
        t = Table(rows_data, colWidths=[col_w*1.4,col_w*0.6]*2)
        t.setStyle(_ts()); story.append(t)
    story.append(PageBreak())
    if metrics_rows:
        story.append(Paragraph("Performance Metrics", sH1))
        mid = (len(metrics_rows)+1)//2
        left = metrics_rows[:mid]; right = metrics_rows[mid:]
        while len(right) < len(left): right.append(("",""))
        rows_data = [["Metric","Value","Metric","Value"]]
        for (la,va),(lb,vb) in zip(left,right): rows_data.append([la,va,lb,vb])
        col_w = usable_w/4
        t = Table(rows_data, colWidths=[col_w*1.4,col_w*0.6]*2)
        t.setStyle(_ts()); story.append(t)
    story.append(PageBreak())
    if status_sections:
        story.append(Paragraph("Status Checks", sH1))
        for sec_title, sec_rows in status_sections:
            story.append(Paragraph(sec_title, ParagraphStyle("secH", fontSize=10,
                textColor=TEAL, spaceBefore=6, spaceAfter=2, fontName="Helvetica-Bold")))
            tdata = [["Metric","Value","Limit","Notes"]]
            ts = _ts()
            for ri,(metric,val,lim,note,tag) in enumerate(sec_rows, 1):
                tdata.append([metric,val,lim,note])
                bg = {"ok":colors.HexColor("#D9F2D9"),"warn":colors.HexColor("#FFF2CC"),
                      "bad":colors.HexColor("#F8D7DA")}.get(tag, colors.white)
                ts.add("BACKGROUND",(0,ri),(-1,ri),bg)
            cw = [usable_w*0.28,usable_w*0.20,usable_w*0.20,usable_w*0.32]
            t = Table(tdata, colWidths=cw); t.setStyle(ts); story.append(t)
            story.append(Spacer(1, 4))
    story.append(PageBreak())
    if figures:
        story.append(Paragraph("Performance Plots", sH1))
        for fig in figures:
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
            buf.seek(0)
            img_w = usable_w
            img_h = img_w * (fig.get_figheight() / max(fig.get_figwidth(), 0.01))
            max_h = PAGE_H - 5*cm
            if img_h > max_h:
                img_h = max_h
                img_w = img_h * (fig.get_figwidth() / max(fig.get_figheight(), 0.01))
            story.append(Image(buf, width=img_w, height=img_h))
            story.append(Spacer(1, 8)); story.append(PageBreak())
    if log_text.strip():
        story.append(Paragraph("Simulation Output Log", sH1))
        mono = ParagraphStyle("mono", fontName="Courier", fontSize=7.5, leading=10, spaceAfter=2)
        for line in log_text.splitlines():
            safe = line.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
            story.append(Paragraph(safe or " ", mono))
    doc.build(story)


# ============================================================
# INLINE HELP  —  tooltips and Simple/Advanced field visibility
# ============================================================

# ------------------------------------------------------------------
# FIELD HELP TEXT   (key -> (what it is, typical value / where to find it))
# ------------------------------------------------------------------
FW_FIELD_HELP = {
    # ---- Airframe ----
    "weight": ("All-up weight WITHOUT payload: airframe, wing, motor, ESC, "
               "servos, battery, receiver — everything you always fly with.",
               "Weigh the finished aircraft. 1.5 m foam trainer ~1200 g, "
               "2 m surveyor ~3000 g."),
    "payload_mass_g": ("Extra mass for this flight only — camera, sensor, cargo.",
                       "0 if flying clean."),
    "num_motors": ("Number of propulsion motors.",
                   "1 for a normal tractor/pusher, 2 for a twin."),
    "cruise_speed": ("The airspeed the single-point run is evaluated at.",
                     "Must be above stall. Typical trainer 15-20 m/s."),
    "periph_cur": ("Steady current drawn DIRECTLY from the main pack by "
                   "anything that is not a motor and not behind a regulator. "
                   "Use this for devices wired straight to pack voltage; use "
                   "the Avionics tab for anything on a regulated rail, where "
                   "converter efficiency matters. Never enter the same device "
                   "in both places.",
                   "A heater, pump, or payload wired to pack voltage. 0 if "
                   "everything runs off a BEC."),
    "wing_span": ("Wing tip to wing tip, in metres.",
                  "Trainer 1.2-1.8 m, surveyor 2-4 m."),
    "wing_area": ("Projected wing planform area in square metres.",
                  "Roughly span x average chord. A 1.5 m span with 0.2 m chord "
                  "is about 0.30 m2."),
    "CD0": ("Zero-lift drag coefficient — the drag the airframe makes even at "
            "zero lift. The single biggest driver of cruise power.",
            "Clean glider 0.015-0.025, typical foam model 0.03-0.05, draggy "
            "airframe with struts and fixed gear 0.05-0.08."),
    "CL_max": ("Maximum lift coefficient before the wing stalls. Sets stall speed.",
               "Flat-plate foam wing 0.8-1.0, cambered airfoil 1.2-1.5, with "
               "flaps 1.8-2.2."),
    "oswald": ("Oswald span efficiency — how close the wing is to an ideal "
               "elliptical lift distribution. Affects induced drag.",
               "Rectangular wing 0.7-0.8, tapered 0.8-0.9, elliptical ~0.95."),
    "mu_roll": ("Rolling friction coefficient during the takeoff roll.",
                "Paved 0.02-0.04, short grass 0.05-0.08, long grass 0.10+."),
    "mu_brake": ("Braking friction coefficient during the landing rollout.",
                 "No brakes (skid only) 0.1-0.2, wheel brakes on pavement "
                 "0.3-0.5."),
    "CL_takeoff": ("Lift coefficient held during the takeoff ground roll and "
                   "rotation.",
                   "Typically 0.7-0.9, below CL_max so you do not stall on "
                   "rotation."),
    "cruise_altitude": ("Height you actually fly at (m ASL). Used for the glide "
                        "distance estimate. Blank means the same as Altitude.",
                        "Altitude is your FIELD elevation; this is your cruise height."),
    "prop_eff_model": ("curve = propeller efficiency varies with advance ratio, "
                       "peaking near 60% of pitch speed (realistic). constant = the "
                       "flat value above at every speed.",
                       "Leave on curve unless comparing against older results."),
    "prop_eff": ("PEAK combined motor + propeller efficiency, reached near the prop's "
                 "design speed. Off-design speeds get less (see Prop Eff Model).",
                 "0.65-0.80 for a well matched setup. 0.75 is a fair default."),

    # ---- Battery ----
    "batt_unit_mode": ("cell = you specify one CELL and how many are wired "
                       "together. pack = you specify a whole PACK and how many "
                       "packs you wire together.",
                       "Use pack if you bought ready-made LiPos."),
    "batt_vmin": ("Lowest safe voltage PER CELL — the discharge cutoff.",
                  "LiPo 3.0-3.5, Li-ion 2.8-3.0."),
    "batt_vnom": ("Nominal voltage PER CELL. Used for energy in Wh.",
                  "LiPo 3.7, Li-ion 3.6."),
    "batt_vmax": ("Fully charged voltage PER CELL.", "LiPo 4.2, Li-ion 4.2."),
    "batt_cell_cap": ("Capacity of ONE cell in mAh. Cell mode only.",
                      "18650 = 3000-3500 mAh."),
    "batt_pack_cap": ("Capacity of ONE pack in mAh, from its label. Pack mode only.",
                      "Wiring packs in SERIES does not change this number."),
    "batt_cell_wt": ("Weight of ONE cell in grams.", "18650 ~45-50 g."),
    "batt_pack_wt": ("Weight of ONE pack in grams.", "4S 5000 mAh LiPo ~500 g."),
    "batt_series": ("Cells (or packs) wired in SERIES. This sets voltage.",
                    "4S LiPo = 4 x 4.2 = 16.8 V fully charged."),
    "batt_parallel": ("Cells (or packs) wired in PARALLEL. This sets capacity.",
                      "2 packs in parallel = double mAh, same voltage."),
    "batt_cells_s": ("Cells in series INSIDE one pack. Pack mode only.",
                     "A 4S LiPo has 4."),
    "batt_cells_p": ("Cells in parallel INSIDE one pack. Pack mode only.",
                     "Most hobby LiPos have 1."),
    "batt_dens": ("Optional override for Wh/kg. Blank = computed from capacity "
                  "and weight.",
                  "LiPo 130-200, Li-ion 200-260. Above 300 means an input is wrong."),
    "batt_dischg_pct": ("How much of the pack you actually use before landing.",
                        "80% is normal."),
    "batt_r": ("Internal resistance of ONE cell in milliohms. Drives voltage sag.",
               "Fresh LiPo 2-5 mOhm/cell, aged 8-15."),
    "batt_a_cont": ("Continuous current the pack can deliver, in amps.",
                    "Leave blank if you prefer to enter C-rate."),
    "batt_a_max": ("Burst current limit in amps.", "Usually 2x continuous."),
    "batt_c_cont": ("Continuous discharge C-rate from the label.",
                    "Amps = C-rate x capacity in Ah."),
    "batt_c_max": ("Burst C-rate from the label.", "Often optimistic."),
    "batt_chg": ("Max charge current in amps. Not used in flight physics.", "Typically 1C."),
    "batt_chem": ("Chemistry label. Also selects the SoC curve when SoC model "
                  "is set to auto.", "LiPo, Li-ion, LiFePO4."),
    "batt_soc_model": ("How pack voltage falls as the battery empties. auto picks "
                       "a curve from the chemistry above; linear anchors voltage "
                       "at full charge (the old behaviour).",
                       "auto, linear, lipo, liion, lifepo4. Leave on auto."),
    "batt_soc_curve_csv": ("Optional CSV of your own measured discharge curve, "
                           "which overrides the preset.",
                           "Columns: soc, ocv_cell, r_scale."),
    "batt_soc_bp": ("State-of-charge breakpoints for a custom curve. Accepts "
                    "0-1 fractions or percentages.",
                    "Advanced. Leave blank to use the preset."),
    "batt_ocv_cell_bp": ("Open-circuit voltage PER CELL at each breakpoint above.",
                         "Advanced. Must be the same length as the SoC list."),
    "batt_r_scale_bp": ("Resistance multiplier at each breakpoint. Cells get far "
                        "more resistive when nearly empty.",
                        "Advanced. 1.0 = the Rcell value you entered."),

    # ---- Motor ----
    "motor_kv": ("Motor RPM per volt with no load.",
                 "Printed on the motor. Fixed-wing 700-1200 is common."),
    "motor_r": ("Winding resistance in OHMS (not milliohms).",
                "Typical 0.02-0.2. If a datasheet gives mOhm, divide by 1000."),
    "motor_i0": ("No-load current spinning free with no prop.", "0.3-1.5 A."),
    "motor_v0": ("Voltage at which the no-load current was measured.",
                 "From the datasheet, usually 10 V."),
    "motor_rated_v": ("Voltage the manufacturer rates the motor for.", "Reference only."),
    "motor_imax": ("Maximum continuous current per motor, in amps.", "From the datasheet."),
    "motor_pmax": ("Maximum continuous power per motor, in watts.", "From the datasheet."),
    "motor_poles": ("Magnet pole count. Only affects the ERPM figure.",
                    "Most hobby outrunners are 14."),
    "motor_wt": ("Weight of ONE motor in grams.", "2826 ~80 g, 3536 ~200 g."),
    "motor_size": ("Stator size label, for reference.", "e.g. 2826."),

    # ---- ESC ----
    "esc_vrating": ("ESC voltage rating as a CELL COUNT (the S number), not volts.",
                    "A 6S ESC has a rating of 6. Must be >= your pack's series count."),
    "esc_cont": ("Continuous current per ESC in amps.", "A '60A ESC' means 60."),
    "esc_max": ("Burst current per ESC in amps.", "Usually 1.3-2x continuous."),
    "esc_idle": ("Current the ESC itself draws doing nothing.", "0.02-0.1 A."),
    "esc_r": ("ESC internal resistance in ohms.", "Typical 0.001-0.005."),
    "esc_wt": ("Weight of ONE ESC in grams.", "60 A ESC ~60 g."),

    # ---- Avionics ----
    "avionics_str": ("All non-motor electronics grouped by supply rail.",
                     "One row per rail. See the Avionics tab."),

    # ---- Propeller ----
    "prop_d": ("Propeller diameter in inches.", "First number: an 11x7 prop is 11 in."),
    "prop_pitch": ("Propeller pitch in inches — travel per revolution.",
                   "Second number: an 11x7 prop is 7 in. Higher pitch suits "
                   "faster cruise."),
    "prop_blades": ("Number of blades.", "2 is the efficient default for fixed-wing."),
    "prop_maxrpm": ("Manufacturer RPM limit, for a status check.", "0 = unknown."),
    "prop_maxthr": ("Maximum static thrust from one motor+prop, in grams.",
                    "From a thrust test table."),
    "prop_table": ("Optional CSV of measured thrust vs power. Much more accurate "
                   "than the theoretical model.",
                   "Columns: Thrust_g, Power_W, RPM."),
    "prop_tconst": ("Thrust coefficient C_T, if known.", "Advanced. Leave blank."),
    "prop_pconst": ("Power coefficient C_P, if known.", "Advanced. Leave blank."),
    "prop_wt": ("Weight of ONE propeller in grams.", "11x7 APC ~20 g."),

    # ---- Mission / environment ----
    "mission": ("Optional mission JSON describing a multi-phase flight.",
                "Use Browse to load one of the bundled examples."),
    "altitude": ("Altitude above sea level in metres. Also used as the height "
                 "for the glide-distance estimate.",
                 "Enter your cruise altitude, not field elevation, if you want a "
                 "meaningful glide distance."),
    "temp": ("Air temperature in Celsius. Blank = standard atmosphere.",
             "Hot air is thinner, so hot days cost thrust and raise stall speed."),
    "pressure": ("Air pressure in pascals. Blank = derived from altitude.",
                 "Sea level standard is 101325 Pa."),
    "wind": ("Wind speed in m/s.", "1 m/s = 2.24 mph = 1.94 knots."),
    "wind_dir": ("Direction the wind is coming FROM, in degrees.",
                 "0 = from the north, 90 = from the east."),
    "course_deg": ("The direction you are flying TOWARD, in degrees.",
                   "Matching wind direction gives a pure headwind."),
    "bank_deg": ("Bank angle for the turning-flight metrics.",
                 "0 = straight and level. 30 deg is a standard turn."),
    "climb_rate": ("Commanded climb rate in m/s. Costs extra power.", "0 for level flight."),
    "descent_rate": ("Commanded descent rate in m/s.", "0 for level flight."),
    "reserve_percent": ("Fraction of usable energy held back as a landing reserve.",
                        "20% is a common minimum."),
    "rth_reserve_Wh": ("Energy reserved for return-to-home, in watt-hours.",
                       "Estimate: cruise power x return time."),
    "diversion_reserve_Wh": ("Extra energy for diverting to another field.",
                             "0 unless operating under a flight plan."),
    "transient_dt_s": ("Integration step for the acceleration ramp during a "
                       "mission run. Smaller is more accurate and slower.",
                       "0.5 s is fine. Only used by Run Mission."),
    "max_accel_mps2": ("How hard the aircraft may accelerate when a mission "
                       "phase commands a higher speed. The ramp costs time, "
                       "distance and energy out of that phase's budget.",
                       "1.0-2.5 m/s2 for most fixed-wing models."),
    "max_decel_mps2": ("How hard the aircraft may decelerate when a phase "
                       "commands a lower speed.",
                       "1.5-3.0 m/s2, usually a little above the accel limit."),
    "decel_regen_eff": ("Fraction of braking energy recovered while slowing, "
                        "0 to 1.",
                        "Leave at 0 — a fixed-pitch propeller recovers almost "
                        "nothing while windmilling."),
        "max_v_plot": ("Highest speed shown on the performance charts.",
                   "Set above your expected top speed."),
}

# ------------------------------------------------------------------
# SIMPLE MODE FIELD SET
# ------------------------------------------------------------------
# The Avionics tab is deliberately included in Simple mode: knowing what
# the receiver, servos, autopilot and payload draw is essential to an
# honest endurance figure, and omitting it is a common beginner error.
# ------------------------------------------------------------------
FW_SIMPLE_FIELDS = {
    # Airframe
    "weight", "payload_mass_g", "num_motors", "cruise_speed",
    "wing_span", "wing_area", "CD0", "CL_max", "oswald",
    "mu_roll", "mu_brake", "CL_takeoff", "prop_eff", "prop_eff_model",
    # Battery
    "batt_unit_mode", "batt_vmin", "batt_vnom", "batt_vmax",
    "batt_cell_cap", "batt_pack_cap", "batt_cell_wt", "batt_pack_wt",
    "batt_series", "batt_parallel", "batt_cells_s", "batt_cells_p",
    "batt_dischg_pct", "batt_r", "batt_c_cont", "batt_a_cont", "batt_chem",
    "batt_soc_model",
    # Motor
    "motor_kv", "motor_r", "motor_i0", "motor_imax", "motor_pmax", "motor_wt",
    # ESC
    "esc_vrating", "esc_cont", "esc_max", "esc_wt",
    # Avionics (kept in Simple mode by request)
    "avionics_str",
    # Propeller
    "prop_d", "prop_pitch", "prop_blades", "prop_maxthr", "prop_table", "prop_wt",
    # Mission / environment
    "mission", "altitude", "cruise_altitude", "temp", "wind", "wind_dir",
    "periph_cur", "max_accel_mps2", "max_decel_mps2",
    "course_deg", "bank_deg",
    "reserve_percent", "max_v_plot",
}


def launch_gui():
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

    # ---------- helpers ----------
    def parse_float(label: str, val: str) -> Optional[float]:
        s = str(val).strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            raise ValueError(f"{label}: expected a number, got {s!r}")

    def parse_int(label: str, val: str) -> Optional[int]:
        s = str(val).strip()
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            raise ValueError(f"{label}: expected an integer, got {s!r}")

    def safe_float(val, default=0.0):
        try:
            return float(str(val).strip())
        except Exception:
            return default

    def choose_file(var, filetypes):
        p = filedialog.askopenfilename(filetypes=filetypes)
        if p:
            var.set(p)

    def exit_app():
        try:
            plt.close("all")
        except Exception:
            pass
        root.quit()
        root.destroy()

    # ---------- root window ----------
    root = tk.Tk()
    root.title(f"Fixed-Wing UAV Power Simulator  v{SIM_VERSION}")
    root.minsize(1100, 700)
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)

    # ------------------------------------------------------------------ #
    #  VIEW STATE — mutable container shared by all scaling callbacks     #
    # ------------------------------------------------------------------ #
    # Capture Tk's default DPI scaling so we can multiply from it cleanly.
    try:
        _base_tk_scale = float(root.tk.call("tk", "scaling"))
    except Exception:
        _base_tk_scale = 1.333          # 96 dpi / 72 pt — safe fallback

    _view = {
        "scale_pct":    100,            # window/widget DPI scale  (%)
        "plot_w":       15.0,           # matplotlib figure width  (inches)
        "plot_h":        9.0,           # matplotlib figure height (inches)
        "mpl_fontsize":  9,             # base font size for plots (pt)
        "ui_fontsize":   9,             # ttk widget font size     (pt)
    }

    # Holds the last successful run so we can re-render on plot-size change.
    _last_run: dict = {}   # keys: "cfg", "max_v", "metrics", "wind"

    # ------------------------------------------------------------------ #
    #  SCALING HELPERS                                                    #
    # ------------------------------------------------------------------ #
    def _apply_tk_scale(pct: int) -> None:
        """
        Scale all Tk/ttk widgets by adjusting Tk's DPI scaling factor.

        Tk uses a "scaling" value = pixels-per-point.  The default is
        typically screen_dpi / 72.  Multiplying it by (pct/100) makes
        every widget — fonts, buttons, entries, paddings — proportionally
        larger or smaller without any style surgery.

        We also nudge the root minsize so the window doesn't become
        unusably small at 75 % or clip content at 200 %.
        """
        _view["scale_pct"] = pct
        factor = _base_tk_scale * (pct / 100.0)
        try:
            root.tk.call("tk", "scaling", factor)
        except Exception:
            pass
        # Adjust minimum window size proportionally
        base_w, base_h = 1100, 700
        root.minsize(int(base_w * pct / 100), int(base_h * pct / 100))
        # Force geometry refresh so widgets reflow immediately
        root.update_idletasks()

    def _apply_ui_font(size: int) -> None:
        """
        Override the font size on every ttk style that carries text.
        Uses TkDefaultFont as the family so it follows the OS default.
        """
        _view["ui_fontsize"] = size
        sty = ttk.Style()
        font_spec = ("TkDefaultFont", size)
        for widget_style in (
            "TLabel", "TButton", "TEntry", "TCombobox",
            "TNotebook.Tab", "Treeview", "Treeview.Heading",
            "TLabelframe.Label", "TLabelframe",
        ):
            try:
                sty.configure(widget_style, font=font_spec)
            except Exception:
                pass
        # Treeview row height should track font size
        row_h = max(18, size + 8)
        try:
            sty.configure("Treeview", rowheight=row_h)
        except Exception:
            pass
        root.update_idletasks()

    def _apply_mpl_font(size: int) -> None:
        """
        Set matplotlib's global base font size, then re-render the
        current plot if one exists, so the change is visible immediately.
        """
        _view["mpl_fontsize"] = size
        import matplotlib as mpl
        mpl.rcParams.update({
            "font.size":        size,
            "axes.titlesize":   size + 1,
            "axes.labelsize":   size,
            "xtick.labelsize":  size - 1,
            "ytick.labelsize":  size - 1,
            "legend.fontsize":  size - 1,
        })
        _rerender_if_possible()

    def _apply_plot_size(w: float, h: float) -> None:
        """Store new figure dimensions and re-render immediately if possible."""
        _view["plot_w"] = w
        _view["plot_h"] = h
        _rerender_if_possible()

    def _rerender_if_possible() -> None:
        """Re-run the plot pipeline with the last known config, if any."""
        if not _last_run:
            return
        try:
            cfg   = _last_run["cfg"]
            max_v = _last_run["max_v"]
            m     = _last_run.get("metrics", {})
            fig   = make_performance_figure(
                cfg,
                max_speed=max_v,
                figsize=(_view["plot_w"], _view["plot_h"]),
            )
            # Generate motor operating point figure if available
            motor_fig = None
            if cfg.propeller.table is not None and m:
                try:
                    motor_fig = make_motor_operating_point_figure(cfg, m, figsize=(_view["plot_w"], 6))
                except Exception:
                    pass
            # Display both figures
            if motor_fig:
                show_figure([fig, motor_fig])
            else:
                show_figure(fig)
        except Exception:
            pass   # silently ignore re-render errors

    # ------------------------------------------------------------------ #
    #  MENU BAR                                                           #
    # ------------------------------------------------------------------ #
    menubar   = tk.Menu(root)
    file_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="File", menu=file_menu)

    # ---- View menu ----
    view_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="View", menu=view_menu)

    # Help > About — makes the running version unambiguous.
    help_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Help", menu=help_menu)
    help_menu.add_command(
        label="About / Version",
        command=lambda: messagebox.showinfo(
            "About",
            f"Fixed-Wing UAV Power Simulator\n"
            f"Version {SIM_VERSION}\n"
            f"{SIM_BUILD_NOTE}\n\n"
            "If you do not see the 'Input detail: Simple / Advanced' selector\n"
            "above the input tabs, or the blue ? help markers beside each\n"
            "field, you are running an older copy of this script."),
    )

    # -- Window Scale sub-menu --
    _scale_var = tk.IntVar(value=100)

    scale_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="Window Scale", menu=scale_menu)
    for pct, label in [(75, "75 %  – Compact"),
                       (90, "90 %  – Smaller"),
                       (100,"100 % – Default"),
                       (115,"115 % – Slightly Larger"),
                       (125,"125 % – Large"),
                       (150,"150 % – Extra Large"),
                       (175,"175 % – Very Large"),
                       (200,"200 % – Max")]:
        scale_menu.add_radiobutton(
            label=label,
            variable=_scale_var,
            value=pct,
            command=lambda p=pct: _apply_tk_scale(p),
        )

    view_menu.add_separator()

    # -- Plot Size sub-menu --
    _plot_size_var = tk.StringVar(value="medium")

    plot_size_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="Plot Size", menu=plot_size_menu)
    for key, label, (w, h) in [
        ("small",    "Small  (12 × 7)",   (12.0,  7.0)),
        ("medium",   "Medium (15 × 9)  ← default", (15.0,  9.0)),
        ("large",    "Large  (18 × 11)",  (18.0, 11.0)),
        ("xlarge",   "X-Large (22 × 13)", (22.0, 13.0)),
        ("xxlarge",  "XX-Large (26 × 15)",(26.0, 15.0)),
    ]:
        plot_size_menu.add_radiobutton(
            label=label,
            variable=_plot_size_var,
            value=key,
            command=lambda pw=w, ph=h: _apply_plot_size(pw, ph),
        )

    view_menu.add_separator()

    # -- UI Font Size sub-menu --
    _ui_font_var = tk.IntVar(value=9)

    ui_font_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="UI Font Size", menu=ui_font_menu)
    for sz, lbl in [(8,  "8 pt  – Tiny"),
                    (9,  "9 pt  – Default"),
                    (10, "10 pt – Comfortable"),
                    (11, "11 pt – Large"),
                    (13, "13 pt – Extra Large"),
                    (15, "15 pt – Accessibility")]:
        ui_font_menu.add_radiobutton(
            label=lbl,
            variable=_ui_font_var,
            value=sz,
            command=lambda s=sz: _apply_ui_font(s),
        )

    view_menu.add_separator()

    # -- Plot Font Size sub-menu --
    _mpl_font_var = tk.IntVar(value=9)

    mpl_font_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="Plot Font Size", menu=mpl_font_menu)
    for sz, lbl in [(7,  "7 pt  – Tiny"),
                    (8,  "8 pt  – Small"),
                    (9,  "9 pt  – Default"),
                    (10, "10 pt – Medium"),
                    (12, "12 pt – Large"),
                    (14, "14 pt – Extra Large")]:
        mpl_font_menu.add_radiobutton(
            label=lbl,
            variable=_mpl_font_var,
            value=sz,
            command=lambda s=sz: _apply_mpl_font(s),
        )

    view_menu.add_separator()

    # -- Quick preset commands --
    def _preset_compact():
        _scale_var.set(85);       _apply_tk_scale(85)
        _ui_font_var.set(8);      _apply_ui_font(8)
        _mpl_font_var.set(8);     _apply_mpl_font(8)
        _plot_size_var.set("small"); _apply_plot_size(12.0, 7.0)

    def _preset_default():
        _scale_var.set(100);      _apply_tk_scale(100)
        _ui_font_var.set(9);      _apply_ui_font(9)
        _mpl_font_var.set(9);     _apply_mpl_font(9)
        _plot_size_var.set("medium"); _apply_plot_size(15.0, 9.0)

    def _preset_presentation():
        _scale_var.set(140);      _apply_tk_scale(140)
        _ui_font_var.set(12);     _apply_ui_font(12)
        _mpl_font_var.set(12);    _apply_mpl_font(12)
        _plot_size_var.set("large"); _apply_plot_size(18.0, 11.0)

    def _preset_accessibility():
        _scale_var.set(160);      _apply_tk_scale(160)
        _ui_font_var.set(14);     _apply_ui_font(14)
        _mpl_font_var.set(13);    _apply_mpl_font(13)
        _plot_size_var.set("xlarge"); _apply_plot_size(22.0, 13.0)

    presets_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="Quick Presets", menu=presets_menu)
    presets_menu.add_command(label="🗜  Compact",          command=_preset_compact)
    presets_menu.add_command(label="⚙  Default",           command=_preset_default)
    presets_menu.add_command(label="📊  Presentation",     command=_preset_presentation)
    presets_menu.add_command(label="♿  Accessibility",     command=_preset_accessibility)

    view_menu.add_separator()
    view_menu.add_command(label="Reset All to Default",    command=_preset_default)

    root.config(menu=menubar)

    # ---------- main panes ----------
    main = ttk.Frame(root, padding=8)
    main.grid(sticky="nsew")
    main.columnconfigure(0, weight=1, minsize=370)
    main.columnconfigure(1, weight=3)
    main.rowconfigure(0, weight=1)

    # ===== LEFT: input notebook =====
    left = ttk.Frame(main)
    left.grid(row=0, column=0, sticky="nsew")
    left.columnconfigure(0, weight=1)
    # Row 0 = Simple/Advanced mode bar, row 1 = the input notebook.
    left.rowconfigure(1, weight=1)

    input_nb = ttk.Notebook(left)
    input_nb.grid(row=1, column=0, sticky="nsew")

    # Create scrollable tab helper
    # --- Scrollable-tab registry so the single wheel binding can find the
    #     active canvas without iterating or calling bind_all repeatedly.
    _tab_canvases: list = []

    def make_scrollable_tab(nb, title):
        """
        Wrap a notebook tab in a Canvas + vertical Scrollbar.

        Key fixes vs the original:
          1. canvas.<Configure> — fires when the canvas itself is resized
             (e.g. when the tab is first shown).  Without this, winfo_width()
             returns 1 at startup so the inner frame is rendered 1 px wide
             and appears blank until the user switches tabs twice.
          2. bind_all is NOT used here — a single wheel binding is attached
             to the notebook after all tabs are created (see below), routing
             scroll events only to the currently-visible canvas.  Calling
             bind_all inside a loop creates N overlapping handlers on every
             widget in the application, which stalls the Tk event loop and
             causes the multi-minute render delay.
        """
        outer = ttk.Frame(nb)
        nb.add(outer, text=title)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)
        cv  = tk.Canvas(outer, highlightthickness=0)
        sb  = ttk.Scrollbar(outer, orient="vertical", command=cv.yview)
        inn = ttk.Frame(cv)
        cv.configure(yscrollcommand=sb.set)
        cv.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        win_id = cv.create_window((0, 0), window=inn, anchor="nw")

        def _resize_inner(width):
            """Set the embedded-window width to match the canvas width."""
            if width > 1:                     # ignore the startup 1-px dummy value
                cv.itemconfig(win_id, width=width)

        def _on_inner_configure(evt):
            """Called when widgets are added to / removed from the inner frame."""
            cv.configure(scrollregion=cv.bbox("all"))
            _resize_inner(cv.winfo_width())

        def _on_canvas_configure(evt):
            """
            Called when the canvas is resized — critically, this fires when
            the tab is *first shown*, giving us the real pixel width.
            """
            cv.configure(scrollregion=cv.bbox("all"))
            _resize_inner(evt.width)

        inn.bind("<Configure>", _on_inner_configure)
        cv.bind("<Configure>",  _on_canvas_configure)   # ← the essential new line

        inn.columnconfigure(0, weight=1)
        inn.columnconfigure(1, weight=1)
        _tab_canvases.append(cv)   # register so wheel handler can find it
        return inn

    tab_airframe = make_scrollable_tab(input_nb, "Airframe")
    tab_batt     = make_scrollable_tab(input_nb, "Battery")
    tab_motor    = make_scrollable_tab(input_nb, "Motor")
    tab_esc      = make_scrollable_tab(input_nb, "ESC")
    tab_avionics = make_scrollable_tab(input_nb, "Avionics")
    tab_prop     = make_scrollable_tab(input_nb, "Propeller")
    tab_env      = make_scrollable_tab(input_nb, "Mission/Environment")
    # --- Single notebook-level mouse-wheel binding ----------------------------
    # Replaces the old bind_all-per-tab pattern.  We find whichever canvas
    # belongs to the currently-selected tab and scroll only that one.
    def _on_nb_mousewheel(evt):
        try:
            idx = input_nb.index(input_nb.select())
            if 0 <= idx < len(_tab_canvases):
                _tab_canvases[idx].yview_scroll(
                    int(-1 * (evt.delta / 120)), "units")
        except Exception:
            pass
    input_nb.bind("<MouseWheel>", _on_nb_mousewheel)
    # Also bind to each outer frame so the wheel works when hovering over
    # the tab content area (not just the notebook header strip).
    for _cv in _tab_canvases:
        _cv.bind("<MouseWheel>", lambda evt, c=_cv:
                 c.yview_scroll(int(-1 * (evt.delta / 120)), "units"))
    # --------------------------------------------------------------------------


    # ---------- StringVars ----------
    def sv(default=""):
        return tk.StringVar(value=str(default))

    # Airframe
    v_weight        = sv(2500)     # base aircraft grams (excluding payload)
    v_payload_mass  = sv(0)        # payload grams
    v_num_motors    = sv(1)
    v_wing_span     = sv(1.6)      # m
    v_wing_area     = sv(0.45)     # m²
    v_CD0           = sv(0.028)
    v_CL_max        = sv(1.30)
    v_oswald        = sv(0.82)
    v_mu_roll       = sv(0.04)
    v_mu_brake      = sv(0.08)
    v_CL_takeoff    = sv(0.80)
    v_prop_eff      = sv(0.75)
    v_cruise_speed  = sv(18.0)     # m/s

    # Battery
    v_batt_chem       = sv("LiPo")
    # --- SoC (state-of-charge) model, mirrors the multicopter simulator ---
    v_batt_soc_model      = sv("auto")
    v_batt_soc_curve_csv  = sv("")
    v_batt_soc_bp         = sv("")
    v_batt_ocv_cell_bp    = sv("")
    v_batt_r_scale_bp     = sv("")
    v_batt_vmin       = sv(3.5)
    v_batt_vnom       = sv(3.8)
    v_batt_vmax       = sv(4.2)
    v_batt_cell_cap   = sv(5000)
    v_batt_pack_cap   = sv("")
    v_batt_cell_wt    = sv(75)
    v_batt_pack_wt    = sv("")
    v_batt_dens       = sv("")
    v_batt_chg        = sv(5)
    v_batt_a_cont     = sv("")
    v_batt_a_max      = sv("")
    v_batt_c_cont     = sv(25)
    v_batt_c_max      = sv(50)
    v_batt_dischg_pct = sv(80)
    v_batt_r          = sv(5)
    v_batt_series     = sv(4)
    v_batt_parallel   = sv(1)
    v_batt_cells_s    = sv(1)
    v_batt_cells_p    = sv(1)
    v_batt_unit_mode  = sv("cell")

    # Motor
    v_motor_kv       = sv(920)
    v_motor_i0       = sv(0.8)
    v_motor_v0       = sv(7.0)
    v_motor_rated_v  = sv(16)
    v_motor_r        = sv(0.06)
    v_motor_imax     = sv(40)
    v_motor_pmax     = sv(500)
    v_motor_poles    = sv(14)
    v_motor_wt       = sv(120)
    v_motor_size     = sv("")

    # ESC
    v_esc_vrating  = sv("")
    v_esc_cont     = sv("")
    v_esc_max      = sv("")
    v_esc_idle     = sv("")
    v_esc_r        = sv("")
    v_esc_wt       = sv("")

    # Avionics
    v_avionics_str = sv("")

    # Propeller
    v_prop_d        = sv(10)
    v_prop_pitch    = sv(4.5)
    v_prop_blades   = sv(2)
    v_prop_maxrpm   = sv(0)
    v_prop_maxthr   = sv(0)
    v_prop_table    = sv("")
    v_prop_tconst   = sv("")
    v_prop_pconst   = sv("")
    v_prop_wt       = sv("20")

    # Environment / Mission
    v_mission          = sv("")
    v_altitude         = sv(0)
    v_cruise_altitude = sv("")
    v_prop_eff_model  = sv("curve")
    v_temp             = sv("")
    v_pressure         = sv("")
    v_wind             = sv(0)
    v_wind_dir         = sv(0)
    v_course_deg       = sv(0)
    v_bank_deg         = sv(0)
    v_climb_rate       = sv(0)
    v_descent_rate     = sv(0)
    v_reserve_percent  = sv(20)
    v_rth_reserve_Wh   = sv(0)
    v_div_reserve_Wh   = sv(0)
    v_max_v_plot       = sv(40)
    v_periph_cur       = sv(0.5)
    # Transient (acceleration) model — mirrors the multicopter simulator.
    v_transient_dt_s   = sv(0.5)
    v_max_accel_mps2   = sv(1.5)
    v_max_decel_mps2   = sv(2.0)
    v_decel_regen_eff  = sv(0.0)

    # UI complexity mode — "Simple" hides advanced inputs, "Advanced" shows all.
    # View setting only; it never changes a computed result.
    v_ui_mode = tk.StringVar(value="Simple")
    # Name of the configuration currently loaded, shown in the mode bar.
    # The Output pane is overwritten by the first run, so the loaded config
    # needs somewhere permanent to live.
    v_loaded_cfg = tk.StringVar(value="(no config loaded)")

    config_vars = dict(
        weight=v_weight, payload_mass_g=v_payload_mass, num_motors=v_num_motors,
        wing_span=v_wing_span, wing_area=v_wing_area,
        CD0=v_CD0, CL_max=v_CL_max, oswald=v_oswald,
        mu_roll=v_mu_roll, mu_brake=v_mu_brake, CL_takeoff=v_CL_takeoff,
        prop_eff=v_prop_eff, cruise_speed=v_cruise_speed,
        batt_chem=v_batt_chem,
        batt_soc_model     = v_batt_soc_model,
        batt_soc_curve_csv = v_batt_soc_curve_csv,
        batt_soc_bp        = v_batt_soc_bp,
        batt_ocv_cell_bp   = v_batt_ocv_cell_bp,
        batt_r_scale_bp    = v_batt_r_scale_bp, batt_vmin=v_batt_vmin,
        batt_vnom=v_batt_vnom, batt_vmax=v_batt_vmax,
        batt_cell_cap=v_batt_cell_cap, batt_pack_cap=v_batt_pack_cap,
        batt_cell_wt=v_batt_cell_wt, batt_pack_wt=v_batt_pack_wt,
        batt_dens=v_batt_dens, batt_chg=v_batt_chg,
        batt_a_cont=v_batt_a_cont, batt_a_max=v_batt_a_max,
        batt_c_cont=v_batt_c_cont, batt_c_max=v_batt_c_max,
        batt_dischg_pct=v_batt_dischg_pct, batt_r=v_batt_r,
        batt_series=v_batt_series, batt_parallel=v_batt_parallel,
        batt_cells_s=v_batt_cells_s, batt_cells_p=v_batt_cells_p,
        batt_unit_mode=v_batt_unit_mode,
        motor_kv=v_motor_kv, motor_i0=v_motor_i0, motor_v0=v_motor_v0,
        motor_rated_v=v_motor_rated_v, motor_r=v_motor_r,
        motor_imax=v_motor_imax, motor_pmax=v_motor_pmax,
        motor_poles=v_motor_poles, motor_wt=v_motor_wt,
        motor_size=v_motor_size,
        esc_vrating=v_esc_vrating, esc_cont=v_esc_cont,
        esc_max=v_esc_max, esc_idle=v_esc_idle,
        esc_r=v_esc_r, esc_wt=v_esc_wt,
        avionics_str=v_avionics_str,
        prop_d=v_prop_d, prop_pitch=v_prop_pitch, prop_blades=v_prop_blades,
        prop_maxrpm=v_prop_maxrpm, prop_maxthr=v_prop_maxthr,
        prop_table=v_prop_table, prop_tconst=v_prop_tconst,
        prop_pconst=v_prop_pconst, prop_wt=v_prop_wt,
        altitude=v_altitude,
        cruise_altitude = v_cruise_altitude,
        prop_eff_model  = v_prop_eff_model, temp=v_temp, pressure=v_pressure,
        mission=v_mission, wind=v_wind, wind_dir=v_wind_dir, course_deg=v_course_deg,
        bank_deg=v_bank_deg, climb_rate=v_climb_rate, descent_rate=v_descent_rate,
        reserve_percent=v_reserve_percent, rth_reserve_Wh=v_rth_reserve_Wh,
        diversion_reserve_Wh=v_div_reserve_Wh, max_v_plot=v_max_v_plot, periph_cur=v_periph_cur,
        transient_dt_s=v_transient_dt_s, max_accel_mps2=v_max_accel_mps2,
        max_decel_mps2=v_max_decel_mps2, decel_regen_eff=v_decel_regen_eff,
    )

    # ---- row helper ----
    # Registry of input rows so Simple/Advanced mode can show/hide them.
    _field_rows = []
    _section_rows = []

    def _register_row(key, widgets, parent, row):
        _field_rows.append({"key": key, "widgets": widgets,
                            "parent": parent, "row": row})

    def add_row(parent, row, label, var, key=None, **kwargs):
        """
        Add one labelled entry row with an inline "?" help marker.

        `key` links the row to FW_FIELD_HELP (tooltip text) and to
        FW_SIMPLE_FIELDS (Simple/Advanced visibility).  Rows without a key
        are always shown and carry no tooltip.
        """
        lbl = ttk.Label(parent, text=label)
        lbl.grid(row=row, column=0, sticky="w", padx=6, pady=3)
        e = ttk.Entry(parent, textvariable=var, width=14, **kwargs)
        e.grid(row=row, column=1, sticky="ew", padx=6, pady=3)

        widgets = [lbl, e]
        help_entry = FW_FIELD_HELP.get(key) if key else None
        if help_entry:
            what, typical = help_entry
            marker = ttk.Label(parent, text=" ? ", foreground="#0B6BCB",
                               cursor="question_arrow",
                               font=("TkDefaultFont", 9, "bold"))
            marker.grid(row=row, column=2, sticky="w", padx=(0, 6))
            _Tooltip(marker, f"{what}\n\nTypical: {typical}")
            _Tooltip(lbl, f"{what}\n\nTypical: {typical}")
            widgets.append(marker)

        if key:
            _register_row(key, widgets, parent, row)
        return e

    # ===== AIRFRAME TAB =====
    r = 0
    add_row(tab_airframe, r, "Base Aircraft Weight (g)",    v_weight, key="weight");       r += 1
    add_row(tab_airframe, r, "Payload Mass (g)",            v_payload_mass, key="payload_mass_g"); r += 1
    add_row(tab_airframe, r, "Number of Motors",            v_num_motors, key="num_motors");   r += 1
    add_row(tab_airframe, r, "Cruise Speed (m/s)",          v_cruise_speed, key="cruise_speed"); r += 1
    add_row(tab_airframe, r, "Peripheral Current (A)",      v_periph_cur, key="periph_cur");   r += 1
    ttk.Separator(tab_airframe, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_airframe, text="── Wing Geometry ──",
              foreground="gray").grid(row=r, column=0, columnspan=2, sticky="w", padx=6); r += 1
    add_row(tab_airframe, r, "Wing Span (m)",               v_wing_span, key="wing_span");    r += 1
    add_row(tab_airframe, r, "Wing Area (m²)",              v_wing_area, key="wing_area");    r += 1
    # chord is derived; shown in metrics
    ttk.Separator(tab_airframe, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_airframe, text="── Aerodynamics ──",
              foreground="gray").grid(row=r, column=0, columnspan=2, sticky="w", padx=6); r += 1
    add_row(tab_airframe, r, "CD0 (zero-lift drag coeff)", v_CD0, key="CD0");           r += 1
    add_row(tab_airframe, r, "CL_max",                     v_CL_max, key="CL_max");        r += 1
    add_row(tab_airframe, r, "Oswald Efficiency (e)",      v_oswald, key="oswald");        r += 1
    ttk.Separator(tab_airframe, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_airframe, text="── Takeoff / Ground Roll ──",
              foreground="gray").grid(row=r, column=0, columnspan=2, sticky="w", padx=6); r += 1
    add_row(tab_airframe, r, "Rolling Friction μ",         v_mu_roll, key="mu_roll");       r += 1
    add_row(tab_airframe, r, "Braking Friction μ",         v_mu_brake, key="mu_brake");      r += 1
    add_row(tab_airframe, r, "CL at Takeoff Rotation",     v_CL_takeoff, key="CL_takeoff");    r += 1
    ttk.Separator(tab_airframe, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_airframe, text="── Propulsion ──",
              foreground="gray").grid(row=r, column=0, columnspan=2, sticky="w", padx=6); r += 1
    add_row(tab_airframe, r, "Prop Efficiency η",          v_prop_eff, key="prop_eff");      r += 1
    add_row(tab_airframe, r, "Prop Eff Model",             v_prop_eff_model, key="prop_eff_model"); r += 1

    # ===== BATTERY TAB =====
    ttk.Label(tab_batt, text="Unit mode:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    unit_mode_cb = ttk.Combobox(tab_batt, textvariable=v_batt_unit_mode,
                                values=["cell","pack"], state="readonly", width=10)
    unit_mode_cb.grid(row=0, column=1, sticky="w", padx=6, pady=4)
    _um_help = FW_FIELD_HELP["batt_unit_mode"]
    _um_mark = ttk.Label(tab_batt, text=" ? ", foreground="#0B6BCB",
                         cursor="question_arrow", font=("TkDefaultFont", 9, "bold"))
    _um_mark.grid(row=0, column=2, sticky="w", padx=(0, 6))
    _Tooltip(_um_mark, f"{_um_help[0]}\n\nTypical: {_um_help[1]}")
    r = 1
    add_row(tab_batt, r, "Vmin/cell (V)",            v_batt_vmin, key="batt_vmin");      r += 1
    add_row(tab_batt, r, "Vnom/cell (V)",            v_batt_vnom, key="batt_vnom");      r += 1
    add_row(tab_batt, r, "Vmax/cell (V)",            v_batt_vmax, key="batt_vmax");      r += 1
    cell_cap_e = add_row(tab_batt, r, "Cell Capacity (mAh)",  v_batt_cell_cap, key="batt_cell_cap"); r += 1
    pack_cap_e = add_row(tab_batt, r, "Pack Capacity (mAh)",  v_batt_pack_cap, key="batt_pack_cap"); r += 1
    cell_wt_e  = add_row(tab_batt, r, "Cell Weight (g)",      v_batt_cell_wt, key="batt_cell_wt");  r += 1
    pack_wt_e  = add_row(tab_batt, r, "Pack Weight (g)",      v_batt_pack_wt, key="batt_pack_wt");  r += 1
    add_row(tab_batt, r, "Energy Density (Wh/kg)",   v_batt_dens, key="batt_dens");      r += 1
    add_row(tab_batt, r, "Max Charge Current (A)",   v_batt_chg, key="batt_chg");       r += 1
    add_row(tab_batt, r, "Cont Discharge Current (A)", v_batt_a_cont, key="batt_a_cont");  r += 1
    add_row(tab_batt, r, "Max Discharge Current (A)", v_batt_a_max, key="batt_a_max");    r += 1
    add_row(tab_batt, r, "Cont C-rate",              v_batt_c_cont, key="batt_c_cont");    r += 1
    add_row(tab_batt, r, "Max C-rate",               v_batt_c_max, key="batt_c_max");     r += 1
    add_row(tab_batt, r, "Usable Discharge (%)",     v_batt_dischg_pct, key="batt_dischg_pct"); r += 1
    add_row(tab_batt, r, "Rcell (mΩ)",               v_batt_r, key="batt_r");         r += 1
    add_row(tab_batt, r, "Series Cells/Packs",       v_batt_series, key="batt_series");    r += 1
    add_row(tab_batt, r, "Parallel Cells/Packs",     v_batt_parallel, key="batt_parallel");  r += 1
    cells_s_e = add_row(tab_batt, r, "Cells in series/pack",  v_batt_cells_s, key="batt_cells_s"); r += 1
    cells_p_e = add_row(tab_batt, r, "Cells in parallel/pack",v_batt_cells_p, key="batt_cells_p"); r += 1
    add_row(tab_batt, r, "Chemistry",                v_batt_chem, key="batt_chem");      r += 1
    add_row(tab_batt, r, "SoC model",                v_batt_soc_model, key="batt_soc_model");    r += 1
    add_row(tab_batt, r, "SoC curve CSV",            v_batt_soc_curve_csv, key="batt_soc_curve_csv"); r += 1
    add_row(tab_batt, r, "SoC breakpoints (0..1)",   v_batt_soc_bp, key="batt_soc_bp");       r += 1
    add_row(tab_batt, r, "OCV/cell breakpoints (V)", v_batt_ocv_cell_bp, key="batt_ocv_cell_bp");  r += 1
    add_row(tab_batt, r, "R-scale breakpoints",      v_batt_r_scale_bp, key="batt_r_scale_bp");   r += 1

    def on_unit_mode(event=None):
        mode = v_batt_unit_mode.get()
        if mode == "cell":
            cell_cap_e.configure(state="normal"); cell_wt_e.configure(state="normal")
            pack_cap_e.configure(state="disabled"); pack_wt_e.configure(state="disabled")
            cells_s_e.configure(state="disabled"); cells_p_e.configure(state="disabled")
        else:
            cell_cap_e.configure(state="disabled"); cell_wt_e.configure(state="disabled")
            pack_cap_e.configure(state="normal"); pack_wt_e.configure(state="normal")
            cells_s_e.configure(state="normal"); cells_p_e.configure(state="normal")
    unit_mode_cb.bind("<<ComboboxSelected>>", on_unit_mode)
    on_unit_mode()

    # ===== MOTOR TAB =====
    r = 0
    add_row(tab_motor, r, "Kv (RPM/V)",        v_motor_kv, key="motor_kv");    r += 1
    add_row(tab_motor, r, "Idle Current I0 (A)",v_motor_i0, key="motor_i0");   r += 1
    add_row(tab_motor, r, "Idle Voltage V0 (V)",v_motor_v0, key="motor_v0");   r += 1
    add_row(tab_motor, r, "Rated Voltage (V)",  v_motor_rated_v, key="motor_rated_v"); r += 1
    add_row(tab_motor, r, "Resistance Rm (Ω)",  v_motor_r, key="motor_r");    r += 1
    add_row(tab_motor, r, "Max Current (A)",    v_motor_imax, key="motor_imax"); r += 1
    add_row(tab_motor, r, "Max Power (W)",      v_motor_pmax, key="motor_pmax"); r += 1
    add_row(tab_motor, r, "Pole Count",         v_motor_poles, key="motor_poles"); r += 1
    add_row(tab_motor, r, "Weight (g)",         v_motor_wt, key="motor_wt");   r += 1
    add_row(tab_motor, r, "Size (e.g. 2826)",   v_motor_size, key="motor_size"); r += 1

    # ===== ESC TAB =====
    r = 0
    add_row(tab_esc, r, "Voltage Rating (S cells)",    v_esc_vrating, key="esc_vrating"); r += 1
    add_row(tab_esc, r, "Continuous Current (A)", v_esc_cont, key="esc_cont");   r += 1
    add_row(tab_esc, r, "Max Current (A)",        v_esc_max, key="esc_max");    r += 1
    add_row(tab_esc, r, "Idle Current (A)",       v_esc_idle, key="esc_idle");   r += 1
    add_row(tab_esc, r, "Resistance (Ω)",         v_esc_r, key="esc_r");      r += 1
    add_row(tab_esc, r, "Weight (g)",             v_esc_wt, key="esc_wt");     r += 1

    # ===== AVIONICS TAB =====
    # Each row = one BEC-regulated rail: voltage (V), current (A), efficiency (0–1].
    # The table is the single source of truth; v_avionics_str is kept in sync
    # so save/load and build_config can share one code path.

    tab_avionics.columnconfigure(0, weight=1)
    tab_avionics.rowconfigure(1, weight=1)   # treeview row expands

    # ---- header label ----
    ttk.Label(
        tab_avionics,
        text="BEC / Avionics voltage rails  —  one row per regulated output bus.\n"
             "Double-click any cell to edit it in-place.",
        wraplength=340, justify="left", foreground="#555555",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=6, pady=(6, 2))

    # ---- Treeview ----
    av_tree_frame = ttk.Frame(tab_avionics)
    av_tree_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=6, pady=(0, 4))
    av_tree_frame.columnconfigure(0, weight=1)
    av_tree_frame.rowconfigure(0, weight=1)

    _AV_COLS = ("voltage", "current", "efficiency")
    av_tree = ttk.Treeview(
        av_tree_frame,
        columns=_AV_COLS,
        show="headings",
        height=8,
        selectmode="browse",
    )
    for col, heading, width in [
        ("voltage",    "Rail Voltage (V)",      130),
        ("current",    "Rail Current (A)",      130),
        ("efficiency", "BEC Efficiency (0–1]",  150),
    ]:
        av_tree.heading(col, text=heading)
        av_tree.column(col, width=width, anchor="center", stretch=True)

    av_tree.grid(row=0, column=0, sticky="nsew")
    av_sb = ttk.Scrollbar(av_tree_frame, orient="vertical", command=av_tree.yview)
    av_sb.grid(row=0, column=1, sticky="ns")
    av_tree.configure(yscrollcommand=av_sb.set)

    # ---- helper: read all rows from treeview → voltage_tree dict ----
    def _av_tree_to_dict() -> dict:
        out = {}
        for iid in av_tree.get_children():
            vals = av_tree.item(iid, "values")
            try:
                v, i, e = float(vals[0]), float(vals[1]), float(vals[2])
                if v <= 0 or i < 0 or not (0 < e <= 1):
                    raise ValueError
                out[v] = (i, e)
            except Exception:
                raise ValueError(
                    f"Invalid avionics row {vals!r}.\n"
                    "Voltage > 0, Current ≥ 0, Efficiency in (0, 1]."
                )
        return out

    # ---- helper: serialise treeview → canonical string (kept in sync) ----
    def _sync_av_str():
        try:
            d = _av_tree_to_dict()
            v_avionics_str.set(
                ", ".join(f"{v}:({i},{e})" for v, (i, e) in sorted(d.items()))
            )
        except Exception:
            pass   # leave string var unchanged if table has a bad row mid-edit

    # ---- helper: load a list of {"voltage", "current", "eff"} dicts into tree ----
    def _av_load_rows(rows: list):
        av_tree.delete(*av_tree.get_children())
        for r in rows:
            try:
                av_tree.insert("", "end", values=(
                    str(r.get("voltage", "")),
                    str(r.get("current", "")),
                    str(r.get("eff", "")),
                ))
            except Exception:
                pass
        _sync_av_str()

    # ---- Double-click in-place cell editor ----
    def _av_begin_edit(event):
        region = av_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = av_tree.identify_row(event.y)
        col_id = av_tree.identify_column(event.x)   # "#1", "#2", "#3"
        if not row_id or not col_id:
            return
        col_idx = int(col_id.replace("#", "")) - 1
        bbox    = av_tree.bbox(row_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox
        old_vals = list(av_tree.item(row_id, "values"))

        ed = tk.Entry(av_tree, justify="center")
        ed.insert(0, old_vals[col_idx])
        ed.select_range(0, tk.END)
        ed.focus_set()
        ed.place(x=x, y=y, width=w, height=h)

        def _commit(_evt=None):
            new_val = ed.get().strip()
            old_vals[col_idx] = new_val
            av_tree.item(row_id, values=tuple(old_vals))
            ed.destroy()
            _sync_av_str()

        def _cancel(_evt=None):
            ed.destroy()

        ed.bind("<Return>",   _commit)
        ed.bind("<Tab>",      _commit)
        ed.bind("<FocusOut>", _commit)
        ed.bind("<Escape>",   _cancel)

    av_tree.bind("<Double-1>", _av_begin_edit)

    # ---- Add / edit row controls ----
    av_entry_frame = ttk.Frame(tab_avionics)
    av_entry_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=6, pady=(2, 2))
    for c in range(6):
        av_entry_frame.columnconfigure(c, weight=1)

    ttk.Label(av_entry_frame, text="Voltage (V):").grid(row=0, column=0, sticky="e", padx=(0, 2))
    _av_v_var = tk.StringVar()
    ttk.Entry(av_entry_frame, textvariable=_av_v_var, width=7).grid(row=0, column=1, sticky="ew", padx=(0, 6))

    ttk.Label(av_entry_frame, text="Current (A):").grid(row=0, column=2, sticky="e", padx=(0, 2))
    _av_i_var = tk.StringVar()
    ttk.Entry(av_entry_frame, textvariable=_av_i_var, width=7).grid(row=0, column=3, sticky="ew", padx=(0, 6))

    ttk.Label(av_entry_frame, text="Efficiency:").grid(row=0, column=4, sticky="e", padx=(0, 2))
    _av_e_var = tk.StringVar(value="0.90")
    ttk.Entry(av_entry_frame, textvariable=_av_e_var, width=7).grid(row=0, column=5, sticky="ew")

    # ---- Button row ----
    av_btn_frame = ttk.Frame(tab_avionics)
    av_btn_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=6, pady=(0, 6))

    def _av_add_or_update():
        try:
            v = float(_av_v_var.get())
            i = float(_av_i_var.get())
            e = float(_av_e_var.get())
        except ValueError:
            messagebox.showerror("Invalid input",
                "Please enter numeric values for Voltage, Current, and Efficiency.")
            return
        if v <= 0:
            messagebox.showerror("Invalid input", "Rail voltage must be > 0 V.")
            return
        if i < 0:
            messagebox.showerror("Invalid input", "Rail current must be ≥ 0 A.")
            return
        if not (0 < e <= 1.0):
            messagebox.showerror("Invalid input", "BEC efficiency must be in (0, 1].")
            return

        # If a row with this voltage already exists, update it in-place
        for iid in av_tree.get_children():
            existing = av_tree.item(iid, "values")
            try:
                if abs(float(existing[0]) - v) < 1e-9:
                    av_tree.item(iid, values=(f"{v:g}", f"{i:g}", f"{e:g}"))
                    _sync_av_str()
                    _av_v_var.set(""); _av_i_var.set(""); _av_e_var.set("0.90")
                    return
            except Exception:
                pass

        # Otherwise append a new row
        av_tree.insert("", "end", values=(f"{v:g}", f"{i:g}", f"{e:g}"))
        _sync_av_str()
        _av_v_var.set(""); _av_i_var.set(""); _av_e_var.set("0.90")

    def _av_remove_selected():
        sel = av_tree.selection()
        for iid in sel:
            av_tree.delete(iid)
        _sync_av_str()

    def _av_clear_all():
        av_tree.delete(*av_tree.get_children())
        _sync_av_str()

    # Clicking a row populates the entry fields for quick editing
    def _av_on_select(event):
        sel = av_tree.selection()
        if not sel:
            return
        vals = av_tree.item(sel[0], "values")
        try:
            _av_v_var.set(vals[0])
            _av_i_var.set(vals[1])
            _av_e_var.set(vals[2])
        except Exception:
            pass

    av_tree.bind("<<TreeviewSelect>>", _av_on_select)

    ttk.Button(av_btn_frame, text="➕  Add / Update Rail",
               command=_av_add_or_update).grid(row=0, column=0, sticky="w", padx=(0, 6))
    ttk.Button(av_btn_frame, text="🗑  Remove Selected",
               command=_av_remove_selected).grid(row=0, column=1, sticky="w", padx=(0, 6))
    ttk.Button(av_btn_frame, text="✖  Clear All",
               command=_av_clear_all).grid(row=0, column=2, sticky="w")

    # ---- Seed the table from v_avionics_str if it has a value ----
    try:
        _initial = parse_voltage_tree(v_avionics_str.get().strip() or None)
        _av_load_rows([{"voltage": v, "current": ci[0], "eff": ci[1]}
                       for v, ci in sorted(_initial.items())])
    except Exception:
        pass  # start with an empty table

    # ===== PROPELLER TAB =====
    r = 0
    add_row(tab_prop, r, "Diameter (in)",    v_prop_d, key="prop_d");      r += 1
    add_row(tab_prop, r, "Pitch (in)",       v_prop_pitch, key="prop_pitch");  r += 1
    add_row(tab_prop, r, "Blades",           v_prop_blades, key="prop_blades"); r += 1
    add_row(tab_prop, r, "Max RPM (0=auto)", v_prop_maxrpm, key="prop_maxrpm"); r += 1
    add_row(tab_prop, r, "Max Thrust (g)",   v_prop_maxthr, key="prop_maxthr"); r += 1
    ttk.Separator(tab_prop, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_prop, text="Prop/Motor CSV table (optional)").grid(
        row=r, column=0, sticky="w", padx=6, pady=2)
    frow = ttk.Frame(tab_prop); frow.grid(row=r, column=1, sticky="ew")
    frow.columnconfigure(0, weight=1)
    ttk.Entry(frow, textvariable=v_prop_table).grid(row=0, column=0, sticky="ew", padx=(6,4))
    ttk.Button(frow, text="Browse…",
               command=lambda: choose_file(v_prop_table, [("CSV","*.csv"),("All","*.*")])).grid(
        row=0, column=1, padx=(0,6)); r += 1
    add_row(tab_prop, r, "TConst (optional)", v_prop_tconst, key="prop_tconst"); r += 1
    add_row(tab_prop, r, "PConst (optional)", v_prop_pconst, key="prop_pconst"); r += 1
    add_row(tab_prop, r, "Weight (g)",        v_prop_wt, key="prop_wt");     r += 1

    # ===== ENVIRONMENT / MISSION TAB =====
    r = 0
    ttk.Label(tab_env, text="Mission JSON (optional)").grid(
        row=r, column=0, sticky="w", padx=6, pady=2)
    mrow = ttk.Frame(tab_env); mrow.grid(row=r, column=1, sticky="ew")
    mrow.columnconfigure(0, weight=1)
    ttk.Entry(mrow, textvariable=v_mission).grid(row=0, column=0, sticky="ew", padx=(6,4))
    ttk.Button(mrow, text="Browse…",
               command=lambda: choose_file(v_mission, [("JSON","*.json"),("All","*.*")])).grid(
        row=0, column=1, padx=(0,6)); r += 1
    ttk.Separator(tab_env, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    add_row(tab_env, r, "Altitude (m)",            v_altitude, key="altitude");    r += 1
    add_row(tab_env, r, "Cruise Altitude (m)",     v_cruise_altitude, key="cruise_altitude"); r += 1
    add_row(tab_env, r, "Temperature (°C, optional)", v_temp, key="temp");     r += 1
    add_row(tab_env, r, "Pressure (Pa, optional)", v_pressure, key="pressure");    r += 1
    add_row(tab_env, r, "Wind speed (m/s)",        v_wind, key="wind");        r += 1
    add_row(tab_env, r, "Wind direction FROM (deg)", v_wind_dir, key="wind_dir");  r += 1
    add_row(tab_env, r, "Course heading (deg)",    v_course_deg, key="course_deg");  r += 1
    add_row(tab_env, r, "Bank angle (deg)",        v_bank_deg, key="bank_deg");    r += 1
    add_row(tab_env, r, "Climb rate cmd (m/s)",    v_climb_rate, key="climb_rate");  r += 1
    add_row(tab_env, r, "Descent rate cmd (m/s)",  v_descent_rate, key="descent_rate"); r += 1
    add_row(tab_env, r, "Reserve percent (%)",     v_reserve_percent, key="reserve_percent"); r += 1
    add_row(tab_env, r, "RTH reserve (Wh)",        v_rth_reserve_Wh, key="rth_reserve_Wh"); r += 1
    add_row(tab_env, r, "Diversion reserve (Wh)",  v_div_reserve_Wh, key="diversion_reserve_Wh"); r += 1
    ttk.Separator(tab_env, orient="horizontal").grid(
        row=r, column=0, columnspan=3, sticky="ew", pady=6)
    ttk.Label(tab_env, text="—— Transients (mission runs) ——",
              foreground="gray").grid(row=r+1, column=0, columnspan=3, sticky="w", padx=6)
    r += 2
    add_row(tab_env, r, "Transient step dt (s)",       v_transient_dt_s,  key="transient_dt_s");  r += 1
    add_row(tab_env, r, "Max accel (m/s²)",            v_max_accel_mps2,  key="max_accel_mps2");  r += 1
    add_row(tab_env, r, "Max decel (m/s²)",            v_max_decel_mps2,  key="max_decel_mps2");  r += 1
    add_row(tab_env, r, "Decel regen efficiency (0-1)",v_decel_regen_eff, key="decel_regen_eff"); r += 1
    add_row(tab_env, r, "Max speed for plot (m/s)", v_max_v_plot, key="max_v_plot"); r += 1

    # ===== RIGHT: output panels =====
    right = ttk.Frame(main)
    right.grid(row=0, column=1, sticky="nsew", padx=(8,0))
    right.columnconfigure(0, weight=1)
    right.rowconfigure(0, weight=2)
    right.rowconfigure(1, weight=1)

    display_nb = ttk.Notebook(right)
    display_nb.grid(row=0, column=0, sticky="nsew")

    tab_plots        = ttk.Frame(display_nb, padding=0)
    tab_status       = ttk.Frame(display_nb, padding=0)
    tab_metrics      = ttk.Frame(display_nb, padding=0)
    tab_mission_plots= ttk.Frame(display_nb, padding=0)
    for t in (tab_plots, tab_status, tab_metrics, tab_mission_plots):
        t.columnconfigure(0, weight=1); t.rowconfigure(0, weight=1)
    display_nb.add(tab_plots,         text="Plots")
    display_nb.add(tab_status,        text="Status")
    display_nb.add(tab_metrics,       text="Metrics")
    display_nb.add(tab_mission_plots, text="Mission Plots")
    tab_weight_budget = ttk.Frame(display_nb, padding=0)
    tab_weight_budget.columnconfigure(0, weight=1)
    tab_weight_budget.rowconfigure(0, weight=1)
    display_nb.add(tab_weight_budget, text="Weight Budget")

    # ---- Airframe Diagram tab ----------------------------------------
    # A plan-view sketch drawn from the entered dimensions. Its job is to make
    # propeller overlap and tip clearance obvious at a glance, which numbers
    # in a table do not.
    tab_airframe_diagram = ttk.Frame(display_nb, padding=0)
    tab_airframe_diagram.columnconfigure(0, weight=1)
    tab_airframe_diagram.rowconfigure(0, weight=1)
    display_nb.add(tab_airframe_diagram, text="Airframe Diagram")

    ad_frame = ttk.LabelFrame(tab_airframe_diagram,
                              text="Plan View (to scale)", padding=4)
    ad_frame.grid(row=0, column=0, sticky="nsew")
    ad_frame.columnconfigure(0, weight=1)
    ad_frame.rowconfigure(0, weight=1)
    ad_holder = ttk.Frame(ad_frame)
    ad_holder.grid(row=0, column=0, sticky="nsew")
    ad_holder.columnconfigure(0, weight=1)
    ad_holder.rowconfigure(0, weight=1)
    ad_placeholder = ttk.Label(
        ad_holder, text="Press Run Single-Point to draw the airframe.",
        foreground="#888888")
    ad_placeholder.grid(row=0, column=0)
    _ad_canvas = {"widget": None}

    # ================================================================
    # SENSITIVITY  and  CONFIG COMPARISON tabs
    # ================================================================
    # Both work off the configuration from the most recent run, so they never
    # re-read the input fields and cannot disagree with what is on screen.

    tab_sensitivity = ttk.Frame(display_nb, padding=0)
    tab_sensitivity.columnconfigure(0, weight=1)
    tab_sensitivity.rowconfigure(1, weight=1)
    display_nb.add(tab_sensitivity, text="Sensitivity")

    sens_bar = ttk.Frame(tab_sensitivity, padding=(6, 6, 6, 0))
    sens_bar.grid(row=0, column=0, sticky="ew")
    ttk.Label(sens_bar, text="Output:").pack(side="left")
    v_sens_metric = tk.StringVar(value="Flight time (min)")
    ttk.Combobox(sens_bar, textvariable=v_sens_metric, state="readonly", width=22,
                 values=["Flight time (min)", "Range (km)",
                         "Total power (W)"]).pack(side="left", padx=(4, 10))
    ttk.Label(sens_bar, text="Vary each input by ±10% and ±20%",
              foreground="#666666", font=("TkDefaultFont", 8)).pack(side="left")

    sens_body = ttk.Frame(tab_sensitivity, padding=4)
    sens_body.grid(row=1, column=0, sticky="nsew")
    sens_body.columnconfigure(0, weight=1)
    sens_body.rowconfigure(0, weight=1)
    sens_body.rowconfigure(1, weight=1)

    _sens_cols = ("param", "m20", "m10", "base", "p10", "p20", "span")
    sens_tv = ttk.Treeview(sens_body, columns=_sens_cols, show="headings", height=8)
    for col, heading, width in [
            ("param", "Input", 190), ("m20", "-20%", 85), ("m10", "-10%", 85),
            ("base", "baseline", 85), ("p10", "+10%", 85), ("p20", "+20%", 85),
            ("span", "swing", 95)]:
        sens_tv.heading(col, text=heading)
        sens_tv.column(col, width=width,
                       anchor="w" if col == "param" else "center", stretch=True)
    sens_tv.grid(row=0, column=0, sticky="nsew")
    sens_tv.tag_configure("strong", font=("TkDefaultFont", 9, "bold"))

    sens_plot = ttk.Frame(sens_body)
    sens_plot.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
    sens_plot.columnconfigure(0, weight=1)
    sens_plot.rowconfigure(0, weight=1)
    sens_ph = ttk.Label(sens_plot,
                        text="Run a single point, then press Run Sensitivity.",
                        foreground="#888888")
    sens_ph.grid(row=0, column=0)
    _sens_canvas = {"widget": None}

    def _sens_levers():
        """
        Inputs worth testing, each with a mutator that scales it on a copy.

        Aerodynamic cleanliness (CD0, Oswald) is included because on a
        fixed-wing it usually dominates, which is exactly the sort of thing a
        tornado chart should make obvious.
        """
        def _scale_capacity(cfg, f):
            cfg.battery.capacity_mAh *= f
            cfg.battery.capacity_Ah *= f          # Wh derives from this

        def _scale_wing_area(cfg, f):
            cfg.airframe.wing_area_m2 *= f

        return [
            ("All-up weight",     lambda c, f: setattr(c, "aircraft_weight_g", c.aircraft_weight_g * f)),
            ("Battery capacity",  _scale_capacity),
            ("CD0 (parasite)",    lambda c, f: setattr(c.airframe, "CD0", c.airframe.CD0 * f)),
            ("Wing area",         _scale_wing_area),
            ("Oswald efficiency", lambda c, f: setattr(c.airframe, "oswald", min(c.airframe.oswald * f, 1.0))),
            ("Prop efficiency",   lambda c, f: setattr(c.airframe, "prop_efficiency", min(c.airframe.prop_efficiency * f, 0.95))),
            ("Cruise speed",      lambda c, f: setattr(c, "cruise_speed_mps", c.cruise_speed_mps * f)),
            ("Motor resistance",  lambda c, f: setattr(c.motor, "resistance", c.motor.resistance * f)),
        ]

    def run_sensitivity():
        if not _last_run.get("cfg"):
            messagebox.showinfo("Sensitivity", "Run a single point first.")
            return
        base_cfg = _last_run["cfg"]
        speed = float(_last_run.get("speed", base_cfg.cruise_speed_mps))
        choice = v_sens_metric.get()

        def evaluate(cfg):
            cfg = base_cfg if cfg is None else cfg
            spd = float(getattr(cfg, "cruise_speed_mps", speed) or speed)
            try:
                m = compute_metrics(cfg, spd)
                if choice.startswith("Flight time"):
                    return float(m["flight_time_min"])
                if choice.startswith("Range"):
                    return float(m["flight_range_km"])
                return float(m["total_power_W"])
            except Exception:
                return None

        evaluate.base_config = base_cfg
        rows = core.sensitivity_sweep(_sens_levers(), evaluate)

        for iid in sens_tv.get_children():
            sens_tv.delete(iid)
        if not rows:
            messagebox.showinfo("Sensitivity",
                                "The baseline configuration produced no usable result.")
            return

        decimals = 0 if choice.startswith("Total power") else 2
        for i, row in enumerate(rows):
            res = row["results"]

            def cell(f):
                v = res.get(f)
                return "—" if v is None else f"{v:.{decimals}f}"

            sens_tv.insert("", "end", tags=("strong",) if i == 0 else (), values=(
                row["name"], cell(0.8), cell(0.9), f"{row['baseline']:.{decimals}f}",
                cell(1.1), cell(1.2),
                f"{row['span']:.{decimals}f}  ({row['span_pct']:.0f}%)"))

        fig = Figure(figsize=(max(_view["plot_w"] * 0.7, 6.0), 4.2))
        ax = fig.add_subplot(111)
        names = [r["name"] for r in rows][::-1]
        lows = [r["low"] - r["baseline"] for r in rows][::-1]
        highs = [r["high"] - r["baseline"] for r in rows][::-1]
        ypos = range(len(names))
        ax.barh(list(ypos), [h - l for l, h in zip(lows, highs)],
                left=lows, color="#A5D6A7", edgecolor="#2E7D32")
        ax.axvline(0.0, color="#37474F", linewidth=1.2)
        ax.set_yticks(list(ypos))
        ax.set_yticklabels(names, fontsize=8)
        ax.set_xlabel(f"change in {choice} vs baseline")
        ax.set_title("Sensitivity (±20% on each input)", fontsize=10)
        ax.grid(True, axis="x", linestyle=":", alpha=0.4)
        fig.tight_layout()

        old = _sens_canvas.get("widget")
        if old is not None:
            try:
                old.get_tk_widget().destroy()
            except Exception:
                pass
        sens_ph.grid_remove()
        canvas = FigureCanvasTkAgg(fig, master=sens_plot)
        canvas.draw()
        canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        _sens_canvas["widget"] = canvas

    ttk.Button(sens_bar, text="⚡  Run Sensitivity",
               command=run_sensitivity).pack(side="right")

    # ---------------- Config comparison ----------------
    tab_compare = ttk.Frame(display_nb, padding=0)
    tab_compare.columnconfigure(0, weight=1)
    tab_compare.rowconfigure(1, weight=1)
    display_nb.add(tab_compare, text="Compare")

    cmp_bar = ttk.Frame(tab_compare, padding=(6, 6, 6, 0))
    cmp_bar.grid(row=0, column=0, sticky="ew")
    v_cmp_baseline = tk.StringVar(value="No baseline pinned")
    ttk.Label(cmp_bar, text="Baseline:").pack(side="left")
    ttk.Label(cmp_bar, textvariable=v_cmp_baseline, foreground="#0B6BCB",
              font=("TkDefaultFont", 9, "bold")).pack(side="left", padx=(4, 12))

    _cmp_cols = ("metric", "base", "curr", "delta", "pct")
    cmp_tv = ttk.Treeview(tab_compare, columns=_cmp_cols, show="headings", height=18)
    for col, heading, width in [("metric", "Metric", 220), ("base", "Baseline", 110),
                                ("curr", "Current", 110), ("delta", "Change", 110),
                                ("pct", "Change %", 100)]:
        cmp_tv.heading(col, text=heading)
        cmp_tv.column(col, width=width,
                      anchor="w" if col == "metric" else "center", stretch=True)
    cmp_tv.grid(row=1, column=0, sticky="nsew", padx=4, pady=4)
    cmp_tv.tag_configure("better", foreground="#1B5E20")
    cmp_tv.tag_configure("worse", foreground="#B71C1C")
    cmp_tv.tag_configure("flat", foreground="#616161")

    _cmp_state = {"baseline": None, "label": None}

    # (key, label, decimals, higher_is_better)
    _CMP_KEYS = [
        ("flight_time_min",  "Flight time (min)",    2,  1),
        ("flight_range_km",  "Range (km)",           2,  1),
        ("total_power_W",    "Total power (W)",      1, -1),
        ("pack_current_A",   "Pack current (A)",     2, -1),
        ("v_load_V",         "Loaded voltage (V)",   2,  1),
        ("LD_ratio",         "L/D ratio",            2,  1),
        ("CL",               "CL at cruise",         3, -1),
        ("stall_speed_mps",  "Stall speed (m/s)",    2, -1),
        ("thrust_required_N", "Thrust required (N)", 2, -1),
        ("rate_of_climb_mps", "Rate of climb (m/s)", 2,  1),
        ("wing_loading_N_m2", "Wing loading (N/m²)", 1, -1),
        ("reserve_margin_Wh", "Reserve margin (Wh)", 2,  1),
    ]

    def _current_comparison_metrics():
        return dict(_last_run.get("metrics") or {}) or None

    def pin_comparison_baseline():
        metrics = _current_comparison_metrics()
        if not metrics:
            messagebox.showinfo("Compare", "Run a single point first.")
            return
        _cmp_state["baseline"] = metrics
        label = str(v_loaded_cfg.get())
        _cmp_state["label"] = label
        v_cmp_baseline.set(f"{label}  (pinned)")
        refresh_comparison()

    def clear_comparison_baseline():
        _cmp_state["baseline"] = None
        _cmp_state["label"] = None
        v_cmp_baseline.set("No baseline pinned")
        for iid in cmp_tv.get_children():
            cmp_tv.delete(iid)

    def refresh_comparison():
        """Redraw the delta table. Called after every run."""
        for iid in cmp_tv.get_children():
            cmp_tv.delete(iid)
        base = _cmp_state.get("baseline")
        if not base:
            return
        current = _current_comparison_metrics()
        if not current:
            return
        rows = core.compare_metric_sets(
            base, current, [(k, lbl, d) for k, lbl, d, _ in _CMP_KEYS])
        better_map = {k: s for k, _, _, s in _CMP_KEYS}
        for row in rows:
            if not row["comparable"]:
                cmp_tv.insert("", "end", tags=("flat",), values=(
                    row["label"], "—", "—", "—", "—"))
                continue
            d = row["decimals"]
            sign = better_map.get(row["key"], 1) * row["direction"]
            tag = "better" if sign > 0 else ("worse" if sign < 0 else "flat")
            cmp_tv.insert("", "end", tags=(tag,), values=(
                row["label"],
                f"{row['baseline']:.{d}f}", f"{row['current']:.{d}f}",
                core.format_delta(row["delta"], d),
                core.format_delta(row["delta_pct"], 1) + " %"
                if row["delta_pct"] is not None else "—"))

    ttk.Button(cmp_bar, text="📌  Pin Current as Baseline",
               command=pin_comparison_baseline).pack(side="right")
    ttk.Button(cmp_bar, text="✖  Clear Baseline",
               command=clear_comparison_baseline).pack(side="right", padx=(0, 6))

    def _refresh_airframe_diagram(cfg_obj):
        """Redraw the plan view for the configuration just simulated."""
        try:
            fig = make_airframe_diagram_figure(
                cfg_obj, figsize=(max(_view["plot_w"] * 0.7, 6.0), 6.0))
        except Exception as exc:
            ad_placeholder.configure(text=f"Could not draw the airframe: {exc}")
            ad_placeholder.grid(row=0, column=0)
            return
        old = _ad_canvas.get("widget")
        if old is not None:
            try:
                old.get_tk_widget().destroy()
            except Exception:
                pass
        ad_placeholder.grid_remove()
        canvas = FigureCanvasTkAgg(fig, master=ad_holder)
        canvas.draw()
        canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        _ad_canvas["widget"] = canvas


    # ---- Plots panel ----
    plot_frame = ttk.LabelFrame(tab_plots, text="Performance Plots", padding=4)
    plot_frame.grid(row=0, column=0, sticky="nsew")
    plot_frame.columnconfigure(0, weight=1)
    plot_frame.rowconfigure(0, weight=1)

    plot_canvas = tk.Canvas(plot_frame, highlightthickness=0)
    plot_canvas.grid(row=0, column=0, sticky="nsew")
    plot_scroll = ttk.Scrollbar(plot_frame, orient="vertical", command=plot_canvas.yview)
    plot_scroll.grid(row=0, column=1, sticky="ns")
    plot_canvas.configure(yscrollcommand=plot_scroll.set)

    plot_inner = ttk.Frame(plot_canvas)
    plot_inner_id = plot_canvas.create_window((0, 0), window=plot_inner, anchor="nw")
    plot_inner.columnconfigure(0, weight=1)

    def _update_plot_scrollregion(event=None):
        plot_canvas.configure(scrollregion=plot_canvas.bbox("all"))

    def _match_plot_inner_width(event):
        plot_canvas.itemconfigure(plot_inner_id, width=event.width)

    plot_inner.bind("<Configure>", lambda event: _update_plot_scrollregion(event))
    plot_canvas.bind("<Configure>", lambda event: _match_plot_inner_width(event))
    plot_inner.bind("<Enter>", lambda _: plot_canvas.bind_all("<MouseWheel>", _on_plot_mousewheel))
    plot_inner.bind("<Leave>", lambda _: plot_canvas.unbind_all("<MouseWheel>"))

    _plot_canvases = []  # Track multiple figure canvases
    _current_plot_figs = []  # Track multiple figures

    def _on_plot_mousewheel(evt):
        plot_canvas.yview_scroll(int(-1 * (evt.delta / 120)), "units")
        return "break"

    plot_canvas.bind("<Enter>", lambda _: plot_canvas.bind_all("<MouseWheel>", _on_plot_mousewheel))
    plot_canvas.bind("<Leave>", lambda _: plot_canvas.unbind_all("<MouseWheel>"))

    def show_figure(fig_or_figs):
        """Display one or more figures in the scrollable plot area."""
        # Clear previous figures
        for canvas in _plot_canvases:
            try:
                canvas.get_tk_widget().destroy()
            except:
                pass
        for fig in _current_plot_figs:
            try:
                plt.close(fig)
            except:
                pass
        _plot_canvases.clear()
        _current_plot_figs.clear()
        
        # Handle both single figure and list of figures
        figures = fig_or_figs if isinstance(fig_or_figs, list) else [fig_or_figs]
        
        for idx, fig in enumerate(figures):
            if fig is None:
                continue
            fc = FigureCanvasTkAgg(fig, master=plot_inner)
            fc.draw()
            fc.get_tk_widget().grid(row=idx, column=0, sticky="nsew")
            _plot_canvases.append(fc)
            _current_plot_figs.append(fig)
        
        # Update scroll region
        plot_inner.update_idletasks()
        plot_canvas.configure(scrollregion=plot_canvas.bbox("all"))

    # ---- Mission Plots panel (matches multicopter layout) ----
    # Series data stored as a one-element list so closures capture the reference.
    last_mission_series = [None]

    # Variables: (series_key, display_label, y-axis_unit)
    # Variables that share the same unit string are plotted on the same y-axis.
    MISSION_VARS = [
        ("airspeed_mps",      "Airspeed",                 "m/s"),
        ("groundspeed_mps",   "Groundspeed",              "m/s"),
        ("headwind_mps",      "Headwind",                 "m/s"),
        ("crosswind_mps",     "Crosswind",                "m/s"),
        ("altitude_m",        "Altitude",                 "m"),
        ("bank_deg",          "Bank angle",               "deg"),
        ("load_factor",       "Load factor",              "g"),
        ("turn_radius_m",     "Turn radius",              "m"),
        ("stall_turn_mps",    "Turn stall speed",         "m/s"),
        ("climb_rate_cmd_mps","Climb rate command",       "m/s"),
        ("descent_rate_cmd_mps","Descent rate command",   "m/s"),
        ("climb_power_add_W", "Climb/descent power",      "W"),
        ("battery_voltage_V", "Battery voltage (loaded)", "V"),
        ("battery_current_A", "Battery current",          "A"),
        ("battery_energy_Wh", "Battery energy remaining", "Wh"),
        ("reserve_target_Wh", "Reserve target",           "Wh"),
        ("reserve_margin_Wh", "Reserve margin",           "Wh"),
        ("reserve_breach",    "Reserve breach flag",      "bool"),
        ("total_power_W",     "Total power",              "W"),
        ("motor_power_W",     "Motor power",              "W"),
        ("drag_N",            "Drag force",               "N"),
        ("thrust_avail_N",    "Thrust available",         "N"),
        ("rate_of_climb_mps", "Rate of climb",            "m/s"),
        ("lift_drag_ratio",   "L/D ratio",                "—"),
        ("cl_cruise",         "CL at cruise speed",       "—"),
        ("motor_temp_est_C",  "Motor temp estimate",      "°C"),
        ("esc_temp_est_C",    "ESC temp estimate",        "°C"),
        ("battery_temp_est_C","Battery temp estimate",    "°C"),
    ]

    # Two-column layout: left controls, right canvas  (same as multicopter)
    mission_container = ttk.Frame(tab_mission_plots, padding=4)
    mission_container.grid(row=0, column=0, sticky="nsew")
    mission_container.columnconfigure(0, weight=0)
    mission_container.columnconfigure(1, weight=1)
    mission_container.rowconfigure(0, weight=1)

    mission_controls   = ttk.LabelFrame(mission_container, text="Y-axis variables", padding=4)
    mission_controls.grid(row=0, column=0, sticky="ns", padx=(0, 8))
    mission_plot_frame = ttk.LabelFrame(mission_container, text="Mission plot", padding=4)
    mission_plot_frame.grid(row=0, column=1, sticky="nsew")
    mission_plot_frame.columnconfigure(0, weight=1)
    mission_plot_frame.rowconfigure(0, weight=1)

    ttk.Label(mission_controls,
              text="Select variables to plot vs mission time.").grid(
        row=0, column=0, sticky="w")

    mission_var_list = tk.Listbox(mission_controls, selectmode="extended",
                                  height=16, exportselection=False)
    mission_var_list.grid(row=1, column=0, sticky="nsew", pady=(4, 4))
    mission_controls.rowconfigure(1, weight=1)
    mission_controls.columnconfigure(0, weight=1)

    ml_sb = ttk.Scrollbar(mission_controls, orient="vertical",
                           command=mission_var_list.yview)
    ml_sb.grid(row=1, column=1, sticky="ns", pady=(4, 4))
    mission_var_list.configure(yscrollcommand=ml_sb.set)

    _mission_items = []
    for _k, _lbl, _unit in MISSION_VARS:
        mission_var_list.insert(tk.END, f"{_lbl} ({_unit})")
        _mission_items.append((_k, _lbl, _unit))

    mission_canvas_ref = [None]

    def _clear_mission_plot():
        for w in mission_plot_frame.winfo_children():
            w.destroy()
        mission_canvas_ref[0] = None

    def _update_mission_plot():
        ms = last_mission_series[0]
        if ms is None:
            messagebox.showinfo("Mission plot", "Run a mission first.")
            return
        sel = list(mission_var_list.curselection())
        if not sel:
            messagebox.showinfo("Mission plot", "Select at least one variable.")
            return

        t_min = [x / 60.0 for x in ms.get("t_s", [])]
        if not t_min:
            return

        # Group selected variables by unit so same-unit series share a y-axis
        selected = [_mission_items[i] for i in sel]
        by_unit: dict = {}
        for k, lbl, unit in selected:
            by_unit.setdefault(unit, []).append((k, lbl))

        import matplotlib.pyplot as _plt
        fig = _plt.Figure(figsize=(7.5, 4.5), dpi=100)
        ax0 = fig.add_subplot(111)
        unit_list = list(by_unit.keys())

        axes = [(ax0, unit_list[0])]
        ax0.set_ylabel(unit_list[0])
        for ui, unit in enumerate(unit_list[1:], 1):
            axn = ax0.twinx()
            axn.spines["right"].set_position(("outward", 55 * (ui - 1)))
            axn.set_ylabel(unit)
            axes.append((axn, unit))

        lines, labels = [], []
        for ax, unit in axes:
            for key, lbl in by_unit.get(unit, []):
                y = ms.get(key, [])
                if not y:
                    continue
                yy = []
                for v in y:
                    try:
                        yy.append(float("nan") if isinstance(v, float) and v != v
                                  else float(v))
                    except Exception:
                        yy.append(float("nan"))
                ln, = ax.plot(t_min, yy, label=lbl)
                lines.append(ln)
                labels.append(lbl)

        ax0.set_xlabel("Mission time (min)")
        ax0.grid(True)
        fig.suptitle("Mission variables vs time")
        if lines:
            ax0.legend(lines, labels, loc="best")

        _clear_mission_plot()
        mc = FigureCanvasTkAgg(fig, master=mission_plot_frame)
        mc.draw()
        mc.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        mission_canvas_ref[0] = mc

    mission_btns = ttk.Frame(mission_controls)
    mission_btns.grid(row=2, column=0, columnspan=2, sticky="ew")
    ttk.Button(mission_btns, text="Plot selected",
               command=_update_mission_plot).grid(row=0, column=0, padx=(0, 4))
    ttk.Button(mission_btns, text="Clear",
               command=_clear_mission_plot).grid(row=0, column=1)

    def _on_list_wheel(evt):
        if evt.delta:
            mission_var_list.yview_scroll(int(-1 * (evt.delta / 120)), "units")
        return "break"
    mission_var_list.bind("<MouseWheel>", _on_list_wheel)

    # ---- log helper (appends to output text instead of replacing) ----
    def log(msg: str):
        out_text.configure(state="normal")
        out_text.insert("end", msg + "\n")
        out_text.see("end")
        out_text.configure(state="disabled")

    def clear_log():
        out_text.configure(state="normal")
        out_text.delete("1.0", "end")
        out_text.configure(state="disabled")


    # ---- Weight Budget panel ----
    _wb_canvas_ref = [None]

    def _draw_weight_chart(rows):
        import matplotlib.pyplot as _plt
        data_rows = [r for r in rows if r[0] != "TOTAL"]
        if not data_rows:
            return
        labels = [r[0] for r in data_rows]
        totals = [r[3] for r in data_rows]
        grand  = sum(totals)
        COLORS = ["#2E75B6","#ED7D31","#A9D18E","#FFC000","#5B9BD5","#FF7F7F"]
        fig, axes = _plt.subplots(1, 2, figsize=(7, max(3, len(labels)*0.6+1)))
        fig.patch.set_facecolor("white")
        ax = axes[0]
        left_ = 0.0
        pcts = [t/grand*100 if grand>0 else 0 for t in totals]
        for i, (lbl, pct) in enumerate(zip(labels, pcts)):
            ax.barh(0, pct, left=left_, color=COLORS[i % len(COLORS)],
                    label=lbl, edgecolor="white", linewidth=0.5)
            if pct > 5:
                ax.text(left_+pct/2, 0, f"{pct:.0f}%",
                        ha="center", va="center", fontsize=7.5, color="white")
            left_ += pct
        ax.set_xlim(0, 100); ax.set_yticks([])
        ax.set_xlabel("% of total weight"); ax.set_title("Weight Distribution")
        ax.legend(loc="upper center", bbox_to_anchor=(0.5,-0.18),
                  ncol=2, fontsize=7, frameon=False)
        ax.grid(axis="x", alpha=0.3)
        ax2 = axes[1]
        _, _, ats = ax2.pie(totals, labels=None, autopct="%1.0f%%",
            colors=COLORS[:len(labels)], startangle=90, pctdistance=0.75,
            wedgeprops=dict(edgecolor="white", linewidth=0.8))
        for at in ats: at.set_fontsize(7)
        ax2.set_title(f"Total: {grand:.0f} g")
        ax2.legend(labels, loc="lower center", bbox_to_anchor=(0.5,-0.22),
                   ncol=2, fontsize=7, frameon=False)
        fig.tight_layout()
        if _wb_canvas_ref[0]:
            try: _wb_canvas_ref[0].get_tk_widget().destroy()
            except Exception: pass
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        fc = FigureCanvasTkAgg(fig, master=wb_right)
        fc.draw(); fc.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        _wb_canvas_ref[0] = fc; _plt.close(fig)

    wb_outer = ttk.Frame(tab_weight_budget, padding=4)
    wb_outer.grid(row=0, column=0, sticky="nsew")
    wb_outer.columnconfigure(0, weight=1); wb_outer.columnconfigure(1, weight=3)
    wb_outer.rowconfigure(0, weight=1)

    wb_left = ttk.LabelFrame(wb_outer, text="Component Weights", padding=4)
    wb_left.grid(row=0, column=0, sticky="nsew", padx=(0,6))
    wb_left.columnconfigure(0, weight=1); wb_left.rowconfigure(1, weight=1)

    wb_ph = ttk.Label(wb_left, text="Run a simulation to populate.",
                      foreground="#888888", wraplength=200)
    wb_ph.grid(row=0, column=0, columnspan=2, sticky="w", padx=4, pady=(0,4))

    _wb_tv_cols = ("component","unit_w","count","total_w","pct")
    wb_tv = ttk.Treeview(wb_left, columns=_wb_tv_cols, show="headings", height=10)
    for col, heading, width in [
        ("component","Component",160), ("unit_w","Unit (g)",70),
        ("count","Qty",40), ("total_w","Total (g)",75), ("pct","% of Total",70)]:
        wb_tv.heading(col, text=heading)
        wb_tv.column(col, width=width,
                     anchor="w" if col=="component" else "center", stretch=True)
    wb_tv.grid(row=1, column=0, sticky="nsew")
    wb_sb = ttk.Scrollbar(wb_left, orient="vertical", command=wb_tv.yview)
    wb_sb.grid(row=1, column=1, sticky="ns")
    wb_tv.configure(yscrollcommand=wb_sb.set)
    wb_tv.tag_configure("total", font=("TkDefaultFont", 9, "bold"))

    wb_right = ttk.LabelFrame(wb_outer, text="Weight Distribution", padding=4)
    wb_right.grid(row=0, column=1, sticky="nsew")
    wb_right.columnconfigure(0, weight=1); wb_right.rowconfigure(0, weight=1)

    def update_weight_budget(cfg):
        rows = _extract_weight_budget(cfg)
        for iid in wb_tv.get_children(): wb_tv.delete(iid)
        total_g = rows[-1][3] if rows else 1.0
        for label, unit_w, count, total_w in rows[:-1]:
            pct = f"{total_w/total_g*100:.1f}%" if total_g > 0 else "0%"
            wb_tv.insert("", "end", values=(label, f"{unit_w:.1f}",
                         str(count), f"{total_w:.1f}", pct))
        if rows:
            label, unit_w, count, total_w = rows[-1]
            wb_tv.insert("", "end", values=(label, "", "",
                         f"{total_w:.1f}", "100%"), tags=("total",))
        _draw_weight_chart(rows)

    # ---- Status panel ----
    style_tv = ttk.Style()
    try: style_tv.theme_use(style_tv.theme_use())
    except Exception: pass

    def _color_tag(value, limit, kind):
        try:
            v = float(value); L = float(limit)
        except Exception: return "na"
        if kind == "max": return "bad" if v > L else ("warn" if v > 0.9*L else "ok")
        else:             return "bad" if v < L else ("warn" if v < 1.1*L else "ok")

    def _make_status_tv(parent, title):
        lf = ttk.LabelFrame(parent, text=title, padding=4)
        lf.pack(fill="both", expand=True, padx=4, pady=4)
        cols = ("metric", "value", "limit", "note")
        tv = ttk.Treeview(lf, columns=cols, show="headings", height=6)
        for c, w in zip(cols, (200, 120, 120, 220)):
            tv.heading(c, text=c.capitalize())
            tv.column(c, width=w, anchor="w" if c in ("metric","note") else "center")
        tv.pack(fill="both", expand=True)
        tv.tag_configure("ok",   background="#d9f2d9")
        tv.tag_configure("warn", background="#fff2cc")
        tv.tag_configure("bad",  background="#f8d7da")
        tv.tag_configure("na",   background="#efefef")
        return tv

    status_scroll = ttk.Frame(tab_status)
    status_scroll.grid(row=0, column=0, sticky="nsew")
    status_scroll.columnconfigure(0, weight=1)

    batt_tv  = _make_status_tv(status_scroll, "Battery Status")
    motor_tv = _make_status_tv(status_scroll, "Motor / ESC Status")
    aero_tv  = _make_status_tv(status_scroll, "Aerodynamic Status")

    def _ins_row(tv, metric, val_str, lim_str, tag, note=""):
        tv.insert("", "end", values=(metric, val_str, lim_str, note), tags=(tag,))

    def _clear_status():
        for tv in (batt_tv, motor_tv, aero_tv):
            for iid in tv.get_children(): tv.delete(iid)

    def update_status(cfg: FixedWingConfig, m: dict):
        _clear_status()
        batt = cfg.battery
        af   = cfg.airframe

        Ipack  = float(m.get("pack_current_A", 0.0))
        Vload  = float(m.get("v_load_V",       0.0))
        Ptot   = float(m.get("total_power_W",  0.0))
        Pmotor = float(m.get("motor_power_W",  0.0))
        T_avail= float(m.get("thrust_available_N", 0.0))
        T_req  = float(m.get("thrust_required_N",  0.0))
        V_stall= float(m.get("stall_speed_mps", 0.0))
        V_cruise=float(m.get("airspeed_mps",   cfg.cruise_speed_mps))
        CL     = float(m.get("CL",             0.0))
        RC     = float(m.get("rate_of_climb_mps", 0.0))
        LD     = float(m.get("LD_ratio",       0.0))
        V_tip  = float(m.get("tip_speed_mps",  0.0))
        rpm    = float(m.get("rpm_est",        0.0))

        # ── Battery Status ────────────────────────────────────────────────
        Vdrop = batt.vmax_pack - Vload
        vsag_pct = (Vdrop / max(batt.vmax_pack, 1e-6)) * 100
        _ins_row(batt_tv, "Pack voltage (loaded)",
                 f"{Vload:.2f} V", f">= {batt.vmin_pack:.2f} V",
                 _color_tag(Vload, batt.vmin_pack, "min"),
                 f"Sag: {Vdrop:.2f} V ({vsag_pct:.1f}%)")

        if math.isfinite(batt.discharge_cont_A):
            _ins_row(batt_tv, "Pack current vs cont rating",
                     f"{Ipack:.1f} A", f"<= {batt.discharge_cont_A:.1f} A",
                     _color_tag(Ipack, batt.discharge_cont_A, "max"))
        if math.isfinite(batt.discharge_max_A):
            _ins_row(batt_tv, "Pack current vs max rating",
                     f"{Ipack:.1f} A", f"<= {batt.discharge_max_A:.1f} A",
                     _color_tag(Ipack, batt.discharge_max_A, "max"))

        cap_Ah = batt.capacity_Ah
        if cap_Ah > 0:
            c_rate = Ipack / cap_Ah
            _ins_row(batt_tv, "Discharge C-rate",
                     f"{c_rate:.2f} C",
                     f"<= {batt.discharge_cont_A/cap_Ah:.1f} C (cont)" if math.isfinite(batt.discharge_cont_A) else "—",
                     _color_tag(c_rate, batt.discharge_cont_A/cap_Ah if math.isfinite(batt.discharge_cont_A) else 1e9, "max"),
                     "Continuous C-rate at cruise")

        _ins_row(batt_tv, "Usable energy",
                 f"{batt.usable_Wh:.1f} Wh", "—", "na",
                 f"Total: {batt.capacity_Wh:.1f} Wh ({batt.discharge_percent:.0f}% usable)")
        _ins_row(batt_tv, "Total electrical power", f"{Ptot:.1f} W", "—", "na")

        if Ptot > 0:
            pct_motor = Pmotor / Ptot * 100
            pct_esc   = float(m.get("esc_loss_W", 0)) / Ptot * 100
            pct_av    = float(m.get("avionics_power_W", 0)) / Ptot * 100
            _ins_row(batt_tv, "Power split (M/ESC/Av)",
                     f"{pct_motor:.0f}% / {pct_esc:.0f}% / {pct_av:.0f}%",
                     "—", "na", "Motor / ESC loss / Avionics")

        rsv_margin = float(m.get("reserve_margin_Wh", float("nan")))
        if math.isfinite(rsv_margin):
            rsv_tag = "bad" if m.get("reserve_breached", False) else (
                      "warn" if rsv_margin < 5.0 else "ok")
            _ins_row(batt_tv, "Energy reserve margin",
                     f"{rsv_margin:+.1f} Wh", ">= 0 Wh", rsv_tag,
                     "VIOLATION" if m.get("reserve_breached", False) else "OK")

        # ── Motor / ESC Status ────────────────────────────────────────────
        pmax = float(m.get("max_prop_power_W", cfg.motor.max_power * cfg.num_motors))
        _ins_row(motor_tv, "Motor electrical power",
                 f"{Pmotor:.1f} W", f"<= {pmax:.1f} W",
                 _color_tag(Pmotor, pmax, "max"))

        if cfg.motor.max_current:
            I_motor_est = (Pmotor / max(cfg.num_motors, 1)) / max(Vload, 1.0)
            _ins_row(motor_tv, "Motor current / motor (est)",
                     f"{I_motor_est:.2f} A", f"<= {cfg.motor.max_current:.2f} A",
                     _color_tag(I_motor_est, cfg.motor.max_current, "max"))

        T_margin_pct = (T_avail - T_req) / max(T_avail, 1e-9) * 100
        _ins_row(motor_tv, "Thrust margin",
                 f"{T_margin_pct:.1f}%", ">= 20%",
                 "ok" if T_margin_pct >= 20 else ("warn" if T_margin_pct >= 5 else "bad"),
                 f"T_avail={T_avail:.2f}N  T_req={T_req:.2f}N")

        TWR = T_avail / max(cfg.weight_N, 1e-9)
        _ins_row(motor_tv, "Thrust-to-weight ratio",
                 f"{TWR:.2f}:1", ">= 1.2:1",
                 "ok" if TWR >= 1.4 else ("warn" if TWR >= 1.1 else "bad"),
                 "< 1.2 means marginal climb performance")

        _ins_row(motor_tv, "Estimated RPM", f"{rpm:.0f}", "—", "na")
        _ins_row(motor_tv, "Prop tip speed",
                 f"{V_tip:.1f} m/s", "<= 200 m/s",
                 _color_tag(V_tip, 200, "max"),
                 "Mach 0.6 ≈ 200 m/s — keep below for noise & efficiency")

        V_pitch = float(m.get("pitch_speed_mps", 0.0))
        if V_pitch > 0:
            _ins_row(motor_tv, "Pitch speed vs cruise",
                     f"{V_pitch:.1f} m/s (pitch)  {V_cruise:.1f} m/s (cruise)",
                     "pitch > cruise", "ok" if V_pitch > V_cruise else "warn",
                     "Cruise near/above pitch speed → prop efficiency collapses")

        T_motor = float(m.get("motor_temp_est_C", float("nan")))
        T_esc   = float(m.get("esc_temp_est_C",   float("nan")))
        T_batt  = float(m.get("battery_temp_est_C",float("nan")))
        if math.isfinite(T_motor):
            _ins_row(motor_tv, "Motor temperature (est)",
                     f"{T_motor:.1f} °C", "<= 100 °C",
                     "ok" if T_motor < 80 else ("warn" if T_motor < 100 else "bad"),
                     f"Thermal status: {m.get('thermal_status','n/a')}")
        if math.isfinite(T_esc):
            _ins_row(motor_tv, "ESC temperature (est)",
                     f"{T_esc:.1f} °C", "<= 90 °C",
                     "ok" if T_esc < 70 else ("warn" if T_esc < 90 else "bad"))
        if math.isfinite(T_batt):
            _ins_row(motor_tv, "Battery temperature (est)",
                     f"{T_batt:.1f} °C", "<= 55 °C",
                     "ok" if T_batt < 40 else ("warn" if T_batt < 55 else "bad"))

        if cfg.esc is not None:
            I_esc_est = (Pmotor / max(cfg.num_motors, 1)) / max(Vload, 1.0)
            # ESCConfig.voltage_rating is a CELL COUNT (an "S" rating), not volts.
            # Compare it against the pack's series-cell count, never against a
            # voltage — "6 >= 22.2 V" would always read as a failure.
            esc_S   = int(cfg.esc.voltage_rating)
            pack_S  = int(batt.series_cells)
            _ins_row(motor_tv, "ESC voltage rating",
                     f"{esc_S}S  ({esc_S * batt.operating_voltage_max:.1f} V max)",
                     f">= {pack_S}S  ({batt.vmax_pack:.1f} V)",
                     "ok" if esc_S >= pack_S else "bad",
                     "ESC S-rating must meet or exceed the pack series count")
            _ins_row(motor_tv, "ESC current vs cont",
                     f"{I_esc_est:.2f} A", f"<= {cfg.esc.continuous_rating_A:.1f} A",
                     _color_tag(I_esc_est, cfg.esc.continuous_rating_A, "max"))

        # ── Aerodynamic Status ────────────────────────────────────────────
        speed_margin = (V_cruise - V_stall) / max(V_stall, 1e-9) * 100
        _ins_row(aero_tv, "Cruise vs stall speed",
                 f"{V_cruise:.1f} m/s  ({speed_margin:+.0f}% margin)",
                 f"> {V_stall * 1.1:.1f} m/s  (1.1×Vstall)",
                 _color_tag(V_cruise, V_stall * 1.1, "min"),
                 "< 1.1×Vstall is structurally unsafe")

        CL_margin = (af.CL_max - CL) / max(af.CL_max, 1e-9) * 100
        _ins_row(aero_tv, "CL at cruise vs CL_max",
                 f"{CL:.3f}  ({CL_margin:.0f}% below max)",
                 f"< {af.CL_max:.3f}",
                 _color_tag(CL, af.CL_max, "max"),
                 "CL approaching CL_max → near stall")

        _ins_row(aero_tv, "L/D ratio (cruise)",
                 f"{LD:.2f}", "—", "na",
                 f"Max L/D: {m.get('best_ld_ratio', 0.0):.2f} @ {m.get('best_range_speed_mps',0):.1f} m/s")

        RC_mpm = RC * 60.0
        _ins_row(aero_tv, "Rate of climb",
                 f"{RC_mpm:.1f} m/min", "> 90 m/min (3 fps)",
                 "ok" if RC_mpm >= 90 else ("warn" if RC_mpm >= 0 else "bad"),
                 "< 90 m/min is marginal for most operations")

        max_rc = float(m.get("max_rc_mps", 0.0)) * 60
        _ins_row(aero_tv, "Max rate of climb",
                 f"{max_rc:.1f} m/min", "> 150 m/min",
                 "ok" if max_rc >= 150 else ("warn" if max_rc >= 60 else "bad"),
                 f"At {m.get('v_max_rc_mps',0):.1f} m/s")

        max_aoc = float(m.get("max_aoc_deg", 0.0))
        _ins_row(aero_tv, "Max angle of climb",
                 f"{max_aoc:.1f}°", "> 5°",
                 "ok" if max_aoc >= 10 else ("warn" if max_aoc >= 3 else "bad"))

        wl = float(m.get("wing_loading_N_m2", 0.0))
        _ins_row(aero_tv, "Wing loading",
                 f"{wl:.1f} N/m²  ({wl/9.81*100:.0f} g/dm²)",
                 "20–120 N/m² (typical UAV)",
                 "ok" if 20 <= wl <= 120 else "warn",
                 "High W/S → fast stall; Low W/S → gust sensitive")

        glide = float(m.get("glide_ratio", LD))
        _ins_row(aero_tv, "Glide ratio (L/D at cruise)",
                 f"{glide:.1f}:1", "> 8:1",
                 "ok" if glide >= 10 else ("warn" if glide >= 6 else "bad"),
                 "Used for unpowered glide range estimation")

        ceil_m = float(m.get("service_ceiling_m", float("inf")))
        ceil_s = f"{ceil_m:.0f} m ASL" if math.isfinite(ceil_m) else "> 6000 m"
        _ins_row(aero_tv, "Service ceiling", ceil_s, "—", "na",
                 "Altitude where max RC drops to 0.5 m/s")

        S_to = float(m.get("takeoff_dist_m", float("inf")))
        if math.isfinite(S_to):
            _ins_row(aero_tv, "Takeoff ground roll",
                     f"{S_to:.1f} m", "<= 100 m",
                     "ok" if S_to <= 60 else ("warn" if S_to <= 120 else "bad"))
        else:
            _ins_row(aero_tv, "Takeoff ground roll",
                     "∞", "—", "bad", "Thrust < rolling friction — cannot take off")

        S_ld = float(m.get("landing_dist_m", float("inf")))
        if math.isfinite(S_ld):
            # This figure is the FAA-style distance over a 15 m (50 ft)
            # obstacle, so it is dominated by the 15 m x L/D approach segment.
            # Thresholds are scaled accordingly — a clean glider with a high
            # L/D legitimately needs a long approach.
            _ins_row(aero_tv, "Landing distance (over 15 m obstacle)",
                     f"{S_ld:.1f} m", "<= 400 m",
                     "ok" if S_ld <= 300 else ("warn" if S_ld <= 450 else "bad"),
                     "Approach segment = 15 m x L/D; ground roll is the remainder")

        SR = float(m.get("specific_range_m_per_Wh", 0.0))
        _ins_row(aero_tv, "Specific range",
                 f"{SR:.1f} m/Wh  ({SR/1000:.3f} km/Wh)", "—", "na",
                 "Higher = more distance per Wh — optimize at best-range speed")

        Re = float(m.get("reynolds_number", 0.0))
        re_tag = "ok" if Re >= 200000 else ("warn" if Re >= 70000 else "bad")
        _ins_row(aero_tv, "Reynolds number",
                 f"{Re:,.0f}", ">= 200 000",
                 re_tag,
                 "< 70 k → laminar separation; 70–200 k → transitional regime")

    # ---- Metrics panel ----
    metrics_frame = ttk.Frame(tab_metrics, padding=4)
    metrics_frame.grid(row=0, column=0, sticky="nsew")
    metrics_frame.columnconfigure(0, weight=1)
    metrics_frame.rowconfigure(0, weight=1)

    metrics_tv = ttk.Treeview(metrics_frame,
                               columns=("metric","value"), show="headings", height=30)
    metrics_tv.heading("metric", text="Metric")
    metrics_tv.heading("value",  text="Value")
    metrics_tv.column("metric", width=300, anchor="w")
    metrics_tv.column("value",  width=280, anchor="w")
    metrics_sb = ttk.Scrollbar(metrics_frame, orient="vertical", command=metrics_tv.yview)
    metrics_tv.configure(yscrollcommand=metrics_sb.set)
    metrics_tv.grid(row=0, column=0, sticky="nsew")
    metrics_sb.grid(row=0, column=1, sticky="ns")

    try:
        ttk.Style().configure("Metrics.Treeview", rowheight=24)
        metrics_tv.configure(style="Metrics.Treeview")
        metrics_tv.tag_configure("section", font=("TkDefaultFont", 10, "bold"))
    except Exception:
        pass

    def _clear_metrics():
        for iid in metrics_tv.get_children(): metrics_tv.delete(iid)
        _metrics_section_stack["current"] = ""

    # ---- Collapsible metrics sections --------------------------------
    # Sections are real parent nodes, so Treeview gives expand/collapse for
    # free. With 100+ rows a flat list is hard to scan; the sections a user
    # cares about stay open and the rest fold away.
    _metrics_section_stack = {"current": ""}
    _metrics_open_state = {}          # section title -> open/closed, remembered

    def _sep_metric(title: str):
        """Start a new collapsible section and make later rows its children."""
        node = metrics_tv.insert(
            "", "end", text=str(title), values=(f"── {title} ──", ""),
            tags=("section",), open=_metrics_open_state.get(str(title), True))
        _metrics_section_stack["current"] = node
        return node

    def _metrics_row(metric: str, value: str):
        """Add a row under the current section (or at top level if none)."""
        metrics_tv.insert(_metrics_section_stack["current"], "end",
                          values=(metric, value))

    def _metrics_remember_open_state(_event=None):
        """Persist which sections the user left open across re-runs."""
        for node in metrics_tv.get_children(""):
            title = str(metrics_tv.item(node, "text"))
            if title:
                _metrics_open_state[title] = bool(metrics_tv.item(node, "open"))

    metrics_tv.bind("<<TreeviewOpen>>", _metrics_remember_open_state, add="+")
    metrics_tv.bind("<<TreeviewClose>>", _metrics_remember_open_state, add="+")

    def _metrics_expand_all():
        for node in metrics_tv.get_children(""):
            metrics_tv.item(node, open=True)
        _metrics_remember_open_state()

    def _metrics_collapse_all():
        for node in metrics_tv.get_children(""):
            metrics_tv.item(node, open=False)
        _metrics_remember_open_state()

    def _ins_metric(metric: str, value: str):
        _metrics_row(metric, value)


    def update_metrics(cfg: FixedWingConfig, m: dict):
        _clear_metrics()
        af  = cfg.airframe
        batt= cfg.battery
        prop= cfg.propeller

        # ── Aircraft ─────────────────────────────────────────────────────
        _sep_metric("Aircraft")
        _ins_metric("Total Weight",              f"{cfg.aircraft_weight_g:.0f} g  ({cfg.weight_N:.2f} N)")
        _ins_metric("Wing Span",                 f"{af.wing_span_m:.3f} m  ({af.wing_span_m*39.37:.1f} in)")
        _ins_metric("Wing Area",                 f"{af.wing_area_m2:.4f} m²  ({af.wing_area_m2*1e4:.1f} cm²)")
        _ins_metric("Wing Loading",              f"{m.get('wing_loading_N_m2',0):.1f} N/m²  "
                                                 f"({m.get('wing_loading_kg_m2',0):.2f} kg/m²  "
                                                 f"/ {m.get('wing_loading_kg_m2',0)*100:.1f} g/dm²)")
        _ins_metric("Mean Chord",                f"{af.chord_m:.4f} m  ({af.chord_m*100:.1f} cm)")
        _ins_metric("Aspect Ratio",              f"{af.aspect_ratio:.2f}")
        _ins_metric("Induced Drag Factor k",     f"{af.k:.5f}  (1/πARe)")
        _ins_metric("Oswald Efficiency e",       f"{af.oswald:.3f}")
        _ins_metric("Lift-curve Slope a",        f"{af.lift_curve_slope():.3f} /rad  ({af.lift_curve_slope()*math.pi/180:.4f} /deg)")
        _ins_metric("CD0 (zero-lift drag)",      f"{af.CD0:.5f}")
        _ins_metric("Number of Motors",          f"{cfg.num_motors}")
        _ins_metric("Propeller Efficiency η",    f"{af.prop_efficiency:.3f}  ({af.prop_efficiency*100:.1f}%)")

        # ── Aerodynamics at Cruise ────────────────────────────────────────
        _sep_metric("Aerodynamics at Cruise")
        V  = m.get("airspeed_mps", cfg.cruise_speed_mps)
        _ins_metric("Cruise Airspeed", f"{V:.2f} m/s  ({V*3.6:.1f} km/h  /  {V*1.944:.1f} kt)")
        _ins_metric("Groundspeed", f"{m.get('groundspeed_mps',0):.2f} m/s  ({m.get('groundspeed_mps',0)*3.6:.1f} km/h  /  {m.get('groundspeed_mps',0)*1.944:.1f} kt)")
        _ins_metric("Head / Cross Wind",         f"{m.get('wind_head_mps',0):+.2f} / {m.get('wind_cross_mps',0):+.2f} m/s")
        _ins_metric("Stall Speed", f"{m.get('stall_speed_mps',0):.2f} m/s  ({m.get('stall_speed_mps',0)*3.6:.1f} km/h  /  {m.get('stall_speed_mps',0)*1.944:.1f} kt)")
        _ins_metric("Speed / Vstall margin",     f"{V/max(m.get('stall_speed_mps',1e-9),1e-9):.2f}×  (>1.1 required)")
        _ins_metric("CL at Cruise",              f"{m.get('CL',0):.4f}  (CL_max = {af.CL_max:.3f})")
        _ins_metric("CL margin to stall",        f"{(af.CL_max - m.get('CL',0)):.4f}  ({(af.CL_max - m.get('CL',0))/max(af.CL_max,1e-9)*100:.1f}%)")
        _ins_metric("CD at Cruise",              f"{m.get('CD',0):.5f}")
        _ins_metric("CD induced  (k·CL²)",       f"{af.k * m.get('CL',0)**2:.5f}")
        _ins_metric("CD parasitic (CD0)",        f"{af.CD0:.5f}")
        _ins_metric("Induced / Parasitic ratio", f"{af.k*m.get('CL',0)**2/max(af.CD0,1e-9):.3f}")
        _ins_metric("L/D Ratio",                 f"{m.get('LD_ratio',0):.2f}")
        _ins_metric("Max L/D (analytic)",        f"{m.get('best_ld_ratio',0):.2f}  @ {m.get('best_range_speed_mps',0):.1f} m/s")
        _ins_metric("Induced Drag",              f"{m.get('induced_drag_N',0):.3f} N  ({m.get('induced_drag_N',0)/max(m.get('induced_drag_N',0)+m.get('parasitic_drag_N',1e-9),1e-9)*100:.0f}%)")
        _ins_metric("Parasitic Drag",            f"{m.get('parasitic_drag_N',0):.3f} N  ({m.get('parasitic_drag_N',0)/max(m.get('induced_drag_N',0)+m.get('parasitic_drag_N',1e-9),1e-9)*100:.0f}%)")
        _ins_metric("Total Drag", f"{m.get('drag_N',0):.3f} N  ({m.get('drag_N',0)/9.81*1000:.0f} g)")
        _ins_metric("Angle of Attack",           f"{m.get('aoa_deg',0):.2f} °")
        _ins_metric("Reynolds Number",           f"{m.get('reynolds_number',0):,.0f}")
        _ins_metric("Glide Ratio (at cruise)",   f"{m.get('glide_ratio',0):.2f}:1")
        _ins_metric("Glide Distance (ref alt)",  f"{_format_distance_m(m.get('glide_distance_m',0))}")

        # ── Turning Flight ────────────────────────────────────────────────
        _sep_metric("Turning Flight")
        _ins_metric("Bank Angle",                f"{m.get('bank_deg',0):.1f} °")
        _ins_metric("Load Factor",               f"{m.get('load_factor',1.0):.3f} g")
        _ins_metric("Turn Stall Speed", f"{m.get('stall_speed_turn_mps',0):.2f} m/s  ({m.get('stall_speed_turn_mps',0)*3.6:.1f} km/h  /  {m.get('stall_speed_turn_mps',0)*1.944:.1f} kt)")
        _tr = m.get('turn_radius_m', float('inf'))
        _ins_metric("Turn Radius",               f"{_tr:.1f} m" if math.isfinite(_tr) else "∞ (wings level)")
        _ins_metric("Turn Rate",                 f"{m.get('turn_rate_deg_s',0):.2f} °/s")
        _ins_metric("Turn Period",               f"{m.get('turn_period_s', float('inf')):.1f} s" if math.isfinite(m.get('turn_period_s', float('inf'))) else "∞")
        _ins_metric("Loiter Circles",            f"{m.get('loiter_circles',0):.2f}")

        # ── Thrust & Power ────────────────────────────────────────────────
        _sep_metric("Thrust & Power")
        T_req   = m.get('thrust_required_N', 0.0)
        T_avail = m.get('thrust_available_N', 0.0)
        T_margin_N   = T_avail - T_req
        T_margin_pct = T_margin_N / max(T_avail, 1e-9) * 100
        TWR          = T_avail / max(cfg.weight_N, 1e-9)
        _ins_metric("Thrust Required", f"{T_req:.3f} N  ({T_req/9.81*1000:.0f} g)")
        _ins_metric("Thrust Available", f"{T_avail:.3f} N  ({T_avail/9.81*1000:.0f} g)")
        _ins_metric("Thrust Margin",             f"{T_margin_N:.3f} N  ({T_margin_pct:.1f}%)")
        _ins_metric("Thrust-to-Weight (T/W)",    f"{TWR:.3f}  ({TWR*100:.1f}% of weight)")
        _ins_metric("Specific Thrust (T_avail/W)",f"{m.get('specific_thrust',0):.4f}")
        P_prop  = m.get('power_required_W', 0.0)
        P_elec  = m.get('motor_power_W',   0.0)
        P_esc   = m.get('esc_loss_W',      0.0)
        P_av    = m.get('avionics_power_W',0.0)
        P_total = m.get('total_power_W',   0.0)
        P_max   = m.get('max_prop_power_W',0.0)
        _ins_metric("Propulsive Power (T·V)",    f"{P_prop:.1f} W")
        _ins_metric("Motor Electrical Power",    f"{P_elec:.1f} W")
        _ins_metric("ESC Losses",                f"{P_esc:.1f} W")
        _ins_metric("Avionics / Payload Power",  f"{P_av:.1f} W")
        _ins_metric("Total Electrical Power",    f"{P_total:.1f} W")
        _ins_metric("Max Motor+Prop Power",      f"{P_max:.1f} W  ({P_elec/max(P_max,1e-9)*100:.0f}% utilisation)")
        if P_total > 0:
            _ins_metric("Power breakdown",       f"Motor {P_elec/P_total*100:.0f}%  ESC {P_esc/P_total*100:.0f}%  Avionics {P_av/P_total*100:.0f}%")
        # Propulsive efficiency: useful thrust power / electrical input
        eta_prop = (T_req * V) / max(P_elec, 1e-9) if P_elec > 0 else 0.0
        _ins_metric("Propulsive efficiency η",   f"{eta_prop*100:.1f}%  (T·V / P_electrical)")
        _ins_metric("Motor copper loss (est)",   f"{m.get('motor_copper_loss_W',0):.2f} W")
        _ins_metric("Battery I²R loss",          f"{m.get('battery_loss_W',0):.3f} W")

        # ── Battery ───────────────────────────────────────────────────────
        _sep_metric("Battery")
        Ipack = m.get('pack_current_A', 0.0)
        Vload = m.get('v_load_V', 0.0)
        Vsag  = batt.vmax_pack - Vload
        _ins_metric("Chemistry",                 f"{batt.chemistry or '—'}")
        _ins_metric("SoC model",                 f"{_soc_model_short_label(getattr(batt,'soc_model_source',None))}"
                                                 f"{'  (non-linear)' if getattr(batt,'soc_nonlinear_enabled',False) else '  (linear)'}")
        _ins_metric("Configuration",             f"{batt.series_cells}S × {batt.parallel_cells}P  ({batt.total_cells} cells total)")
        _ins_metric("Pack Voltage (no load)",    f"{batt.vmax_pack:.2f} V")
        _ins_metric("Pack Voltage (min cutoff)", f"{batt.vmin_pack:.2f} V")
        _ins_metric("Pack Voltage (under load)", f"{Vload:.2f} V  (sag: {Vsag:.2f} V / {Vsag/max(batt.vmax_pack,1e-9)*100:.1f}%)")
        _ins_metric("Pack Resistance",           f"{batt.pack_resistance*1000:.1f} mΩ  ({batt.resistance_cell*1000:.1f} mΩ/cell)")
        _ins_metric("Pack Current",              f"{Ipack:.2f} A")
        cap_Ah = batt.capacity_Ah
        if cap_Ah > 0:
            c_rate = Ipack / cap_Ah
            _ins_metric("Discharge C-rate",      f"{c_rate:.2f} C  (cont limit: {batt.discharge_cont_A/cap_Ah:.1f} C)" if math.isfinite(batt.discharge_cont_A) else f"{c_rate:.2f} C")
        _ins_metric("Pack Capacity",             f"{batt.capacity_mAh:.0f} mAh  ({batt.capacity_Ah:.3f} Ah)")
        _ins_metric("Pack Energy",               f"{batt.capacity_Wh:.2f} Wh")
        _ins_metric("Usable Energy",             f"{batt.usable_Wh:.2f} Wh  ({batt.discharge_percent:.0f}% of pack)")
        _ins_metric("Usable Capacity",           f"{batt.capacity_mAh*batt.usable_fraction:.0f} mAh")
        _ins_metric("Energy Density",            f"{batt.energy_density_Wh_per_kg:.0f} Wh/kg")
        _ins_metric("Battery Weight",            f"{batt.weight_g:.0f} g  ({batt.weight_g/max(cfg.aircraft_weight_g,1e-9)*100:.1f}% of AUW)")

        # ── Propeller ─────────────────────────────────────────────────────
        _sep_metric("Propeller")
        rpm_est = m.get('rpm_est', 0.0)
        n_rps   = rpm_est / 60.0 if rpm_est > 0 else 1e-9
        J       = V / (n_rps * prop.diameter_m) if n_rps > 0 and prop.diameter_m > 0 else 0.0
        pd_ratio= prop.pitch_m / max(prop.diameter_m, 1e-9)
        tip_mach= m.get('tip_speed_mps', 0.0) / 340.0
        _ins_metric("Diameter",                  f"{prop.diameter_in:.1f} in  ({prop.diameter_m*100:.1f} cm)")
        _ins_metric("Pitch",                     f"{prop.pitch_in:.1f} in  ({prop.pitch_m*100:.1f} cm)")
        _ins_metric("Blades",                    f"{prop.blades}")
        _ins_metric("Pitch / Diameter ratio",    f"{pd_ratio:.3f}")
        _ins_metric("Estimated RPM",             f"{rpm_est:.0f} rpm")
        _ins_metric("Tip Speed", f"{m.get('tip_speed_mps',0):.2f} m/s  ({m.get('tip_speed_mps',0)*3.6:.1f} km/h  /  {m.get('tip_speed_mps',0)*1.944:.1f} kt)")
        _ins_metric("Tip Mach number",           f"{tip_mach:.4f}{'  ⚠ noise above 0.6' if tip_mach > 0.6 else ''}")
        _ins_metric("Pitch Speed", f"{m.get('pitch_speed_mps',0):.2f} m/s  ({m.get('pitch_speed_mps',0)*3.6:.1f} km/h  /  {m.get('pitch_speed_mps',0)*1.944:.1f} kt)")
        _ins_metric("Advance Ratio J (V/nD)",    f"{J:.3f}  (0 = static, 1 = pitch speed)")
        _ins_metric("Disk Area",                 f"{prop.disk_area()*1e4:.1f} cm²  per rotor")

        # ── Climb Performance ─────────────────────────────────────────────
        _sep_metric("Climb Performance")
        RC = m.get('rate_of_climb_mps', 0.0)
        _ins_metric("Rate of Climb @ Cruise",    f"{RC*60:.1f} m/min  ({RC:.3f} m/s  /  {RC*196.85:.0f} fpm)")
        _ins_metric("Commanded Climb Rate", f"{m.get('climb_rate_cmd_mps',0):.2f} m/s  ({m.get('climb_rate_cmd_mps',0)*3.6:.1f} km/h  /  {m.get('climb_rate_cmd_mps',0)*1.944:.1f} kt)")
        _ins_metric("Commanded Descent Rate", f"{m.get('descent_rate_cmd_mps',0):.2f} m/s  ({m.get('descent_rate_cmd_mps',0)*3.6:.1f} km/h  /  {m.get('descent_rate_cmd_mps',0)*1.944:.1f} kt)")
        _ins_metric("Potential Power Term",      f"{m.get('potential_power_W',0):+.1f} W")
        max_rc = m.get('max_rc_mps', 0.0)
        _ins_metric("Max Rate of Climb",         f"{max_rc*60:.1f} m/min  ({max_rc*196.85:.0f} fpm)  @ {m.get('v_max_rc_mps',0):.1f} m/s")
        _ins_metric("Max Angle of Climb",        f"{m.get('max_aoc_deg',0):.2f} °  @ {m.get('v_max_aoc_mps',0):.1f} m/s")
        ceil_m = m.get('service_ceiling_m', float('inf'))
        _ins_metric("Service Ceiling (ASL)",     f"{ceil_m:.0f} m  ({ceil_m*3.281:.0f} ft)" if math.isfinite(ceil_m) else "> 12 000 m")
        ceil_agl = m.get('service_ceiling_agl_m', float('inf'))
        _ins_metric("Service Ceiling (AGL)",     f"{ceil_agl:.0f} m  ({ceil_agl*3.281:.0f} ft)" if math.isfinite(ceil_agl) else "> 12 000 m")
        S_to = m.get('takeoff_dist_m', float('inf'))
        _ins_metric("Takeoff Ground Roll",       f"{S_to:.1f} m  ({S_to*3.281:.0f} ft)" if math.isfinite(S_to) else "∞ (T < rolling friction)")
        S_ld = m.get('landing_dist_m', float('inf'))
        _ins_metric("Landing Distance (over 15 m obstacle)",
                    f"{S_ld:.1f} m  ({S_ld*3.281:.0f} ft)" if math.isfinite(S_ld) else "∞ (insufficient braking)")
        _ins_metric("Landing Braking μ",         f"{af.mu_brake:.2f}  (V_TD = {m.get('stall_speed_mps',0)*1.15:.1f} m/s)")

        # ── Optimal Speeds ────────────────────────────────────────────────
        _sep_metric("Optimal Speeds")
        Vbe = m.get('best_endurance_speed_mps', 0.0)
        Vbr = m.get('best_range_speed_mps',     0.0)
        Vms = m.get('min_sink_speed_mps',       0.0)
        _ins_metric("Best Endurance (Vy-equiv)", f"{Vbe:.1f} m/s  ({Vbe*3.6:.1f} km/h)  — min power")
        _ins_metric("Best Range (max L/D)",      f"{Vbr:.1f} m/s  ({Vbr*3.6:.1f} km/h)  — max L/D")
        _ins_metric("Min Sink (best glide time)",f"{Vms:.1f} m/s  ({Vms*3.6:.1f} km/h)  — min power/W")
        _ins_metric("Min Sink Rate",             f"{m.get('min_sink_rate_mps',0):.3f} m/s  ({m.get('min_sink_rate_mps',0)*196.85:.1f} fpm)")
        _ins_metric("Max L/D Ratio",             f"{m.get('best_ld_ratio',0):.2f}")
        _ins_metric("Cruise vs Best Endurance",  f"Cruise is {V/max(Vbe,1e-9):.2f}× best endurance speed")
        _ins_metric("Cruise vs Best Range",      f"Cruise is {V/max(Vbr,1e-9):.2f}× best range speed")

        # ── Endurance & Range ─────────────────────────────────────────────
        _sep_metric("Endurance & Range @ Cruise")
        t_min   = m.get('flight_time_min', 0.0)
        d_km    = m.get('flight_range_km', 0.0)
        _ins_metric("Flight Time",               f"{t_min:.1f} min  ({t_min/60:.2f} h)")
        _ins_metric("Flight Range", f"{d_km:.2f} km  ({d_km*0.6214:.2f} mi  /  {d_km*0.5400:.2f} nm)")
        SR = m.get('specific_range_m_per_Wh', 0.0)
        SE = m.get('specific_endurance_min_per_Wh', 0.0)
        _ins_metric("Specific Range",            f"{SR:.1f} m/Wh  ({SR/1000:.3f} km/Wh  /  {m.get('specific_range_km_per_kWh',0):.2f} km/kWh)")
        _ins_metric("Specific Endurance",        f"{SE:.3f} min/Wh  ({m.get('specific_endurance_h_per_kWh',0):.3f} h/kWh)")
        _ins_metric("Reserve Target",            f"{m.get('reserve_target_Wh',0):.1f} Wh")
        _ins_metric("Reserve Margin",            f"{m.get('reserve_margin_Wh',0):+.1f} Wh  ({'OK' if not m.get('reserve_breached',False) else 'VIOLATION'})")

        # ── Thermal & Losses ──────────────────────────────────────────────
        _sep_metric("Thermal & Losses")
        _ins_metric("Motor temperature (est)", f"{m.get('motor_temp_est_C',0):.1f} °C")
        _ins_metric("ESC temperature (est)", f"{m.get('esc_temp_est_C',0):.1f} °C")
        _ins_metric("Battery temperature (est)", f"{m.get('battery_temp_est_C',0):.1f} °C")
        _ins_metric("Thermal status",            str(m.get("thermal_status", "OK")))
        _ins_metric("Motor copper loss (est)",   f"{m.get('motor_copper_loss_W',0):.2f} W")
        _ins_metric("Battery I²R loss",          f"{m.get('battery_loss_W',0):.3f} W")

        # ── Environment ───────────────────────────────────────────────────
        _sep_metric("Environment")
        rho = cfg.air_density
        rho_sl = 1.225
        rho_drop = (rho_sl - rho) / rho_sl * 100
        _ins_metric("Air Density",               f"{rho:.4f} kg/m³  ({rho_drop:+.1f}% vs ISA sea level)")
        _ins_metric("ISA Sea Level Density", f"{rho_sl:.4f} kg/m³")
        # Approximate density altitude from ISA
        try:
            T0_k, L_k = 288.15, 0.0065
            import math as _m
            da_m = T0_k / L_k * (1.0 - (rho / rho_sl) ** (1.0 / (1.0 - L_k * 287.05 / 9.80665)))
            _ins_metric("Density Altitude (est)", f"{da_m:.0f} m  ({da_m*3.281:.0f} ft)")
        except Exception:
            pass
        _ins_metric("Stall Speed at Density",    f"{m.get('stall_speed_mps',0):.2f} m/s  (increases at altitude)")

        # ── Design Summary ────────────────────────────────────────────────
        _sep_metric("Design Summary")
        _ins_metric("Thrust-to-Weight ratio",    f"{T_avail/max(cfg.weight_N,1e-9):.3f}  (available / weight)")
        _ins_metric("Thrust Margin at Cruise",   f"{T_margin_N:.2f} N  ({T_margin_pct:.1f}%)")
        _ins_metric("Battery mass fraction",     f"{batt.weight_g/max(cfg.aircraft_weight_g,1e-9)*100:.1f}%  ({batt.weight_g:.0f} g)")
        _ins_metric("Energy density (battery)",  f"{batt.energy_density_Wh_per_kg:.0f} Wh/kg")
        _ins_metric("Power loading",             f"{P_total/max(cfg.aircraft_weight_g/1000,1e-9):.1f} W/kg  ({P_total/max(cfg.weight_N,1e-9):.2f} W/N)")
        _ins_metric("Propulsive efficiency η",   f"{eta_prop*100:.1f}%  at cruise")
        _ins_metric("System efficiency",         f"{(P_prop/max(P_total,1e-9))*100:.1f}%  (T·V / P_total)")

    # ---- Output text ----
    out_frame = ttk.LabelFrame(right, text="Output", padding=4)
    out_frame.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
    out_frame.columnconfigure(0, weight=1)
    out_frame.rowconfigure(0, weight=1)

    out_text = tk.Text(out_frame, height=8, wrap="word", state="disabled")
    out_text.grid(row=0, column=0, sticky="nsew")
    out_sb   = ttk.Scrollbar(out_frame, orient="vertical", command=out_text.yview)
    out_sb.grid(row=0, column=1, sticky="ns")
    out_text.configure(yscrollcommand=out_sb.set)

    def out_print(msg: str):
        out_text.configure(state="normal")
        out_text.delete("1.0", "end")
        out_text.insert("end", msg)
        out_text.configure(state="disabled")

    # ---- Build config from GUI ----
    def build_config() -> FixedWingConfig:
        alt  = safe_float(v_altitude.get(), 0.0)
        temp = parse_float("Temp", v_temp.get())
        pres = parse_float("Pressure", v_pressure.get())
        rho  = isa_density(alt, temp, pres)

        af = AirframeConfig(
            wing_span_m     = parse_float("Wing span",   v_wing_span.get()),
            wing_area_m2    = parse_float("Wing area",   v_wing_area.get()),
            CD0             = safe_float(v_CD0.get(), 0.030),
            CL_max          = safe_float(v_CL_max.get(), 1.30),
            oswald          = safe_float(v_oswald.get(), 0.82),
            mu_roll         = safe_float(v_mu_roll.get(), 0.04),
            mu_brake        = safe_float(v_mu_brake.get(), 0.08),
            CL_takeoff      = safe_float(v_CL_takeoff.get(), 0.80),
            prop_efficiency = safe_float(v_prop_eff.get(), 0.75),
            prop_eff_model  = (v_prop_eff_model.get().strip() or "curve"),
            num_motors      = parse_int("Num motors", v_num_motors.get()) or 1,
        )

        batt = BatteryConfig(
            chemistry               = v_batt_chem.get().strip() or None,
            operating_voltage_min   = safe_float(v_batt_vmin.get(), 3.5),
            operating_voltage_nominal = safe_float(v_batt_vnom.get(), 3.8),
            operating_voltage_max   = safe_float(v_batt_vmax.get(), 4.2),
            unit_mode               = v_batt_unit_mode.get().strip().lower() or "cell",
            series_units            = parse_int("Series", v_batt_series.get()) or 1,
            parallel_units          = parse_int("Parallel", v_batt_parallel.get()) or 1,
            cells_series_per_unit   = parse_int("Cells series", v_batt_cells_s.get()) or 1,
            cells_parallel_per_unit = parse_int("Cells parallel", v_batt_cells_p.get()) or 1,
            cell_capacity_mAh       = parse_float("Cell cap", v_batt_cell_cap.get()),
            pack_capacity_mAh       = parse_float("Pack cap", v_batt_pack_cap.get()),
            cell_weight_g           = parse_float("Cell wt", v_batt_cell_wt.get()),
            pack_weight_g           = parse_float("Pack wt", v_batt_pack_wt.get()),
            unit_energy_density     = parse_float("Dens", v_batt_dens.get()),
            charge_current_max      = parse_float("Chg", v_batt_chg.get()),
            discharge_cont_A        = parse_float("A_cont", v_batt_a_cont.get()),
            discharge_max_A         = parse_float("A_max", v_batt_a_max.get()),
            discharge_c_cont        = parse_float("C_cont", v_batt_c_cont.get()),
            discharge_c_max         = parse_float("C_max", v_batt_c_max.get()),
            discharge_percent       = safe_float(v_batt_dischg_pct.get(), 80.0),
            resistance_cell_mOhm    = safe_float(v_batt_r.get(), 0.0),
            # SoC model: blank fields fall back to the chemistry preset.
            soc_model               = (v_batt_soc_model.get().strip() or "auto"),
            soc_curve_csv           = (v_batt_soc_curve_csv.get().strip() or None),
            soc_bp                  = parse_soc_breakpoints(v_batt_soc_bp.get()),
            ocv_cell_bp             = parse_float_list(v_batt_ocv_cell_bp.get()),
            r_scale_bp              = parse_float_list(v_batt_r_scale_bp.get()),
        )

        motor = MotorConfig(
            kv            = parse_float("Kv", v_motor_kv.get()),
            idle_current  = safe_float(v_motor_i0.get(), 0.5),
            idle_voltage  = safe_float(v_motor_v0.get(), 7.0),
            rated_voltage = parse_int("Rated V", v_motor_rated_v.get()) or 16,
            resistance    = safe_float(v_motor_r.get(), 0.05),
            max_current   = safe_float(v_motor_imax.get(), 40),
            max_power     = safe_float(v_motor_pmax.get(), 500),
            pole_count    = parse_int("Poles", v_motor_poles.get()),
            weight_g      = parse_float("Motor wt", v_motor_wt.get()),
            size_mm       = v_motor_size.get().strip() or None,
        )

        esc_fields = [v_esc_vrating.get().strip(), v_esc_cont.get().strip(),
                      v_esc_max.get().strip(),    v_esc_idle.get().strip(),
                      v_esc_r.get().strip(),       v_esc_wt.get().strip()]
        esc = None
        if any(esc_fields):
            esc = ESCConfig(
                voltage_rating      = parse_int("ESC V", v_esc_vrating.get()) or 30,
                continuous_current_A= safe_float(v_esc_cont.get(), 30),
                max_current_A       = safe_float(v_esc_max.get(), 40),
                idle_current_A      = safe_float(v_esc_idle.get(), 0.1),
                resistance          = safe_float(v_esc_r.get(), 0.01),
                weight_g            = parse_float("ESC wt", v_esc_wt.get()),
            )

        avionics = AvionicsConfig(
            voltage_tree = _av_tree_to_dict())

        prop_table_path = v_prop_table.get().strip() or None
        prop = PropellerConfig(
            diameter_in  = safe_float(v_prop_d.get(), 10),
            pitch_in     = safe_float(v_prop_pitch.get(), 4.5),
            blades       = parse_int("Blades", v_prop_blades.get()) or 2,
            max_rpm      = safe_float(v_prop_maxrpm.get(), 0),
            max_thrust_g = safe_float(v_prop_maxthr.get(), 0),
            table_csv    = prop_table_path,
            TConst       = parse_float("TConst", v_prop_tconst.get()),
            PConst       = parse_float("PConst", v_prop_pconst.get()),
            weight_g     = parse_float("Prop wt", v_prop_wt.get()),
        )

        return FixedWingConfig(
            airframe            = af,
            battery             = batt,
            motor               = motor,
            propeller           = prop,
            aircraft_weight_g   = safe_float(v_weight.get(), 2500),
            payload_mass_g      = max(safe_float(v_payload_mass.get(), 0.0), 0.0),
            cruise_speed_mps    = safe_float(v_cruise_speed.get(), 18.0),
            periph_current_A    = safe_float(v_periph_cur.get(), 0.0),
            esc                 = esc,
            avionics            = avionics,
            air_density         = rho,
            reference_altitude_m= alt,
            cruise_altitude_m   = (safe_float(v_cruise_altitude.get(), 0.0)
                                   if v_cruise_altitude.get().strip() else None),
        )

    # ---- Run: single-point ----
    def run_single_point():
        clear_log()
        try:
            cfg = build_config()
        except Exception as e:
            messagebox.showerror("Input error", str(e))
            return
        try:
            V_cruise = cfg.cruise_speed_mps
            wind_speed = safe_float(v_wind.get(), 0.0)
            wind_dir = safe_float(v_wind_dir.get(), 0.0)
            course_deg = safe_float(v_course_deg.get(), 0.0)
            bank_deg = safe_float(v_bank_deg.get(), 0.0)
            headwind, crosswind = wind_components_mps(wind_speed, wind_dir, course_deg)
            m = compute_metrics(
                cfg, V_cruise,
                bank_deg=bank_deg,
                ambient_temp_C=safe_float(v_temp.get(), 25.0),
                wind_head_mps=headwind,
                wind_cross_mps=crosswind,
            )
            climb_rate = max(safe_float(v_climb_rate.get(), 0.0), 0.0)
            descent_rate = max(safe_float(v_descent_rate.get(), 0.0), 0.0)
            if climb_rate > 0 and descent_rate > 0:
                descent_rate = 0.0
            potential_power_W = cfg.weight_N * (climb_rate - descent_rate)
            m["climb_rate_cmd_mps"] = climb_rate
            m["descent_rate_cmd_mps"] = descent_rate
            m["potential_power_W"] = potential_power_W
            m["climb_power_add_W"] = potential_power_W
            base_total = float(m.get("total_power_W", 0.0))
            adj_total = max(base_total + potential_power_W, 0.0)
            if base_total > 0:
                scale = adj_total / base_total
                m["motor_power_W"] = float(m.get("motor_power_W", 0.0)) * scale
                m["esc_loss_W"] = float(m.get("esc_loss_W", 0.0)) * scale
            m["total_power_W"] = adj_total
            m["pack_current_A"] = adj_total / max(cfg.battery.vnom_pack, 1.0)
            m["v_load_V"] = cfg.battery.voltage_under_load(m["pack_current_A"])
            m["groundspeed_mps"] = groundspeed_along_track_mps(V_cruise, headwind, crosswind)
            if m["total_power_W"] > 0:
                m["flight_time_min"] = (cfg.battery.usable_Wh / m["total_power_W"]) * 60.0
                m["flight_range_km"] = m["groundspeed_mps"] * (m["flight_time_min"] * 60.0) / 1000.0
            else:
                m["flight_time_min"] = 0.0
                m["flight_range_km"] = 0.0
            reserve_target_Wh = max(
                cfg.battery.usable_Wh * (safe_float(v_reserve_percent.get(), 20.0) / 100.0),
                safe_float(v_rth_reserve_Wh.get(), 0.0) + safe_float(v_div_reserve_Wh.get(), 0.0),
            )
            m["reserve_target_Wh"] = reserve_target_Wh
            m["reserve_margin_Wh"] = cfg.battery.usable_Wh - reserve_target_Wh
            m["reserve_breached"] = bool(m["reserve_margin_Wh"] < 0)

            V_stall = m["stall_speed_mps"]
            V_be    = m["best_endurance_speed_mps"]
            V_br    = m["best_range_speed_mps"]

            update_status(cfg, m)
            update_metrics(cfg, m)

            max_v = safe_float(v_max_v_plot.get(), 40.0)

            # Cache for View-menu re-render
            _last_run["cfg"]     = cfg
            _last_run["max_v"]   = max_v
            # Needed by the Sensitivity tab, which re-evaluates this point.
            _last_run["speed"]   = V_cruise
            _last_run["metrics"] = m
            _last_run["wind"]    = wind_speed

            fig = make_performance_figure(
                cfg,
                max_speed=max_v,
                figsize=(_view["plot_w"], _view["plot_h"]),
            )
            # Generate motor operating point figure if available
            motor_fig = None
            if cfg.propeller.table is not None and m:
                try:
                    motor_fig = make_motor_operating_point_figure(cfg, m, figsize=(_view["plot_w"], 6))
                except Exception:
                    pass
            # Display both figures
            if motor_fig:
                show_figure([fig, motor_fig])
            else:
                show_figure(fig)
            display_nb.select(tab_plots)
            # Capture sweep data for export
            _sp_vs = [stall_speed(cfg)+0.1 + (max_v-stall_speed(cfg)-0.1)*i/300
                       for i in range(301)]
            _sp_vs = [max(v, 0.1) for v in _sp_vs]
            _last_run_sweep.clear()
            _last_run_sweep.update({
                "Speed (m/s)":           _sp_vs,
                "Flight Time (min)":     [flight_time_min(cfg, v) for v in _sp_vs],
                "Range (km)":            [flight_range_km(cfg, v) for v in _sp_vs],
                "Power Required (W)":    [power_required_W(cfg, v) for v in _sp_vs],
                "Drag (N)":              [drag_N(cfg, v) for v in _sp_vs],
                "Induced Drag (N)":      [drag_components_N(cfg, v)[0] for v in _sp_vs],
                "Parasitic Drag (N)":    [drag_components_N(cfg, v)[1] for v in _sp_vs],
                "Rate of Climb (m/s)":   [rate_of_climb_mps(cfg, v) for v in _sp_vs],
                "L/D Ratio":             [cfg.airframe.ld_ratio(
                    cfg.airframe.cl_at_speed(cfg.weight_N, v, cfg.air_density))
                    for v in _sp_vs],
            })
            _last_run_cfg[0] = cfg
            update_weight_budget(cfg)
            _refresh_airframe_diagram(cfg)
            refresh_comparison()

            log(f"=== Fixed-Wing Single-Point @ {V_cruise:.1f} m/s ({V_cruise*3.6:.1f} km/h) ===")
            log(f"Stall Speed   : {V_stall:.1f} m/s")
            log(f"Turn Stall    : {m.get('stall_speed_turn_mps',0):.1f} m/s @ {bank_deg:.1f}° bank")
            log(f"Wind Components: head {headwind:+.1f} m/s, cross {crosswind:+.1f} m/s")
            log(f"Groundspeed   : {m.get('groundspeed_mps',0):.1f} m/s")
            log(f"L/D Ratio     : {m['LD_ratio']:.2f}")
            log(f"Turn Radius   : {m.get('turn_radius_m', float('inf')):.1f} m")
            log(f"Load Factor   : {m.get('load_factor',1.0):.2f} g")
            log(f"Thrust Req    : {m['thrust_required_N']:.2f} N  |  Avail: {m['thrust_available_N']:.2f} N")
            log(f"Total Power   : {m['total_power_W']:.1f} W")
            log(f"Potential Power: {m.get('potential_power_W', 0.0):+.1f} W")
            log(f"Flight Time   : {m['flight_time_min']:.1f} min")
            log(f"Flight Range  : {m['flight_range_km']:.2f} km")
            log(f"Rate of Climb : {m['rate_of_climb_mps']*60:.0f} m/min")
            log(f"Max RC        : {m['max_rc_mps']*60:.0f} m/min  @ {m['v_max_rc_mps']:.1f} m/s")
            log(f"Best Endurance: {V_be:.1f} m/s  |  Best Range: {V_br:.1f} m/s")
            log(f"Takeoff Roll  : {m['takeoff_dist_m']:.1f} m")
            log(f"Landing Dist. : {m['landing_dist_m']:.1f} m")
            log(f"Min Sink Speed: {m['min_sink_speed_mps']:.1f} m/s  |  Sink: {m['min_sink_rate_mps']:.2f} m/s")
            log(f"Specific Range: {m['specific_range_m_per_Wh']:.1f} m/Wh")
            log(f"Reserve Margin: {m.get('reserve_margin_Wh', 0.0):+.1f} Wh")
            log(f"Thermal Est.  : Motor {m.get('motor_temp_est_C',0):.1f}°C, ESC {m.get('esc_temp_est_C',0):.1f}°C, Battery {m.get('battery_temp_est_C',0):.1f}°C [{m.get('thermal_status','OK')}]")
            if math.isfinite(m["service_ceiling_m"]):
                log(f"Service Ceiling: {m['service_ceiling_m']:.0f} m ASL")
            else:
                log("Service Ceiling: > 12000 m ASL")
        except Exception as e:
            import traceback
            messagebox.showerror("Simulation error", traceback.format_exc())

    # ---- Run: mission JSON ----
    def run_mission():
        clear_log()
        mission_path = v_mission.get().strip()
        if not mission_path:
            messagebox.showerror("No mission file",
                "Select a mission JSON file in the Environment tab first.")
            return
        try:
            cfg = build_config()
        except Exception as e:
            messagebox.showerror("Input error", str(e))
            return
        try:
            mission = MissionProfile.from_json(mission_path)
            mission.reserve_percent = safe_float(v_reserve_percent.get(), mission.reserve_percent)
            mission.rth_reserve_Wh = safe_float(v_rth_reserve_Wh.get(), mission.rth_reserve_Wh)
            mission.diversion_reserve_Wh = safe_float(v_div_reserve_Wh.get(), mission.diversion_reserve_Wh)
            mission.wind_direction_deg = safe_float(v_wind_dir.get(), mission.wind_direction_deg)
            # Transient settings from the GUI. A mission JSON that specifies
            # its own values has already set them; these fill in the rest.
            mission.transient_dt_s  = safe_float(v_transient_dt_s.get(),  mission.transient_dt_s)
            mission.max_accel_mps2  = safe_float(v_max_accel_mps2.get(),  mission.max_accel_mps2)
            mission.max_decel_mps2  = safe_float(v_max_decel_mps2.get(),  mission.max_decel_mps2)
            mission.decel_regen_eff = safe_float(v_decel_regen_eff.get(), mission.decel_regen_eff)
            _bank = safe_float(v_bank_deg.get(), 0.0)
            _course = safe_float(v_course_deg.get(), 0.0)
            _climb = safe_float(v_climb_rate.get(), 0.0)
            _descent = safe_float(v_descent_rate.get(), 0.0)
            for _p in mission.phases:
                _p.bank_deg = float(_p.bank_deg if _p.bank_deg else _bank)
                _p.course_deg = float(_p.course_deg if _p.course_deg else _course)
                if _p.climb_rate_mps is None:
                    _p.climb_rate_mps = _climb
                if _p.descent_rate_mps is None:
                    _p.descent_rate_mps = _descent
            wind    = safe_float(v_wind.get(), 0.0)
            temp    = v_temp.get().strip()
            pres    = v_pressure.get().strip()

            results, worst_m, series = simulate_fw_mission(
                cfg, mission,
                wind_mps      = wind,
                temperature_C = float(temp) if temp else None,
                pressure_Pa   = float(pres) if pres else None,
            )

            # Stash series for Mission Plots tab
            last_mission_series[0] = series

            # Populate status with worst-case metrics
            if worst_m:
                update_status(cfg, worst_m)

            # Rebuild metrics from worst-case point
            if worst_m:
                update_metrics(cfg, worst_m)

            max_v = safe_float(v_max_v_plot.get(), 40.0)
            _last_run["cfg"]   = cfg
            _last_run["max_v"] = max_v
            _last_run["metrics"] = worst_m

            _last_run_cfg[0] = cfg
            update_weight_budget(cfg)
            fig = make_performance_figure(
                cfg, max_speed=max_v,
                figsize=(_view["plot_w"], _view["plot_h"]))
            # Generate motor operating point figure if available
            motor_fig = None
            if cfg.propeller.table is not None and worst_m:
                try:
                    motor_fig = make_motor_operating_point_figure(cfg, worst_m, figsize=(_view["plot_w"], 6))
                except Exception:
                    pass
            # Display both figures
            if motor_fig:
                show_figure([fig, motor_fig])
            else:
                show_figure(fig)
            display_nb.select(tab_plots)

            import os as _os
            log(f"=== Fixed-Wing Mission: {_os.path.basename(mission_path)} ===")
            log(f"Wind: {wind:+.1f} m/s  |  Phases: {len(results)}")
            log("")
            total_t, total_d = 0.0, 0.0
            for name, t_m, d_k, status in results:
                total_t += t_m; total_d += d_k
                log(f"  {name:<18} {t_m:6.1f} min  {d_k:6.2f} km  — {status}")
            log("")
            log(f"TOTAL:             {total_t:6.1f} min  {total_d:6.2f} km")
            log("")
            log("Switch to Mission Plots tab and select variables to visualise.")
        except Exception as e:
            import traceback
            messagebox.showerror("Mission error", traceback.format_exc())


    # ---- Status TV pairs (for PDF report) ----
    _status_tv_pairs = [
        (batt_tv,  "Battery Status"),
        (motor_tv, "Motor / ESC Status"),
        (aero_tv,  "Aerodynamic Status"),
    ]
    _report_title = "Fixed-Wing Power Simulator — Performance Analysis"
    _last_run_sweep: dict = {}
    _last_run_cfg   = [None]

    def _get_metrics_rows() -> list:
        rows = []
        for iid in metrics_tv.get_children():
            vals = metrics_tv.item(iid, "values")
            if vals and len(vals) >= 2:
                rows.append((str(vals[0]), str(vals[1])))
        return rows

    def _get_status_sections() -> list:
        result = []
        for tv, title in _status_tv_pairs:
            sec_rows = []
            for iid in tv.get_children():
                vals = tv.item(iid, "values")
                tags = tv.item(iid, "tags")
                tag  = tags[0] if tags else "na"
                if vals and len(vals) >= 4:
                    sec_rows.append((str(vals[0]), str(vals[1]),
                                     str(vals[2]), str(vals[3]), str(tag)))
            if sec_rows:
                result.append((title, sec_rows))
        return result

    def _get_log_text() -> str:
        try:
            out_text.configure(state="normal")
            t = out_text.get("1.0", "end")
            out_text.configure(state="disabled")
            return t
        except Exception:
            return ""

    def _get_inputs_rows() -> list:
        rows = []
        for k, var in config_vars.items():
            try:
                val = var.get()
                if val not in ("", None):
                    rows.append((k.replace("_", " ").title(), str(val)))
            except Exception:
                pass
        return rows

    def _do_export_csv():
        if not _last_run_sweep:
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        path = filedialog.asksaveasfilename(
            title="Export CSV", defaultextension=".csv",
            filetypes=[("CSV files","*.csv"),("All files","*.*")])
        if not path: return
        try:
            _export_csv_file(path, _last_run_sweep, _get_metrics_rows())
            messagebox.showinfo("Exported", f"CSV saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))

    def _do_export_excel():
        if not _last_run_sweep:
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        path = filedialog.asksaveasfilename(
            title="Export Excel", defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx"),("All files","*.*")])
        if not path: return
        try:
            cfg = _last_run_cfg[0]
            wb  = _extract_weight_budget(cfg) if cfg else []
            _export_excel_file(path, _last_run_sweep, _get_metrics_rows(), wb)
            messagebox.showinfo("Exported", f"Excel saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))

    def _do_generate_report():
        if not _last_run_sweep and not _get_metrics_rows():
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        path = filedialog.asksaveasfilename(
            title="Save PDF Report", defaultextension=".pdf",
            filetypes=[("PDF files","*.pdf"),("All files","*.*")])
        if not path: return
        try:
            cfg  = _last_run_cfg[0]
            figs = []
            # Add all current plot figures from the scrollable area
            figs.extend(_current_plot_figs)
            if mission_canvas_ref[0] is not None:
                figs.append(mission_canvas_ref[0].figure)
            for num in plt.get_fignums():
                fig = plt.figure(num)
                if fig not in figs:
                    figs.append(fig)
            wb   = _extract_weight_budget(cfg) if cfg else []
            _generate_pdf_report(
                path            = path,
                report_title    = _report_title,
                inputs_rows     = _get_inputs_rows(),
                metrics_rows    = _get_metrics_rows(),
                status_sections = _get_status_sections(),
                log_text        = _get_log_text(),
                figures         = figs,
                weight_budget   = wb,
            )
            messagebox.showinfo("Report generated", f"PDF report saved to:\n{path}")
        except Exception as e:
            import traceback
            messagebox.showerror("Report error", traceback.format_exc())

    # ---- Buttons ----
    btn_frame = ttk.Frame(main)
    btn_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 0))
    btn_frame.columnconfigure(2, weight=1)  # spacer

    ttk.Button(btn_frame, text="▶  Run Single-Point",
               command=run_single_point).grid(row=0, column=0, padx=(0, 6), pady=4)
    ttk.Button(btn_frame, text="🗺  Run Mission (JSON)",
               command=run_mission).grid(row=0, column=1, padx=(0, 6), pady=4)

    # ------------------------------------------------------------------
    # SIMPLE / ADVANCED MODE
    # ------------------------------------------------------------------
    def _apply_field_mode(*_a):
        """
        Show or hide input rows according to the Simple/Advanced setting.

        Rows are removed with grid_remove(), which keeps their grid options so
        re-showing is a plain grid() call.  Hidden fields keep their values, so
        switching modes never changes a result — only what is visible.
        """
        simple = (v_ui_mode.get() == "Simple")
        for row in _field_rows:
            show = (not simple) or (row["key"] in FW_SIMPLE_FIELDS)
            for w in row["widgets"]:
                try:
                    if show: w.grid()
                    else:    w.grid_remove()
                except Exception:
                    pass
        for sec in _section_rows:
            show = (not simple) or (not sec["keys"]) or bool(sec["keys"] & FW_SIMPLE_FIELDS)
            for w in sec["widgets"]:
                try:
                    if show: w.grid()
                    else:    w.grid_remove()
                except Exception:
                    pass

    # ---- Simple / Advanced selector -------------------------------------
    mode_bar = ttk.Frame(left)
    mode_bar.grid(row=0, column=0, sticky="ew", pady=(0, 4))
    ttk.Label(mode_bar, text="Input detail:").pack(side="left", padx=(2, 6))
    for _m in ("Simple", "Advanced"):
        ttk.Radiobutton(mode_bar, text=_m, value=_m,
                        variable=v_ui_mode).pack(side="left", padx=(0, 8))
    ttk.Label(mode_bar,
              text="Simple hides advanced tuning inputs. Hover any ? for help.",
              foreground="#666666", font=("TkDefaultFont", 8)).pack(side="left", padx=(6, 0))
    ttk.Label(mode_bar, text="Config:", foreground="#666666",
              font=("TkDefaultFont", 8)).pack(side="left", padx=(12, 2))
    ttk.Label(mode_bar, textvariable=v_loaded_cfg, foreground="#0B6BCB",
              font=("TkDefaultFont", 8, "bold")).pack(side="left")


    v_ui_mode.trace_add("write", _apply_field_mode)
    _apply_field_mode()          # apply the default (Simple) at startup


    # Startup banner: states the version and points at the two new UI features,
    # so a stale copy of the script is immediately obvious.

    # ---- Save / Load config ----
    def save_cfg():
        path = filedialog.asksaveasfilename(
            title="Save Config", defaultextension=".json",
            filetypes=[("JSON","*.json"),("All","*.*")])
        if not path: return
        # Persist avionics as a list of row dicts so it round-trips cleanly
        av_rows = []
        for iid in av_tree.get_children():
            vals = av_tree.item(iid, "values")
            try:
                av_rows.append({
                    "voltage":    float(vals[0]),
                    "current":    float(vals[1]),
                    "eff":        float(vals[2]),
                })
            except Exception:
                pass
        data = {
            "schema":        "fixedwing_power_sim_v1",
            "vars":          {k: v.get() for k, v in config_vars.items()},
            "avionics_rows": av_rows,
        }
        with open(path, "w") as f:
            json.dump(data, f, indent=2)
        messagebox.showinfo("Saved", f"Configuration saved to:\n{path}")

    def _apply_config_file(path: str):
        """
        Load a saved configuration from `path` into the GUI fields.

        Split out of load_cfg() so the first-launch example autoloader can
        reuse it without going through a file dialog.
        """
        # Keep the mode-bar label in step with whatever was just loaded.
        try:
            v_loaded_cfg.set(os.path.basename(str(path)))
        except Exception:
            pass
        with open(path) as f:
            data = json.load(f)
        for k, val in data.get("vars", {}).items():
            if k in config_vars:
                try: config_vars[k].set("" if val is None else str(val))
                except Exception: pass
        on_unit_mode()
        # Re-apply visibility so restored values respect the current UI mode.
        _apply_field_mode()
        # Restore avionics table — prefer structured rows, fall back to string var
        av_rows = data.get("avionics_rows", None)
        if isinstance(av_rows, list) and av_rows:
            _av_load_rows(av_rows)
        else:
            # Legacy: parse from the string var that was just restored
            try:
                d = parse_voltage_tree(v_avionics_str.get().strip() or None)
                _av_load_rows([{"voltage": v, "current": ci[0], "eff": ci[1]}
                               for v, ci in sorted(d.items())])
            except Exception:
                pass
        # NOTE: no messagebox here on purpose. This function is reusable and
        # may be called before the event loop is running (e.g. autoloading an
        # example at startup); a modal dialog would block forever waiting for
        # a click that cannot arrive. The interactive wrapper below shows the
        # confirmation instead.




    def load_cfg():
        """Prompt for a config file and apply it."""
        path = filedialog.askopenfilename(
            title="Load Config",
            filetypes=[("JSON","*.json"),("All","*.*")])
        if not path:
            return
        try:
            _apply_config_file(path)
        except Exception as exc:
            messagebox.showerror("Error loading config", str(exc))
            return
        messagebox.showinfo("Loaded", f"Configuration loaded from:\n{path}")


    file_menu.add_command(label="Load Config…", command=load_cfg)
    file_menu.add_command(label="Save Config…", command=save_cfg)
    
    # Wire export items into File menu
    file_menu.add_separator()
    file_menu.add_command(label="Export CSV…",          command=_do_export_csv)
    file_menu.add_command(label="Export Excel…",        command=_do_export_excel)
    file_menu.add_command(label="Generate PDF Report…", command=_do_generate_report)

    file_menu.add_separator()
    file_menu.add_command(label="Exit", command=exit_app)

    ttk.Button(btn_frame, text="💾  Save Config", command=save_cfg).grid(row=0, column=3, padx=4)
    ttk.Button(btn_frame, text="📂  Load Config", command=load_cfg).grid(row=0, column=4, padx=4)
    ttk.Button(btn_frame, text="📊  Export CSV",
               command=_do_export_csv).grid(row=0, column=5, padx=4)
    ttk.Button(btn_frame, text="📗  Export Excel",
               command=_do_export_excel).grid(row=0, column=6, padx=4)
    ttk.Button(btn_frame, text="📄  Generate Report",
               command=_do_generate_report).grid(row=0, column=7, padx=4)

    root.protocol("WM_DELETE_WINDOW", exit_app)


    _banner_lines = [
        f"Fixed-Wing UAV Power Simulator  v{SIM_VERSION}",
        "=" * 46,
        "Input detail is set to Simple (selector above the input tabs).",
        "Hover the blue ? beside any field for an explanation and a",
        "typical value. Switch to Advanced to reveal every input.",
        "Metrics sections collapse - click a section heading to fold it.",
        "",
    ]
    _banner_lines.append("Load Config -> examples/configs/ for ready-made aircraft.")
    _banner_lines.append("Run Mission (JSON) -> examples/missions/ for flight profiles.")
    log("\n".join(_banner_lines) + "\n")

    root.mainloop()


# ============================================================
# CLI
# ============================================================
def build_arg_parser():
    p = argparse.ArgumentParser(
        description="Fixed-Wing UAV Power Simulator",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    p.add_argument("--gui", action="store_true", help="Launch Tkinter GUI")

    # Airframe
    p.add_argument("--weight",         type=float, help="Base aircraft weight excluding payload (g)")
    p.add_argument("--payload_mass_g", type=float, default=0.0, help="Payload mass added to base aircraft weight (g)")
    p.add_argument("--num_motors",     type=int,   default=1)
    p.add_argument("--wing_span",      type=float, help="Wing span (m)")
    p.add_argument("--wing_area",      type=float, help="Wing area (m²)")
    p.add_argument("--CD0",            type=float, default=0.028)
    p.add_argument("--CL_max",         type=float, default=1.30)
    p.add_argument("--oswald",         type=float, default=0.82)
    p.add_argument("--mu_roll",        type=float, default=0.04)
    p.add_argument("--mu_brake",       type=float, default=0.08, help="Landing rollout braking friction coefficient")
    p.add_argument("--CL_takeoff",     type=float, default=0.80)
    p.add_argument("--prop_efficiency",type=float, default=0.75)
    p.add_argument("--cruise_speed",   type=float, default=18.0)
    p.add_argument("--periph_current", type=float, default=0.0)

    # Battery
    p.add_argument("--battery_operating_voltage_min",      type=float, default=3.5)
    p.add_argument("--battery_operating_voltage_nominal",  type=float, default=3.8)
    p.add_argument("--battery_operating_voltage_max",      type=float, default=4.2)
    p.add_argument("--battery_series_units",               type=int,   default=4)
    p.add_argument("--battery_parallel_units",             type=int,   default=1)
    p.add_argument("--battery_cells_series_per_unit",      type=int,   default=1)
    p.add_argument("--battery_cells_parallel_per_unit",    type=int,   default=1)
    p.add_argument("--battery_cell_capacity",              type=float, help="Cell capacity (mAh)")
    p.add_argument("--battery_pack_capacity",              type=float, help="Pack capacity (mAh)")
    p.add_argument("--battery_cell_weight_g",              type=float)
    p.add_argument("--battery_pack_weight_g",              type=float)
    p.add_argument("--battery_unit_mode",                  choices=["cell","pack"], default="cell")
    p.add_argument("--battery_energy_density",             type=float)
    p.add_argument("--battery_charge_current_max",         type=float)
    p.add_argument("--battery_discharge_cont_A",           type=float)
    p.add_argument("--battery_discharge_max_A",            type=float)
    p.add_argument("--battery_discharge_c_cont",           type=float)
    p.add_argument("--battery_discharge_c_max",            type=float)
    p.add_argument("--battery_discharge_percent",          type=float, default=80.0)
    p.add_argument("--battery_resistance_cell",            type=float, default=5.0, help="mΩ")
    p.add_argument("--battery_chemistry",                  type=str,   default="LiPo")

    # Motor
    p.add_argument("--motor_kv",           type=float)
    p.add_argument("--motor_idle_current",  type=float, default=0.5)
    p.add_argument("--motor_idle_voltage",  type=float, default=7.0)
    p.add_argument("--motor_rated_voltage", type=int,   default=16)
    p.add_argument("--motor_resistance",    type=float, default=0.06)
    p.add_argument("--motor_max_current",   type=float, default=40)
    p.add_argument("--motor_max_power",     type=float, default=500)
    p.add_argument("--motor_pole_count",    type=int)
    p.add_argument("--motor_weight",        type=float)

    # Propeller
    p.add_argument("--prop_diameter", type=float, default=10)
    p.add_argument("--prop_pitch",    type=float, default=4.5)
    p.add_argument("--prop_blades",   type=int,   default=2)
    p.add_argument("--prop_table",    type=str,   default=None)

    # Environment
    p.add_argument("--altitude",    type=float, default=0.0)
    p.add_argument("--temperature", type=float, default=None)
    p.add_argument("--wind",        type=float, default=0.0)
    p.add_argument("--wind_direction_deg", type=float, default=0.0,
                   help="Wind direction FROM (deg), meteorological convention")
    p.add_argument("--course_deg", type=float, default=0.0,
                   help="Aircraft course heading (deg)")
    p.add_argument("--bank_deg", type=float, default=0.0,
                   help="Bank angle in coordinated turn (deg)")
    p.add_argument("--climb_rate_mps", type=float, default=0.0,
                   help="Commanded climb rate (m/s)")
    p.add_argument("--descent_rate_mps", type=float, default=0.0,
                   help="Commanded descent rate (m/s)")
    p.add_argument("--reserve_percent", type=float, default=20.0,
                   help="Mission reserve as percent of usable energy")
    p.add_argument("--rth_reserve_Wh", type=float, default=0.0,
                   help="Return-to-home reserve (Wh)")
    p.add_argument("--diversion_reserve_Wh", type=float, default=0.0,
                   help="Diversion reserve (Wh)")
    p.add_argument("--plot",        action="store_true")
    # ---- ESC (parity with the multicopter simulator) -------------------
    # Without these, a CLI or batch run silently omits ESC losses while the
    # GUI includes them, so the same config file gave two different answers.
    p.add_argument("--esc_voltage_rating", type=int, default=None,
                   help="ESC voltage rating as a CELL COUNT (the S number), not volts.")
    p.add_argument("--esc_cont_current", type=float, default=None,
                   help="ESC continuous current rating (A).")
    p.add_argument("--esc_max_current", type=float, default=None,
                   help="ESC burst current rating (A).")
    p.add_argument("--esc_idle_current", type=float, default=None,
                   help="Quiescent current drawn by the ESC itself (A).")
    p.add_argument("--esc_resistance", type=float, default=None,
                   help="ESC internal resistance (ohms).")
    p.add_argument("--esc_weight", type=float, default=None,
                   help="Weight of ONE ESC (g).")

    # ---- Environment / propeller extras --------------------------------
    p.add_argument("--cruise_altitude", type=float, default=None,
                   help="Height actually flown at (m ASL), used for the glide-distance "
                        "estimate. Defaults to --altitude (the field elevation).")
    p.add_argument("--prop_eff_model", type=str, default="curve",
                   choices=["curve", "constant"],
                   help="'curve' varies propeller efficiency with advance ratio "
                        "(default); 'constant' uses the flat --prop_efficiency value.")
    p.add_argument("--pressure", type=float, default=None,
                   help="Ambient pressure (Pa). Blank derives it from altitude.")
    p.add_argument("--prop_max_rpm", type=float, default=None,
                   help="Manufacturer prop RPM limit, used for a status check.")
    p.add_argument("--prop_max_thrust", type=float, default=None,
                   help="Max static thrust of one motor+prop (g), for a status check.")
    p.add_argument("--prop_tconst", type=float, default=None,
                   help="Propeller thrust coefficient C_T, if known.")
    p.add_argument("--prop_pconst", type=float, default=None,
                   help="Propeller power coefficient C_P, if known.")
    p.add_argument("--prop_weight", type=float, default=None,
                   help="Weight of ONE propeller (g).")
    p.add_argument("--motor_size", type=str, default=None,
                   help="Motor stator size label, e.g. 2826. Reference only.")

    p.add_argument("--battery_soc_model", type=str, default="auto",
                   help="Battery SoC model: auto/linear/lipo/liion/lifepo4. "
                        "auto picks a preset from --battery_chemistry.")
    p.add_argument("--battery_soc_curve_csv", type=str, default=None,
                   help="CSV with a measured discharge curve. "
                        "Columns: soc, ocv_cell, r_scale.")
    p.add_argument("--battery_soc_bp", type=str, default=None,
                   help="Comma-separated SoC breakpoints in [0,1] (percentages "
                        "above 1.0 are converted automatically).")
    p.add_argument("--battery_ocv_cell_bp", type=str, default=None,
                   help="Comma-separated open-circuit volts PER CELL at each SoC breakpoint.")
    p.add_argument("--battery_r_scale_bp", type=str, default=None,
                   help="Comma-separated resistance multipliers at each SoC breakpoint.")
    p.add_argument(
        "--mission", type=str, default=None,
        help="Path to a mission profile JSON. If omitted, a single-point run is done. "
             "Same schema the GUI uses — see examples/missions/fw_*.json.",
    )
    p.add_argument(
        "--avionics_voltage_tree", type=str, default=None,
        help="Avionics/BEC rails as 'V:(I,eff), V:(I,eff)' — e.g. "
             "'5.0:(1.4,0.9), 12.0:(0.8,0.88)'. Matches the Avionics tab in the GUI.",
    )

    return p


def validate_required_cli_args(args):
    """
    Check that CLI mode has the arguments the physics genuinely needs.

    Only load-bearing inputs are required.  Ratings used purely for status
    checks (motor max current/power, charge current, energy density) and
    inputs that are alternatives to one another are optional — requiring all
    of them made batch runs and the bundled example configs fail for no
    physical reason.
    """
    required = [
        "weight", "wing_span", "wing_area",
        "battery_resistance_cell",
        "motor_kv", "motor_resistance",
        "prop_diameter", "prop_pitch",
    ]
    missing = [k for k in required if getattr(args, k) is None]

    # Capacity: need whichever pair matches the unit mode.
    unit_mode = str(getattr(args, "battery_unit_mode", "cell") or "cell").strip().lower()
    if unit_mode == "pack":
        if getattr(args, "battery_pack_capacity", None) is None:
            missing.append("battery_pack_capacity")
    else:
        if getattr(args, "battery_cell_capacity", None) is None:
            missing.append("battery_cell_capacity")
    if missing:
        raise SystemExit(
            f"Missing required CLI args: {', '.join('--' + m for m in missing)}\n"
            f"Tip: run with --gui to use the graphical interface."
        )



def _load_mission_or_exit(path: str) -> "MissionProfile":
    """
    Load a mission JSON for a CLI run, failing with a readable message.

    A mistyped path or a malformed file is ordinary user error, so it should
    produce a one-line explanation rather than a Python traceback.
    """
    import os
    if not os.path.exists(path):
        raise SystemExit(
            f"Mission file not found: {path}\n"
            "Check the path, or try one of the bundled examples in "
            "examples/missions/."
        )
    try:
        return MissionProfile.from_json(path)
    except json.JSONDecodeError as e:
        raise SystemExit(f"Mission file {path} is not valid JSON: {e}")
    except (KeyError, TypeError, ValueError) as e:
        raise SystemExit(
            f"Mission file {path} could not be read: {e}\n"
            "Each entry in \"phases\" needs a name and either a duration or a "
            "distance. See examples/missions/ for working files."
        )


def main():
    parser = build_arg_parser()
    args   = parser.parse_args()

    if args.gui:
        launch_gui()
        return

    # ---- CLI mode ----
    validate_required_cli_args(args)
    # Honour an explicit pressure override, matching the GUI and the
    # multicopter simulator; blank falls back to the ISA pressure profile.
    rho = isa_density(args.altitude, args.temperature, getattr(args, "pressure", None))
    print(f"Air density: {rho:.4f} kg/m³  at altitude {args.altitude:.0f} m")

    af = AirframeConfig(
        wing_span_m     = args.wing_span,
        wing_area_m2    = args.wing_area,
        CD0             = args.CD0,
        CL_max          = args.CL_max,
        oswald          = args.oswald,
        mu_roll         = args.mu_roll,
        mu_brake        = args.mu_brake,
        CL_takeoff      = args.CL_takeoff,
        prop_efficiency = args.prop_efficiency,
        prop_eff_model  = getattr(args, "prop_eff_model", "curve"),
        num_motors      = args.num_motors,
    )

    batt = BatteryConfig(
        chemistry               = args.battery_chemistry,
        operating_voltage_min   = args.battery_operating_voltage_min,
        operating_voltage_nominal = args.battery_operating_voltage_nominal,
        operating_voltage_max   = args.battery_operating_voltage_max,
        unit_mode               = args.battery_unit_mode,
        series_units            = args.battery_series_units,
        parallel_units          = args.battery_parallel_units,
        cells_series_per_unit   = args.battery_cells_series_per_unit,
        cells_parallel_per_unit = args.battery_cells_parallel_per_unit,
        cell_capacity_mAh       = args.battery_cell_capacity,
        pack_capacity_mAh       = args.battery_pack_capacity,
        cell_weight_g           = args.battery_cell_weight_g,
        pack_weight_g           = args.battery_pack_weight_g,
        unit_energy_density     = args.battery_energy_density,
        charge_current_max      = args.battery_charge_current_max,
        discharge_cont_A        = args.battery_discharge_cont_A,
        discharge_max_A         = args.battery_discharge_max_A,
        discharge_c_cont        = args.battery_discharge_c_cont,
        discharge_c_max         = args.battery_discharge_c_max,
        discharge_percent       = args.battery_discharge_percent,
        resistance_cell_mOhm    = args.battery_resistance_cell,
        soc_model               = getattr(args, "battery_soc_model", "auto"),
        soc_curve_csv           = getattr(args, "battery_soc_curve_csv", None),
        soc_bp                  = parse_soc_breakpoints(getattr(args, "battery_soc_bp", None)),
        ocv_cell_bp             = parse_float_list(getattr(args, "battery_ocv_cell_bp", None)),
        r_scale_bp              = parse_float_list(getattr(args, "battery_r_scale_bp", None)),
    )

    motor = MotorConfig(
        kv            = args.motor_kv,
        idle_current  = args.motor_idle_current,
        idle_voltage  = args.motor_idle_voltage,
        rated_voltage = args.motor_rated_voltage,
        resistance    = args.motor_resistance,
        max_current   = args.motor_max_current,
        max_power     = args.motor_max_power,
        pole_count    = args.motor_pole_count,
        weight_g      = args.motor_weight,
        size_mm       = args.motor_size,
    )

    prop = PropellerConfig(
        diameter_in  = args.prop_diameter,
        pitch_in     = args.prop_pitch,
        blades       = args.prop_blades,
        max_rpm      = (args.prop_max_rpm    if args.prop_max_rpm    is not None else 0.0),
        max_thrust_g = (args.prop_max_thrust if args.prop_max_thrust is not None else 0.0),
        table_csv    = args.prop_table,
        TConst       = args.prop_tconst,
        PConst       = args.prop_pconst,
        weight_g     = args.prop_weight,
    )

    # ESC is optional: build one only if the user supplied any ESC argument.
    # Omitting it means ESC losses are simply not modelled (loss = 0), which
    # is what happened implicitly before these arguments existed.
    esc_cli = None
    if any(getattr(args, k, None) is not None for k in
           ("esc_voltage_rating", "esc_cont_current", "esc_max_current",
            "esc_idle_current", "esc_resistance", "esc_weight")):
        esc_cli = ESCConfig(
            voltage_rating       = (args.esc_voltage_rating
                                    if args.esc_voltage_rating is not None
                                    else int(args.battery_series_units or 1)),
            continuous_current_A = (args.esc_cont_current  if args.esc_cont_current  is not None else 0.0),
            max_current_A        = (args.esc_max_current   if args.esc_max_current   is not None else 0.0),
            idle_current_A       = (args.esc_idle_current  if args.esc_idle_current  is not None else 0.0),
            resistance           = (args.esc_resistance    if args.esc_resistance    is not None else 0.0),
            weight_g             = args.esc_weight,
        )

    # Avionics rails are optional on the CLI; when given they add real load,
    # which is why the GUI keeps this tab visible even in Simple mode.
    avionics_cli = None
    _av_spec = getattr(args, "avionics_voltage_tree", None)
    if _av_spec:
        avionics_cli = AvionicsConfig(voltage_tree=parse_voltage_tree(_av_spec))

    cfg = FixedWingConfig(
        airframe          = af,
        battery           = batt,
        motor             = motor,
        propeller         = prop,
        aircraft_weight_g = args.weight,
        payload_mass_g    = max(float(args.payload_mass_g or 0.0), 0.0),
        cruise_speed_mps  = args.cruise_speed,
        periph_current_A  = args.periph_current,
        esc               = esc_cli,
        avionics          = avionics_cli,
        air_density       = rho,
        reference_altitude_m = args.altitude,
        cruise_altitude_m    = getattr(args, "cruise_altitude", None),
    )

    # ---- Mission mode -------------------------------------------------
    # When --mission is supplied we run the phase-by-phase simulation instead
    # of a single operating point, mirroring "Run Mission (JSON)" in the GUI.
    if getattr(args, "mission", None):
        mission = _load_mission_or_exit(args.mission)
        # CLI flags act as defaults for any phase that does not set its own.
        mission.reserve_percent      = float(args.reserve_percent)
        mission.rth_reserve_Wh       = float(args.rth_reserve_Wh)
        mission.diversion_reserve_Wh = float(args.diversion_reserve_Wh)
        mission.wind_direction_deg   = float(args.wind_direction_deg)
        for _p in mission.phases:
            _p.bank_deg   = float(_p.bank_deg   if _p.bank_deg   else args.bank_deg)
            _p.course_deg = float(_p.course_deg if _p.course_deg else args.course_deg)
            if _p.climb_rate_mps is None:
                _p.climb_rate_mps = float(args.climb_rate_mps)
            if _p.descent_rate_mps is None:
                _p.descent_rate_mps = float(args.descent_rate_mps)

        results, worst_m, series = simulate_fw_mission(
            cfg, mission,
            wind_mps      = float(args.wind),
            temperature_C = (float(args.temperature) if args.temperature is not None else None),
            pressure_Pa   = (float(getattr(args, "pressure", None))
                             if getattr(args, "pressure", None) is not None else None),
        )

        print("\n=== MISSION RESULTS ===")
        print(f"{'Phase':<28}{'Time (min)':>12}{'Dist (km)':>12}  Status")
        tot_t = tot_d = 0.0
        for name, t_min, d_km, status in results:
            print(f"{str(name):<28}{t_min:>12.2f}{d_km:>12.3f}  {status}")
            tot_t += float(t_min); tot_d += float(d_km)
        print("-" * 66)
        print(f"{'TOTAL':<28}{tot_t:>12.2f}{tot_d:>12.3f}")
        if worst_m:
            print("\n--- Worst-case operating point ---")
            for k in ("total_power_W", "pack_current_A", "v_load_V",
                      "thrust_required_N", "CL", "LD_ratio",
                      "reserve_margin_Wh", "thermal_status"):
                if k in worst_m:
                    val = worst_m[k]
                    print(f"  {k:<24}: {val:.3f}" if isinstance(val, (int, float))
                          else f"  {k:<24}: {val}")
        return

    # ---- Single-point mode --------------------------------------------
    V_cruise = args.cruise_speed
    headwind, crosswind = wind_components_mps(args.wind, args.wind_direction_deg, args.course_deg)
    m = compute_metrics(
        cfg, V_cruise,
        bank_deg=args.bank_deg,
        ambient_temp_C=(float(args.temperature) if args.temperature is not None else 25.0),
        wind_head_mps=headwind,
        wind_cross_mps=crosswind,
        glide_altitude_m=(args.cruise_altitude if getattr(args, "cruise_altitude", None)
                          is not None else args.altitude),
    )
    climb_rate = max(float(args.climb_rate_mps), 0.0)
    descent_rate = max(float(args.descent_rate_mps), 0.0)
    if climb_rate > 0 and descent_rate > 0:
        descent_rate = 0.0
    potential_power_W = cfg.weight_N * (climb_rate - descent_rate)
    base_total = float(m.get("total_power_W", 0.0))
    m["climb_rate_cmd_mps"] = climb_rate
    m["descent_rate_cmd_mps"] = descent_rate
    m["potential_power_W"] = potential_power_W
    m["climb_power_add_W"] = potential_power_W
    m["wind_head_mps"] = headwind
    m["wind_cross_mps"] = crosswind
    m["groundspeed_mps"] = groundspeed_along_track_mps(V_cruise, headwind, crosswind)
    m["reserve_target_Wh"] = max(
        cfg.battery.usable_Wh * (float(args.reserve_percent) / 100.0),
        float(args.rth_reserve_Wh) + float(args.diversion_reserve_Wh),
    )
    m["reserve_margin_Wh"] = cfg.battery.usable_Wh - m["reserve_target_Wh"]
    m["reserve_breached"] = bool(m["reserve_margin_Wh"] < 0)
    adj_total = max(base_total + potential_power_W, 0.0)
    if base_total > 0:
        scale = adj_total / base_total
        m["motor_power_W"] = float(m.get("motor_power_W", 0.0)) * scale
        m["esc_loss_W"] = float(m.get("esc_loss_W", 0.0)) * scale
    m["total_power_W"] = adj_total
    m["pack_current_A"] = adj_total / max(cfg.battery.vnom_pack, 1.0)
    m["v_load_V"] = cfg.battery.voltage_under_load(m["pack_current_A"])
    if adj_total > 0:
        m["flight_time_min"] = (cfg.battery.usable_Wh / adj_total) * 60.0
        m["flight_range_km"] = m["groundspeed_mps"] * (m["flight_time_min"] * 60.0) / 1000.0
    else:
        m["flight_time_min"] = 0.0
        m["flight_range_km"] = 0.0

    print(f"\n{'='*55}")
    print(f"  Fixed-Wing UAV Performance @ {V_cruise:.1f} m/s ({V_cruise*3.6:.1f} km/h)")
    print(f"{'='*55}")
    print(f"  Wing span / area      : {af.wing_span_m:.2f} m / {af.wing_area_m2:.3f} m²")
    print(f"  Chord / AR            : {af.chord_m:.3f} m / {af.aspect_ratio:.2f}")
    print(f"  Reynolds Number       : {m['reynolds_number']:,.0f}")
    print(f"  Stall Speed           : {m['stall_speed_mps']:.2f} m/s ({m['stall_speed_mps']*3.6:.1f} km/h)")
    print(f"  Turn Stall Speed      : {m.get('stall_speed_turn_mps',0):.2f} m/s @ {args.bank_deg:.1f}° bank")
    print(f"  Wing Loading          : {m['wing_loading_N_m2']:.1f} N/m² ({m['wing_loading_kg_m2']:.2f} kg/m²)")
    print(f"  CL / CD               : {m['CL']:.4f} / {m['CD']:.5f}")
    print(f"  L/D Ratio             : {m['LD_ratio']:.2f}")
    print(f"  Groundspeed           : {m.get('groundspeed_mps',0):.2f} m/s")
    print(f"  Headwind / Crosswind  : {m.get('wind_head_mps',0):+.2f} / {m.get('wind_cross_mps',0):+.2f} m/s")
    print(f"  Glide Ratio           : {m['glide_ratio']:.2f}:1")
    print(f"  Glide Distance        : {m['glide_distance_m']:.1f} m (from {args.altitude:.0f} m altitude)")
    print(f"  Angle of Attack       : {m['aoa_deg']:.2f} °")
    print(f"  Drag (thrust req)     : {m['drag_N']:.2f} N")
    print(f"  Induced / Parasitic D : {m['induced_drag_N']:.2f} / {m['parasitic_drag_N']:.2f} N")
    print(f"  Thrust available      : {m['thrust_available_N']:.2f} N")
    print(f"  Specific Thrust (T/W) : {m['specific_thrust']:.3f}")
    print(f"  Turn Radius / Rate    : {m.get('turn_radius_m', float('inf')):.1f} m / {m.get('turn_rate_deg_s',0):.2f} deg/s")
    print(f"  Load Factor           : {m.get('load_factor',1.0):.2f} g")
    print(f"  Tip Speed             : {m['tip_speed_mps']:.1f} m/s")
    print(f"  Pitch Speed           : {m['pitch_speed_mps']:.1f} m/s")
    print(f"  Motor Power           : {m['motor_power_W']:.1f} W")
    print(f"  Total Elec. Power     : {m['total_power_W']:.1f} W")
    print(f"  Potential Power Term  : {m.get('potential_power_W',0):+.1f} W")
    print(f"  Max Prop Power        : {m['max_prop_power_W']:.1f} W")
    print(f"  Rate of Climb         : {m['rate_of_climb_mps']*60:.1f} m/min")
    print(f"  Max Rate of Climb     : {m['max_rc_mps']*60:.1f} m/min @ {m['v_max_rc_mps']:.1f} m/s")
    print(f"  Max Angle of Climb    : {m['max_aoc_deg']:.1f} °  @ {m['v_max_aoc_mps']:.1f} m/s")
    if math.isfinite(m['service_ceiling_m']):
        print(f"  Service Ceiling       : {m['service_ceiling_m']:.0f} m ASL ({m['service_ceiling_agl_m']:.0f} m AGL)")
    else:
        print(f"  Service Ceiling       : > 12000 m ASL")
    print(f"  Takeoff Ground Roll   : {m['takeoff_dist_m']:.1f} m")
    print(f"  Landing Distance      : {m['landing_dist_m']:.1f} m")
    print(f"  Best Endurance Speed  : {m['best_endurance_speed_mps']:.1f} m/s ({m['best_endurance_speed_mps']*3.6:.1f} km/h)")
    print(f"  Best Range Speed      : {m['best_range_speed_mps']:.1f} m/s ({m['best_range_speed_mps']*3.6:.1f} km/h)")
    print(f"  Minimum Sink Speed    : {m['min_sink_speed_mps']:.1f} m/s")
    print(f"  Minimum Sink Rate     : {m['min_sink_rate_mps']:.2f} m/s")
    _b = cfg.battery
    print(f"  SoC model             : {_soc_model_short_label(getattr(_b,'soc_model_source',None))}"
          f" ({'non-linear' if getattr(_b,'soc_nonlinear_enabled',False) else 'linear'})")
    print(f"  Flight Time           : {m['flight_time_min']:.1f} min")
    print(f"  Flight Range          : {m['flight_range_km']:.2f} km")
    print(f"  Specific Range        : {m['specific_range_m_per_Wh']:.1f} m/Wh ({m['specific_range_km_per_kWh']:.1f} km/kWh)")
    print(f"  Specific Endurance    : {m['specific_endurance_min_per_Wh']:.3f} min/Wh ({m['specific_endurance_h_per_kWh']:.3f} h/kWh)")
    print(f"  Reserve Target/Margin : {m.get('reserve_target_Wh',0):.1f} / {m.get('reserve_margin_Wh',0):+.1f} Wh")
    print(f"  Reserve Status        : {'VIOLATION' if m.get('reserve_breached', False) else 'OK'}")
    print(f"  Thermal (M/ESC/Batt)  : {m.get('motor_temp_est_C',0):.1f}/{m.get('esc_temp_est_C',0):.1f}/{m.get('battery_temp_est_C',0):.1f} °C [{m.get('thermal_status','OK')}]")
    print(f"{'='*55}\n")

    V_be, t_best, V_br, d_best = find_optimal_speeds(cfg)
    print(f"  Best endurance: {V_be:.1f} m/s → {t_best:.1f} min")
    print(f"  Best range    : {V_br:.1f} m/s → {d_best:.2f} km")

    if args.plot:
        fig = make_performance_figure(cfg)
        plt.show()


if __name__ == "__main__":
    main()
