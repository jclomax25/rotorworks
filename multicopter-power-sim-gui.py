"""
multicopter-power-sim.py
------------------
Multicopter performance simulator with three modeling modes:
1. Motor test table (CSV input with thrust vs power).
2. Motor electrical model (using KV, idle current, resistance, limits).
3. Theoretical induced velocity model (simplified).

Features:
- Power consumption (hover + forward flight)
- Flight time & distance
- Best endurance & best range speeds
- Plotting of performance curves
- Mission profile simulation (JSON)
- NEW: Optional Tkinter GUI for entering inputs and viewing plots in one window.
      CLI interface is preserved.

Examples (CLI):
    python multicopter-power-sim.py --num_motors 4 --weight 1.5 --area 0.05 \
        --battery_operating_voltage_min 3.0 --battery_operating_voltage_max 4.2 \
        --battery_capacity 5000 --battery_weight 400 --battery_energy_density 200 \
        --battery_charge_current_max 5 --battery_discharge_cont 60 --battery_resistance_cell 20 \
        --battery_cell_count 4 --battery_chemistry LiIon \
        --motor_kv 650 --motor_idle_current 0.5 --motor_resistance 0.2 --motor_max_current 20 --motor_max_power 200 \
        --prop_diameter 12 --prop_pitch 6 \
        --speed 10 --plot

    # Use motor/prop test table for power interpolation:
    python multicopter-power-sim.py ... --prop_table motor_data.csv --plot

    # Run GUI:
    python multicopter-power-sim.py --gui
"""

from __future__ import annotations

from html import parser
import math
import argparse
import json
import os
import re
from dataclasses import dataclass
import sys
from typing import Optional, List, Tuple

import pandas as pd
import numpy as np

# ------------------------------------------------------------------
# Shared core, extracted so the two simulators cannot drift apart.
# See rotorworks_core.py for what lives there and what deliberately
# does not. It must sit beside this file.
# ------------------------------------------------------------------
try:
    import rotorworks_core as core
except ImportError as _exc:      # pragma: no cover - install/deploy problem
    raise SystemExit(
        "rotorworks_core.py could not be imported. It must sit in the same "
        f"folder as this script.\nOriginal error: {_exc}"
    )

# Names re-exported under their historical spellings so existing call sites,
# saved scripts and the test suite keep working unchanged.
SOC_PRESETS = core.SOC_PRESETS
SOC_PRESET_ALIASES = core.SOC_PRESET_ALIASES
Tooltip = _Tooltip = core.Tooltip
parse_float_list = core.parse_float_list
parse_soc_breakpoints = core.parse_soc_breakpoints
wind_components_mps = core.wind_components_mps
groundspeed_along_track_mps = core.groundspeed_along_track_mps
thermal_step = core.thermal_step
_eval_poly = core.eval_poly
_interp_linear_clamped = core.interp_linear_clamped
_battery_preset_key = core.battery_preset_key
_normalize_soc_curves = core.normalize_soc_curves
_load_soc_curve_csv = core.load_soc_curve_csv
_configure_battery_soc_model = core.configure_battery_soc_model
_soc_model_short_label = core.soc_model_short_label
battery_pack_ocv_from_soc = core.pack_ocv_from_soc
battery_pack_resistance_from_soc = core.pack_resistance_from_soc
battery_voltage_under_load = core.pack_voltage_under_load
battery_soc_after_energy_draw = core.soc_after_energy_draw
_fit_propeller_curve = core.fit_propeller_curve
kinetic_power_term_W = core.kinetic_power_term_W
ramp_speed = core.ramp_speed

# Build identifier. Shown in the title bar, the Output pane and Help > About
# so you can always tell which copy of the script you are running.
SIM_VERSION = "2.12.0"
SIM_BUILD_NOTE = "Airspeed-dependent thrust available; climb rates corrected"
import matplotlib.pyplot as plt

# -------------------------------
# Constants
# -------------------------------
AIR_DENSITY = 1.225  # kg/m^3 (sea level ISA)

# Default rotor inflow efficiency map in forward flight.
# eta > 1.0 means translational lift improves rotor efficiency at that advance ratio.
DEFAULT_INFLOW_MU_BP = [0.0, 0.08, 0.16, 0.24, 0.32, 0.40, 0.50]
DEFAULT_INFLOW_EFF_BP = [1.00, 1.04, 1.08, 1.06, 1.00, 0.94, 0.88]

# Gas constant / ISA params
R = 287.05   # J/kg/K (specific gas constant for dry air)
T0 = 288.15  # K (sea level standard temp, 15°C)
P0 = 101325  # Pa (sea level standard pressure)
L = 0.0065   # K/m (temperature lapse rate)
g0 = 9.80665 # m/s^2

# Generic nonlinear SoC templates by chemistry:
# curves are intentionally conservative and only approximate behavior.



# -------------------------------
# Battery Model
# -------------------------------
class BatteryConfig:
    """
    Simple pack model:
      - nominal voltage taken as Vmax_cell * Ncells
      - internal resistance modeled as Rcell * Ncells (series)
      - under-load voltage V = Vnom - I*Rpack (clamped at Vmin_cell*Ncells)
    """
    def __init__(self,
                 chemistry: Optional[str],
                 operating_voltage_min: float,  # V per cell
                 operating_voltage_nominal: float,  # V per cell
                 operating_voltage_max: float,  # V per cell
                 unit_mode: str = "cell",  # "cell" or "pack" (for voltage inputs)
                 series_units: int = 1,  # number of units (cells or packs) in series
                 parallel_units: int = 1,  # number of parallel units (cells or packs)
                 cells_series_per_unit: int = 1,  # for "pack" mode: how many cells in series per pack
                 cells_parallel_per_unit: int = 1,  # for "pack" mode: how many cells in parallel per pack
                 pack_weight_g: float = None,
                 cell_weight_g: float = None,
                 cell_capacity_mAh: float = None,
                 pack_capacity_mAh: float = None,
                 unit_energy_density: float = None,
                 max_operating_temperature_C: Optional[float] = None,
                 min_operating_temperature_C: Optional[float] = None,
                 charge_current_max: float = None,
                 discharge_cont_A: Optional[float] = None,
                 discharge_max_A: Optional[float] = None,
                 discharge_c_cont: Optional[float] = None,
                 discharge_c_max: Optional[float] = None,
                 discharge_percent: float = 100.0,
                 resistance_cell_mOhm: float = 0.0,
                 soc_model: str = "auto",
                 soc_curve_csv: Optional[str] = None,
                 soc_bp: Optional[List[float]] = None,
                 ocv_cell_bp: Optional[List[float]] = None,
                 r_scale_bp: Optional[List[float]] = None):
        self.chemistry = chemistry
        self.operating_voltage_min = float(operating_voltage_min)
        self.operating_voltage_nominal = float(operating_voltage_nominal)
        self.operating_voltage_max = float(operating_voltage_max)
        self.unit_mode = str(unit_mode).strip().lower() if unit_mode is not None else "cell"
        # Counts must be at least 1 — a 0 or negative count would silently
        # produce a zero-volt / zero-capacity pack and divide-by-zero later.
        self.series_units = max(int(series_units), 1)
        self.parallel_units = max(int(parallel_units), 1)

        if self.unit_mode not in ("cell", "pack"):
            # Fallback if the GUI or args provide a non-standard string.
            self.unit_mode = "pack" if pack_weight_g is not None else "cell"

        if self.unit_mode == "cell":
            cells_series_per_unit = 1
            cells_parallel_per_unit = 1

        self.cells_series_per_unit = max(int(cells_series_per_unit), 1)
        self.cells_parallel_per_unit = max(int(cells_parallel_per_unit), 1)

        # Final pack layout in CELLS
        self.series_cells = self.series_units * self.cells_series_per_unit
        self.parallel_cells = self.parallel_units * self.cells_parallel_per_unit
        self.total_cells = self.series_cells * self.parallel_cells

        # Derived pack voltage
        self.vmin_pack = self.operating_voltage_min * self.series_cells
        self.vmax_pack = self.operating_voltage_max * self.series_cells
        self.vnom_pack = self.operating_voltage_nominal * self.series_cells

        self.pack_weight_g = float(pack_weight_g) if pack_weight_g is not None else None
        self.cell_weight_g = float(cell_weight_g) if cell_weight_g is not None else None
        self.cell_capacity_mAh = float(cell_capacity_mAh) if cell_capacity_mAh is not None else None
        self.pack_capacity_mAh = float(pack_capacity_mAh) if pack_capacity_mAh is not None else None
        self.max_operating_temperature_C = float(max_operating_temperature_C) if max_operating_temperature_C is not None else None
        self.min_operating_temperature_C = float(min_operating_temperature_C) if min_operating_temperature_C is not None else None
        self.charge_current_max = float(charge_current_max) if charge_current_max is not None else 0.0

        # ------------------------------------------------------------------
        # Capacity in mAh / Ah  — must be computed before energy density.
        #
        # PHYSICS: capacity (Ah) is set by the number of PARALLEL branches.
        # Connecting units in SERIES raises voltage, NOT capacity.  Total
        # energy still rises with series count because  E = C_Ah x V_pack
        # and V_pack scales with series count.
        #
        #   6S1P of 5000 mAh packs ->  5000 mAh @ 22.2 V =  111 Wh
        #   6S2P of 5000 mAh packs -> 10000 mAh @ 22.2 V =  222 Wh
        #  12S1P of 5000 mAh packs ->  5000 mAh @ 44.4 V =  222 Wh   <- same energy, 2x voltage
        #  12S2P of 5000 mAh packs -> 10000 mAh @ 44.4 V =  444 Wh
        #
        # NOTE: branch on self.unit_mode (normalised/lower-cased), never the
        # raw `unit_mode` argument — "Pack"/"PACK"/" cell" would otherwise
        # fall through and silently zero the capacity and weight.
        # ------------------------------------------------------------------
        if self.unit_mode == "cell":
            # One "unit" is a cell: parallel cells set capacity.
            self.capacity_mAh = (self.cell_capacity_mAh or 0.0) * self.parallel_cells
        else:  # "pack"
            # One "unit" is a pack: only packs wired in PARALLEL add capacity.
            self.capacity_mAh = (self.pack_capacity_mAh or 0.0) * self.parallel_units
        self.capacity_Ah = self.capacity_mAh / 1000.0

        # Weight DOES scale with the total number of units (series x parallel),
        # because every physical unit is carried regardless of how it is wired.
        if self.unit_mode == "cell":
            self.weight_g = (self.cell_weight_g or 0.0) * self.total_cells
        else:  # "pack"
            self.weight_g = (self.pack_weight_g or 0.0) * self.series_units * self.parallel_units

        # Energy density — must follow capacity_Ah and weight_g
        if unit_energy_density is not None:
            self.energy_density_Wh_per_kg = float(unit_energy_density)
        else:
            wkg = self.weight_g / 1000.0
            self.energy_density_Wh_per_kg = (self.capacity_Wh / wkg) if wkg > 0 else 0.0

        # Discharge limits:
        # - If you provide discharge_cont_A (legacy), we use it directly (amps).
        # - Otherwise, if you provide discharge_c_cont, we compute I_cont = C_cont * capacity_Ah.
        # - If you provide discharge_c_max, we compute I_max = C_max * capacity_Ah (burst limit).
        if discharge_cont_A is not None:
            self.discharge_cont_A = float(discharge_cont_A)
        elif discharge_c_cont is not None:
            self.discharge_cont_A = float(discharge_c_cont) * self.capacity_Ah
        else:
            self.discharge_cont_A = float("inf")  # No limit specified (not recommended)

        if discharge_max_A is not None:
            self.discharge_max_A = float(discharge_max_A)
        elif discharge_c_max is not None:
            self.discharge_max_A = float(discharge_c_max) * self.capacity_Ah
        else:
            # Default: max equals continuous if not specified
            self.discharge_max_A = float(self.discharge_cont_A)

        # Store C-rates (if derivable)
        if discharge_c_cont is not None:
            self.discharge_c_cont = float(discharge_c_cont)
        else:
            self.discharge_c_cont = (self.discharge_cont_A / self.capacity_Ah) if self.capacity_Ah > 0 else None

        if discharge_c_max is not None:
            self.discharge_c_max = float(discharge_c_max)
        else:
            self.discharge_c_max = (self.discharge_max_A / self.capacity_Ah) if self.capacity_Ah > 0 else None

        # Usable fraction of the pack energy (e.g., 80% -> stop at 20% remaining)
        self.discharge_percent = float(discharge_percent)
        self.discharge_percent = min(max(self.discharge_percent, 0.0), 100.0)
        self.usable_fraction = self.discharge_percent / 100.0

        self.resistance_cell = float(resistance_cell_mOhm) / 1000.0  # Ω
        self.soc_model = str(soc_model or "auto").strip().lower()
        self.soc_curve_csv = (str(soc_curve_csv).strip() if soc_curve_csv else None)

        # Nonlinear SoC model state (populated by helper).
        self.soc_nonlinear_enabled: bool = False
        self.soc_model_source: str = "linear-fallback"
        self.soc_bp: List[float] = []
        self.ocv_cell_bp: List[float] = []
        self.r_scale_bp: List[float] = []
        _configure_battery_soc_model(
            self,
            model=self.soc_model,
            curve_csv=self.soc_curve_csv,
            soc_bp=soc_bp,
            ocv_cell_bp=ocv_cell_bp,
            r_scale_bp=r_scale_bp,
        )

    # ---- SoC API -----------------------------------------------------
    # Thin wrappers over the module-level helpers below, so that this class
    # exposes the SAME method names as the fixed-wing simulator's
    # BatteryConfig. Same physics either way; having one API makes the two
    # interchangeable in shared code and in tests.
    def ocv_at_soc(self, soc: float) -> float:
        """Pack open-circuit voltage at a given state of charge (0..1)."""
        return battery_pack_ocv_from_soc(self, soc)

    def resistance_at_soc(self, soc: float) -> float:
        """Pack internal resistance at a given state of charge (0..1)."""
        return battery_pack_resistance_from_soc(self, soc)

    def voltage_under_load(self, current_A: float,
                           soc: Optional[float] = None) -> float:
        """Loaded pack voltage; soc defaults to fully charged."""
        return battery_voltage_under_load(self, current_A, soc)

    def soc_after_energy_draw(self, soc_now: float, energy_draw_Wh: float) -> float:
        """Advance SoC after drawing a given amount of energy."""
        return battery_soc_after_energy_draw(self, soc_now, energy_draw_Wh)

    @property
    def pack_resistance(self) -> float:
        return self.resistance_cell * self.series_cells / self.parallel_cells

    @property
    def capacity_Wh(self) -> float:
        return self.capacity_Ah * self.vnom_pack

    @property
    def usable_Wh(self) -> float:
        return self.capacity_Wh * self.usable_fraction


def battery_ocv_pack(battery: BatteryConfig, soc: float) -> float:
    return battery_pack_ocv_from_soc(battery, soc)


def battery_pack_resistance(battery: BatteryConfig, soc: float) -> float:
    return battery_pack_resistance_from_soc(battery, soc)


# -------------------------------
# Motor Model
# -------------------------------
class MotorConfig:
    def __init__(self,
                 kv: Optional[float],
                 idle_current: float,
                 idle_voltage: float,
                 rated_voltage: int, #S rating for ESC compatibility checks
                 resistance: float,
                 max_current: Optional[float] = None,
                 max_power: Optional[float] = None,
                 pole_count: Optional[int] = None,
                 weight_g: Optional[float] = None,
                 size_mm: Optional[str] = None):
        self.kv = None if kv is None else float(kv)   # RPM/V
        self.idle_current = float(idle_current or 0.0)   # A
        self.idle_voltage = float(idle_voltage or 10.0)
        self.resistance = float(resistance)              # Ω
        # max_current / max_power are rating limits used only for status
        # checks.  None means "no rating supplied" — the check is skipped
        # rather than the whole run failing.
        self.max_current = None if max_current is None else float(max_current)
        self.max_power   = None if max_power   is None else float(max_power)
        self.rated_voltage = int(rated_voltage or 0)
        self.pole_count = pole_count
        self.weight_g = weight_g
        self.size_mm = size_mm

# -------------------------------
# ESC Model
# -------------------------------
class ESCConfig:
    def __init__(self,
                 voltage_rating: int,  # S rating for ESC compatibility checks
                 continuous_current_A: float,
                 max_current_A: float,
                 idle_current_A: float,
                 resistance: float,
                 weight_g: Optional[float] = None):
        self.voltage_rating = int(voltage_rating)
        self.continuous_rating_A = float(continuous_current_A)
        self.max_current_A = float(max_current_A)
        self.idle_current_A = float(idle_current_A)
        self.weight_g = weight_g
        self.resistance = float(resistance)           # Ω

# -------------------------------
# Avionics/Peripherals Model
# -------------------------------
class AvionicsConfig:
    def __init__(self,
                voltage_tree: Optional[dict] = None):
        self.voltage_tree = voltage_tree or {}  # e.g., {5.0: (2, 0.9), 12.0: (1.5, 0.85)} translates to 5V with 2A load, 90% BEC efficiency, and 12V with 1.5A load, 85% BEC efficiency}


def parse_voltage_tree(spec: Optional[str]) -> dict:
    """Parse an avionics voltage tree specification into {V_rail: (I_rail, eff)}.

    Accepted formats (comma-separated rails, commas inside parentheses are ignored):
      - "5.0:(2,0.9), 12.0:(1.5,0.85)"
      - "5.0:2:0.9, 12.0:1.5:0.85"

    Returns:
      dict[float, tuple[float, float]]
    """
    if spec is None:
        return {}
    if isinstance(spec, dict):
        out = {}
        for k, v in spec.items():
            try:
                vk = float(k)
                if isinstance(v, (list, tuple)) and len(v) == 2:
                    out[vk] = (float(v[0]), float(v[1]))
                else:
                    raise ValueError
            except Exception as e:
                raise ValueError(f"Invalid voltage_tree entry {k!r}: {v!r}") from e
        return out

    s = str(spec).strip()
    if not s:
        return {}

    # Split on commas that are NOT inside parentheses
    parts, depth, buf = [], 0, ""
    for ch in s:
        if ch == "(":
            depth += 1; buf += ch
        elif ch == ")":
            depth -= 1; buf += ch
        elif ch == "," and depth == 0:
            parts.append(buf.strip()); buf = ""
        else:
            buf += ch
    if buf.strip():
        parts.append(buf.strip())

    out: dict = {}
    for p in parts:
        if not p:
            continue
        m = re.match(r"^\s*([0-9]*\.?[0-9]+)\s*:\s*\(\s*([0-9]*\.?[0-9]+)\s*,\s*([0-9]*\.?[0-9]+)\s*\)\s*$", p)
        if m:
            v, i, eff = map(float, m.groups())
        else:
            m2 = re.match(r"^\s*([0-9]*\.?[0-9]+)\s*:\s*([0-9]*\.?[0-9]+)\s*:\s*([0-9]*\.?[0-9]+)\s*$", p)
            if not m2:
                raise ValueError(
                    "Invalid --avionics_voltage_tree format. "
                    "Use e.g. \"5.0:(2,0.9), 12.0:(1.5,0.85)\" or \"5.0:2:0.9, 12.0:1.5:0.85\"."
                )
            v, i, eff = map(float, m2.groups())

        if v <= 0:
            raise ValueError(f"Avionics rail voltage must be > 0, got {v}.")
        if i < 0:
            raise ValueError(f"Avionics rail current must be >= 0, got {i}.")
        if eff <= 0 or eff > 1.0:
            raise ValueError(f"BEC efficiency must be in (0, 1], got {eff}.")
        out[float(v)] = (float(i), float(eff))

    return out


def parse_optional_bool(spec: Optional[object]) -> Optional[bool]:
    """
    Parse permissive CLI-style booleans.
    Returns None when the value is omitted/blank so callers can apply defaults.
    """
    if spec is None:
        return None
    if isinstance(spec, bool):
        return spec
    s = str(spec).strip().lower()
    if not s:
        return None
    if s in ("1", "true", "yes", "y", "on", "enable", "enabled"):
        return True
    if s in ("0", "false", "no", "n", "off", "disable", "disabled"):
        return False
    raise ValueError(f"Invalid boolean value: {spec!r}")


def _split_csv_tokens(spec: str) -> List[str]:
    return [tok.strip() for tok in str(spec).split(",") if tok.strip()]


def avionics_input_power_W(avionics: Optional[AvionicsConfig]) -> float:
    """Return total input power drawn from the battery to supply avionics rails (W).

    For each rail:
      P_in = (V_rail * I_rail) / efficiency
    """
    if avionics is None or not getattr(avionics, "voltage_tree", None):
        return 0.0
    total = 0.0
    for v, (i, eff) in avionics.voltage_tree.items():
        total += (float(v) * float(i)) / max(float(eff), 1e-9)
    return float(total)


def esc_loss_and_checks(config: "DroneConfig", v_pack: float, motor_power_total_W: float) -> tuple[float, str, float]:
    """Compute ESC electrical loss and simple current-limit checks.

    Returns:
      (esc_loss_W, status_note, motor_current_per_esc_A)

    Model:
      - Conduction loss: I^2 * R for each ESC
      - Idle/overhead draw: I_idle * V for each ESC (if provided)
    Assumption:
      Motor/ESC sees pack voltage (no separate motor rail).
    """
    esc = getattr(config, "esc", None)
    if esc is None:
        return 0.0, "", 0.0

    v = max(float(v_pack), 1e-9)
    p_per_motor = float(motor_power_total_W) / max(int(config.num_motors), 1)
    i_motor = p_per_motor / v  # A per ESC (approx)

    # Losses per ESC
    p_loss_cond = (i_motor ** 2) * max(float(esc.resistance), 0.0)
    p_loss_idle = max(float(esc.idle_current_A), 0.0) * v
    esc_loss_total = (p_loss_cond + p_loss_idle) * int(config.num_motors)

    note_parts = []
    if i_motor > float(esc.max_current_A):
        note_parts.append(f"ESC OVER MAX: {i_motor:.1f}A > {esc.max_current_A:.1f}A")
    elif i_motor > float(esc.continuous_rating_A):
        note_parts.append(f"ESC over continuous: {i_motor:.1f}A > {esc.continuous_rating_A:.1f}A")

    return float(esc_loss_total), ("; ".join(note_parts) if note_parts else ""), float(i_motor)


def total_power_with_esc(config: "DroneConfig",
                         motor_power_W: float,
                         periph_power_W: float,
                         iters: int = 6,
                         soc: Optional[float] = None) -> tuple[float, float, float, str, float]:
    """Iteratively solve pack voltage/current while accounting for ESC loss.

    Returns:
      (total_power_W, v_load_V, pack_current_A, esc_note, motor_current_per_esc_A)
    """
    total_power = float(motor_power_W) + float(periph_power_W)
    v_load = battery_ocv_pack(config.battery, soc if soc is not None else 1.0)
    pack_current = total_power / max(v_load, 1e-9)
    esc_note = ""
    i_motor = 0.0

    # Fixed-point iteration (ESC loss depends on v_load)
    for _ in range(max(int(iters), 1)):
        v_load, pack_current = solve_pack_voltage_and_current(
            config.battery, total_power, soc=soc
        )
        esc_loss_W, esc_note, i_motor = esc_loss_and_checks(config, v_load, motor_power_W)
        total_power = float(motor_power_W) + float(periph_power_W) + float(esc_loss_W)

    return float(total_power), float(v_load), float(pack_current), esc_note, float(i_motor)


def solve_pack_voltage_and_current(battery: "BatteryConfig",
                                   total_power_W: float,
                                   iters: int = 12,
                                   soc: Optional[float] = None) -> tuple[float, float]:
    """Solve V_load and I_pack for a load that draws a (roughly) constant electrical power.

    We iterate:
      I = P / V
      V = Vmax - I*Rpack (clamped at Vmin)

    Returns (V_load, I_pack).
    """
    soc_eval = 1.0 if soc is None else min(max(float(soc), 0.0), 1.0)
    if total_power_W <= 0:
        return (battery_ocv_pack(battery, soc_eval), 0.0)

    v = battery_ocv_pack(battery, soc_eval)
    i = total_power_W / max(v, 1e-9)
    for _ in range(max(1, int(iters))):
        v = battery_voltage_under_load(battery, i, soc=soc_eval)
        i = total_power_W / max(v, 1e-9)
    return (float(v), float(i))


# -------------------------------
# Propeller Model
# -------------------------------
def load_motor_prop_table_csv(path: str) -> pd.DataFrame:
    """Load a motor/prop test table CSV.

    Supports two formats:
    1) Simple header with columns like: Thrust_g, Power_W, Current_A, RPM, Voltage_V, Throttle_pct, Efficiency_gW, Temp_C
    2) eCalc-style exported CSV where the *real* header is on a later row (e.g. first row says 'Test Data') and
       columns are named like 'Thrust (g)', 'Power (W)', 'Current (A)', 'RPM', 'Voltage (V)', 'Throttle', etc.

    Returns a cleaned DataFrame sorted by Thrust_g with standardized column names.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    # Try normal header first
    try:
        df0 = pd.read_csv(path)
    except Exception:
        df0 = pd.DataFrame()

    def _looks_good(df: pd.DataFrame) -> bool:
        cols = {str(c).strip().lower() for c in df.columns}
        return ("thrust_g" in cols) or ("thrust (g)" in cols) or ("thrust" in cols and "(g)" in " ".join(cols))

    if df0 is not None and len(df0.columns) > 1 and _looks_good(df0):
        df = df0
    else:
        # Read without header, find the header row
        raw = pd.read_csv(path, header=None)
        header_row = None
        for i in range(min(len(raw), 25)):
            # Coerce every cell to a string explicitly. Do NOT rely on
            # Series.astype(str): with pandas' newer "str" dtype it leaves
            # NaN as a real float, so a sparse row (a title line followed by
            # empty cells) yields floats here and .startswith() below raises
            # "'float' object has no attribute 'startswith'".
            row = ["" if pd.isna(v) else str(v).strip().lower()
                   for v in raw.iloc[i].tolist()]
            if any("throttle" == x or x.startswith("throttle") for x in row) and any("thrust" in x for x in row):
                header_row = i
                break
            # Some exports put "Voltage (V)" and "Thrust (g)" on same line
            if any("voltage" in x for x in row) and any("thrust" in x for x in row) and any("power" in x for x in row):
                header_row = i
                break
        if header_row is None:
            raise ValueError(f"Could not locate header row in CSV: {path}")

        df = pd.read_csv(path, header=header_row)

    # Standardize column names
    rename_map = {}
    for c in df.columns:
        c0 = str(c).strip()
        cl = c0.lower()

        if cl in ("thrust_g", "thrust (g)"):
            rename_map[c] = "Thrust_g"
        elif cl in ("power_w", "power (w)"):
            rename_map[c] = "Power_W"
        elif cl in ("current_a", "current (a)"):
            rename_map[c] = "Current_A"
        elif cl in ("voltage_v", "voltage (v)"):
            rename_map[c] = "Voltage_V"
        elif cl in ("rpm",):
            rename_map[c] = "RPM"
        elif cl in ("efficiency (g/w)", "efficiency_gw", "efficiency (g/w) "):
            rename_map[c] = "Efficiency_gW"
        elif "operating temperature" in cl or "temperature" in cl:
            rename_map[c] = "Temp_C"
        elif cl.startswith("throttle"):
            rename_map[c] = "Throttle_pct"
        elif cl == "propeller":
            rename_map[c] = "Propeller"
        elif cl == "type":
            rename_map[c] = "Type"
        elif "torque" in cl:
            rename_map[c] = "Torque_Nm"

    df = df.rename(columns=rename_map)

    # Clean / coerce types
    def _to_float_series(s: pd.Series) -> pd.Series:
        return pd.to_numeric(s.astype(str).str.replace("%", "", regex=False).str.strip(), errors="coerce")

    for col in ("Thrust_g", "Power_W", "Current_A", "Voltage_V", "RPM", "Efficiency_gW", "Temp_C", "Throttle_pct", "Torque_Nm"):
        if col in df.columns:
            df[col] = _to_float_series(df[col])

    # Drop rows without thrust/power
    if "Thrust_g" not in df.columns or "Power_W" not in df.columns:
        raise ValueError("Motor/prop CSV must contain thrust and power columns (e.g., 'Thrust (g)' and 'Power (W)')")

    df = df.dropna(subset=["Thrust_g", "Power_W"]).copy()

    # Sort & de-duplicate by thrust (keep last)
    df = df.sort_values("Thrust_g")
    df = df.drop_duplicates(subset=["Thrust_g"], keep="last").reset_index(drop=True)

    return df

class PropellerConfig:
    def __init__(self,
                 diameter_in: float,
                 pitch_in: float,
                 max_rpm: int,
                 max_thrust_g: float,
                 blades: int = 2,
                 table_csv: Optional[str] = None,
                 TConst: Optional[float] = None,
                 PConst: Optional[float] = None,
                 weight_g: Optional[float] = None):
        self.diameter_in = float(diameter_in)
        self.pitch_in = float(pitch_in)
        self.blades = int(blades)
        self.max_rpm = int(max_rpm)
        self.max_thrust_g = float(max_thrust_g)
        self.table_csv = table_csv
        self.TConst = TConst
        self.PConst = PConst
        self.weight_g = weight_g

        self.table: Optional[pd.DataFrame] = None
        if table_csv:
            self.table = load_motor_prop_table_csv(table_csv)
            self._cache_table_arrays()

    def _cache_table_arrays(self) -> None:
        """
        Precompute the scalars and arrays the hot paths need.

        Every one of these was previously recomputed from the DataFrame on
        each call. The climb-rate and best-speed searches evaluate thrust
        hundreds of times per run, so a pandas reduction per evaluation turned
        a 5 ms calculation into 50 ms and made the window stop responding.
        """
        self._thrust_g_arr = None
        self._power_w_arr = None
        self._thrust_g_min = None
        self._thrust_g_max = None
        if self.table is None or "Thrust_g" not in self.table:
            return
        self._thrust_g_arr = self.table["Thrust_g"].to_numpy(dtype=float)
        if "Power_W" in self.table:
            self._power_w_arr = self.table["Power_W"].to_numpy(dtype=float)
        if self._thrust_g_arr.size:
            self._thrust_g_min = float(self._thrust_g_arr.min())
            self._thrust_g_max = float(self._thrust_g_arr.max())


# -------------------------------
# Drone Config
# -------------------------------
class DroneConfig:
    def __init__(self,
                 num_motors: int,
                 battery: BatteryConfig,
                 motor: MotorConfig,
                 propeller: PropellerConfig,
                 drone_weight_g: float,
                 profile_drag_coefficient: float,
                 profile_area: float,
                 parasite_drag_coefficient: float,
                 parasite_area: float,
                 frontal_area: float,
                 cruise_speed: float,
                 periph_current: float,
                 esc: Optional[ESCConfig] = None,
                 avionics: Optional[AvionicsConfig] = None,
                 air_density: float = AIR_DENSITY,
                 drag_model: str = "auto",  # "auto", "manual", or "geometry"
                 # --- Vehicle geometry / mechanical params (new) ---
                 body_length_m: Optional[float] = None,
                 body_width_m: Optional[float] = None,
                 body_height_m: Optional[float] = None,
                 arm_length_m: Optional[float] = None,
                 arm_width_m: Optional[float] = None,
                 coaxial_spacing_m: Optional[float] = None,
                 max_tilt_deg: Optional[float] = None,
                 motor_configuration: str = "flat",  # "flat" or "coaxial"
                 # --- Mission-transient dynamics ---
                 transient_dt_s: float = 0.5,
                 max_accel_mps2: float = 2.0,
                 max_decel_mps2: float = 2.5,
                 decel_regen_eff: float = 0.0,
                 # --- Rotor inflow map ---
                 inflow_map_enabled: bool = True,
                 inflow_mu_bp: Optional[List[float]] = None,
                 inflow_eff_bp: Optional[List[float]] = None,
                 ):
        self.num_motors = int(num_motors)
        self.battery = battery
        self.motor = motor
        self.propeller = propeller
        self.drone_weight_g = float(drone_weight_g)

        # Drag parameters (may be overridden / derived from geometry if not provided)
        self.profile_drag_coefficient = float(profile_drag_coefficient) if profile_drag_coefficient is not None else 0.0
        self.profile_area = float(profile_area) if profile_area is not None else 0.0
        self.parasite_drag_coefficient = float(parasite_drag_coefficient) if parasite_drag_coefficient is not None else 0.0
        self.parasite_area = float(parasite_area) if parasite_area is not None else 0.0
        self.frontal_area = float(frontal_area) if frontal_area is not None else 0.0

        self.cruise_speed = float(cruise_speed)
        self.periph_current = float(periph_current)
        self.esc = esc
        self.avionics = avionics
        self.air_density = float(air_density)
        drag_mode_raw = str(drag_model or "auto").strip().lower()
        if drag_mode_raw in ("manual", "direct", "exact"):
            self.drag_model = "manual"
        elif drag_mode_raw in ("geometry", "rect", "derived", "derive"):
            self.drag_model = "geometry"
        else:
            self.drag_model = "auto"

        # Geometry / limits
        self.body_length_m = float(body_length_m) if body_length_m not in (None, "") else None
        self.body_width_m = float(body_width_m) if body_width_m not in (None, "") else None
        self.body_height_m = float(body_height_m) if body_height_m not in (None, "") else None
        self.arm_length_m = float(arm_length_m) if arm_length_m not in (None, "") else None
        self.arm_width_m = float(arm_width_m) if arm_width_m not in (None, "") else None
        self.coaxial_spacing_m = float(coaxial_spacing_m) if coaxial_spacing_m not in (None, "") else None
        self.max_tilt_deg = float(max_tilt_deg) if max_tilt_deg not in (None, "") else None
        self.motor_configuration = (motor_configuration or "flat").strip().lower()
        if self.motor_configuration not in ("flat", "coaxial"):
            self.motor_configuration = "flat"

        # Transient acceleration/deceleration model used by mission simulation.
        self.transient_dt_s = max(float(transient_dt_s), 0.05)
        self.max_accel_mps2 = max(float(max_accel_mps2), 0.01)
        self.max_decel_mps2 = max(float(max_decel_mps2), 0.01)
        self.decel_regen_eff = min(max(float(decel_regen_eff), 0.0), 1.0)

        # Rotor inflow / forward-flight efficiency map:
        # mu (advance ratio) -> eta_inflow (dimensionless efficiency).
        self.inflow_map_enabled = bool(inflow_map_enabled)
        mu_raw = list(inflow_mu_bp) if inflow_mu_bp else list(DEFAULT_INFLOW_MU_BP)
        eff_raw = list(inflow_eff_bp) if inflow_eff_bp else list(DEFAULT_INFLOW_EFF_BP)
        if len(mu_raw) != len(eff_raw) or len(mu_raw) < 2:
            mu_raw = list(DEFAULT_INFLOW_MU_BP)
            eff_raw = list(DEFAULT_INFLOW_EFF_BP)
        pairs = sorted((max(float(mu), 0.0), max(float(eta), 0.20))
                       for mu, eta in zip(mu_raw, eff_raw))
        self.inflow_mu_bp: List[float] = []
        self.inflow_eff_bp: List[float] = []
        for mu, eta in pairs:
            if self.inflow_mu_bp and abs(mu - self.inflow_mu_bp[-1]) < 1e-9:
                self.inflow_eff_bp[-1] = eta
            else:
                self.inflow_mu_bp.append(mu)
                self.inflow_eff_bp.append(eta)
        if len(self.inflow_mu_bp) < 2:
            self.inflow_mu_bp = list(DEFAULT_INFLOW_MU_BP)
            self.inflow_eff_bp = list(DEFAULT_INFLOW_EFF_BP)

        # Internal flag so we only derive drag once (unless user edits values)
        self._derived_drag_from_geometry = False

    def derive_drag_from_geometry_if_missing(self) -> None:
        """
        Depending on drag_model:
          - manual: never derive (always use user-entered Cd/area terms)
          - auto: derive only when drag terms are missing
          - geometry: always derive from geometry (once)

        Geometry derivation uses a simple box body + square-tube arm model.

        This is intentionally simple and meant as a fallback, not a substitute
        for measured CdA.
        """
        if self.drag_model == "manual":
            return
        if self._derived_drag_from_geometry:
            return
        if self.drag_model == "auto":
            # Frontal area is not part of forward-drag terms, so don't use it to
            # decide whether profile/parasite drag was explicitly provided.
            provided = any(x > 0 for x in (
                self.profile_drag_coefficient, self.profile_area,
                self.parasite_drag_coefficient, self.parasite_area,
            ))
            if provided:
                return

        # Need enough geometry to do anything useful
        if self.body_width_m is None or self.body_height_m is None:
            return
        if self.arm_length_m is None:
            return

        # --- Assumptions (fallback constants) ---
        # Rectangular box Cd normal to flow ~ 1.0–1.2; use 1.05
        CD_BOX = 1.05
        # Square tube / cylinder-ish boom Cd ~ 1.0–1.3; use 1.1
        CD_ARM = 1.10
        # Arm tube width (square tube outer width). User-configurable; fallback to 20 mm.
        ARM_TUBE_SIDE_M = float(self.arm_width_m) if self.arm_width_m not in (None,) else 0.02

        # Determine number of arms (coaxial usually has fewer arms for the same motor count)
        if self.motor_configuration == "coaxial":
            num_arms = max(1, self.num_motors // 2)
        else:
            num_arms = self.num_motors

        # Body frontal area for forward flight (assume length is forward axis)
        A_body = self.body_width_m * self.body_height_m

        # Arms projected area (very rough). Each arm contributes ~ (tube_side * arm_length)
        # with a projection factor to account for non-perfect alignment (X layout etc).
        PROJ = 0.7
        A_arms = num_arms * (ARM_TUBE_SIDE_M * self.arm_length_m) * PROJ

        # Parasite drag: body + arms
        A_total = A_body + A_arms
        CdA_total = CD_BOX * A_body + CD_ARM * A_arms
        if A_total > 0:
            self.parasite_area = A_total
            self.parasite_drag_coefficient = CdA_total / A_total
        else:
            self.parasite_area = 0.0
            self.parasite_drag_coefficient = 0.0

        # Profile term = the SIDE silhouette, used when the vehicle translates
        # laterally while level (the "hover" orientation).  Approximate it as
        # the body seen side-on (length x height) plus the arms, since an arm
        # presents roughly the same tube area from the side as from the front.
        A_body_side = self.body_length_m * self.body_height_m
        A_side_total = A_body_side + A_arms
        if A_side_total > 0:
            self.profile_area = A_side_total
            self.profile_drag_coefficient = (
                (CD_BOX * A_body_side + CD_ARM * A_arms) / A_side_total)
        else:
            self.profile_area = 0.0
            self.profile_drag_coefficient = 0.0

        # Frontal area stored separately if you want to use it elsewhere
        self.frontal_area = A_body

        self._derived_drag_from_geometry = True


# -------------------------------
# Mission Profile
# -------------------------------
@dataclass
class MissionPhase:
    name: str
    speed: float
    duration: Optional[float] = None  # seconds
    distance: Optional[float] = None  # meters
    altitude: float = 0.0             # meters
    course_deg: float = 0.0
    climb_rate_mps: Optional[float] = None
    descent_rate_mps: Optional[float] = None


class MissionProfile:
    def __init__(self,
                 phases: List[MissionPhase],
                 reserve_percent: float = 20.0,
                 rth_reserve_Wh: float = 0.0,
                 diversion_reserve_Wh: float = 0.0,
                 wind_direction_deg: float = 0.0,
                 transient_dt_s: float = 0.5,
                 max_accel_mps2: float = 2.0,
                 max_decel_mps2: float = 2.5,
                 decel_regen_eff: float = 0.0):
        self.phases = phases
        self.reserve_percent = float(reserve_percent)
        self.rth_reserve_Wh = float(rth_reserve_Wh)
        self.diversion_reserve_Wh = float(diversion_reserve_Wh)
        self.wind_direction_deg = float(wind_direction_deg)
        self.transient_dt_s = max(float(transient_dt_s), 0.05)
        self.max_accel_mps2 = max(float(max_accel_mps2), 0.01)
        self.max_decel_mps2 = max(float(max_decel_mps2), 0.01)
        self.decel_regen_eff = min(max(float(decel_regen_eff), 0.0), 1.0)

    @staticmethod
    def from_json(path: str) -> "MissionProfile":
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)

        phases = []
        for p in data.get("phases", []):
            phases.append(MissionPhase(
                name=p["name"],
                speed=float(p["speed"]),
                duration=p.get("duration"),
                distance=p.get("distance"),
                altitude=float(p.get("altitude", 0.0)),
                course_deg=float(p.get("course_deg", 0.0)),
                climb_rate_mps=(float(p["climb_rate_mps"]) if "climb_rate_mps" in p else None),
                descent_rate_mps=(float(p["descent_rate_mps"]) if "descent_rate_mps" in p else None),
            ))
        return MissionProfile(
            phases,
            reserve_percent=float(data.get("reserve_percent", 20.0)),
            rth_reserve_Wh=float(data.get("rth_reserve_Wh", 0.0)),
            diversion_reserve_Wh=float(data.get("diversion_reserve_Wh", 0.0)),
            wind_direction_deg=float(data.get("wind_direction_deg", 0.0)),
            transient_dt_s=float(data.get("transient_dt_s", 0.5)),
            max_accel_mps2=float(data.get("max_accel_mps2", 2.0)),
            max_decel_mps2=float(data.get("max_decel_mps2", 2.5)),
            decel_regen_eff=float(data.get("decel_regen_eff", 0.0)),
        )


# -------------------------------
# Physics Helpers
# -------------------------------

def drag_force_required(config: DroneConfig, speed_mps: float, orientation: str) -> float:
    """Compute aerodynamic drag force (N) for the given airspeed and orientation."""
    # If enabled and drag parameters were not provided, derive from geometry fallback.
    # In manual mode, user-specified Cd/area values are used directly.
    config.derive_drag_from_geometry_if_missing()

    q = 0.5 * config.air_density * speed_mps ** 2      # dynamic pressure [Pa]

    if orientation == "hover":
        # Level attitude, translating sideways/laterally: the vehicle presents
        # its SIDE profile to the airflow.
        return q * config.profile_area * config.profile_drag_coefficient

    # Forward flight, nose pitched down: the vehicle presents its FRONTAL
    # silhouette.  parasite_area is the FULL frontal area (body + arms +
    # legs), so it already accounts for every frontal component.
    #
    # Previously this returned  parasite + profile, which double-counted the
    # arms: the geometry fallback puts the arms into BOTH parasite_area
    # (body + arms) and profile_area (arms), so their drag was charged twice.
    return q * config.parasite_area * config.parasite_drag_coefficient


def required_tilt_deg(config: DroneConfig, speed_mps: float, orientation: str) -> float:
    """Approximate tilt angle (deg) required in forward flight to balance drag."""
    if orientation != "forward":
        return 0.0
    weight_force = config.drone_weight_g * 9.81 / 1000.0  # Convert grams to kg
    drag_force = drag_force_required(config, speed_mps, orientation="forward")
    return math.degrees(math.atan2(drag_force, max(weight_force, 1e-9)))

def disk_area(diameter_in: float) -> float:
    d_m = diameter_in * 0.0254
    return math.pi * (d_m / 2.0) ** 2


def advance_ratio_mu(config: DroneConfig, airspeed_mps: float, rpm: Optional[float]) -> float:
    """Rotor advance ratio mu = V / (Omega * R) for one rotor."""
    if rpm is None:
        return 0.0
    try:
        r_m = max(float(config.propeller.diameter_in) * 0.0254 / 2.0, 1e-9)
        omega = max(2.0 * math.pi * float(rpm) / 60.0, 1e-9)
        return max(float(airspeed_mps), 0.0) / (omega * r_m)
    except Exception:
        return 0.0


def inflow_efficiency_from_mu(config: DroneConfig, mu: float, orientation: str) -> float:
    """Inflow efficiency factor from configured mu map."""
    if orientation != "forward":
        return 1.0
    if not bool(getattr(config, "inflow_map_enabled", True)):
        return 1.0
    mu_bp = list(getattr(config, "inflow_mu_bp", []) or DEFAULT_INFLOW_MU_BP)
    eff_bp = list(getattr(config, "inflow_eff_bp", []) or DEFAULT_INFLOW_EFF_BP)
    return max(_interp_linear_clamped(max(float(mu), 0.0), mu_bp, eff_bp), 0.2)


def rotor_inflow_power_multiplier(config: DroneConfig,
                                  airspeed_mps: float,
                                  rpm: Optional[float],
                                  orientation: str) -> Tuple[float, float, float]:
    """Return (power_multiplier, mu, eta_inflow) for forward-flight inflow map."""
    mu = advance_ratio_mu(config, airspeed_mps, rpm)
    eta = inflow_efficiency_from_mu(config, mu, orientation)
    # eta>1 reduces required power, eta<1 increases it.
    multiplier = 1.0 / max(eta, 0.2)
    return float(multiplier), float(mu), float(eta)




def total_disk_area(config: DroneConfig) -> float:
    """Total rotor disk area over all motors [m²]."""
    return disk_area(config.propeller.diameter_in) * max(int(config.num_motors), 1)


def disk_loading_N_m2(config: DroneConfig) -> float:
    """Disk loading: DL = W / (N * A_disk) [N/m²]."""
    weight_N = config.drone_weight_g * 9.81 / 1000.0
    area = total_disk_area(config)
    return weight_N / area if area > 0 else 0.0


def available_total_thrust_N(config: DroneConfig) -> float:
    """Estimate maximum total thrust available [N] from prop/motor data."""
    nm = max(int(config.num_motors), 1)
    if getattr(config.propeller, "table", None) is not None:
        try:
            _cached = getattr(config.propeller, "_thrust_g_max", None)
            _max_g = (_cached if _cached is not None
                      else float(config.propeller.table["Thrust_g"].max()))
            return float(_max_g) * 9.81 / 1000.0 * nm
        except Exception:
            pass

    max_thr_g = float(getattr(config.propeller, "max_thrust_g", 0.0) or 0.0)
    if max_thr_g > 0:
        return max_thr_g * 9.81 / 1000.0 * nm

    # Fallback from motor max power + actuator-disk estimate.
    p_max_pm = max(float(getattr(config.motor, "max_power", 0.0) or 0.0), 0.0)
    if p_max_pm <= 0:
        return 0.0
    rho = max(float(config.air_density), 1e-9)
    A_pm = max(disk_area(config.propeller.diameter_in), 1e-9)
    # P = T^(3/2) / sqrt(2*rho*A)  => T = (P*sqrt(2*rho*A))^(2/3)
    t_pm = (p_max_pm * math.sqrt(2.0 * rho * A_pm)) ** (2.0 / 3.0)
    return t_pm * nm


def tip_speed_mps_from_rpm(diameter_in: float, rpm: float) -> float:
    """Blade tip speed from prop diameter and RPM [m/s]."""
    d_m = max(float(diameter_in), 0.0) * 0.0254
    n = max(float(rpm), 0.0) / 60.0
    return math.pi * d_m * n


def estimate_blade_chord_m(diameter_in: float, blades: int) -> float:
    """
    Estimate average blade chord from diameter and blade count.
    This is a coarse empirical estimate for small multicopter props.
    """
    d_m = max(float(diameter_in), 0.0) * 0.0254
    b = max(int(blades), 1)
    # Base chord scales with diameter; slightly narrower blades when blade count increases.
    return 0.11 * d_m / (1.0 + 0.08 * (b - 2))


def propeller_solidity(diameter_in: float, blades: int) -> float:
    """
    Propeller solidity:
      sigma = N_blades * c_blade / (pi * R)
    """
    d_m = max(float(diameter_in), 0.0) * 0.0254
    r_m = d_m / 2.0
    if r_m <= 0:
        return 0.0
    c_blade = estimate_blade_chord_m(diameter_in, blades)
    return max(int(blades), 1) * c_blade / (math.pi * r_m)


def hover_ideal_induced_power_W(config: DroneConfig,
                                total_thrust_N: float) -> float:
    """
    Ideal actuator-disk induced power in hover:
      P_ideal = T^(3/2) / sqrt(2 * rho * A_total)
    """
    rho = max(float(config.air_density), 1e-9)
    area = total_disk_area(config)
    if area <= 0:
        return 0.0
    T = max(float(total_thrust_N), 0.0)
    return (T ** 1.5) / math.sqrt(2.0 * rho * area)


def hover_figure_of_merit(config: DroneConfig,
                          total_thrust_N: float,
                          actual_induced_power_W: float) -> float:
    """
    Figure of merit based on induced power:
      FM = P_ideal / P_actual_induced
    """
    p_actual = max(float(actual_induced_power_W), 1e-9)
    p_ideal = hover_ideal_induced_power_W(config, total_thrust_N)
    return min(p_ideal / p_actual, 1.5)


def hover_wind_resistance_mps(config: DroneConfig) -> float:
    """
    Estimate maximum hover wind resistance from available horizontal thrust:
      V_max = sqrt(2 * T_horizontal / (rho * C_D * A_frontal))
    Uses thrust margin, tilt limit, and frontal drag model.
    """
    rho = max(float(config.air_density), 1e-9)
    Cd = max(float(getattr(config, "parasite_drag_coefficient", 0.0)), 0.2)

    # Reference area: prefer the area the drag model actually uses
    # (parasite_area, which the geometry fallback populates), then an
    # explicitly entered frontal_area.
    #
    # If neither is known we must NOT substitute a token value: the old
    # 1e-6 m2 floor turned an unknown area into a divide-by-almost-zero and
    # reported wind resistances of thousands of m/s. Return NaN instead so
    # the figure displays as "n/a" rather than a confident wrong number.
    A_frontal = 0.0
    for attr in ("parasite_area", "frontal_area"):
        try:
            val = float(getattr(config, attr, 0.0) or 0.0)
        except (TypeError, ValueError):
            val = 0.0
        if val > 0.0:
            A_frontal = val
            break
    if A_frontal < 1e-4:          # smaller than 1 cm2 is not a real airframe
        return float("nan")

    weight_N = config.drone_weight_g * 9.81 / 1000.0

    t_avail = max(available_total_thrust_N(config), 0.0)
    if t_avail <= weight_N:
        return 0.0

    tilt_lim = float(getattr(config, "max_tilt_deg", 35.0) or 35.0)
    tilt_lim = max(min(tilt_lim, 85.0), 5.0)
    t_h_balanced = math.sqrt(max(t_avail ** 2 - weight_N ** 2, 0.0))
    t_h_tilt_cap = t_avail * math.sin(math.radians(tilt_lim))
    t_horizontal = min(t_h_balanced, t_h_tilt_cap)

    if t_horizontal <= 0:
        return 0.0
    return math.sqrt((2.0 * t_horizontal) / (rho * Cd * A_frontal))


def compute_air_density(altitude_m: float,
                        temperature_C: Optional[float] = None,
                        pressure_Pa: Optional[float] = None) -> float:
    """
    Air density [kg/m^3]. Thin wrapper over the shared core implementation.

    Kept under this name because both simulators, their CLIs and saved
    scripts refer to it. The two used to carry separate implementations
    that disagreed about pressure overrides; there is now exactly one.
    """
    return core.air_density(altitude_m, temperature_C, pressure_Pa)


def thrust_required(config: DroneConfig, speed_mps: float, orientation: str) -> float:
    """
    Simplified force model.

    Hover:
      Required thrust ≈ weight + (small profile drag term if you model it)

    Forward flight:
      Drag is primarily horizontal and weight is vertical, so required thrust magnitude is:
          T = sqrt(weight^2 + drag^2)
      and required tilt is:
          tilt = atan(drag / weight)

    This function returns total thrust magnitude (N).
    """
    weight_force = config.drone_weight_g * 9.81 / 1000.0  # Convert grams to kg

    drag_force = drag_force_required(config, speed_mps, orientation)

    if orientation == "hover":
        return weight_force + drag_force

    # forward flight (vector sum)
    return math.sqrt(weight_force**2 + drag_force**2)


def _extrapolate_motor_value(df: pd.DataFrame, thrust_g: float, column: str) -> Optional[float]:
    """Extrapolate a motor property value using fitted curve if thrust is below minimum.
    
    Args:
        df: Propeller table dataframe
        thrust_g: Target thrust in grams
        column: Column name to extrapolate (Power_W, Current_A, etc.)
    
    Returns:
        Extrapolated value or None if not available
    """
    if column not in df.columns:
        return None
    
    thrust_g_data = df["Thrust_g"].values
    try:
        y_data = pd.to_numeric(df[column], errors='coerce').values
    except:
        return None
    
    min_thrust = thrust_g_data.min()
    
    # If we're within data range, don't extrapolate
    if thrust_g >= min_thrust:
        return None
    
    # Fit a 2nd degree polynomial to the data
    fit_result = _fit_propeller_curve(thrust_g_data, y_data, degree=2)
    if fit_result is None:
        return None
    
    coeffs, _, _ = fit_result
    extrapolated = _eval_poly(coeffs, thrust_g)
    
    # For some properties, clamp to minimum reasonable values
    if column == "Current_A" and extrapolated is not None:
        extrapolated = max(extrapolated, 0.0)
    elif column == "Power_W" and extrapolated is not None:
        extrapolated = max(extrapolated, 0.0)
    elif column == "Efficiency_gW" and extrapolated is not None:
        extrapolated = max(extrapolated, 0.0)
    elif column == "RPM" and extrapolated is not None:
        extrapolated = max(extrapolated, 0.0)
    
    return extrapolated


def interpolate_motor_point(config: DroneConfig, thrust_per_motor_N: float) -> dict:
    """Interpolate a full operating point from a motor/prop test table.

    Returns a dict with (when available):
      Power_W, RPM, Current_A, Voltage_V, Throttle_pct, Efficiency_gW, Temp_C, Torque_Nm

    The lookup key is thrust (grams-force). Interpolation is linear between nearest points.
    """
    if config.propeller.table is None:
        raise ValueError("No prop_table loaded, cannot interpolate.")

    thrust_g = thrust_per_motor_N * 1000.0 / 9.81
    df = config.propeller.table

    if "Thrust_g" not in df.columns or "Power_W" not in df.columns:
        raise ValueError("prop_table CSV must have columns: Thrust_g, Power_W (after parsing)")

    min_thrust = float(df["Thrust_g"].min())
    max_thrust = float(df["Thrust_g"].max())
    
    # Handle below minimum: extrapolate below the measured range.
    #
    # A polynomial fitted to the measured band and run downward is unsafe — on
    # a table spanning 1426-6733 g it crosses zero near 250 g and turns
    # negative, which surfaces as an operating point at negative power and an
    # absurd efficiency. Quantities with a known physical scaling use it
    # instead, anchored on the lowest MEASURED row so the fit is exact at the
    # edge of the table:
    #
    #     power    proportional to T^1.5   (static momentum theory)
    #     RPM      proportional to T^0.5   (thrust goes as omega^2)
    #     current  proportional to torque, which goes as T^1.5
    #
    # Anything without a clean scaling law falls back to the fitted curve, and
    # is then floored so it can never go negative.
    if thrust_g < min_thrust:
        out = {}
        first = df.iloc[0]
        ratio = (thrust_g / min_thrust) if min_thrust > 0 else 0.0
        PHYSICAL_EXPONENT = {
            "Power_W": 1.5, "Current_A": 1.5, "Torque_Nm": 1.5,
            "RPM": 0.5, "Throttle_pct": 0.5,
        }
        for col in ("Power_W","RPM","Current_A","Voltage_V","Throttle_pct","Efficiency_gW","Temp_C","Torque_Nm"):
            if col not in df.columns or not pd.notna(first[col]):
                continue
            anchor = float(first[col])
            exponent = PHYSICAL_EXPONENT.get(col)
            if exponent is not None and anchor > 0 and ratio > 0:
                out[col] = max(anchor * (ratio ** exponent), 1e-9)
                continue
            if col in ("Voltage_V", "Temp_C"):
                out[col] = anchor          # roughly flat; holding is honest
                continue
            extrapolated = _extrapolate_motor_value(df, thrust_g, col)
            if extrapolated is not None and float(extrapolated) > 0:
                out[col] = float(extrapolated)
            else:
                out[col] = anchor
        # Efficiency must stay consistent with the power we just produced.
        if "Power_W" in out and out["Power_W"] > 0:
            out["Efficiency_gW"] = thrust_g / out["Power_W"]
        return out

    if thrust_g >= max_thrust:
        row = df.iloc[-1]
        return {k: float(row[k]) for k in df.columns if k in (
            "Thrust_g","Power_W","RPM","Current_A","Voltage_V","Throttle_pct","Efficiency_gW","Temp_C","Torque_Nm"
        ) and pd.notna(row[k])}

    lower = df[df["Thrust_g"] <= thrust_g].iloc[-1]
    upper = df[df["Thrust_g"] >= thrust_g].iloc[0]

    denom = float(upper["Thrust_g"] - lower["Thrust_g"])
    frac = 0.0 if denom == 0 else float((thrust_g - lower["Thrust_g"]) / denom)

    out = {}
    for col in ("Power_W","RPM","Current_A","Voltage_V","Throttle_pct","Efficiency_gW","Temp_C","Torque_Nm"):
        if col in df.columns and pd.notna(lower[col]) and pd.notna(upper[col]):
            out[col] = float(lower[col] + (upper[col] - lower[col]) * frac)
        elif col in df.columns and pd.notna(lower[col]) and pd.isna(upper[col]):
            out[col] = float(lower[col])
        elif col in df.columns and pd.isna(lower[col]) and pd.notna(upper[col]):
            out[col] = float(upper[col])
    return out


def interpolate_motor_power(config: DroneConfig, thrust_per_motor_N: float) -> float:
    """Interpolate electrical input power from a motor/prop test table."""
    pt = interpolate_motor_point(config, thrust_per_motor_N)
    if "Power_W" not in pt:
        raise ValueError("prop_table interpolation failed to produce Power_W")
    return float(pt["Power_W"])

def induced_velocity_forward_flight(v_hover: float,
                                    airspeed_mps: float,
                                    disk_incidence_rad: float = 0.0,
                                    iters: int = 40) -> float:
    """
    Rotor induced velocity in forward flight (Glauert's momentum theory).

    A hovering rotor must accelerate still air, so it works hard:
        v_h = sqrt( T / (2 * rho * A) )

    Once the vehicle is moving, the rotor meets air that is ALREADY moving,
    so far less velocity has to be added and induced power falls sharply.
    Glauert's relation for a rotor at incidence alpha to the freestream is:

        vi = v_h^2 / sqrt( (V*cos a)^2 + (V*sin a + vi)^2 )

    Special cases this reduces to:
      * V = 0            ->  vi = v_h                     (hover)
      * a = 90 deg       ->  vi = -V/2 + sqrt((V/2)^2 + v_h^2)
                                                          (axial / propeller)
      * a = 0 deg        ->  vi = v_h^2 / sqrt(V^2 + vi^2)
                                                          (edgewise / multirotor)

    A multicopter in forward flight sits near the edgewise case: its rotors
    stay nearly horizontal and only tilt by the angle needed to balance drag,
    so the freestream is almost in the disk plane.

    Why this matters
    ----------------
    Using the hover value at all speeds overstates induced power by roughly
    2x at 10 m/s and 4x at 20 m/s for a typical quad.  That removes the
    "power bucket" — the 10-25% dip below hover power that real multirotors
    show around 8-14 m/s — and leaves a monotonically rising power curve.
    With no minimum in the curve, best-endurance and best-range speed
    searches simply pin to the ends of their search range.

    This returns vi only.  Shaft power is then
        P = T * (V*sin(a) + vi)
    where the first term is the propulsive power overcoming airframe drag.

    Solved by damped fixed-point iteration, which converges quickly because
    the right-hand side is a contraction for all physical inputs.
    """
    vh = max(float(v_hover), 0.0)
    V = max(float(airspeed_mps), 0.0)
    if vh <= 0.0:
        return 0.0
    if V <= 1e-9:
        return vh                       # hover

    a = float(disk_incidence_rad)
    v_par = V * math.cos(a)             # in-plane component
    v_perp = V * math.sin(a)            # through-disk component

    vi = vh                             # start from the hover value
    for _ in range(max(int(iters), 1)):
        denom = math.sqrt(v_par ** 2 + (v_perp + vi) ** 2)
        vi_new = (vh ** 2) / max(denom, 1e-9)
        # Damping keeps the iteration stable near the vortex-ring region.
        vi_next = 0.5 * (vi + vi_new)
        if abs(vi_next - vi) < 1e-9:
            vi = vi_next
            break
        vi = vi_next
    return max(vi, 0.0)


def motor_power_from_params(config: DroneConfig, thrust_per_motor_N: float,
                            airspeed_mps: float = 0.0,
                            disk_incidence_rad: float = 0.0) -> float:
    """
    Estimate electrical input power required for a given thrust per motor.

    Two options:
      - If propeller TConst/PConst provided: solve for RPM via thrust model:
            T = C_T * rho * n^2 * D^4
        Then mechanical shaft power:
            P_mech = C_P * rho * n^3 * D^5
      - Else: momentum-theory induced power, corrected for forward flight:
            v_h = sqrt(T / (2*rho*A))
            vi  = induced_velocity_forward_flight(v_h, V, incidence)
            P_mech ≈ T * vi
        At V = 0 this is the familiar hover result; at speed the rotor meets
        already-moving air, so vi (and induced power) drop sharply.

    Electrical conversion (very simplified):
      - torque constant: Kt = 60 / (2π Kv)  [Nm/A]
      - approximate motor torque from mech power and omega
      - current ≈ torque/Kt + I0
      - copper loss via motor resistance
      - clamp by motor max current/power
    """
    if config.motor.kv is None:
        raise ValueError("motor_kv must be set to use motor electrical model.")

    D = config.propeller.diameter_in * 0.0254  # meters
    rho = config.air_density

    # ---- Mechanical shaft power ----
    if config.propeller.TConst and config.propeller.PConst:
        low, high = 100.0, 40000.0  # RPM bounds
        rpm_solution = None

        for _ in range(40):
            mid = 0.5 * (low + high)
            n = mid / 60.0  # rev/s
            thrust = config.propeller.TConst * rho * (n**2) * (D**4)
            if thrust < thrust_per_motor_N:
                low = mid
            else:
                high = mid
                rpm_solution = mid

        if rpm_solution is None:
            return 0.0

        n = rpm_solution / 60.0
        mech_power_W = config.propeller.PConst * rho * (n**3) * (D**5)
        omega = 2.0 * math.pi * n
        torque_Nm = mech_power_W / max(omega, 1e-9)
    else:
        A = disk_area(config.propeller.diameter_in)
        v_hover = math.sqrt(max(thrust_per_motor_N, 0.0) / max(2.0 * rho * A, 1e-9))
        vi = induced_velocity_forward_flight(v_hover, airspeed_mps, disk_incidence_rad)
        # Shaft power = thrust x (through-disk freestream + induced velocity).
        #   P = T * (V*sin(a) + vi)
        # The first term is the propulsive power that overcomes airframe drag:
        # by the tilt balance T*sin(a) = D, so T*V*sin(a) = D*V exactly.
        # Dropping it (as an induced-only model does) makes power fall without
        # bound at speed; at V = 0 it vanishes and hover is unchanged.
        v_through = max(float(airspeed_mps), 0.0) * math.sin(float(disk_incidence_rad))
        mech_power_W = thrust_per_motor_N * (v_through + vi)
        # crude omega/torque estimate from Kv and voltage
        omega = (config.battery.vnom_pack * config.motor.kv) * (2.0 * math.pi / 60.0)
        torque_Nm = mech_power_W / max(omega, 1e-9)

    # ---- Electrical model ----
    kt = 60.0 / (2.0 * math.pi * config.motor.kv)  # Nm/A
    current_A = torque_Nm / max(kt, 1e-9) + config.motor.idle_current

    # motor terminal voltage ~ pack voltage - I*R (motor copper)
    v_drop = current_A * config.motor.resistance
    v_eff = max(config.battery.vnom_pack - v_drop, 0.0)
    input_power_W = v_eff * current_A

    # Enforce current/power limits
    if config.motor.max_current and current_A > config.motor.max_current:
        current_A = config.motor.max_current
        input_power_W = config.battery.vnom_pack * current_A
    if config.motor.max_power and input_power_W > config.motor.max_power:
        input_power_W = config.motor.max_power

    return float(input_power_W)



def motor_configuration_power_multiplier(config: DroneConfig, orientation: str) -> float:
    """Return a multiplier applied to per-motor electrical power based on motor configuration.

    Coaxial (stacked) rotors suffer aerodynamic interference that depends strongly on the
    vertical spacing between rotors (relative to rotor diameter). This models that effect
    as a smooth penalty multiplier on *per-motor* electrical power for a given required thrust.

    - For motor_configuration == "flat": returns 1.0.
    - For motor_configuration == "coaxial": returns > 1.0 depending on spacing.

    The intent is to provide a reasonable default without requiring detailed coaxial aero data.
    """
    cfg = getattr(config, "motor_configuration", "flat")
    if cfg != "coaxial":
        return 1.0

    # Rotor diameter (m) from prop config (inches -> meters)
    try:
        diameter_m = float(config.propeller.diameter_in) * 0.0254
    except Exception:
        diameter_m = 0.0

    # If we can't compute a ratio, fall back to a conservative default.
    if diameter_m <= 0.0:
        return 1.18 if orientation == "hover" else 1.12

    # Use user spacing if provided; otherwise assume a typical ~0.2D spacing.
    spacing_m = getattr(config, "coaxial_spacing_m", None)
    if spacing_m is None or float(spacing_m) <= 0.0:
        spacing_m = 0.20 * diameter_m

    spacing_ratio = max(0.0, float(spacing_m) / diameter_m)

    # Penalty model:
    # - Higher penalty at small spacing, diminishing as spacing increases.
    # - Hover interference tends to be worse than in forward flight.
    # Tuning targets: ~1.18 at ~0.2D in hover; smaller in forward flight.
    inc = 0.25 * math.exp(-3.0 * spacing_ratio) + 0.03  # 3% floor + spacing-dependent term
    if orientation != "hover":
        inc *= 0.70  # reduced penalty in forward flight

    return 1.0 + inc

def power_required(config: DroneConfig,
                   speed_mps: float,
                   orientation: str,
                   inflow_multiplier: Optional[float] = None) -> float:
    """
    Total electrical power for all motors (W), not including peripheral current.
    """
    total_thrust_N = thrust_required(config, speed_mps, orientation)
    thrust_per_motor_N = total_thrust_N / config.num_motors

    # Disk incidence: the angle between the freestream and the rotor DISK
    # PLANE.  A multicopter only tilts by the amount needed to balance drag,
    # so the airflow is close to edgewise (incidence ~ tilt angle).  In hover
    # there is no freestream, so the incidence is irrelevant.
    if orientation == "forward":
        incidence_rad = math.radians(required_tilt_deg(config, speed_mps, "forward"))
        airspeed_for_inflow = max(float(speed_mps), 0.0)
    else:
        incidence_rad = 0.0
        airspeed_for_inflow = max(float(speed_mps), 0.0)

    if config.propeller.table is not None:
        # A measured table already embeds the real operating point, so no
        # analytic inflow correction is applied on top of it.
        motor_power_W = interpolate_motor_power(config, thrust_per_motor_N)
    elif config.motor.kv is not None:
        motor_power_W = motor_power_from_params(
            config, thrust_per_motor_N,
            airspeed_mps=airspeed_for_inflow,
            disk_incidence_rad=incidence_rad)
    else:
        A = disk_area(config.propeller.diameter_in)
        v_hover = math.sqrt(max(thrust_per_motor_N, 0.0) /
                            max(2.0 * config.air_density * A, 1e-9))
        vi = induced_velocity_forward_flight(v_hover, airspeed_for_inflow, incidence_rad)
        v_through = airspeed_for_inflow * math.sin(incidence_rad)
        motor_power_W = (thrust_per_motor_N * (v_through + vi)) / 0.85

    # Apply motor-configuration penalty (e.g., coaxial interference)
    motor_power_W *= motor_configuration_power_multiplier(config, orientation)

    if inflow_multiplier is not None:
        motor_power_W *= max(float(inflow_multiplier), 0.2)

    return motor_power_W * config.num_motors


def _compute_operating_metrics_core(drone: DroneConfig,
                                    speed_mps: float,
                                    orientation: str,
                                    wind_mps: float = 0.0,
                                    wind_direction_deg: float = 0.0,
                                    course_deg: float = 0.0,
                                    ambient_temp_C: float = 25.0,
                                    soc: Optional[float] = None) -> dict:
    input_speed = float(speed_mps)
    headwind_mps, crosswind_mps = wind_components_mps(wind_mps, wind_direction_deg, course_deg)
    airspeed = input_speed
    groundspeed_mps = input_speed
    if orientation == "forward":
        airspeed = max(0.0, float(input_speed))
        groundspeed_mps = groundspeed_along_track_mps(airspeed, headwind_mps, crosswind_mps)
    drone.derive_drag_from_geometry_if_missing()
    # First pass for RPM estimate (needed to compute mu for inflow map).
    motor_power_W = power_required(drone, airspeed, orientation)
    periph_power_W = avionics_input_power_W(getattr(drone, "avionics", None))
    if periph_power_W <= 0.0:
        periph_power_W = drone.battery.vnom_pack * max(drone.periph_current, 0.0)
    soc_eval = 1.0 if soc is None else min(max(float(soc), 0.0), 1.0)
    total_power_W, v_load, pack_current_A, esc_note, motor_I_esc_A = total_power_with_esc(
        drone, motor_power_W=motor_power_W, periph_power_W=periph_power_W, soc=soc_eval)
    total_thrust_N     = thrust_required(drone, airspeed, orientation)
    thrust_per_motor_N = total_thrust_N / max(int(drone.num_motors), 1)
    esc_loss_W  = max(0.0, float(total_power_W) - float(motor_power_W) - float(periph_power_W))
    rpm_est     = None
    if drone.propeller.TConst and drone.propeller.PConst:
        try:
            D   = drone.propeller.diameter_in * 0.0254
            rho = drone.air_density
            lo, hi = 100.0, 40000.0
            sol = None
            for _ in range(40):
                mid = 0.5*(lo+hi)
                n   = mid/60.0
                T   = float(drone.propeller.TConst)*rho*(n**2)*(D**4)
                if T < thrust_per_motor_N:
                    lo = mid
                else:
                    hi = mid
                    sol = mid
            rpm_est = float(sol) if sol is not None else None
        except Exception:
            rpm_est = None
    motor_table = None
    if getattr(drone, 'propeller', None) is not None and drone.propeller.table is not None:
        try:
            motor_table = interpolate_motor_point(drone, thrust_per_motor_N)
            if 'RPM' in motor_table:
                rpm_est = float(motor_table['RPM'])
        except Exception:
            motor_table = None
    inflow_mult, mu_adv, eta_inflow = rotor_inflow_power_multiplier(
        drone, airspeed_mps=airspeed, rpm=rpm_est, orientation=orientation
    )
    if orientation == "forward":
        # Recompute propulsion power with inflow multiplier applied.
        motor_power_W = power_required(drone, airspeed, orientation, inflow_multiplier=inflow_mult)
        total_power_W, v_load, pack_current_A, esc_note, motor_I_esc_A = total_power_with_esc(
            drone, motor_power_W=motor_power_W, periph_power_W=periph_power_W, soc=soc_eval
        )
        esc_loss_W = max(0.0, float(total_power_W) - float(motor_power_W) - float(periph_power_W))
    tilt_req = required_tilt_deg(drone, airspeed, orientation)

    hover_thrust_total_N = thrust_required(drone, 0.0, "hover")
    hover_thrust_pm_N = hover_thrust_total_N / max(int(drone.num_motors), 1)
    hover_motor_power_W = power_required(drone, 0.0, "hover")
    hover_total_power_W, _, _, _, _ = total_power_with_esc(
        drone, motor_power_W=hover_motor_power_W, periph_power_W=periph_power_W
    )
    hover_propulsion_power_W = max(float(hover_total_power_W) - float(periph_power_W), 0.0)
    hover_thrust_total_g = hover_thrust_total_N * 1000.0 / 9.81
    hover_efficiency_gW = hover_thrust_total_g / hover_propulsion_power_W if hover_propulsion_power_W > 0 else 0.0

    A_total = total_disk_area(drone)
    p_ideal_hover = hover_ideal_induced_power_W(drone, hover_thrust_total_N)
    p_actual_induced = hover_propulsion_power_W
    fm_hover = hover_figure_of_merit(drone, hover_thrust_total_N, p_actual_induced)

    tip_speed = tip_speed_mps_from_rpm(drone.propeller.diameter_in, float(rpm_est)) if rpm_est is not None else float("nan")
    tip_mach = tip_speed / 340.0 if tip_speed == tip_speed else float("nan")

    p_copper = (motor_I_esc_A ** 2) * float(getattr(drone.motor, "resistance", 0.0))
    # The motor thermal estimate is anchored to the motor's rated current:
    # at the rating we assume a 55 C rise, scaling with I^2.  If no rating was
    # supplied there is nothing to scale against, so report ambient rather
    # than crashing or inventing a limit.
    _i_max_raw = getattr(drone.motor, "max_current", None)
    if _i_max_raw is None:
        thermal_rise_C = 0.0
    else:
        motor_i_max = max(float(_i_max_raw), 1e-9)
        thermal_rise_C = 55.0 * (motor_I_esc_A / motor_i_max) ** 2 if motor_I_esc_A >= 0 else 0.0
    motor_temp_est_C = float(ambient_temp_C) + thermal_rise_C
    esc_temp_est_C = float(ambient_temp_C) + float(esc_loss_W) * 0.75
    battery_loss_W = (float(pack_current_A) ** 2) * max(
        float(battery_pack_resistance(drone.battery, soc_eval)),
        0.0
    )
    battery_temp_est_C = float(ambient_temp_C) + battery_loss_W * 0.25
    motor_thermal_headroom_C = 120.0 - motor_temp_est_C
    max_temp_c = max(motor_temp_est_C, esc_temp_est_C, battery_temp_est_C)
    thermal_status = "OK" if max_temp_c < 95.0 else ("WARN" if max_temp_c < 115.0 else "HOT")

    rotor_solidity = propeller_solidity(drone.propeller.diameter_in, drone.propeller.blades)
    blade_chord_est_m = estimate_blade_chord_m(drone.propeller.diameter_in, drone.propeller.blades)
    dl = disk_loading_N_m2(drone)
    v_max_wind = hover_wind_resistance_mps(drone)

    return {
        "airspeed_mps":        float(airspeed),
        "groundspeed_mps":     float(groundspeed_mps),
        "wind_head_mps":       float(headwind_mps),
        "wind_cross_mps":      float(crosswind_mps),
        "wind_direction_deg":  float(wind_direction_deg),
        "course_deg":          float(course_deg),
        "soc":                float(soc_eval),
        "soc_percent":        float(soc_eval * 100.0),
        "soc_model_source":   str(getattr(drone.battery, "soc_model_source", "linear-fallback")),
        "tilt_required_deg":   float(tilt_req),
        "tilt_limit_deg":      (float(drone.max_tilt_deg) if getattr(drone,'max_tilt_deg',None) is not None else None),
        "motor_power_W":       float(motor_power_W),
        "periph_power_W":      float(periph_power_W),
        "total_power_W":       float(total_power_W),
        "v_load_V":            float(v_load),
        "pack_current_A":      float(pack_current_A),
        "esc_loss_W":          float(esc_loss_W),
        "esc_note":            str(esc_note),
        "motor_I_per_esc_A":   float(motor_I_esc_A),
        "thrust_total_N":      float(total_thrust_N),
        "thrust_per_motor_N":  float(thrust_per_motor_N),
        "prop_rpm":            (float(rpm_est) if rpm_est is not None else None),
        "motor_table_throttle_pct": (float(motor_table["Throttle_pct"]) if motor_table and "Throttle_pct" in motor_table else None),
        "motor_table_eff_gW":  (float(motor_table["Efficiency_gW"])    if motor_table and "Efficiency_gW" in motor_table else None),
        "motor_table_temp_C":  (float(motor_table["Temp_C"])           if motor_table and "Temp_C" in motor_table else None),
        "hover_efficiency_gW": float(hover_efficiency_gW),
        "hover_thrust_total_N": float(hover_thrust_total_N),
        "hover_thrust_per_motor_N": float(hover_thrust_pm_N),
        "hover_propulsion_power_W": float(hover_propulsion_power_W),
        "hover_ideal_power_W": float(p_ideal_hover),
        "actual_induced_power_W": float(p_actual_induced),
        "figure_of_merit":     float(fm_hover),
        "disk_loading_N_m2":   float(dl),
        "total_disk_area_m2":  float(A_total),
        "tip_speed_mps":       (float(tip_speed) if math.isfinite(tip_speed) else None),
        "tip_mach":            (float(tip_mach) if math.isfinite(tip_mach) else None),
        "advance_ratio_mu":    float(mu_adv),
        "inflow_efficiency":   float(eta_inflow),
        "inflow_power_multiplier": float(inflow_mult),
        "noise_significant":   bool(math.isfinite(tip_mach) and tip_mach > 0.6),
        "motor_copper_loss_W_per_motor": float(p_copper),
        "battery_loss_W":      float(battery_loss_W),
        "motor_temp_est_C":    float(motor_temp_est_C),
        "esc_temp_est_C":      float(esc_temp_est_C),
        "battery_temp_est_C":  float(battery_temp_est_C),
        "motor_thermal_headroom_C": float(motor_thermal_headroom_C),
        "thermal_status":      thermal_status,
        "hover_wind_resistance_mps": float(v_max_wind),
        "prop_solidity_sigma": float(rotor_solidity),
        "blade_chord_est_m":   float(blade_chord_est_m),
    }


def compute_operating_metrics(config: DroneConfig,
                              speed_mps: float,
                              orientation: str,
                              wind_mps: float = 0.0,
                              wind_direction_deg: float = 0.0,
                              course_deg: float = 0.0,
                              ambient_temp_C: float = 25.0,
                              soc: Optional[float] = None) -> dict:
    """Shared operating-point metrics for GUI, mission simulation, and CLI."""
    return _compute_operating_metrics_core(
        config,
        speed_mps=speed_mps,
        orientation=orientation,
        wind_mps=wind_mps,
        wind_direction_deg=wind_direction_deg,
        course_deg=course_deg,
        ambient_temp_C=ambient_temp_C,
        soc=soc,
    )


# -------------------------------
# Flight Performance
# -------------------------------
def estimate_flight_time_minutes(config: DroneConfig, speed_mps: float, orientation: str = "forward") -> float:
    """
    Returns minutes of flight time based on:
      usable_energy_Wh / total_power_W

    Total power includes motor power plus avionics/peripheral draw.

    Avionics draw can be provided either as:
      - Legacy: config.periph_current (A at pack input), OR
      - Voltage rails: config.avionics.voltage_tree = {Vrail: (Irail, eff), ...}

    When voltage rails are used, we treat the avionics *output* loads as constant,
    so the BEC input power is:
        P_avionics_in = sum(Vrail * Irail / eff)
    and the pack current increases as pack voltage sags.
    """
    # Enforce max tilt in forward flight (if provided)
    if orientation == "forward" and getattr(config, "max_tilt_deg", None) is not None:
        tilt_req = required_tilt_deg(config, speed_mps, orientation="forward")
        if tilt_req > float(config.max_tilt_deg) + 1e-9:
            return 0.0
    soc = 1.0
    motor_power_W = power_required(config, speed_mps, orientation)
    periph_power_W = avionics_input_power_W(getattr(config, "avionics", None))
    if periph_power_W <= 0.0:
        # Legacy behavior: constant current draw at the pack input
        periph_power_W = config.battery.vnom_pack * max(config.periph_current, 0.0)

    total_power_W, v_load, pack_current_A, esc_note, motor_I_esc_A = total_power_with_esc(
        config,
        motor_power_W=motor_power_W,
        periph_power_W=periph_power_W,
        soc=soc,
    )

    if total_power_W <= 0:
        return 0.0

    # Discharge limit checks (battery)
    if pack_current_A > config.battery.discharge_max_A:
        return 0.0
    
    if v_load < config.battery.vmin_pack:
        return 0.0
    
    # ESC limit checks
    if getattr(config, "esc", None) is not None and motor_I_esc_A > config.esc.max_current_A:
        return 0.0

    time_h = config.battery.usable_Wh / total_power_W
    return float(time_h * 60.0)


def estimate_flight_distance_km(config: DroneConfig, speed_mps: float, orientation: str = "forward") -> float:
    t_min = estimate_flight_time_minutes(config, speed_mps, orientation)
    return float((speed_mps * (t_min * 60.0)) / 1000.0)


def find_optimal_speeds(config: DroneConfig, min_speed: float = 1.0, max_speed: float = 30.0, step: float = 0.5):
    """
    Returns:
      - best_endurance_speed (max minutes)
      - best_range_speed (max km)
    """
    best_endurance_speed, best_minutes = 0.0, -1.0
    best_range_speed, best_km = 0.0, -1.0

    n_steps = int((max_speed - min_speed) / step) + 1
    for i in range(n_steps):
        v = min_speed + i * step
        t = estimate_flight_time_minutes(config, v, orientation="forward")
        d = estimate_flight_distance_km(config, v, orientation="forward")
        if t > best_minutes:
            best_minutes = t
            best_endurance_speed = v
        if d > best_km:
            best_km = d
            best_range_speed = v

    return best_endurance_speed, best_minutes, best_range_speed, best_km


def simulate_mission(config: DroneConfig,
                     mission: MissionProfile,
                     orientation: str = "forward",
                     temperature_C: Optional[float] = None,
                     pressure_Pa: Optional[float] = None,
                     wind_mps: float = 0.0) -> Tuple[List[Tuple[str, float, float, str]], Optional[dict], Optional[dict]]:
    """
    Simulate mission phases, draining remaining energy (Wh).

    Wind sign convention:
      +wind_mps = headwind (airspeed > groundspeed)
      -wind_mps = tailwind

    Each result tuple:
      (phase_name, time_minutes, distance_km, status_string)

    Returns:
      (results, worst_metrics) where worst_metrics aggregates worst-case values across phases
      for limit/status checks.
    """
    usable_wh = config.battery.usable_Wh
    reserve_target_wh = max(
        usable_wh * (float(mission.reserve_percent) / 100.0),
        float(mission.rth_reserve_Wh) + float(mission.diversion_reserve_Wh),
    )
    remaining_wh = usable_wh
    results: List[Tuple[str, float, float, str]] = []
    worst_metrics: Optional[dict] = None
    reserve_breached = False
    reserve_min_wh = remaining_wh
    soc_state = 1.0
    ambient_c = float(temperature_C) if temperature_C is not None else 25.0
    motor_temp_c = ambient_c
    esc_temp_c = ambient_c
    battery_temp_c = ambient_c
    prev_alt_m = float(mission.phases[0].altitude) if mission.phases else 0.0
    current_speed_mps = 0.0
    soc_state = 1.0

    base_dt_s = max(
        float(getattr(mission, "transient_dt_s", getattr(config, "transient_dt_s", 0.5))),
        0.05
    )
    accel_up_mps2 = max(
        float(getattr(mission, "max_accel_mps2", getattr(config, "max_accel_mps2", 2.0))),
        0.01
    )
    accel_down_mps2 = max(
        float(getattr(mission, "max_decel_mps2", getattr(config, "max_decel_mps2", 2.5))),
        0.01
    )
    regen_eff = min(max(
        float(getattr(mission, "decel_regen_eff", getattr(config, "decel_regen_eff", 0.0))),
        0.0
    ), 1.0)
    soc_state = 1.0

    mission_series: dict = {
        't_s': [],
        'phase': [],
        'segment_type': [],
        'airspeed_mps': [],
        'commanded_airspeed_mps': [],
        'accel_mps2': [],
        'groundspeed_mps': [],
        'headwind_mps': [],
        'crosswind_mps': [],
        'distance_km': [],
        'altitude_m': [],
        'tilt_deg': [],
        'climb_rate_cmd_mps': [],
        'descent_rate_cmd_mps': [],
        'climb_power_add_W': [],
        'potential_power_W': [],
        'kinetic_power_W': [],
        'battery_voltage_V': [],
        'battery_current_A': [],
        'battery_energy_Wh': [],
        'battery_soc_frac': [],
        'battery_soc_percent': [],
        'reserve_target_Wh': [],
        'reserve_margin_Wh': [],
        'reserve_breach': [],
        'battery_capacity_mAh': [],
        'total_power_W': [],
        'motor_power_W': [],
        'motor_power_per_motor_W': [],
        'motor_current_A': [],
        'motor_rpm': [],
        'motor_thrust_N': [],
        'thrust_total_N': [],
        'periph_power_W': [],
        'esc_loss_W': [],
        'hover_efficiency_gW': [],
        'figure_of_merit': [],
        'disk_loading_N_m2': [],
        'tip_mach': [],
        'advance_ratio_mu': [],
        'inflow_efficiency': [],
        'inflow_power_multiplier': [],
        'motor_temp_est_C': [],
        'esc_temp_est_C': [],
        'battery_temp_est_C': [],
        'thermal_status': [],
        'hover_wind_resistance_mps': [],
        'prop_solidity_sigma': [],
    }

    t_s = 0.0
    dist_km = 0.0

    def _append_point(phase_name: str,
                      phase_alt_m: float,
                      segment_type: str,
                      m: dict,
                      t_s_now: float,
                      dist_km_now: float,
                      remaining_wh_now: float,
                      reserve_hit: bool):
        mission_series['t_s'].append(float(t_s_now))
        mission_series['phase'].append(str(phase_name))
        mission_series['segment_type'].append(str(segment_type))
        mission_series['airspeed_mps'].append(float(m.get('airspeed_mps', 0.0)))
        mission_series['commanded_airspeed_mps'].append(float(m.get('commanded_airspeed_mps', 0.0)))
        mission_series['accel_mps2'].append(float(m.get('accel_mps2', 0.0)))
        mission_series['groundspeed_mps'].append(float(m.get('groundspeed_mps', 0.0)))
        mission_series['headwind_mps'].append(float(m.get('wind_head_mps', 0.0)))
        mission_series['crosswind_mps'].append(float(m.get('wind_cross_mps', 0.0)))
        mission_series['distance_km'].append(float(dist_km_now))
        mission_series['altitude_m'].append(float(phase_alt_m))
        mission_series['tilt_deg'].append(float(m.get('tilt_required_deg', 0.0)))
        mission_series['climb_rate_cmd_mps'].append(float(m.get('climb_rate_cmd_mps', 0.0)))
        mission_series['descent_rate_cmd_mps'].append(float(m.get('descent_rate_cmd_mps', 0.0)))
        mission_series['climb_power_add_W'].append(float(m.get('climb_power_add_W', 0.0)))
        mission_series['potential_power_W'].append(float(m.get('potential_power_W', 0.0)))
        mission_series['kinetic_power_W'].append(float(m.get('kinetic_power_W', 0.0)))
        mission_series['battery_voltage_V'].append(float(m.get('v_load_V', 0.0)))
        mission_series['battery_current_A'].append(float(m.get('pack_current_A', 0.0)))
        mission_series['battery_energy_Wh'].append(float(remaining_wh_now))
        soc_now = float(m.get('soc', soc_state))
        mission_series['battery_soc_frac'].append(float(soc_now))
        mission_series['battery_soc_percent'].append(float(soc_now * 100.0))
        mission_series['reserve_target_Wh'].append(float(reserve_target_wh))
        mission_series['reserve_margin_Wh'].append(float(remaining_wh_now - reserve_target_wh))
        mission_series['reserve_breach'].append(1 if reserve_hit else 0)
        vnom = float(config.battery.vnom_pack) if float(config.battery.vnom_pack) > 1e-9 else 1.0
        mission_series['battery_capacity_mAh'].append(float(remaining_wh_now) * 1000.0 / vnom)
        mission_series['total_power_W'].append(float(m.get('total_power_W', 0.0)))
        mp = float(m.get('motor_power_W', 0.0))
        mission_series['motor_power_W'].append(mp)
        mission_series['motor_power_per_motor_W'].append(mp / max(int(config.num_motors), 1))
        mission_series['motor_current_A'].append(float(m.get('motor_I_per_esc_A', 0.0)))
        mission_series['motor_rpm'].append(float(m.get('prop_rpm')) if m.get('prop_rpm') is not None else float('nan'))
        mission_series['motor_thrust_N'].append(float(m.get('thrust_per_motor_N', 0.0)))
        mission_series['thrust_total_N'].append(float(m.get('thrust_total_N', 0.0)))
        mission_series['periph_power_W'].append(float(m.get('periph_power_W', 0.0)))
        mission_series['esc_loss_W'].append(float(m.get('esc_loss_W', 0.0)))
        mission_series['hover_efficiency_gW'].append(float(m.get('hover_efficiency_gW', 0.0)))
        mission_series['figure_of_merit'].append(float(m.get('figure_of_merit', 0.0)))
        mission_series['disk_loading_N_m2'].append(float(m.get('disk_loading_N_m2', 0.0)))
        mission_series['tip_mach'].append(float(m.get('tip_mach')) if m.get('tip_mach') is not None else float('nan'))
        mission_series['advance_ratio_mu'].append(float(m.get('advance_ratio_mu', 0.0)))
        mission_series['inflow_efficiency'].append(float(m.get('inflow_efficiency', 1.0)))
        mission_series['inflow_power_multiplier'].append(float(m.get('inflow_power_multiplier', 1.0)))
        mission_series['motor_temp_est_C'].append(float(motor_temp_c))
        mission_series['esc_temp_est_C'].append(float(esc_temp_c))
        mission_series['battery_temp_est_C'].append(float(battery_temp_c))
        mission_series['thermal_status'].append(str(m.get('thermal_status', 'OK')))
        mission_series['hover_wind_resistance_mps'].append(float(m.get('hover_wind_resistance_mps', 0.0)))
        mission_series['prop_solidity_sigma'].append(float(m.get('prop_solidity_sigma', 0.0)))

    def _merge_worst(worst: Optional[dict], m: dict) -> dict:
        if worst is None:
            return dict(m)
        if m.get("soc") is not None:
            worst["soc"] = min(float(worst.get("soc", 1.0)), float(m.get("soc", 1.0)))
            worst["soc_percent"] = min(float(worst.get("soc_percent", 100.0)), float(m.get("soc_percent", 100.0)))
            worst["soc_model_source"] = str(m.get("soc_model_source", worst.get("soc_model_source", "linear-fallback")))
        for k in ("pack_current_A", "total_power_W", "motor_power_W", "periph_power_W",
                  "esc_loss_W", "motor_I_per_esc_A", "thrust_total_N", "thrust_per_motor_N",
                  "motor_temp_est_C", "esc_temp_est_C", "battery_temp_est_C",
                  "accel_mps2", "commanded_airspeed_mps", "advance_ratio_mu"):
            worst[k] = max(float(worst.get(k, 0.0)), float(m.get(k, 0.0)))
        worst["v_load_V"] = min(float(worst.get("v_load_V", 1e9)), float(m.get("v_load_V", 1e9)))
        worst["reserve_margin_Wh"] = min(float(worst.get("reserve_margin_Wh", 1e9)),
                                         float(m.get("reserve_margin_Wh", 1e9)))
        if worst.get("prop_rpm") is None:
            worst["prop_rpm"] = m.get("prop_rpm")
        elif m.get("prop_rpm") is not None:
            worst["prop_rpm"] = max(float(worst["prop_rpm"]), float(m["prop_rpm"]))
        if str(m.get("esc_note", "")).strip():
            worst["esc_note"] = (str(worst.get("esc_note", "")) + "; " + str(m.get("esc_note", ""))).strip("; ")
        return worst

    depleted = False
    for phase in mission.phases:
        rho = compute_air_density(
            altitude_m=phase.altitude,
            temperature_C=temperature_C,
            pressure_Pa=pressure_Pa
        )
        config.air_density = rho

        headwind_mps, crosswind_mps = wind_components_mps(
            wind_speed_mps=wind_mps,
            wind_direction_deg=float(mission.wind_direction_deg),
            course_deg=float(phase.course_deg),
        )

        climb_cmd = max(float(phase.climb_rate_mps or 0.0), 0.0)
        descent_cmd = max(float(phase.descent_rate_mps or 0.0), 0.0)
        if climb_cmd > 0 and descent_cmd > 0:
            descent_cmd = 0.0
        if phase.climb_rate_mps is None and phase.descent_rate_mps is None:
            dh = float(phase.altitude) - prev_alt_m
            if phase.duration is not None and float(phase.duration) > 0:
                est_vz = dh / float(phase.duration)
                if est_vz > 0:
                    climb_cmd = est_vz
                elif est_vz < 0:
                    descent_cmd = -est_vz

        target_speed_mps = max(float(phase.speed), 0.0)
        phase_elapsed_s = 0.0
        phase_distance_m = 0.0
        phase_status = "OK"
        step_guard = 0

        _append_point(
            phase_name=phase.name,
            phase_alt_m=float(phase.altitude),
            segment_type="phase-start",
            m={
                "airspeed_mps": current_speed_mps,
                "commanded_airspeed_mps": target_speed_mps,
                "groundspeed_mps": groundspeed_along_track_mps(current_speed_mps, headwind_mps, crosswind_mps),
                "wind_head_mps": headwind_mps,
                "wind_cross_mps": crosswind_mps,
                "accel_mps2": 0.0,
                "climb_rate_cmd_mps": climb_cmd,
                "descent_rate_cmd_mps": descent_cmd,
                "potential_power_W": 0.0,
                "climb_power_add_W": 0.0,
                "kinetic_power_W": 0.0,
                "advance_ratio_mu": 0.0,
                "inflow_efficiency": 1.0,
                "inflow_power_multiplier": 1.0,
            },
            t_s_now=t_s,
            dist_km_now=dist_km,
            remaining_wh_now=remaining_wh,
            reserve_hit=reserve_breached,
        )

        while True:
            step_guard += 1
            if step_guard > 200000:
                phase_status = "Aborted: mission step limit reached"
                break

            if phase.duration is not None:
                remain_s = float(phase.duration) - phase_elapsed_s
                if remain_s <= 1e-9:
                    break
                dt_s = min(base_dt_s, remain_s)
            elif phase.distance is not None:
                remain_m = float(phase.distance) - phase_distance_m
                if remain_m <= 1e-6:
                    break
                dt_s = base_dt_s
            else:
                phase_status = "Invalid: phase missing duration/distance"
                break

            dv_cmd = target_speed_mps - current_speed_mps
            if dv_cmd >= 0.0:
                dv = min(dv_cmd, accel_up_mps2 * dt_s)
            else:
                dv = max(dv_cmd, -accel_down_mps2 * dt_s)
            v_next = max(0.0, current_speed_mps + dv)
            accel_mps2 = dv / dt_s if dt_s > 0 else 0.0
            v_eval = 0.5 * (current_speed_mps + v_next)
            segment_type = "transient" if abs(accel_mps2) > 1e-4 else "steady"

            m = compute_operating_metrics(
                config,
                speed_mps=v_eval,
                orientation=orientation,
                wind_mps=wind_mps,
                wind_direction_deg=float(mission.wind_direction_deg),
                course_deg=float(phase.course_deg),
                ambient_temp_C=ambient_c,
                soc=soc_state,
            )

            potential_power_w = (config.drone_weight_g * 9.81 / 1000.0) * (climb_cmd - descent_cmd)
            kinetic_power_w = kinetic_power_term_W(
                config.drone_weight_g,
                current_speed_mps,
                v_next,
                dt_s,
                regen_eff,
            )
            base_total_power_w = float(m.get("total_power_W", 0.0))
            total_power_w = max(base_total_power_w + potential_power_w + kinetic_power_w, 0.0)
            scale = (total_power_w / base_total_power_w) if base_total_power_w > 1e-9 else 1.0
            if base_total_power_w > 1e-9:
                m["motor_power_W"] = float(m.get("motor_power_W", 0.0)) * scale
                m["esc_loss_W"] = float(m.get("esc_loss_W", 0.0)) * scale
                m["motor_I_per_esc_A"] = float(m.get("motor_I_per_esc_A", 0.0)) * max(scale, 0.0)
            m["climb_rate_cmd_mps"] = climb_cmd
            m["descent_rate_cmd_mps"] = descent_cmd
            m["potential_power_W"] = potential_power_w
            m["kinetic_power_W"] = kinetic_power_w
            m["climb_power_add_W"] = potential_power_w + kinetic_power_w
            m["total_power_W"] = total_power_w
            m["pack_current_A"] = total_power_w / max(float(config.battery.vnom_pack), 1.0)
            v_solve, i_solve = solve_pack_voltage_and_current(
                config.battery, total_power_w, soc=soc_state
            )
            m["v_load_V"] = float(v_solve)
            m["pack_current_A"] = float(i_solve)
            gs_mps = groundspeed_along_track_mps(v_eval, headwind_mps, crosswind_mps)
            m["airspeed_mps"] = v_eval
            m["commanded_airspeed_mps"] = target_speed_mps
            m["accel_mps2"] = accel_mps2
            m["groundspeed_mps"] = gs_mps
            m["wind_head_mps"] = headwind_mps
            m["wind_cross_mps"] = crosswind_mps
            m["reserve_margin_Wh"] = remaining_wh - reserve_target_wh
            m["soc_model_source"] = str(getattr(config.battery, "soc_model_source", "linear-fallback"))

            if phase.distance is not None:
                remain_m = float(phase.distance) - phase_distance_m
                step_dist_m = gs_mps * dt_s
                if step_dist_m <= 1e-9 and remain_m > 1e-3:
                    phase_status = "Invalid: zero groundspeed with distance phase"
                    break
                if step_dist_m > remain_m:
                    frac = max(min(remain_m / max(step_dist_m, 1e-9), 1.0), 0.0)
                    dt_s *= frac
                    if dt_s <= 1e-9:
                        break
                    v_next = current_speed_mps + (v_next - current_speed_mps) * frac
                    accel_mps2 = (v_next - current_speed_mps) / dt_s
                    v_eval = 0.5 * (current_speed_mps + v_next)
                    m = compute_operating_metrics(
                        config,
                        speed_mps=v_eval,
                        orientation=orientation,
                        wind_mps=wind_mps,
                        wind_direction_deg=float(mission.wind_direction_deg),
                        course_deg=float(phase.course_deg),
                        ambient_temp_C=ambient_c,
                        soc=soc_state,
                    )
                    kinetic_power_w = kinetic_power_term_W(
                        config.drone_weight_g,
                        current_speed_mps,
                        v_next,
                        dt_s,
                        regen_eff,
                    )
                    base_total_power_w = float(m.get("total_power_W", 0.0))
                    total_power_w = max(base_total_power_w + potential_power_w + kinetic_power_w, 0.0)
                    scale = (total_power_w / base_total_power_w) if base_total_power_w > 1e-9 else 1.0
                    if base_total_power_w > 1e-9:
                        m["motor_power_W"] = float(m.get("motor_power_W", 0.0)) * scale
                        m["esc_loss_W"] = float(m.get("esc_loss_W", 0.0)) * scale
                        m["motor_I_per_esc_A"] = float(m.get("motor_I_per_esc_A", 0.0)) * max(scale, 0.0)
                    m["climb_rate_cmd_mps"] = climb_cmd
                    m["descent_rate_cmd_mps"] = descent_cmd
                    m["potential_power_W"] = potential_power_w
                    m["kinetic_power_W"] = kinetic_power_w
                    m["climb_power_add_W"] = potential_power_w + kinetic_power_w
                    m["total_power_W"] = total_power_w
                    m["pack_current_A"] = total_power_w / max(float(config.battery.vnom_pack), 1.0)
                    v_solve, i_solve = solve_pack_voltage_and_current(
                        config.battery, total_power_w, soc=soc_state
                    )
                    m["v_load_V"] = float(v_solve)
                    m["pack_current_A"] = float(i_solve)
                    gs_mps = groundspeed_along_track_mps(v_eval, headwind_mps, crosswind_mps)
                    m["airspeed_mps"] = v_eval
                    m["commanded_airspeed_mps"] = target_speed_mps
                    m["accel_mps2"] = accel_mps2
                    m["groundspeed_mps"] = gs_mps
                    m["wind_head_mps"] = headwind_mps
                    m["wind_cross_mps"] = crosswind_mps
                    m["reserve_margin_Wh"] = remaining_wh - reserve_target_wh
                    m["soc_model_source"] = str(getattr(config.battery, "soc_model_source", "linear-fallback"))

            pack_current_A = float(m.get("pack_current_A", 0.0))
            v_load = float(m.get("v_load_V", 0.0))
            motor_I_esc_A = float(m.get("motor_I_per_esc_A", 0.0))
            esc_note = str(m.get("esc_note", "")).strip()
            if pack_current_A > config.battery.discharge_max_A:
                phase_status = "Battery depleted (discharge limit exceeded)"
                break
            if v_load < config.battery.vmin_pack:
                phase_status = "Battery depleted (voltage under load)"
                break
            if getattr(config, "esc", None) is not None and motor_I_esc_A > config.esc.max_current_A:
                phase_status = f"ESC over max current: {motor_I_esc_A:.1f}A > {config.esc.max_current_A:.1f}A"
                break

            energy_used_Wh = total_power_w * (dt_s / 3600.0)
            if energy_used_Wh > remaining_wh and total_power_w > 1e-9:
                dt_s = (remaining_wh * 3600.0) / total_power_w
                energy_used_Wh = remaining_wh
                v_next = current_speed_mps + (v_next - current_speed_mps) * min(max(dt_s / max(base_dt_s, 1e-9), 0.0), 1.0)
                gs_mps = groundspeed_along_track_mps(0.5 * (current_speed_mps + v_next), headwind_mps, crosswind_mps)
                phase_status = "Battery depleted"
                depleted = True

            remaining_wh = max(remaining_wh - energy_used_Wh, 0.0)
            soc_state = battery_soc_after_energy_draw(config.battery, soc_state, energy_used_Wh)
            m["soc"] = soc_state
            m["soc_percent"] = soc_state * 100.0
            reserve_breached = reserve_breached or (remaining_wh < reserve_target_wh)
            reserve_min_wh = min(reserve_min_wh, remaining_wh)

            step_distance_m = gs_mps * dt_s
            phase_distance_m += step_distance_m
            phase_elapsed_s += dt_s
            t_s += dt_s
            dist_km += step_distance_m / 1000.0

            motor_copper_total_w = (motor_I_esc_A ** 2) * float(getattr(config.motor, "resistance", 0.0)) * max(int(config.num_motors), 1)
            battery_loss_w = (pack_current_A ** 2) * max(float(getattr(config.battery, "pack_resistance", 0.0)), 0.0)
            m["motor_copper_loss_W"] = motor_copper_total_w
            m["battery_loss_W"] = battery_loss_w

            motor_temp_c = thermal_step(motor_temp_c, ambient_c, m.get("motor_copper_loss_W", 0.0), 0.35, 240.0, dt_s)
            esc_temp_c = thermal_step(esc_temp_c, ambient_c, m.get("esc_loss_W", 0.0), 0.75, 180.0, dt_s)
            battery_temp_c = thermal_step(battery_temp_c, ambient_c, m.get("battery_loss_W", 0.0), 0.25, 500.0, dt_s)
            mt = max(motor_temp_c, esc_temp_c, battery_temp_c)
            m["motor_temp_est_C"] = motor_temp_c
            m["esc_temp_est_C"] = esc_temp_c
            m["battery_temp_est_C"] = battery_temp_c
            m["thermal_status"] = "OK" if mt < 95.0 else ("WARN" if mt < 115.0 else "HOT")
            m["reserve_margin_Wh"] = remaining_wh - reserve_target_wh

            worst_metrics = _merge_worst(worst_metrics, m)
            _append_point(
                phase_name=phase.name,
                phase_alt_m=float(phase.altitude),
                segment_type=segment_type,
                m=m,
                t_s_now=t_s,
                dist_km_now=dist_km,
                remaining_wh_now=remaining_wh,
                reserve_hit=reserve_breached,
            )

            current_speed_mps = v_next
            if depleted:
                break

        results.append((
            phase.name,
            phase_elapsed_s / 60.0,
            phase_distance_m / 1000.0,
            phase_status if phase_status else "OK"
        ))
        prev_alt_m = float(phase.altitude)
        if depleted or (phase_status and phase_status != "OK"):
            break

    if worst_metrics is not None:
        worst_metrics["reserve_target_Wh"] = reserve_target_wh
        worst_metrics["reserve_min_Wh"] = reserve_min_wh
        worst_metrics["reserve_margin_Wh"] = reserve_min_wh - reserve_target_wh
        worst_metrics["reserve_breached"] = bool(reserve_breached)
        worst_metrics["soc"] = soc_state
        worst_metrics["soc_percent"] = soc_state * 100.0
        worst_metrics["soc_model_source"] = str(getattr(config.battery, "soc_model_source", "linear-fallback"))
        worst_metrics["motor_temp_est_C"] = motor_temp_c
        worst_metrics["esc_temp_est_C"] = esc_temp_c
        worst_metrics["battery_temp_est_C"] = battery_temp_c
        mt = max(motor_temp_c, esc_temp_c, battery_temp_c)
        worst_metrics["thermal_status"] = "OK" if mt < 95.0 else ("WARN" if mt < 115.0 else "HOT")
        worst_metrics["segment_type"] = "worst-case"

    return results, worst_metrics, mission_series



# -------------------------------
# Plotting
# -------------------------------
def make_performance_figure(config: DroneConfig,
                             max_speed: float = 30.0,
                             figsize: tuple = (15, 9)):
    """
    Four-panel multicopter performance figure:
      1. Flight Time & Range vs Speed  (forward)
      2. Power Required vs Speed       (hover + forward)
      3. Thrust Required vs Speed      (hover + forward)
      4. Power Breakdown bar chart     (at cruise speed)
    """
    v_lo  = 0.5
    steps = 200
    speeds = [v_lo + (max_speed - v_lo) * i / steps for i in range(steps + 1)]

    times_fwd, dists_fwd = [], []
    pwr_fwd, pwr_hov     = [], []
    thr_fwd, thr_hov     = [], []
    cruise = float(config.cruise_speed)

    for V in speeds:
        times_fwd.append(estimate_flight_time_minutes(config, V, orientation="forward"))
        dists_fwd.append(estimate_flight_distance_km(config, V, orientation="forward"))
        pwr_fwd.append(power_required(config, V, "forward"))
        pwr_hov.append(power_required(config, V, "hover"))
        thr_fwd.append(thrust_required(config, V, "forward"))
        thr_hov.append(thrust_required(config, V, "hover"))

    be_v, be_t, br_v, br_d = find_optimal_speeds(config, min_speed=v_lo, max_speed=max_speed)

    fig, axes = plt.subplots(2, 2, figsize=figsize)
    fig.suptitle("Multicopter Performance", fontsize=13, fontweight="bold")

    # 1. Time & Range
    ax = axes[0, 0]
    ax2 = ax.twinx()
    l1, = ax.plot(speeds, times_fwd, color="royalblue",  label="Flight Time (min)")
    l2, = ax2.plot(speeds, dists_fwd, color="darkorange", label="Range (km)", linestyle="--")
    ax.axvline(be_v,  color="royalblue",  linestyle=":", linewidth=1.2)
    ax.axvline(br_v,  color="darkorange", linestyle=":", linewidth=1.2)
    ax.axvline(cruise,color="gray",       linestyle="-.", linewidth=1.0, alpha=0.7)
    ax.set_xlabel("Speed (m/s)"); ax.set_ylabel("Time (min)"); ax2.set_ylabel("Range (km)")
    ax.set_title("Flight Time & Range vs Speed")
    ax.legend(handles=[l1, l2], loc="upper right", fontsize=8)
    ax.grid(True, alpha=0.4)

    # 2. Power
    ax = axes[0, 1]
    ax.plot(speeds, [p/1000 for p in pwr_fwd], color="crimson",   label="Forward Flight")
    ax.plot(speeds, [p/1000 for p in pwr_hov], color="steelblue", label="Hover", linestyle="--")
    ax.axvline(cruise, color="gray", linestyle="-.", linewidth=1.0, alpha=0.7)
    ax.set_xlabel("Speed (m/s)"); ax.set_ylabel("Power (kW)")
    ax.set_title("Power Required vs Speed")
    ax.legend(fontsize=8); ax.grid(True, alpha=0.4)

    # 3. Thrust
    ax = axes[1, 0]
    ax.plot(speeds, thr_fwd, color="teal",      label="Forward Flight")
    ax.plot(speeds, thr_hov, color="slategray", label="Hover", linestyle="--")
    ax.axvline(cruise, color="gray", linestyle="-.", linewidth=1.0, alpha=0.7)
    ax.set_xlabel("Speed (m/s)"); ax.set_ylabel("Thrust (N)")
    ax.set_title("Thrust Required vs Speed")
    ax.legend(fontsize=8); ax.grid(True, alpha=0.4)

    # 4. Power breakdown at cruise
    ax = axes[1, 1]
    try:
        motor_P  = power_required(config, cruise, "forward")
        periph_P = avionics_input_power_W(getattr(config, "avionics", None))
        if periph_P <= 0.0:
            periph_P = config.battery.vnom_pack * max(config.periph_current, 0.0)
        total_P, _, _, _, _ = total_power_with_esc(
            config, motor_power_W=motor_P, periph_power_W=periph_P)
        esc_P = max(0.0, total_P - motor_P - periph_P)
        labels = ["Motor", "ESC Loss", "Avionics"]
        values = [motor_P, esc_P, periph_P]
        colors = ["#4c72b0", "#dd8452", "#55a868"]
        valid  = [(l, v, c) for l, v, c in zip(labels, values, colors) if v > 0.1]
        if valid:
            ls, vs, cs = zip(*valid)
            bars = ax.bar(ls, [v/1000 for v in vs], color=cs, edgecolor="white", linewidth=0.8)
            ax.bar_label(bars, fmt=lambda x: f"{x*1000:.0f} W", padding=3, fontsize=8)
    except Exception:
        pass
    ax.set_ylabel("Power (kW)"); ax.set_title(f"Power Breakdown @ {cruise:.1f} m/s")
    ax.grid(True, axis="y", alpha=0.4)

    fig.tight_layout(rect=[0, 0, 1, 0.95])
    return fig


def plot_performance(config: DroneConfig, max_speed: float = 30.0):
    fig = make_performance_figure(config, max_speed=max_speed)
    plt.show()


def make_airframe_diagram_figure(config: DroneConfig, figsize: tuple = (9, 8)):
    """
    Plan-view stick diagram of the airframe, with dimensions labelled.

    Layout rules:
      * The body is an equilateral polygon with one vertex per motor.
      * Each arm runs radially outward from a vertex.
      * A propeller disc is drawn at every rotor position, to scale.

    A coaxial X8 has two motors per arm, so it is drawn with N/2 arm positions
    and a second dashed disc at each — otherwise eight separate arms would be
    shown for a four-arm aircraft.

    The point of the drawing is to make propeller overlap obvious. Discs that
    intersect are drawn in red and the tip gap is reported as a negative
    number.
    """
    import matplotlib.pyplot as _plt
    from matplotlib.patches import Circle as _Circle, Polygon as _Polygon

    n_motors = max(int(config.num_motors), 1)
    coaxial = str(getattr(config, "motor_configuration", "flat")).lower() == "coaxial"
    n_arms = max(n_motors // 2, 1) if coaxial else n_motors

    prop_d = float(config.propeller.diameter_in) * 0.0254

    # Body size: use the entered plan dimensions when available, else fall
    # back to something proportionate so the sketch is still readable.
    body_l = float(getattr(config, "body_length_m", 0.0) or 0.0)
    body_w = float(getattr(config, "body_width_m", 0.0) or 0.0)
    if body_l > 0 or body_w > 0:
        body_r = max(body_l, body_w) / 2.0
        body_known = True
    else:
        body_r = prop_d * 0.45
        body_known = False

    arm_len = float(getattr(config, "arm_length_m", 0.0) or 0.0)
    if arm_len <= 0:
        arm_len = prop_d * 0.75          # enough to clear the body if unstated
        arm_known = False
    else:
        arm_known = True

    arm_w = float(getattr(config, "arm_width_m", 0.0) or 0.0) or prop_d * 0.03

    layout = core.rotor_ring_layout(
        num_positions=n_arms, body_circumradius_m=body_r,
        arm_length_m=arm_len, prop_diameter_m=prop_d,
        rotation_rad=math.pi / n_arms if n_arms % 2 == 0 else math.pi / 2.0)

    fig, ax = _plt.subplots(figsize=figsize)
    overlap = layout["overlaps"]
    disc_colour = "#C62828" if overlap else "#2E7D32"

    # Body polygon
    ax.add_patch(_Polygon(layout["vertices"], closed=True, fill=True,
                          facecolor="#CFD8DC", edgecolor="#37474F",
                          linewidth=1.8, zorder=2))

    # Arms and rotors
    for (vx, vy), (mx, my) in zip(layout["vertices"], layout["rotors"]):
        ax.plot([vx, mx], [vy, my], color="#37474F",
                linewidth=max(arm_w * 400.0, 2.0), solid_capstyle="round", zorder=1)
        ax.add_patch(_Circle((mx, my), layout["prop_radius_m"], fill=False,
                             edgecolor=disc_colour, linewidth=1.6, zorder=3))
        if coaxial:
            ax.add_patch(_Circle((mx, my), layout["prop_radius_m"] * 0.97,
                                 fill=False, edgecolor=disc_colour,
                                 linewidth=1.0, linestyle="--", zorder=3))
        ax.plot([mx], [my], marker="o", markersize=5,
                color="#37474F", zorder=4)

    # Dimension annotations
    r_rot = layout["rotor_radius_m"]
    ax.annotate("", xy=layout["rotors"][0], xytext=(0, 0),
                arrowprops=dict(arrowstyle="<->", color="#1565C0", lw=1.2))
    ax.text(layout["rotors"][0][0] * 0.5, layout["rotors"][0][1] * 0.5,
            f"  centre to rotor\n  {r_rot * 1000:.0f} mm",
            color="#1565C0", fontsize=8, va="bottom")

    if n_arms >= 2:
        a, b = layout["rotors"][0], layout["rotors"][1]
        ax.annotate("", xy=a, xytext=b,
                    arrowprops=dict(arrowstyle="<->", color="#6A1B9A", lw=1.2))
        ax.text((a[0] + b[0]) / 2, (a[1] + b[1]) / 2,
                f"  {layout['motor_spacing_m'] * 1000:.0f} mm\n  motor pitch",
                color="#6A1B9A", fontsize=8, ha="center")

    ax.set_aspect("equal", adjustable="box")
    limit = (r_rot + layout["prop_radius_m"]) * 1.25
    ax.set_xlim(-limit, limit)
    ax.set_ylim(-limit, limit)
    ax.grid(True, linestyle=":", alpha=0.4)
    ax.set_xlabel("metres")
    ax.set_ylabel("metres")

    gap_mm = layout["adjacent_gap_m"] * 1000.0
    verdict = (f"OVERLAP {abs(gap_mm):.0f} mm" if overlap
               else f"tip gap {gap_mm:.0f} mm")
    title = (f"{n_motors} motors"
             + (f" on {n_arms} coaxial arms" if coaxial else "")
             + f"  |  {config.propeller.diameter_in:g} in props  |  {verdict}")
    ax.set_title(title, fontsize=11,
                 color="#C62828" if overlap else "#000000")

    notes = [
        f"Prop diameter      {prop_d * 1000:.0f} mm",
        f"Body across flats  {body_r * 2000:.0f} mm" + ("" if body_known else "  (assumed)"),
        f"Arm length         {arm_len * 1000:.0f} mm" + ("" if arm_known else "  (assumed)"),
        f"Motor pitch        {layout['motor_spacing_m'] * 1000:.0f} mm",
        f"Tip-to-tip gap     {gap_mm:+.0f} mm",
        f"Overall span       {layout['span_m'] * 1000:.0f} mm",
    ]
    if not (body_known and arm_known):
        notes.append("")
        notes.append("Enter Body and Arm dimensions on the")
        notes.append("Airframe tab for a to-scale drawing.")
    ax.text(0.02, 0.02, "\n".join(notes), transform=ax.transAxes,
            fontsize=8, family="monospace", va="bottom",
            bbox=dict(boxstyle="round", facecolor="#FFFDE7", edgecolor="#BDBDBD"))

    fig.tight_layout()
    return fig


def make_motor_operating_point_figure(config: DroneConfig, metrics: dict, figsize: tuple = (12, 8)):
    """
    Create a figure showing motor/propeller operating curves with the current operating point marked.
    Uses propeller table data if available.
    
    Two subplots:
      1. Thrust vs Power (left), Thrust vs Current (right) on same plot
      2. Thrust vs Efficiency g/W (left), Thrust vs RPM (right) on same plot
    """
    if config.propeller.table is None:
        # Return empty figure if no propeller table
        fig, ax = plt.subplots(1, 1, figsize=figsize)
        ax.text(0.5, 0.5, "Propeller table not available\nCannot plot operating curves",
                ha="center", va="center", transform=ax.transAxes, fontsize=12)
        ax.axis("off")
        return fig
    
    df = config.propeller.table
    thrust_pm_N = float(metrics.get("thrust_per_motor_N", 0.0))
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=figsize)
    
    # Get data from propeller table
    thrust_g = df["Thrust_g"].values if "Thrust_g" in df.columns else []
    
    # Subplot 1: Thrust vs Power & Thrust vs Current
    ax1_1 = ax1
    ax1_2 = None
    if "Power_W" in df.columns and len(thrust_g) > 0:
        power_W = df["Power_W"].values
        ax1_1.plot(thrust_g, power_W, "b-", linewidth=2, label="Power")
        ax1_1.set_xlabel("Thrust per motor (gf)", fontsize=10)
        ax1_1.set_ylabel("Power (W)", fontsize=10, color="b")
        ax1_1.tick_params(axis="y", labelcolor="b")
        ax1_1.grid(True, alpha=0.3)
    
    if "Current_A" in df.columns and len(thrust_g) > 0:
        current_A = df["Current_A"].values
        ax1_2 = ax1.twinx()
        ax1_2.plot(thrust_g, current_A, "r--", linewidth=2, label="Current")
        ax1_2.set_ylabel("Current (A)", fontsize=10, color="r")
        ax1_2.tick_params(axis="y", labelcolor="r")
    
    # Mark operating point on subplot 1 (both y-axes)
    thrust_g_op = thrust_pm_N * 1000.0 / 9.81
    point = interpolate_motor_point(config, thrust_pm_N)
    
    if "Power_W" in df.columns and len(thrust_g) > 0:
        power_op = point.get("Power_W", 0.0)
        ax1_1.plot(thrust_g_op, power_op, "b*", markersize=15, label=f"Pow: {power_op:.1f}W", markeredgewidth=0.5, markeredgecolor="darkblue")
    
    if "Current_A" in df.columns and len(thrust_g) > 0 and ax1_2:
        current_op = point.get("Current_A", 0.0)
        ax1_2.plot(thrust_g_op, current_op, "r*", markersize=15, label=f"Cur: {current_op:.2f}A", markeredgewidth=0.5, markeredgecolor="darkred")
    
    ax1.set_title("Thrust vs Power & Current", fontsize=11, fontweight="bold")
    # Combine legends from both axes
    lines1, labels1 = ax1_1.get_legend_handles_labels()
    lines2, labels2 = (ax1_2.get_legend_handles_labels() if ax1_2 else ([], []))
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left", fontsize=9)
    
    # Subplot 2: Thrust vs Efficiency & Thrust vs RPM
    ax2_1 = ax2
    ax2_2 = None
    if "Efficiency_gW" in df.columns and len(thrust_g) > 0:
        eff_gW = df["Efficiency_gW"].values
        ax2_1.plot(thrust_g, eff_gW, "g-", linewidth=2, label="Efficiency")
        ax2_1.set_xlabel("Thrust per motor (gf)", fontsize=10)
        ax2_1.set_ylabel("Efficiency (g/W)", fontsize=10, color="g")
        ax2_1.tick_params(axis="y", labelcolor="g")
        ax2_1.grid(True, alpha=0.3)
    
    if "RPM" in df.columns and len(thrust_g) > 0:
        rpm = df["RPM"].values
        ax2_2 = ax2.twinx()
        ax2_2.plot(thrust_g, rpm, "m--", linewidth=2, label="RPM")
        ax2_2.set_ylabel("RPM", fontsize=10, color="m")
        ax2_2.tick_params(axis="y", labelcolor="m")
    
    # Mark operating point on subplot 2 (both y-axes)
    if "Efficiency_gW" in df.columns and len(thrust_g) > 0:
        eff_op = point.get("Efficiency_gW", 0.0)
        ax2_1.plot(thrust_g_op, eff_op, "g*", markersize=15, label=f"Eff: {eff_op:.2f}g/W", markeredgewidth=0.5, markeredgecolor="darkgreen")
    
    if "RPM" in df.columns and len(thrust_g) > 0 and ax2_2:
        rpm_op = point.get("RPM", 0.0)
        ax2_2.plot(thrust_g_op, rpm_op, "m*", markersize=15, label=f"RPM: {rpm_op:.0f}", markeredgewidth=0.5, markeredgecolor="darkmagenta")
    
    ax2.set_title("Thrust vs Efficiency & RPM", fontsize=11, fontweight="bold")
    # Combine legends from both axes
    lines1, labels1 = ax2_1.get_legend_handles_labels()
    lines2, labels2 = (ax2_2.get_legend_handles_labels() if ax2_2 else ([], []))
    ax2.legend(lines1 + lines2, labels1 + labels2, loc="upper left", fontsize=9)
    
    # Say plainly when the operating point sits outside the measured data.
    # Anything outside the table is an extrapolation, and a reader deserves to
    # know that before trusting the marker's position.
    _t_lo = float(df["Thrust_g"].min()) if "Thrust_g" in df else None
    _t_hi = float(df["Thrust_g"].max()) if "Thrust_g" in df else None
    _outside = (_t_lo is not None
                and (thrust_g_op < _t_lo or thrust_g_op > _t_hi))
    _range_note = ""
    if _outside:
        _where = "below" if thrust_g_op < _t_lo else "above"
        _range_note = (f"   —  EXTRAPOLATED, {_where} the measured "
                       f"{_t_lo:.0f}-{_t_hi:.0f} g range")
    fig.suptitle(f"Motor/Propeller Operating Curves (Thrust/Motor: {thrust_g_op:.0f}g)"
                 + _range_note,
                 color=("#C62828" if _outside else "black"), 
                 fontsize=12, fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    return fig


# -------------------------------
# Config builders
# -------------------------------
def build_drone_from_args(args) -> DroneConfig:
    inflow_mu_bp = parse_float_list(getattr(args, "inflow_mu_bp", None))
    inflow_eff_bp = parse_float_list(getattr(args, "inflow_eff_bp", None))
    soc_bp = parse_float_list(getattr(args, "battery_soc_bp", None))
    ocv_cell_bp = parse_float_list(getattr(args, "battery_ocv_cell_bp", None))
    r_scale_bp = parse_float_list(getattr(args, "battery_r_scale_bp", None))
    inflow_map_raw = parse_optional_bool(getattr(args, "inflow_map_enabled", None))
    inflow_map_enabled = True if inflow_map_raw is None else bool(inflow_map_raw)
    drag_model_raw = str(getattr(args, "drag_model_mode", "auto")).strip().lower()
    if drag_model_raw not in ("auto", "manual", "geometry"):
        drag_model_raw = "auto"
    battery = BatteryConfig(
        operating_voltage_min=args.battery_operating_voltage_min,
        operating_voltage_nominal=args.battery_operating_voltage_nominal,
        operating_voltage_max=args.battery_operating_voltage_max,
        unit_energy_density=args.battery_energy_density,
        chemistry=args.battery_chemistry,
        charge_current_max=args.battery_charge_current_max,
        discharge_cont_A=args.battery_discharge_cont_A,
        discharge_max_A=args.battery_discharge_max_A,
        discharge_c_cont=args.battery_discharge_c_cont,
        discharge_c_max=args.battery_discharge_c_max,
        discharge_percent=args.battery_discharge_percent,
        resistance_cell_mOhm=args.battery_resistance_cell,
        unit_mode=args.battery_unit_mode,
        series_units=args.battery_series_units,
        parallel_units=args.battery_parallel_units,
        cells_series_per_unit=args.battery_cells_series_per_unit,
        cells_parallel_per_unit=args.battery_cells_parallel_per_unit,
        pack_weight_g=args.battery_pack_weight_g,
        cell_weight_g=args.battery_cell_weight_g,
        cell_capacity_mAh=args.battery_cell_capacity,
        pack_capacity_mAh=args.battery_pack_capacity,
        soc_model=getattr(args, "battery_soc_model", "auto"),
        soc_curve_csv=getattr(args, "battery_soc_curve_csv", None),
        soc_bp=soc_bp,
        ocv_cell_bp=ocv_cell_bp,
        r_scale_bp=r_scale_bp,
    )

    # Reference-only motor fields fall back to safe defaults when omitted, so
    # a CLI/batch run does not need every datasheet number just to get an answer.
    motor = MotorConfig(
        kv=args.motor_kv,
        idle_current=args.motor_idle_current if args.motor_idle_current is not None else 0.0,
        idle_voltage=args.motor_idle_voltage if args.motor_idle_voltage is not None else 10.0,
        rated_voltage=args.motor_rated_voltage if args.motor_rated_voltage is not None else 0,
        resistance=args.motor_resistance,
        max_current=args.motor_max_current,      # None = no limit checked
        max_power=args.motor_max_power,          # None = no limit checked
        pole_count=args.motor_pole_count if args.motor_pole_count is not None else 14,
        weight_g=args.motor_weight,
        size_mm=args.motor_size,
    )

    prop = PropellerConfig(
        diameter_in=args.prop_diameter,
        pitch_in=args.prop_pitch,
        max_rpm=getattr(args, "prop_max_rpm", 0) or 0,
        max_thrust_g=getattr(args, "prop_max_thrust", 0) or 0,
        blades=args.prop_blades if args.prop_blades is not None else 2,
        table_csv=args.prop_table,
        PConst=args.prop_pconst,
        TConst=args.prop_tconst,
        weight_g=getattr(args, "prop_weight", None),
    )

    avionics = AvionicsConfig(
        voltage_tree=parse_voltage_tree(args.avionics_voltage_tree),
    )

    esc = None
    if any(x is not None for x in [args.esc_voltage_rating, args.esc_cont_current, args.esc_max_current, args.esc_idle_current, args.esc_resistance, args.esc_weight]):
        # Provide safe defaults if some fields omitted
        esc = ESCConfig(
            voltage_rating=int(args.esc_voltage_rating) if args.esc_voltage_rating is not None else int(args.battery_series_units or 1),
            continuous_current_A=float(args.esc_cont_current) if args.esc_cont_current is not None else 0.0,
            max_current_A=float(args.esc_max_current) if args.esc_max_current is not None else float(args.esc_cont_current or 0.0),
            idle_current_A=float(args.esc_idle_current) if args.esc_idle_current is not None else 0.0,
            resistance=float(args.esc_resistance) if args.esc_resistance is not None else 0.0,
            weight_g=float(args.esc_weight) if args.esc_weight is not None else None,
        )

    base_weight_g = float(args.weight)
    payload_mass_g = max(float(getattr(args, "payload_mass_g", 0.0) or 0.0), 0.0)
    drone = DroneConfig(
        num_motors=args.num_motors,
        battery=battery,
        motor=motor,
        propeller=prop,
        drone_weight_g=base_weight_g + payload_mass_g,
        profile_drag_coefficient=args.profile_drag,
        profile_area=args.profile_area,
        parasite_drag_coefficient=args.parasite_drag,
        parasite_area=args.parasite_area,
        frontal_area=args.area,
        cruise_speed=args.speed,
        periph_current=args.periph_current,
        esc=esc,
        avionics=avionics,
        air_density=AIR_DENSITY,
        body_length_m=args.body_length_m,
        body_width_m=args.body_width_m,
        body_height_m=args.body_height_m,
        arm_length_m=args.arm_length_m,
        arm_width_m=args.arm_width_m,
        coaxial_spacing_m=args.coaxial_spacing_m,
        max_tilt_deg=args.max_tilt_deg,
        motor_configuration=args.motor_configuration,
        drag_model=drag_model_raw,
        transient_dt_s=float(getattr(args, "transient_dt_s", 0.5)),
        max_accel_mps2=float(getattr(args, "max_accel_mps2", 2.0)),
        max_decel_mps2=float(getattr(args, "max_decel_mps2", 2.5)),
        decel_regen_eff=float(getattr(args, "decel_regen_eff", 0.0)),
        inflow_map_enabled=inflow_map_enabled,
        inflow_mu_bp=inflow_mu_bp,
        inflow_eff_bp=inflow_eff_bp,
    )
    drone.payload_mass_g = payload_mass_g

    # Initialize air density at user-specified conditions
    drone.air_density = compute_air_density(
        altitude_m=args.altitude,
        temperature_C=args.temperature,
        pressure_Pa=args.pressure,
    )
    return drone


# -------------------------------
# GUI
# -------------------------------

# -------------------------------
# GUI
# -------------------------------
# ============================================================
# SHARED REPORTING / EXPORT UTILITIES
# ============================================================
import csv as _csv
import io  as _io
import datetime as _dt
import tempfile as _tmpfile

def _fmt_g(x, nd=2):
    try:    return f"{float(x):.{nd}f}"
    except: return "n/a"

def _extract_weight_budget(cfg) -> list:
    rows = []
    total_g = float(getattr(cfg, "drone_weight_g",
                    getattr(cfg, "aircraft_weight_g", 0.0)))
    payload_g = max(float(getattr(cfg, "payload_mass_g", 0.0) or 0.0), 0.0)
    num_motors = int(getattr(cfg, "num_motors",
                    getattr(getattr(cfg, "airframe", None), "num_motors", 1)))
    batt  = getattr(cfg, "battery",   None)
    motor = getattr(cfg, "motor",     None)
    esc   = getattr(cfg, "esc",       None)
    prop  = getattr(cfg, "propeller", None)
    accounted = 0.0
    # Battery: weight_g already includes all series/parallel units
    if batt:
        w = float(getattr(batt, "weight_g", 0.0) or 0.0)
        rows.append(("Battery", w, 1, w)); accounted += w
    # Motor: per-motor weight, multiply by num_motors
    if motor:
        w = float(getattr(motor, "weight_g", 0.0) or 0.0)
        rows.append(("Motor", w, num_motors, w * num_motors)); accounted += w * num_motors
    # ESC: per-ESC weight, multiply by num_motors
    if esc:
        w = float(getattr(esc, "weight_g", 0.0) or 0.0)
        rows.append(("ESC", w, num_motors, w * num_motors)); accounted += w * num_motors
    # Propeller: per-propeller weight, multiply by num_motors
    if prop:
        w = float(getattr(prop, "weight_g", 0.0) or 0.0)
        rows.append(("Propeller", w, num_motors, w * num_motors)); accounted += w * num_motors
    if payload_g > 0:
        rows.append(("Payload", payload_g, 1, payload_g)); accounted += payload_g
    airframe_g = max(0.0, total_g - accounted)
    rows.append(("Airframe / Structure", airframe_g, 1, airframe_g))
    rows.append(("TOTAL", total_g, 1, total_g))
    return rows

def _export_csv_file(path: str, sweep: dict, metrics: list) -> None:
    import csv
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["[Performance Sweep]"])
        if sweep:
            headers = list(sweep.keys())
            w.writerow(headers)
            n = max(len(v) for v in sweep.values())
            for i in range(n):
                w.writerow([sweep[h][i] if i < len(sweep[h]) else "" for h in headers])
        w.writerow([])
        w.writerow(["[Metrics]"])
        w.writerow(["Metric", "Value"])
        for label, value in metrics:
            w.writerow([label, value])

def _export_excel_file(path: str, sweep: dict, metrics: list, weight_budget: list) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active; ws.title = "Performance Sweep"
    if sweep:
        headers = list(sweep.keys())
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
        n = max(len(v) for v in sweep.values())
        for ri in range(n):
            for ci, h in enumerate(headers, 1):
                ws.cell(row=ri+2, column=ci, value=sweep[h][ri] if ri < len(sweep[h]) else None)
    ws2 = wb.create_sheet("Metrics")
    ws2.append(["Metric", "Value"])
    for r in [ws2["A1"], ws2["B1"]]: r.font = Font(bold=True)
    for label, value in metrics: ws2.append([label, value])
    ws3 = wb.create_sheet("Weight Budget")
    ws3.append(["Component", "Unit Weight (g)", "Count", "Total Weight (g)", "% of Total"])
    for c in ws3[1]: c.font = Font(bold=True)
    total_g = weight_budget[-1][3] if weight_budget else 1.0
    for label, uw, cnt, tw in weight_budget[:-1]:
        pct = round(tw/total_g*100, 1) if total_g > 0 else 0
        ws3.append([label, round(uw,1), cnt, round(tw,1), pct])
    if weight_budget:
        label, uw, cnt, tw = weight_budget[-1]
        ws3.append([label, "", "", round(tw,1), 100.0])
        for c in list(ws3.rows)[-1]: c.font = Font(bold=True)
    wb.save(path)

def _generate_pdf_report(path: str, report_title: str,
                          inputs_rows: list, metrics_rows: list,
                          status_sections: list, log_text: str,
                          figures: list, weight_budget: list) -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, PageBreak, Image,
                                    HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units  import cm
    from reportlab.lib        import colors
    from reportlab.lib.enums  import TA_CENTER
    import io
    PAGE_W, PAGE_H = A4
    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2.0*cm,  bottomMargin=2.0*cm)
    styles = getSampleStyleSheet()
    story  = []
    NAVY  = colors.HexColor("#1F3864")
    TEAL  = colors.HexColor("#2E75B6")
    LGREY = colors.HexColor("#F2F2F2")
    DGREY = colors.HexColor("#595959")
    sTitle = ParagraphStyle("rTitle", fontSize=22, textColor=NAVY,
                             spaceAfter=6, alignment=TA_CENTER, fontName="Helvetica-Bold")
    sSub   = ParagraphStyle("rSub", fontSize=11, textColor=DGREY,
                             spaceAfter=20, alignment=TA_CENTER, fontName="Helvetica")
    sH1    = ParagraphStyle("rH1", fontSize=13, textColor=NAVY,
                             spaceBefore=12, spaceAfter=4, fontName="Helvetica-Bold")
    def _ts(header_bg=TEAL):
        return TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), header_bg),
            ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LGREY]),
            ("GRID",          (0,0),(-1,-1), 0.3, colors.lightgrey),
            ("LEFTPADDING",   (0,0),(-1,-1), 4),
            ("RIGHTPADDING",  (0,0),(-1,-1), 4),
            ("TOPPADDING",    (0,0),(-1,-1), 2),
            ("BOTTOMPADDING", (0,0),(-1,-1), 2),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ])
    usable_w = PAGE_W - 3.6*cm
    # Title page
    story.append(Spacer(1, 3*cm))
    story.append(Paragraph(report_title, sTitle))
    story.append(Paragraph(
        _dt.datetime.now().strftime("Generated %d %B %Y  %H:%M"), sSub))
    story.append(HRFlowable(width="100%", thickness=1.5, color=TEAL, spaceAfter=12))
    if weight_budget:
        story.append(Paragraph("Component Weight Summary", sH1))
        total_g = weight_budget[-1][3] if weight_budget else 1.0
        wb_data = [["Component", "Unit (g)", "Qty", "Total (g)", "%"]]
        for label, uw, cnt, tw in weight_budget[:-1]:
            pct = f"{tw/total_g*100:.1f}" if total_g > 0 else "0"
            wb_data.append([label, f"{uw:.1f}", str(cnt), f"{tw:.1f}", pct])
        last = weight_budget[-1]
        wb_data.append(["TOTAL", "", "", f"{last[3]:.1f}", "100.0"])
        ts = _ts()
        ts.add("FONTNAME",(0,len(wb_data)-1),(-1,len(wb_data)-1),"Helvetica-Bold")
        t = Table(wb_data, colWidths=[usable_w*0.38, usable_w*0.15,
                                       usable_w*0.1, usable_w*0.17, usable_w*0.1])
        t.setStyle(ts); story.append(t)
    story.append(PageBreak())
    # Inputs
    if inputs_rows:
        story.append(Paragraph("Design Inputs", sH1))
        mid = (len(inputs_rows)+1)//2
        left = inputs_rows[:mid]; right = inputs_rows[mid:]
        while len(right) < len(left): right.append(("",""))
        rows_data = [["Parameter","Value","Parameter","Value"]]
        for (la,va),(lb,vb) in zip(left,right): rows_data.append([la,va,lb,vb])
        col_w = usable_w/4
        t = Table(rows_data, colWidths=[col_w*1.4,col_w*0.6]*2)
        t.setStyle(_ts()); story.append(t)
    story.append(PageBreak())
    # Metrics
    if metrics_rows:
        story.append(Paragraph("Performance Metrics", sH1))
        mid = (len(metrics_rows)+1)//2
        left = metrics_rows[:mid]; right = metrics_rows[mid:]
        while len(right) < len(left): right.append(("",""))
        rows_data = [["Metric","Value","Metric","Value"]]
        for (la,va),(lb,vb) in zip(left,right): rows_data.append([la,va,lb,vb])
        col_w = usable_w/4
        t = Table(rows_data, colWidths=[col_w*1.4,col_w*0.6]*2)
        t.setStyle(_ts()); story.append(t)
    story.append(PageBreak())
    # Status
    if status_sections:
        story.append(Paragraph("Status Checks", sH1))
        for sec_title, sec_rows in status_sections:
            story.append(Paragraph(sec_title, ParagraphStyle("secH", fontSize=10,
                textColor=TEAL, spaceBefore=6, spaceAfter=2, fontName="Helvetica-Bold")))
            tdata = [["Metric","Value","Limit","Notes"]]
            ts = _ts()
            for ri,(metric,val,lim,note,tag) in enumerate(sec_rows,1):
                tdata.append([metric,val,lim,note])
                bg = {"ok":colors.HexColor("#D9F2D9"),"warn":colors.HexColor("#FFF2CC"),
                      "bad":colors.HexColor("#F8D7DA")}.get(tag, colors.white)
                ts.add("BACKGROUND",(0,ri),(-1,ri),bg)
            cw = [usable_w*0.28,usable_w*0.20,usable_w*0.20,usable_w*0.32]
            t = Table(tdata, colWidths=cw); t.setStyle(ts); story.append(t)
            story.append(Spacer(1,4))
    story.append(PageBreak())
    # Plots
    if figures:
        story.append(Paragraph("Performance Plots", sH1))
        for fig in figures:
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=120, bbox_inches="tight")
            buf.seek(0)
            img_w = usable_w
            img_h = img_w * (fig.get_figheight() / max(fig.get_figwidth(), 0.01))
            max_h = PAGE_H - 5*cm
            if img_h > max_h:
                img_h = max_h
                img_w = img_h * (fig.get_figwidth() / max(fig.get_figheight(), 0.01))
            story.append(Image(buf, width=img_w, height=img_h))
            story.append(Spacer(1, 8)); story.append(PageBreak())
    # Log
    if log_text.strip():
        story.append(Paragraph("Simulation Output Log", sH1))
        mono = ParagraphStyle("mono", fontName="Courier", fontSize=7.5, leading=10, spaceAfter=2)
        for line in log_text.splitlines():
            safe = line.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
            story.append(Paragraph(safe or " ", mono))
    doc.build(story)


# ============================================================
# INLINE HELP  —  tooltips and Simple/Advanced field visibility
# ============================================================

# ------------------------------------------------------------------
# FIELD HELP TEXT
# ------------------------------------------------------------------
# One entry per input.  Each value is:
#     (short_label_help, typical_range_or_where_to_find_it)
# Shown on hover over the "?" marker beside each field.
#
# Keep these in plain language — the audience is someone building their
# first drone, not someone who already knows what an Oswald factor is.
# ------------------------------------------------------------------
MC_FIELD_HELP = {
    # ---- Drone ----
    "num_motors": ("How many motors the aircraft has.",
                   "4 = quad, 6 = hex, 8 = octo. Coaxial X8 counts all 8."),
    "weight": ("All-up weight WITHOUT payload: frame, motors, ESCs, props, "
               "battery, flight controller, wiring — everything you always fly with.",
               "Weigh the assembled aircraft on a kitchen scale. 5in quad ~700 g, "
               "7in ~1200 g, 450-class ~1500 g."),
    "payload_mass_g": ("Extra mass carried on this particular flight — camera, "
                       "gimbal, sensor, cargo. Kept separate so you can sweep it.",
                       "0 if you are flying clean."),
    "area": ("Frontal cross-section area of the airframe, used only if you have "
             "no better drag data.",
             "Roughly body width x height. A 5in quad is about 0.01 m2."),
    "speed": ("The forward airspeed the single-point run is evaluated at.",
              "0 = hover. Typical cruise 8-15 m/s."),
    "periph_current": ("Steady current drawn from the main pack by anything that "
                       "is not a motor and not on a BEC rail.",
                       "Usually 0 if you have listed your rails on the Avionics tab."),
    "motor_configuration": ("flat = all motors on one plane (quad, hex, flat octo). "
                            "coaxial = stacked motor pairs (X8).",
                            "Coaxial pairs lose 10-20% efficiency from the lower "
                            "prop working in the upper prop's wash."),
    "max_tilt_deg": ("Maximum pitch angle the controller will command in forward "
                     "flight. Limits top speed and wind resistance.",
                     "Typical 25-45 deg. Cinematic rigs are often limited to 25-30."),
    "drag_model_mode": ("auto = estimate drag from the body dimensions below. "
                        "manual = use the Cd and area numbers you type in.",
                        "Use manual with numbers from the drag coefficient "
                        "calculator for a real measurement."),
    "profile_drag": ("Drag coefficient for SIDE-ON airflow (moving sideways "
                     "while hovering level).",
                     "Blunt multirotor bodies are about 1.0-1.3."),
    "profile_area": ("SIDE silhouette area, seen from the side of the aircraft.",
                     "From the drag calculator's Side View, or length x height."),
    "parasite_drag": ("Drag coefficient for FORWARD flight, nose-on.",
                      "Blunt multirotor bodies are about 1.0-1.3."),
    "parasite_area": ("FRONT silhouette area, seen from directly ahead.",
                      "From the drag calculator's Front View, or width x height."),
    "body_length_m": ("Length of the central body, front to back. Used by the "
                      "auto drag estimate.", "Excludes arms. 5in quad ~0.12 m."),
    "body_width_m": ("Width of the central body, side to side.",
                     "Excludes arms. 5in quad ~0.08 m."),
    "body_height_m": ("Height of the central body including the battery stack.",
                      "5in quad ~0.05 m."),
    "arm_length_m": ("Length of ONE arm, from body edge to motor centre.",
                     "Roughly half the motor-to-motor diagonal."),
    "arm_width_m": ("Width or diameter of one arm tube.",
                    "Typical 0.01-0.03 m."),
    "coaxial_spacing_m": ("Vertical gap between the upper and lower prop discs "
                          "in a coaxial pair.",
                          "Closer spacing = worse efficiency. Typical 0.05-0.12 m."),

    # ---- Battery ----
    "batt_unit_mode": ("cell = you specify one CELL and how many are wired "
                       "together. pack = you specify a whole PACK and how many "
                       "packs you wire together.",
                       "Use pack if you bought ready-made LiPos."),
    "batt_vmin": ("Lowest safe voltage PER CELL — the discharge cutoff.",
                  "LiPo 3.0-3.5, Li-ion 2.8-3.0, LiFePO4 2.5."),
    "batt_vnom": ("Nominal (average) voltage PER CELL. Used for energy in Wh.",
                  "LiPo 3.7, Li-ion 3.6, LiFePO4 3.2."),
    "batt_vmax": ("Fully charged voltage PER CELL.",
                  "LiPo 4.2, Li-ion 4.2, LiFePO4 3.65."),
    "batt_cell_capacity": ("Capacity of ONE cell in mAh. Used in cell mode.",
                           "Common 18650 = 3000-3500 mAh."),
    "batt_pack_capacity": ("Capacity of ONE pack in mAh, as printed on its label. "
                           "Used in pack mode.",
                           "Wiring packs in SERIES does not change this number."),
    "batt_series": ("How many cells (or packs) are wired in SERIES. This sets "
                    "voltage.", "6S LiPo = 6 x 4.2 = 25.2 V fully charged."),
    "batt_parallel": ("How many cells (or packs) are wired in PARALLEL. This sets "
                      "capacity.",
                      "2 packs in parallel = double the mAh, same voltage."),
    "batt_cells_series": ("Cells in series INSIDE one pack. Pack mode only.",
                          "A 6S LiPo has 6."),
    "batt_cells_parallel": ("Cells in parallel INSIDE one pack. Pack mode only.",
                            "Most hobby LiPos have 1."),
    "batt_cell_weight": ("Weight of ONE cell in grams. Cell mode only.",
                         "18650 ~45-50 g."),
    "batt_pack_weight": ("Weight of ONE pack in grams. Pack mode only.",
                         "From the label or a scale. 6S 5000 mAh ~750 g."),
    "batt_energy_density": ("Optional override for Wh/kg. Leave blank to let the "
                            "sim compute it from capacity and weight.",
                            "LiPo 130-200, Li-ion 200-260 Wh/kg. Above 300 means "
                            "one of your inputs is wrong."),
    "batt_dischg_pct": ("How much of the pack you actually use before landing.",
                        "80% is normal. Going to 100% shortens pack life badly."),
    "batt_r": ("Internal resistance of ONE cell, in milliohms. Drives voltage "
               "sag and heating.",
               "Fresh LiPo 2-5 mOhm/cell, aged 8-15. Higher = more sag."),
    "batt_a_cont": ("Continuous current the pack can safely deliver, in amps.",
                    "Leave blank and use C-rate instead if that is what the label gives."),
    "batt_a_max": ("Short burst current limit, in amps.", "Usually 2x the continuous."),
    "batt_c_cont": ("Continuous discharge C-rate from the label.",
                    "Amps = C-rate x capacity in Ah. A 5 Ah 30C pack = 150 A."),
    "batt_c_max": ("Burst C-rate from the label.", "Often marketing-inflated."),
    "batt_chg": ("Maximum charge current in amps. Not used in flight physics.",
                 "Typically 1C."),
    "batt_chem": ("Chemistry label, for your own reference.", "LiPo, Li-ion, LiFePO4."),
    "batt_soc_model": ("How pack voltage falls as it empties. auto uses a built-in "
                       "curve, linear uses a straight line.",
                       "Leave on auto unless you have measured your own curve."),
    "batt_soc_curve_csv": ("Optional CSV of your own measured discharge curve.",
                           "Columns: soc, ocv_cell, r_scale."),
    "batt_soc_bp": ("State-of-charge breakpoints, 0 to 1, for a custom curve.",
                    "Advanced. Leave blank for the built-in curve."),
    "batt_ocv_cell_bp": ("Open-circuit cell voltage at each breakpoint above.",
                         "Advanced. Leave blank."),
    "batt_r_scale_bp": ("Resistance multiplier at each breakpoint. Cells get more "
                        "resistive when nearly empty.",
                        "Advanced. Leave blank."),

    # ---- Motor ----
    "motor_kv": ("Motor RPM per volt with no load. The single most important "
                 "motor number.",
                 "Printed on the motor. 5in racing 1800-2800, 7in 1200-1600, "
                 "heavy-lift 200-500."),
    "motor_r": ("Winding resistance in OHMS (not milliohms). Drives copper loss "
                "and heat.",
                "Small motors 0.05-0.2, large 0.01-0.05. If a datasheet gives "
                "mOhm, divide by 1000."),
    "motor_i0": ("No-load current: what the motor draws spinning free with no "
                 "prop.", "Typical 0.3-1.5 A."),
    "motor_v0": ("Voltage at which the no-load current was measured.",
                 "From the datasheet, usually 10 V."),
    "motor_rated_v": ("Voltage the manufacturer rates the motor for.",
                      "Reference only."),
    "motor_imax": ("Maximum continuous current per motor, in amps.",
                   "From the datasheet. Used for the status check."),
    "motor_pmax": ("Maximum continuous electrical power per motor, in watts.",
                   "From the datasheet."),
    "motor_pole_count": ("Number of magnet poles. Only affects the ERPM figure.",
                         "Almost all hobby outrunners are 14."),
    "motor_weight": ("Weight of ONE motor in grams.", "2207 ~32 g, 5010 ~180 g."),
    "motor_size": ("Stator size label, for your reference.",
                   "e.g. 2207 = 22 mm wide, 7 mm tall."),

    # ---- ESC ----
    "esc_voltage_rating": ("ESC voltage rating as a CELL COUNT (the S number), "
                           "not volts.",
                           "A 6S ESC has a rating of 6. Must be >= your pack's "
                           "series count."),
    "esc_cont_current": ("Continuous current per ESC in amps.",
                         "From the label. A '35A ESC' means 35."),
    "esc_max_current": ("Burst current per ESC in amps.", "Usually 1.3-2x continuous."),
    "esc_idle_current": ("Current the ESC itself draws doing nothing.",
                         "Typically 0.02-0.1 A."),
    "esc_r": ("ESC internal resistance in ohms.",
              "Typical 0.001-0.005. Leave default if unknown."),
    "esc_weight": ("Weight of ONE ESC in grams.", "4-in-1 boards: divide by 4."),

    # ---- Avionics ----
    "avionics_voltage_tree": ("All the non-motor electronics, grouped by the "
                              "voltage rail that feeds them.",
                              "One row per rail. See the Avionics tab."),

    # ---- Propeller ----
    "prop_d": ("Propeller diameter in inches.",
               "First number on the prop: a 5045 prop is 5.0 in."),
    "prop_pitch": ("Propeller pitch in inches — how far it would screw forward "
                   "in one turn.",
                   "Second number: a 5045 prop is 4.5 in pitch."),
    "prop_blades": ("Number of blades.",
                    "2 = efficient, 3 = more thrust and grip, 5+ = loud and draggy."),
    "prop_max_rpm": ("Manufacturer RPM limit. Used for a status check.",
                     "Leave 0 if unknown."),
    "prop_max_thrust": ("Maximum thrust ONE motor+prop can make, in grams.",
                        "From a thrust test table. Used for the status check."),
    "prop_table": ("Optional CSV of measured thrust vs power. Overrides the "
                   "theoretical model and is far more accurate.",
                   "Columns: Thrust_g, Power_W, RPM. From the motor datasheet."),
    "prop_tconst": ("Thrust coefficient C_T, if you know it.",
                    "Advanced. Leave blank to use momentum theory."),
    "prop_pconst": ("Power coefficient C_P, if you know it.",
                    "Advanced. Leave blank."),
    "prop_weight": ("Weight of ONE propeller in grams.", "5in tri-blade ~4 g."),

    # ---- Mission / environment ----
    "mission": ("Optional mission JSON describing a multi-phase flight.",
                "Use Browse to load one of the bundled examples."),
    "orientation": ("forward = flying nose-first. hover = station-keeping or "
                    "translating sideways.",
                    "Use hover with speed 0 for pure hover endurance."),
    "alt": ("Altitude above sea level, in metres. Thinner air means less thrust.",
            "Your field elevation, not height above the ground."),
    "temp": ("Air temperature in Celsius. Optional — blank uses the standard "
             "atmosphere.", "Hot air is thinner, so hot days cost you thrust."),
    "press": ("Air pressure in pascals. Optional — blank derives it from altitude.",
              "Sea level standard is 101325 Pa."),
    "wind": ("Wind speed in m/s.", "1 m/s = 2.24 mph = 1.94 knots."),
    "wind_dir": ("Direction the wind is coming FROM, in degrees.",
                 "Meteorological convention: 0 = from the north, 90 = from the east."),
    "course_deg": ("The direction you are flying TOWARD, in degrees.",
                   "If this matches wind direction you have a pure headwind."),
    "climb_rate": ("Commanded climb rate in m/s. Costs extra power.", "0 for level flight."),
    "descent_rate": ("Commanded descent rate in m/s.", "0 for level flight."),
    "reserve_percent": ("Fraction of usable energy held back as a landing reserve.",
                        "20% is a common minimum. Never plan to land at 0."),
    "rth_reserve_Wh": ("Energy reserved for return-to-home, in watt-hours.",
                       "Estimate: cruise power x return time."),
    "diversion_reserve_Wh": ("Extra energy held for diverting to another landing "
                             "site.", "0 unless operating under a flight plan."),
    "transient_dt_s": ("Time step for acceleration modelling in missions.",
                       "Advanced. 0.5 s is fine."),
    "max_accel_mps2": ("Maximum acceleration used when changing speed between "
                       "mission phases.", "Advanced. Typical 2 m/s2."),
    "max_decel_mps2": ("Maximum deceleration between mission phases.",
                       "Advanced. Typical 2.5 m/s2."),
    "decel_regen_eff": ("Energy recovered while slowing down, 0 to 1.",
                        "Advanced. Leave 0 — props are poor regenerators."),
    "inflow_map_enabled": ("1 = correct rotor efficiency for forward speed. 0 = off.",
                           "Advanced. Leave at 1 if you have breakpoints set."),
    "inflow_mu_bp": ("Advance-ratio breakpoints for the inflow efficiency map.",
                     "Advanced. Leave blank."),
    "inflow_eff_bp": ("Efficiency at each advance ratio above.", "Advanced. Leave blank."),
    "max_speed_plot": ("Highest speed shown on the performance charts.",
                       "Set a bit above your expected top speed."),
}

# ------------------------------------------------------------------
# SIMPLE MODE FIELD SET
# ------------------------------------------------------------------
# Fields visible in Simple mode.  Everything not listed here is hidden
# until the user switches to Advanced.
#
# The rule used to pick these: a beginner sizing a first powertrain needs
# weight, pack, motor Kv/resistance, prop size, and the environment.  They
# do NOT need SoC breakpoint curves, inflow maps, or transient tuning.
#
# The whole Avionics tab is deliberately included: knowing what the
# flight controller, VTX, camera and payload draw is essential to an
# honest endurance number, and it is a common beginner mistake to omit it.
# ------------------------------------------------------------------
MC_SIMPLE_FIELDS = {
    # Drone
    "num_motors", "weight", "payload_mass_g", "speed", "motor_configuration",
    "max_tilt_deg", "drag_model_mode", "coaxial_spacing_m",
    "body_length_m", "body_width_m", "body_height_m", "arm_length_m", "arm_width_m",
    "parasite_drag", "parasite_area", "profile_drag", "profile_area",
    # Battery
    "batt_unit_mode", "batt_vmin", "batt_vnom", "batt_vmax",
    "batt_cell_capacity", "batt_pack_capacity",
    "batt_cell_weight", "batt_pack_weight",
    "batt_series", "batt_parallel", "batt_cells_series", "batt_cells_parallel",
    "batt_dischg_pct", "batt_r", "batt_c_cont", "batt_a_cont", "batt_chem",
    # Motor
    "motor_kv", "motor_r", "motor_i0", "motor_imax", "motor_pmax", "motor_weight",
    # ESC
    "esc_voltage_rating", "esc_cont_current", "esc_max_current", "esc_weight",
    # Avionics (whole tab stays in Simple mode by request)
    "avionics_voltage_tree",
    # Propeller
    "prop_d", "prop_pitch", "prop_blades", "prop_max_thrust", "prop_table",
    "prop_weight",
    # Mission / environment
    "mission", "orientation", "alt", "temp", "wind", "wind_dir", "course_deg",
    "reserve_percent", "max_speed_plot",
}


def launch_gui():
    """
    Tkinter GUI — styled to match the fixed-wing simulator:
      - Scrollable input tabs (Drone / Battery / Motor / ESC / Avionics / Propeller / Mission+Env)
      - Right panel: Plots | Status | Metrics | Mission Plots
      - View menu with Window Scale, Plot Size, UI Font Size, Plot Font Size, Quick Presets
      - Save / Load config (JSON)
    """
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

    # ------------------------------------------------------------------ #
    #  HELPERS                                                            #
    # ------------------------------------------------------------------ #
    def parse_float(name, s):
        try:
            return float(s)
        except Exception:
            raise ValueError(f"Invalid {name}: {s!r}")

    def parse_float_opt(name, s, default=None):
        """
        Parse an OPTIONAL numeric field: a blank entry returns `default`
        (None unless told otherwise) instead of raising.

        Needed because many inputs are either mutually exclusive alternatives
        — cell vs pack capacity/weight (chosen by unit mode), amp limits vs
        C-rate limits — or are reference-only values used solely for status
        checks.  A beginner should be able to leave those blank and still get
        a working answer rather than a wall of validation errors.
        """
        t = str(s).strip()
        if not t:
            return default
        try:
            return float(t)
        except Exception:
            raise ValueError(f"Invalid {name}: {s!r}")

    def parse_int_opt(name, s, default=None):
        """Integer counterpart of parse_float_opt."""
        t = str(s).strip()
        if not t:
            return default
        try:
            return int(float(t))
        except Exception:
            raise ValueError(f"Invalid {name}: {s!r}")

    def parse_int(name, s):
        try:
            return int(float(s))
        except Exception:
            raise ValueError(f"Invalid {name}: {s!r}")

    def safe_float(val, default=0.0):
        try:
            return float(str(val).strip())
        except Exception:
            return default

    def choose_file(var, filetypes):
        p = filedialog.askopenfilename(filetypes=filetypes)
        if p:
            var.set(p)

    def exit_app():
        try:
            plt.close("all")
        except Exception:
            pass
        root.quit()
        root.destroy()

    # ------------------------------------------------------------------ #
    #  ROOT WINDOW                                                        #
    # ------------------------------------------------------------------ #
    root = tk.Tk()
    root.title(f"Multicopter Power Simulator  v{SIM_VERSION}")
    root.minsize(1100, 700)
    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)

    # ------------------------------------------------------------------ #
    #  VIEW STATE  (shared by all scaling callbacks)                      #
    # ------------------------------------------------------------------ #
    try:
        _base_tk_scale = float(root.tk.call("tk", "scaling"))
    except Exception:
        _base_tk_scale = 1.333

    _view = {
        "scale_pct":    100,
        "plot_w":       15.0,
        "plot_h":        9.0,
        "mpl_fontsize":  9,
        "ui_fontsize":   9,
    }
    _last_run: dict = {}   # cached after each single-point run for View re-renders

    # ---- scaling helpers ----
    def _apply_tk_scale(pct: int) -> None:
        _view["scale_pct"] = pct
        factor = _base_tk_scale * (pct / 100.0)
        try:
            root.tk.call("tk", "scaling", factor)
        except Exception:
            pass
        root.minsize(int(1100 * pct / 100), int(700 * pct / 100))
        root.update_idletasks()

    def _apply_ui_font(size: int) -> None:
        _view["ui_fontsize"] = size
        sty = ttk.Style()
        font_spec = ("TkDefaultFont", size)
        for ws in ("TLabel", "TButton", "TEntry", "TCombobox",
                   "TNotebook.Tab", "Treeview", "Treeview.Heading",
                   "TLabelframe.Label", "TLabelframe"):
            try:
                sty.configure(ws, font=font_spec)
            except Exception:
                pass
        try:
            sty.configure("Treeview", rowheight=max(18, size + 8))
        except Exception:
            pass
        root.update_idletasks()

    def _apply_mpl_font(size: int) -> None:
        _view["mpl_fontsize"] = size
        import matplotlib as mpl
        mpl.rcParams.update({
            "font.size":        size,
            "axes.titlesize":   size + 1,
            "axes.labelsize":   size,
            "xtick.labelsize":  size - 1,
            "ytick.labelsize":  size - 1,
            "legend.fontsize":  size - 1,
        })
        _rerender_if_possible()

    def _apply_plot_size(w: float, h: float) -> None:
        _view["plot_w"] = w
        _view["plot_h"] = h
        _rerender_if_possible()

    def _rerender_if_possible() -> None:
        if not _last_run:
            return
        try:
            drone   = _last_run["drone"]
            max_spd = _last_run["max_spd"]
            metrics = _last_run.get("metrics", {})
            fig = make_performance_figure(
                drone,
                max_speed=max_spd,
                figsize=(_view["plot_w"], _view["plot_h"]),
            )
            # Generate motor operating point figure if available
            motor_fig = None
            if drone.propeller.table is not None and metrics:
                try:
                    motor_fig = make_motor_operating_point_figure(drone, metrics, figsize=(_view["plot_w"], 6))
                except Exception:
                    pass
            # Display both figures
            if motor_fig:
                _show_figure([fig, motor_fig])
            else:
                _show_figure(fig)
        except Exception:
            pass

    # ------------------------------------------------------------------ #
    #  MENU BAR                                                           #
    # ------------------------------------------------------------------ #
    menubar   = tk.Menu(root)
    file_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="File", menu=file_menu)

    view_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="View", menu=view_menu)

    # Help > About — makes the running version unambiguous.
    help_menu = tk.Menu(menubar, tearoff=0)
    menubar.add_cascade(label="Help", menu=help_menu)
    help_menu.add_command(
        label="About / Version",
        command=lambda: messagebox.showinfo(
            "About",
            f"Multicopter Power Simulator\n"
            f"Version {SIM_VERSION}\n"
            f"{SIM_BUILD_NOTE}\n\n"
            "If you do not see the 'Input detail: Simple / Advanced' selector\n"
            "above the input tabs, or the blue ? help markers beside each\n"
            "field, you are running an older copy of this script."),
    )

    # -- Window Scale --
    _scale_var = tk.IntVar(value=100)
    scale_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="Window Scale", menu=scale_menu)
    for pct, lbl in [(75,  "75 %  – Compact"),
                     (90,  "90 %  – Smaller"),
                     (100, "100 % – Default"),
                     (115, "115 % – Slightly Larger"),
                     (125, "125 % – Large"),
                     (150, "150 % – Extra Large"),
                     (175, "175 % – Very Large"),
                     (200, "200 % – Max")]:
        scale_menu.add_radiobutton(label=lbl, variable=_scale_var, value=pct,
                                   command=lambda p=pct: _apply_tk_scale(p))
    view_menu.add_separator()

    # -- Plot Size --
    _plot_size_var = tk.StringVar(value="medium")
    plot_size_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="Plot Size", menu=plot_size_menu)
    for key, lbl, (w, h) in [
        ("small",   "Small  (12 × 7)",          (12.0,  7.0)),
        ("medium",  "Medium (15 × 9)  ← default",(15.0,  9.0)),
        ("large",   "Large  (18 × 11)",          (18.0, 11.0)),
        ("xlarge",  "X-Large (22 × 13)",         (22.0, 13.0)),
        ("xxlarge", "XX-Large (26 × 15)",        (26.0, 15.0)),
    ]:
        plot_size_menu.add_radiobutton(label=lbl, variable=_plot_size_var, value=key,
                                       command=lambda pw=w, ph=h: _apply_plot_size(pw, ph))
    view_menu.add_separator()

    # -- UI Font Size --
    _ui_font_var = tk.IntVar(value=9)
    ui_font_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="UI Font Size", menu=ui_font_menu)
    for sz, lbl in [(8, "8 pt  – Tiny"), (9, "9 pt  – Default"),
                    (10,"10 pt – Comfortable"), (11,"11 pt – Large"),
                    (13,"13 pt – Extra Large"), (15,"15 pt – Accessibility")]:
        ui_font_menu.add_radiobutton(label=lbl, variable=_ui_font_var, value=sz,
                                     command=lambda s=sz: _apply_ui_font(s))
    view_menu.add_separator()

    # -- Plot Font Size --
    _mpl_font_var = tk.IntVar(value=9)
    mpl_font_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="Plot Font Size", menu=mpl_font_menu)
    for sz, lbl in [(7, "7 pt  – Tiny"), (8, "8 pt  – Small"),
                    (9, "9 pt  – Default"), (10,"10 pt – Medium"),
                    (12,"12 pt – Large"), (14,"14 pt – Extra Large")]:
        mpl_font_menu.add_radiobutton(label=lbl, variable=_mpl_font_var, value=sz,
                                      command=lambda s=sz: _apply_mpl_font(s))
    view_menu.add_separator()

    # -- Quick Presets --
    def _preset_compact():
        _scale_var.set(85);       _apply_tk_scale(85)
        _ui_font_var.set(8);      _apply_ui_font(8)
        _mpl_font_var.set(8);     _apply_mpl_font(8)
        _plot_size_var.set("small"); _apply_plot_size(12.0, 7.0)

    def _preset_default():
        _scale_var.set(100);      _apply_tk_scale(100)
        _ui_font_var.set(9);      _apply_ui_font(9)
        _mpl_font_var.set(9);     _apply_mpl_font(9)
        _plot_size_var.set("medium"); _apply_plot_size(15.0, 9.0)

    def _preset_presentation():
        _scale_var.set(140);      _apply_tk_scale(140)
        _ui_font_var.set(12);     _apply_ui_font(12)
        _mpl_font_var.set(12);    _apply_mpl_font(12)
        _plot_size_var.set("large"); _apply_plot_size(18.0, 11.0)

    def _preset_accessibility():
        _scale_var.set(160);      _apply_tk_scale(160)
        _ui_font_var.set(14);     _apply_ui_font(14)
        _mpl_font_var.set(13);    _apply_mpl_font(13)
        _plot_size_var.set("xlarge"); _apply_plot_size(22.0, 13.0)

    presets_menu = tk.Menu(view_menu, tearoff=0)
    view_menu.add_cascade(label="Quick Presets", menu=presets_menu)
    presets_menu.add_command(label="🗜  Compact",       command=_preset_compact)
    presets_menu.add_command(label="⚙  Default",        command=_preset_default)
    presets_menu.add_command(label="📊  Presentation",  command=_preset_presentation)
    presets_menu.add_command(label="♿  Accessibility",  command=_preset_accessibility)
    view_menu.add_separator()
    view_menu.add_command(label="Reset All to Default", command=_preset_default)

    root.config(menu=menubar)

    # ------------------------------------------------------------------ #
    #  STRING VARS (with sensible defaults)                               #
    # ------------------------------------------------------------------ #
    def sv(val=""):
        return tk.StringVar(value=str(val))

    # Drone
    v_num_motors        = sv(4)
    v_weight            = sv(1500)
    v_payload_mass      = sv(0)
    v_area              = sv(0.05)
    v_speed             = sv(10)
    v_periph_current    = sv(0.0)
    v_profile_drag      = sv(0.02)
    v_profile_area      = sv(0.01)
    v_parasite_drag     = sv(0.9)
    v_parasite_area     = sv(0.05)
    v_drag_model_mode   = sv("auto")
    v_body_length_m     = sv("")
    v_body_width_m      = sv("")
    v_body_height_m     = sv("")
    v_arm_length_m      = sv("")
    v_arm_width_m       = sv("")
    v_coaxial_spacing_m = sv("")
    v_max_tilt_deg      = sv("")
    v_motor_configuration = sv("flat")

    # Battery
    v_batt_vmin         = sv(3.0)
    v_batt_vnom         = sv(3.7)
    v_batt_vmax         = sv(4.2)
    v_batt_unit_mode    = sv("cell")
    v_batt_cell_capacity= sv(5000)
    v_batt_pack_capacity= sv(5000)
    v_batt_energy_density = sv(200)
    v_batt_chg          = sv(5)
    v_batt_a_cont       = sv(50)
    v_batt_a_max        = sv(100)
    v_batt_c_cont       = sv(15)
    v_batt_c_max        = sv(25)
    v_batt_dischg_pct   = sv(80)
    v_batt_r            = sv(20)
    v_batt_chem         = sv("LiPo")
    v_batt_soc_model    = sv("auto")
    v_batt_soc_curve_csv= sv("")
    v_batt_soc_bp       = sv("")
    v_batt_ocv_cell_bp  = sv("")
    v_batt_r_scale_bp   = sv("")
    v_batt_series       = sv(4)
    v_batt_parallel     = sv(1)
    v_batt_cells_series = sv(1)
    v_batt_cells_parallel = sv(1)
    v_batt_pack_weight  = sv(0)
    v_batt_cell_weight  = sv(0)

    # Motor
    v_motor_kv          = sv(650)
    v_motor_i0          = sv(0.5)
    v_motor_v0          = sv(10)
    v_motor_rated_v     = sv(6)
    v_motor_r           = sv(0.2)
    v_motor_imax        = sv(20)
    v_motor_pmax        = sv(200)
    v_motor_pole_count  = sv(14)
    v_motor_weight      = sv(168)
    v_motor_size        = sv("28x28mm")

    # ESC
    v_esc_voltage_rating= sv(6)
    v_esc_cont_current  = sv(30)
    v_esc_max_current   = sv(60)
    v_esc_idle_current  = sv(0.5)
    v_esc_r             = sv(0.01)
    v_esc_weight        = sv(36)

    # Avionics (string var kept in sync with the treeview)
    v_avionics_voltage_tree = sv("5.0:(2,0.9), 12.0:(1.5,0.85)")

    # Prop
    v_prop_d            = sv(12)
    v_prop_pitch        = sv(6)
    v_prop_max_rpm      = sv(10000)
    v_prop_max_thrust   = sv(3000)
    v_prop_blades       = sv(2)
    v_prop_table        = sv("")
    v_prop_tconst       = sv("")
    v_prop_pconst       = sv("")
    v_prop_weight       = sv(20)

    # Mission / env
    v_mission           = sv("")
    v_alt               = sv(0)
    v_temp              = sv("")
    v_press             = sv("")
    v_wind              = sv(0)
    v_wind_dir          = sv(0)
    v_course_deg        = sv(0)
    v_climb_rate        = sv(0)
    v_descent_rate      = sv(0)
    v_reserve_percent   = sv(20)
    v_rth_reserve_Wh    = sv(0)
    v_div_reserve_Wh    = sv(0)
    v_transient_dt_s    = sv(0.5)
    v_max_accel_mps2    = sv(2.0)
    v_max_decel_mps2    = sv(2.5)
    v_decel_regen_eff   = sv(0.0)
    v_inflow_map_enabled = sv("1")
    v_inflow_mu_bp      = sv(",".join(f"{x:g}" for x in DEFAULT_INFLOW_MU_BP))
    v_inflow_eff_bp     = sv(",".join(f"{x:g}" for x in DEFAULT_INFLOW_EFF_BP))
    v_orientation       = sv("forward")
    v_max_speed_plot    = sv(30)
    # UI complexity mode — "Simple" hides advanced inputs, "Advanced" shows all.
    # This is a view setting only; it never changes a computed result.
    v_ui_mode           = sv("Simple")
    # Name of the configuration currently loaded, shown in the mode bar.
    # The Output pane is overwritten by the first run, so the loaded config
    # needs somewhere permanent to live.
    v_loaded_cfg = sv("(no config loaded)")

    # config_vars collected for save/load  (populated after tab widgets are built)
    config_vars: dict = {}

    # ------------------------------------------------------------------ #
    #  MAIN LAYOUT                                                        #
    # ------------------------------------------------------------------ #
    main = ttk.Frame(root, padding=8)
    main.grid(sticky="nsew")
    main.columnconfigure(0, weight=1, minsize=370)
    main.columnconfigure(1, weight=3)
    main.rowconfigure(0, weight=1)

    # ===== LEFT: scrollable input notebook =====
    left = ttk.Frame(main)
    left.grid(row=0, column=0, sticky="nsew")
    left.columnconfigure(0, weight=1)
    # Row 0 = Simple/Advanced mode bar (fixed height)
    # Row 1 = the input notebook (takes all remaining vertical space)
    left.rowconfigure(1, weight=1)

    input_nb = ttk.Notebook(left)

    # ---- Simple / Advanced selector -------------------------------------
    # Sits directly above the input tabs so it is the first thing a new user
    # sees.  Simple mode is the default: it shows only the inputs needed to
    # size a powertrain, which is roughly a third of the full set.
    mode_bar = ttk.Frame(left)
    mode_bar.grid(row=0, column=0, sticky="ew", pady=(0, 4))
    ttk.Label(mode_bar, text="Input detail:").pack(side="left", padx=(2, 6))
    for _m in ("Simple", "Advanced"):
        ttk.Radiobutton(mode_bar, text=_m, value=_m,
                        variable=v_ui_mode).pack(side="left", padx=(0, 8))
    _mode_hint = ttk.Label(
        mode_bar,
        text="Simple hides advanced tuning inputs. Hover any ? for help.",
        foreground="#666666", font=("TkDefaultFont", 8))
    _mode_hint.pack(side="left", padx=(6, 0))
    ttk.Label(mode_bar, text="Config:", foreground="#666666",
              font=("TkDefaultFont", 8)).pack(side="left", padx=(12, 2))
    ttk.Label(mode_bar, textvariable=v_loaded_cfg, foreground="#0B6BCB",
              font=("TkDefaultFont", 8, "bold")).pack(side="left")
    input_nb.grid(row=1, column=0, sticky="nsew")

    # --- Scrollable-tab registry so the single wheel binding can find the
    #     active canvas without iterating or calling bind_all repeatedly.
    _tab_canvases: list = []

    def make_scrollable_tab(nb, title):
        """
        Wrap a notebook tab in a Canvas + vertical Scrollbar.

        Key fixes vs the original:
          1. canvas.<Configure> — fires when the canvas itself is resized
             (e.g. when the tab is first shown).  Without this, winfo_width()
             returns 1 at startup so the inner frame is rendered 1 px wide
             and appears blank until the user switches tabs twice.
          2. bind_all is NOT used here — a single wheel binding is attached
             to the notebook after all tabs are created (see below), routing
             scroll events only to the currently-visible canvas.  Calling
             bind_all inside a loop creates N overlapping handlers on every
             widget in the application, which stalls the Tk event loop and
             causes the multi-minute render delay.
        """
        outer = ttk.Frame(nb)
        nb.add(outer, text=title)
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)
        cv  = tk.Canvas(outer, highlightthickness=0)
        sb  = ttk.Scrollbar(outer, orient="vertical", command=cv.yview)
        inn = ttk.Frame(cv)
        cv.configure(yscrollcommand=sb.set)
        cv.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        win_id = cv.create_window((0, 0), window=inn, anchor="nw")

        def _resize_inner(width):
            """Set the embedded-window width to match the canvas width."""
            if width > 1:                     # ignore the startup 1-px dummy value
                cv.itemconfig(win_id, width=width)

        def _on_inner_configure(evt):
            """Called when widgets are added to / removed from the inner frame."""
            cv.configure(scrollregion=cv.bbox("all"))
            _resize_inner(cv.winfo_width())

        def _on_canvas_configure(evt):
            """
            Called when the canvas is resized — critically, this fires when
            the tab is *first shown*, giving us the real pixel width.
            """
            cv.configure(scrollregion=cv.bbox("all"))
            _resize_inner(evt.width)

        inn.bind("<Configure>", _on_inner_configure)
        cv.bind("<Configure>",  _on_canvas_configure)   # <- the essential new line

        inn.columnconfigure(0, weight=1)
        inn.columnconfigure(1, weight=1)
        _tab_canvases.append(cv)   # register so wheel handler can find it
        return inn

    # Tab label is "Airframe" to match the fixed-wing simulator; the internal
    # variable stays tab_drone so existing references keep working.
    tab_drone    = make_scrollable_tab(input_nb, "Airframe")
    tab_batt     = make_scrollable_tab(input_nb, "Battery")
    tab_motor    = make_scrollable_tab(input_nb, "Motor")
    tab_esc      = make_scrollable_tab(input_nb, "ESC")
    tab_avionics = make_scrollable_tab(input_nb, "Avionics")
    tab_prop     = make_scrollable_tab(input_nb, "Propeller")
    tab_mission  = make_scrollable_tab(input_nb, "Mission/Environment")

    # --- Single notebook-level mouse-wheel binding ----------------------------
    # Replaces the old bind_all-per-tab pattern.  We find whichever canvas
    # belongs to the currently-selected tab and scroll only that one.
    def _on_nb_mousewheel(evt):
        try:
            idx = input_nb.index(input_nb.select())
            if 0 <= idx < len(_tab_canvases):
                _tab_canvases[idx].yview_scroll(
                    int(-1 * (evt.delta / 120)), "units")
        except Exception:
            pass
    input_nb.bind("<MouseWheel>", _on_nb_mousewheel)
    # Also bind directly to each canvas so hovering over tab content works.
    for _cv in _tab_canvases:
        _cv.bind("<MouseWheel>", lambda evt, c=_cv:
                 c.yview_scroll(int(-1 * (evt.delta / 120)), "units"))
    # --------------------------------------------------------------------------


    # ------------------------------------------------------------------
    # Registry of every input row, so Simple/Advanced mode can show or hide
    # rows without rebuilding the GUI.  Each entry records the widgets that
    # make up one row plus the grid position to restore it to.
    # ------------------------------------------------------------------
    _field_rows = []          # list of dicts: key, widgets, parent, row
    _section_rows = []        # separators / section headings, with their keys

    def _register_row(key, widgets, parent, row):
        """Record a row so _apply_field_mode() can grid/ungrid it later."""
        _field_rows.append({"key": key, "widgets": widgets,
                            "parent": parent, "row": row})

    def add_row(parent, r, label, var, key=None, **kw):
        """
        Add one labelled entry row, with an inline "?" help marker.

        `key` ties the row to MC_FIELD_HELP (for the tooltip) and to
        MC_SIMPLE_FIELDS (for Simple/Advanced visibility).  Rows with no key
        are always shown and get no tooltip.
        """
        lbl = ttk.Label(parent, text=label)
        lbl.grid(row=r, column=0, sticky="w", padx=6, pady=3)
        e = ttk.Entry(parent, textvariable=var, width=14, **kw)
        e.grid(row=r, column=1, sticky="ew", padx=6, pady=3)

        widgets = [lbl, e]
        help_entry = MC_FIELD_HELP.get(key) if key else None
        if help_entry:
            what, typical = help_entry
            marker = ttk.Label(parent, text=" ? ", foreground="#0B6BCB",
                               cursor="question_arrow",
                               font=("TkDefaultFont", 9, "bold"))
            marker.grid(row=r, column=2, sticky="w", padx=(0, 6))
            _Tooltip(marker, f"{what}\n\nTypical: {typical}")
            _Tooltip(lbl, f"{what}\n\nTypical: {typical}")
            widgets.append(marker)

        if key:
            _register_row(key, widgets, parent, r)
        return e

    def add_section(parent, r, text, keys=()):
        """
        Add a separator + grey section heading.  If `keys` is given, the whole
        section is hidden in Simple mode when none of those keys are visible.
        """
        sep = ttk.Separator(parent, orient="horizontal")
        sep.grid(row=r, column=0, columnspan=3, sticky="ew", pady=6)
        lab = ttk.Label(parent, text=text, foreground="gray")
        lab.grid(row=r + 1, column=0, columnspan=3, sticky="w", padx=6)
        _section_rows.append({"keys": set(keys), "widgets": [sep, lab],
                              "parent": parent, "row": r})
        return r + 2

    # ===== DRONE TAB =====
    ttk.Label(tab_drone, text="Motor Configuration:").grid(
        row=0, column=0, sticky="w", padx=6, pady=4)
    motor_cfg_cb = ttk.Combobox(tab_drone, textvariable=v_motor_configuration,
                                values=["flat","coaxial"], state="readonly", width=12)
    motor_cfg_cb.grid(row=0, column=1, sticky="w", padx=6, pady=4)
    _mc_help = MC_FIELD_HELP["motor_configuration"]
    _mc_mark = ttk.Label(tab_drone, text=" ? ", foreground="#0B6BCB",
                         cursor="question_arrow", font=("TkDefaultFont", 9, "bold"))
    _mc_mark.grid(row=0, column=2, sticky="w", padx=(0, 6))
    _Tooltip(_mc_mark, f"{_mc_help[0]}\n\nTypical: {_mc_help[1]}")

    r = 1
    add_row(tab_drone, r, "Num Motors",              v_num_motors, key="num_motors");        r += 1
    add_row(tab_drone, r, "Base Weight (g)",         v_weight, key="weight");            r += 1
    add_row(tab_drone, r, "Payload Mass (g)",        v_payload_mass, key="payload_mass_g");      r += 1
    add_row(tab_drone, r, "Frontal Area (m²)",       v_area, key="area");              r += 1
    add_row(tab_drone, r, "Cruise Speed (m/s)",      v_speed, key="speed");             r += 1
    add_row(tab_drone, r, "Peripheral Current (A)",  v_periph_current, key="periph_current");    r += 1

    ttk.Separator(tab_drone, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_drone, text="── Drag Parameters ──",
              foreground="gray").grid(row=r, column=0, columnspan=2, sticky="w", padx=6); r += 1
    ttk.Label(tab_drone, text="Drag model mode").grid(row=r, column=0, sticky="w", padx=6, pady=3)
    drag_mode_cb = ttk.Combobox(
        tab_drone,
        textvariable=v_drag_model_mode,
        values=["auto", "manual"],
        state="readonly",
        width=12,
    )
    drag_mode_cb.grid(row=r, column=1, sticky="w", padx=6, pady=3)
    _dm_help = MC_FIELD_HELP["drag_model_mode"]
    _dm_mark = ttk.Label(tab_drone, text=" ? ", foreground="#0B6BCB",
                         cursor="question_arrow", font=("TkDefaultFont", 9, "bold"))
    _dm_mark.grid(row=r, column=2, sticky="w", padx=(0, 6))
    _Tooltip(_dm_mark, f"{_dm_help[0]}\n\nTypical: {_dm_help[1]}"); r += 1
    add_row(tab_drone, r, "Profile Cd",              v_profile_drag, key="profile_drag");      r += 1
    add_row(tab_drone, r, "Profile Area (m²)",       v_profile_area, key="profile_area");      r += 1
    add_row(tab_drone, r, "Parasite Cd",             v_parasite_drag, key="parasite_drag");     r += 1
    add_row(tab_drone, r, "Parasite Area (m²)",      v_parasite_area, key="parasite_area");     r += 1

    ttk.Separator(tab_drone, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_drone, text="── Body Geometry (optional) ──",
              foreground="gray").grid(row=r, column=0, columnspan=2, sticky="w", padx=6); r += 1
    body_len_e = add_row(tab_drone, r, "Body Length (m)",         v_body_length_m, key="body_length_m");     r += 1
    body_w_e = add_row(tab_drone, r, "Body Width (m)",          v_body_width_m, key="body_width_m");      r += 1
    body_h_e = add_row(tab_drone, r, "Body Height (m)",         v_body_height_m, key="body_height_m");     r += 1
    arm_len_e = add_row(tab_drone, r, "Arm Length (m)",          v_arm_length_m, key="arm_length_m");      r += 1
    arm_w_e = add_row(tab_drone, r, "Arm Width (m)",           v_arm_width_m, key="arm_width_m");       r += 1
    add_row(tab_drone, r, "Max Tilt (deg)",          v_max_tilt_deg, key="max_tilt_deg");      r += 1
    coax_entry = add_row(tab_drone, r, "Coaxial Spacing (m)", v_coaxial_spacing_m, key="coaxial_spacing_m"); r += 1
    coax_entry.configure(state="disabled")

    def _update_coax_state(event=None):
        if v_motor_configuration.get().strip().lower() == "coaxial":
            coax_entry.configure(state="normal")
        else:
            v_coaxial_spacing_m.set("")
            coax_entry.configure(state="disabled")
    motor_cfg_cb.bind("<<ComboboxSelected>>", _update_coax_state)

    def _update_drag_model_state(event=None):
        manual = (v_drag_model_mode.get().strip().lower() == "manual")
        state = "disabled" if manual else "normal"
        for e in (body_len_e, body_w_e, body_h_e, arm_len_e, arm_w_e):
            e.configure(state=state)
    drag_mode_cb.bind("<<ComboboxSelected>>", _update_drag_model_state)
    _update_drag_model_state()

    # ===== BATTERY TAB =====
    ttk.Label(tab_batt, text="Unit mode:").grid(row=0, column=0, sticky="w", padx=6, pady=4)
    unit_mode_cb = ttk.Combobox(tab_batt, textvariable=v_batt_unit_mode,
                                values=["cell","pack"], state="readonly", width=10)
    unit_mode_cb.grid(row=0, column=1, sticky="w", padx=6, pady=4)
    _um_help = MC_FIELD_HELP["batt_unit_mode"]
    _um_mark = ttk.Label(tab_batt, text=" ? ", foreground="#0B6BCB",
                         cursor="question_arrow", font=("TkDefaultFont", 9, "bold"))
    _um_mark.grid(row=0, column=2, sticky="w", padx=(0, 6))
    _Tooltip(_um_mark, f"{_um_help[0]}\n\nTypical: {_um_help[1]}")

    r = 1
    add_row(tab_batt, r, "Vmin/cell (V)",            v_batt_vmin, key="batt_vmin");         r += 1
    add_row(tab_batt, r, "Vnom/cell (V)",            v_batt_vnom, key="batt_vnom");         r += 1
    add_row(tab_batt, r, "Vmax/cell (V)",            v_batt_vmax, key="batt_vmax");         r += 1
    cell_cap_e = add_row(tab_batt, r, "Cell Capacity (mAh)",   v_batt_cell_capacity, key="batt_cell_capacity"); r += 1
    pack_cap_e = add_row(tab_batt, r, "Pack Capacity (mAh)",   v_batt_pack_capacity, key="batt_pack_capacity"); r += 1
    cell_wt_e  = add_row(tab_batt, r, "Cell Weight (g)",       v_batt_cell_weight, key="batt_cell_weight");   r += 1
    pack_wt_e  = add_row(tab_batt, r, "Pack Weight (g)",       v_batt_pack_weight, key="batt_pack_weight");   r += 1
    add_row(tab_batt, r, "Energy Density (Wh/kg)",   v_batt_energy_density, key="batt_energy_density"); r += 1
    add_row(tab_batt, r, "Max Charge Current (A)",   v_batt_chg, key="batt_chg");          r += 1
    add_row(tab_batt, r, "Cont Discharge (A)",       v_batt_a_cont, key="batt_a_cont");       r += 1
    add_row(tab_batt, r, "Max Discharge (A)",        v_batt_a_max, key="batt_a_max");        r += 1
    add_row(tab_batt, r, "Cont C-rate",              v_batt_c_cont, key="batt_c_cont");       r += 1
    add_row(tab_batt, r, "Max C-rate",               v_batt_c_max, key="batt_c_max");        r += 1
    add_row(tab_batt, r, "Usable Discharge (%)",     v_batt_dischg_pct, key="batt_dischg_pct");   r += 1
    add_row(tab_batt, r, "Rcell (mΩ)",               v_batt_r, key="batt_r");            r += 1
    add_row(tab_batt, r, "SoC model",                v_batt_soc_model, key="batt_soc_model");    r += 1
    add_row(tab_batt, r, "SoC curve CSV",            v_batt_soc_curve_csv, key="batt_soc_curve_csv"); r += 1
    add_row(tab_batt, r, "SoC breakpoints (0..1)",   v_batt_soc_bp, key="batt_soc_bp");       r += 1
    add_row(tab_batt, r, "OCV/cell breakpoints (V)", v_batt_ocv_cell_bp, key="batt_ocv_cell_bp");  r += 1
    add_row(tab_batt, r, "R-scale breakpoints",      v_batt_r_scale_bp, key="batt_r_scale_bp");   r += 1
    add_row(tab_batt, r, "Series Cells/Packs",       v_batt_series, key="batt_series");       r += 1
    add_row(tab_batt, r, "Parallel Cells/Packs",     v_batt_parallel, key="batt_parallel");     r += 1
    cells_s_e = add_row(tab_batt, r, "Cells in Series/Pack",  v_batt_cells_series, key="batt_cells_series");  r += 1
    cells_p_e = add_row(tab_batt, r, "Cells in Parallel/Pack",v_batt_cells_parallel, key="batt_cells_parallel"); r += 1
    add_row(tab_batt, r, "Chemistry",                v_batt_chem, key="batt_chem");         r += 1

    def on_unit_mode_change(event=None):
        mode = v_batt_unit_mode.get()
        if mode == "cell":
            cell_cap_e.configure(state="normal");  cell_wt_e.configure(state="normal")
            pack_cap_e.configure(state="disabled"); pack_wt_e.configure(state="disabled")
            cells_s_e.configure(state="disabled");  cells_p_e.configure(state="disabled")
        else:
            cell_cap_e.configure(state="disabled"); cell_wt_e.configure(state="disabled")
            pack_cap_e.configure(state="normal");   pack_wt_e.configure(state="normal")
            cells_s_e.configure(state="normal");    cells_p_e.configure(state="normal")
    unit_mode_cb.bind("<<ComboboxSelected>>", on_unit_mode_change)
    on_unit_mode_change()

    # ===== MOTOR TAB =====
    r = 0
    add_row(tab_motor, r, "Kv (RPM/V)",          v_motor_kv, key="motor_kv");         r += 1
    add_row(tab_motor, r, "Idle Current I0 (A)",  v_motor_i0, key="motor_i0");         r += 1
    add_row(tab_motor, r, "Idle Voltage V0 (V)",  v_motor_v0, key="motor_v0");         r += 1
    add_row(tab_motor, r, "Rated Voltage (V)",    v_motor_rated_v, key="motor_rated_v");    r += 1
    add_row(tab_motor, r, "Resistance Rm (Ω)",    v_motor_r, key="motor_r");          r += 1
    add_row(tab_motor, r, "Max Current (A)",      v_motor_imax, key="motor_imax");       r += 1
    add_row(tab_motor, r, "Max Power (W)",        v_motor_pmax, key="motor_pmax");       r += 1
    add_row(tab_motor, r, "Pole Count",           v_motor_pole_count, key="motor_pole_count"); r += 1
    add_row(tab_motor, r, "Weight (g)",           v_motor_weight, key="motor_weight");     r += 1
    add_row(tab_motor, r, "Size (e.g. 28x28mm)",  v_motor_size, key="motor_size");       r += 1

    # ===== ESC TAB =====
    r = 0
    add_row(tab_esc, r, "Voltage Rating (S cells)",     v_esc_voltage_rating, key="esc_voltage_rating"); r += 1
    add_row(tab_esc, r, "Continuous Current (A)",  v_esc_cont_current, key="esc_cont_current");   r += 1
    add_row(tab_esc, r, "Max Current (A)",         v_esc_max_current, key="esc_max_current");    r += 1
    add_row(tab_esc, r, "Idle Current (A)",        v_esc_idle_current, key="esc_idle_current");   r += 1
    add_row(tab_esc, r, "Resistance (Ω)",          v_esc_r, key="esc_r");              r += 1
    add_row(tab_esc, r, "Weight (g)",              v_esc_weight, key="esc_weight");         r += 1

    # ===== AVIONICS TAB =====
    tab_avionics.columnconfigure(0, weight=1)
    tab_avionics.rowconfigure(1, weight=1)

    ttk.Label(
        tab_avionics,
        text="BEC / Avionics voltage rails  —  one row per regulated output bus.\n"
             "Double-click any cell to edit it in-place.",
        wraplength=340, justify="left", foreground="#555555",
    ).grid(row=0, column=0, columnspan=2, sticky="w", padx=6, pady=(6, 2))

    av_tree_frame = ttk.Frame(tab_avionics)
    av_tree_frame.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=6, pady=(0, 4))
    av_tree_frame.columnconfigure(0, weight=1)
    av_tree_frame.rowconfigure(0, weight=1)

    _AV_COLS = ("voltage", "current", "efficiency")
    avionics_tree = ttk.Treeview(
        av_tree_frame, columns=_AV_COLS, show="headings", height=8, selectmode="browse")
    for col, heading, width in [
        ("voltage",    "Rail Voltage (V)",      130),
        ("current",    "Rail Current (A)",      130),
        ("efficiency", "BEC Efficiency (0–1]",  150),
    ]:
        avionics_tree.heading(col, text=heading)
        avionics_tree.column(col, width=width, anchor="center", stretch=True)
    avionics_tree.grid(row=0, column=0, sticky="nsew")

    av_sb = ttk.Scrollbar(av_tree_frame, orient="vertical", command=avionics_tree.yview)
    av_sb.grid(row=0, column=1, sticky="ns")
    avionics_tree.configure(yscrollcommand=av_sb.set)

    # ---- Helpers for the avionics table ----
    def _canonical_voltage_tree_string(d: dict) -> str:
        items = sorted(((float(v), float(i), float(e)) for v, (i, e) in d.items()),
                       key=lambda t: t[0])
        return ", ".join(f"{v:.3g}:({i:.3g},{e:.3g})" for v, i, e in items)

    def _sync_voltage_tree_var_from_table() -> None:
        d = {}
        for iid in avionics_tree.get_children():
            v_s, i_s, e_s = avionics_tree.item(iid, "values")
            try:
                d[float(v_s)] = (float(i_s), float(e_s))
            except Exception:
                continue
        v_avionics_voltage_tree.set(_canonical_voltage_tree_string(d))

    def _get_voltage_tree_from_table() -> dict:
        d = {}
        for iid in avionics_tree.get_children():
            v_s, i_s, e_s = avionics_tree.item(iid, "values")
            v, i, e = float(v_s), float(i_s), float(e_s)
            if v <= 0:   raise ValueError(f"Rail voltage must be > 0, got {v}.")
            if i < 0:    raise ValueError(f"Rail current must be ≥ 0, got {i}.")
            if e <= 0 or e > 1.0: raise ValueError(f"BEC efficiency must be in (0,1], got {e}.")
            d[v] = (i, e)
        return d

    def _av_load_rows(rows: list) -> None:
        avionics_tree.delete(*avionics_tree.get_children())
        for r in sorted(rows, key=lambda x: float(x.get("voltage", 0))):
            try:
                v = float(r.get("voltage")); i = float(r.get("current")); e = float(r.get("eff"))
                avionics_tree.insert("", "end", values=(f"{v:g}", f"{i:g}", f"{e:g}"))
            except Exception:
                continue
        _sync_voltage_tree_var_from_table()

    # ---- Double-click in-place editing ----
    def _av_begin_edit(event):
        region = avionics_tree.identify("region", event.x, event.y)
        if region != "cell": return
        row_id = avionics_tree.identify_row(event.y)
        col_id = avionics_tree.identify_column(event.x)
        if not row_id or not col_id: return
        col_idx = int(col_id.replace("#", "")) - 1
        bbox = avionics_tree.bbox(row_id, col_id)
        if not bbox: return
        x, y, w, h = bbox
        old_vals = list(avionics_tree.item(row_id, "values"))

        ed = tk.Entry(avionics_tree, justify="center")
        ed.insert(0, old_vals[col_idx])
        ed.select_range(0, tk.END)
        ed.focus_set()
        ed.place(x=x, y=y, width=w, height=h)

        def _commit(_evt=None):
            old_vals[col_idx] = ed.get().strip()
            avionics_tree.item(row_id, values=tuple(old_vals))
            ed.destroy()
            _sync_voltage_tree_var_from_table()

        def _cancel(_evt=None): ed.destroy()

        ed.bind("<Return>",   _commit)
        ed.bind("<Tab>",      _commit)
        ed.bind("<FocusOut>", _commit)
        ed.bind("<Escape>",   _cancel)

    avionics_tree.bind("<Double-1>", _av_begin_edit)

    # ---- Click row → populate entry fields ----
    _av_v_var = tk.StringVar()
    _av_i_var = tk.StringVar()
    _av_e_var = tk.StringVar(value="0.90")

    def _av_on_select(event):
        sel = avionics_tree.selection()
        if not sel: return
        vals = avionics_tree.item(sel[0], "values")
        try: _av_v_var.set(vals[0]); _av_i_var.set(vals[1]); _av_e_var.set(vals[2])
        except Exception: pass
    avionics_tree.bind("<<TreeviewSelect>>", _av_on_select)

    # ---- Entry fields for add/update ----
    av_entry_frame = ttk.Frame(tab_avionics)
    av_entry_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=6, pady=(2, 2))
    for c in range(6): av_entry_frame.columnconfigure(c, weight=1)

    ttk.Label(av_entry_frame, text="Voltage (V):").grid(row=0, column=0, sticky="e", padx=(0,2))
    ttk.Entry(av_entry_frame, textvariable=_av_v_var, width=7).grid(row=0, column=1, sticky="ew", padx=(0,6))
    ttk.Label(av_entry_frame, text="Current (A):").grid(row=0, column=2, sticky="e", padx=(0,2))
    ttk.Entry(av_entry_frame, textvariable=_av_i_var, width=7).grid(row=0, column=3, sticky="ew", padx=(0,6))
    ttk.Label(av_entry_frame, text="Efficiency:").grid(row=0, column=4, sticky="e", padx=(0,2))
    ttk.Entry(av_entry_frame, textvariable=_av_e_var, width=7).grid(row=0, column=5, sticky="ew")

    # ---- Buttons ----
    av_btn_frame = ttk.Frame(tab_avionics)
    av_btn_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=6, pady=(0, 6))

    def _av_add_or_update():
        try:
            v = float(_av_v_var.get()); i = float(_av_i_var.get()); e = float(_av_e_var.get())
        except ValueError:
            messagebox.showerror("Invalid input", "Voltage, Current and Efficiency must be numbers.")
            return
        if v <= 0:
            messagebox.showerror("Invalid input", "Rail voltage must be > 0 V."); return
        if i < 0:
            messagebox.showerror("Invalid input", "Rail current must be ≥ 0 A."); return
        if not (0 < e <= 1.0):
            messagebox.showerror("Invalid input", "BEC efficiency must be in (0, 1]."); return
        for iid in avionics_tree.get_children():
            try:
                if abs(float(avionics_tree.item(iid,"values")[0]) - v) < 1e-9:
                    avionics_tree.item(iid, values=(f"{v:g}", f"{i:g}", f"{e:g}"))
                    _sync_voltage_tree_var_from_table()
                    _av_v_var.set(""); _av_i_var.set(""); _av_e_var.set("0.90")
                    return
            except Exception: pass
        avionics_tree.insert("", "end", values=(f"{v:g}", f"{i:g}", f"{e:g}"))
        _sync_voltage_tree_var_from_table()
        _av_v_var.set(""); _av_i_var.set(""); _av_e_var.set("0.90")

    def _av_remove():
        sel = avionics_tree.selection()
        for iid in sel: avionics_tree.delete(iid)
        _sync_voltage_tree_var_from_table()

    def _av_clear():
        avionics_tree.delete(*avionics_tree.get_children())
        _sync_voltage_tree_var_from_table()

    ttk.Button(av_btn_frame, text="➕  Add / Update Rail",
               command=_av_add_or_update).grid(row=0, column=0, sticky="w", padx=(0, 6))
    ttk.Button(av_btn_frame, text="🗑  Remove Selected",
               command=_av_remove).grid(row=0, column=1, sticky="w", padx=(0, 6))
    ttk.Button(av_btn_frame, text="✖  Clear All",
               command=_av_clear).grid(row=0, column=2, sticky="w")

    # Seed the table from the default string var
    try:
        _initial_tree = parse_voltage_tree(v_avionics_voltage_tree.get())
        _av_load_rows([{"voltage": v, "current": ci[0], "eff": ci[1]}
                       for v, ci in sorted(_initial_tree.items())])
    except Exception:
        _av_clear()

    # ===== PROPELLER TAB =====
    r = 0
    add_row(tab_prop, r, "Diameter (in)",         v_prop_d, key="prop_d");          r += 1
    add_row(tab_prop, r, "Pitch (in)",            v_prop_pitch, key="prop_pitch");      r += 1
    add_row(tab_prop, r, "Blades",                v_prop_blades, key="prop_blades");     r += 1
    add_row(tab_prop, r, "Max RPM",               v_prop_max_rpm, key="prop_max_rpm");    r += 1
    add_row(tab_prop, r, "Max Thrust (g)",         v_prop_max_thrust, key="prop_max_thrust"); r += 1
    ttk.Separator(tab_prop, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_prop, text="Prop/Motor CSV table (optional)").grid(
        row=r, column=0, sticky="w", padx=6, pady=2)
    frow = ttk.Frame(tab_prop); frow.grid(row=r, column=1, sticky="ew")
    frow.columnconfigure(0, weight=1)
    ttk.Entry(frow, textvariable=v_prop_table).grid(row=0, column=0, sticky="ew", padx=(6,4))
    ttk.Button(frow, text="Browse…",
               command=lambda: choose_file(v_prop_table, [("CSV","*.csv"),("All","*.*")])).grid(
        row=0, column=1, padx=(0,6)); r += 1
    add_row(tab_prop, r, "TConst (optional)",     v_prop_tconst, key="prop_tconst");     r += 1
    add_row(tab_prop, r, "PConst (optional)",     v_prop_pconst, key="prop_pconst");     r += 1
    add_row(tab_prop, r, "Weight (g)",            v_prop_weight, key="prop_weight");     r += 1

    # ===== MISSION / ENV TAB =====
    r = 0
    ttk.Label(tab_mission, text="Mission JSON (optional)").grid(
        row=r, column=0, sticky="w", padx=6, pady=2)
    mrow = ttk.Frame(tab_mission); mrow.grid(row=r, column=1, sticky="ew")
    mrow.columnconfigure(0, weight=1)
    ttk.Entry(mrow, textvariable=v_mission).grid(row=0, column=0, sticky="ew", padx=(6,4))
    ttk.Button(mrow, text="Browse…",
               command=lambda: choose_file(v_mission, [("JSON","*.json"),("All","*.*")])).grid(
        row=0, column=1, padx=(0,6)); r += 1
    ttk.Separator(tab_mission, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1

    orientation_cb = ttk.Combobox(tab_mission, textvariable=v_orientation,
                                  values=["forward","hover"], state="readonly", width=12)
    ttk.Label(tab_mission, text="Orientation:").grid(row=r, column=0, sticky="w", padx=6, pady=3)
    orientation_cb.grid(row=r, column=1, sticky="w", padx=6, pady=3)
    _or_help = MC_FIELD_HELP["orientation"]
    _or_mark = ttk.Label(tab_mission, text=" ? ", foreground="#0B6BCB",
                         cursor="question_arrow", font=("TkDefaultFont", 9, "bold"))
    _or_mark.grid(row=r, column=2, sticky="w", padx=(0, 6))
    _Tooltip(_or_mark, f"{_or_help[0]}\n\nTypical: {_or_help[1]}"); r += 1

    add_row(tab_mission, r, "Altitude (m)",              v_alt, key="alt");             r += 1
    add_row(tab_mission, r, "Temperature (°C, optional)",v_temp, key="temp");            r += 1
    add_row(tab_mission, r, "Pressure (Pa, optional)",   v_press, key="press");           r += 1
    add_row(tab_mission, r, "Wind speed (m/s)",           v_wind, key="wind");            r += 1
    add_row(tab_mission, r, "Wind direction FROM (deg)",  v_wind_dir, key="wind_dir");        r += 1
    add_row(tab_mission, r, "Course heading (deg)",       v_course_deg, key="course_deg");      r += 1
    add_row(tab_mission, r, "Climb rate cmd (m/s)",       v_climb_rate, key="climb_rate");      r += 1
    add_row(tab_mission, r, "Descent rate cmd (m/s)",     v_descent_rate, key="descent_rate");    r += 1
    add_row(tab_mission, r, "Reserve percent (%)",        v_reserve_percent, key="reserve_percent"); r += 1
    add_row(tab_mission, r, "RTH reserve (Wh)",           v_rth_reserve_Wh, key="rth_reserve_Wh");  r += 1
    add_row(tab_mission, r, "Diversion reserve (Wh)",     v_div_reserve_Wh, key="diversion_reserve_Wh");  r += 1
    ttk.Separator(tab_mission, orient="horizontal").grid(
        row=r, column=0, columnspan=2, sticky="ew", pady=6); r += 1
    ttk.Label(tab_mission, text="── Transients & Inflow Map ──",
              foreground="gray").grid(row=r, column=0, columnspan=2, sticky="w", padx=6); r += 1
    add_row(tab_mission, r, "Transient step dt (s)",      v_transient_dt_s, key="transient_dt_s");  r += 1
    add_row(tab_mission, r, "Max accel (m/s²)",           v_max_accel_mps2, key="max_accel_mps2");  r += 1
    add_row(tab_mission, r, "Max decel (m/s²)",           v_max_decel_mps2, key="max_decel_mps2");  r += 1
    add_row(tab_mission, r, "Decel regen efficiency (0-1)", v_decel_regen_eff, key="decel_regen_eff"); r += 1
    add_row(tab_mission, r, "Inflow map enabled (1/0)",   v_inflow_map_enabled, key="inflow_map_enabled"); r += 1
    add_row(tab_mission, r, "Inflow mu breakpoints",      v_inflow_mu_bp, key="inflow_mu_bp");    r += 1
    add_row(tab_mission, r, "Inflow eta breakpoints",     v_inflow_eff_bp, key="inflow_eff_bp");   r += 1
    add_row(tab_mission, r, "Max Speed for Plot (m/s)",   v_max_speed_plot, key="max_speed_plot");  r += 1

    # ------------------------------------------------------------------ #
    #  Collect config_vars for save/load                                  #
    # ------------------------------------------------------------------ #
    config_vars = {
        "num_motors": v_num_motors, "weight": v_weight, "payload_mass_g": v_payload_mass, "area": v_area,
        "speed": v_speed, "periph_current": v_periph_current,
        "profile_drag": v_profile_drag, "profile_area": v_profile_area,
        "parasite_drag": v_parasite_drag, "parasite_area": v_parasite_area,
        "body_length_m": v_body_length_m, "body_width_m": v_body_width_m,
        "body_height_m": v_body_height_m, "arm_length_m": v_arm_length_m,
        "arm_width_m": v_arm_width_m, "coaxial_spacing_m": v_coaxial_spacing_m,
        "max_tilt_deg": v_max_tilt_deg, "motor_configuration": v_motor_configuration,
        "drag_model_mode": v_drag_model_mode,
        "batt_vmin": v_batt_vmin, "batt_vnom": v_batt_vnom, "batt_vmax": v_batt_vmax,
        "batt_unit_mode": v_batt_unit_mode,
        "batt_cell_capacity": v_batt_cell_capacity, "batt_pack_capacity": v_batt_pack_capacity,
        "batt_energy_density": v_batt_energy_density,
        "batt_chg": v_batt_chg, "batt_a_cont": v_batt_a_cont, "batt_a_max": v_batt_a_max,
        "batt_c_cont": v_batt_c_cont, "batt_c_max": v_batt_c_max,
        "batt_dischg_pct": v_batt_dischg_pct, "batt_r": v_batt_r, "batt_chem": v_batt_chem,
        "batt_soc_model": v_batt_soc_model, "batt_soc_curve_csv": v_batt_soc_curve_csv,
        "batt_soc_bp": v_batt_soc_bp, "batt_ocv_cell_bp": v_batt_ocv_cell_bp,
        "batt_r_scale_bp": v_batt_r_scale_bp,
        "batt_series": v_batt_series, "batt_parallel": v_batt_parallel,
        "batt_cells_series": v_batt_cells_series, "batt_cells_parallel": v_batt_cells_parallel,
        "batt_pack_weight": v_batt_pack_weight, "batt_cell_weight": v_batt_cell_weight,
        "motor_kv": v_motor_kv, "motor_i0": v_motor_i0, "motor_v0": v_motor_v0,
        "motor_rated_v": v_motor_rated_v, "motor_r": v_motor_r,
        "motor_imax": v_motor_imax, "motor_pmax": v_motor_pmax,
        "motor_pole_count": v_motor_pole_count, "motor_weight": v_motor_weight,
        "motor_size": v_motor_size,
        "esc_voltage_rating": v_esc_voltage_rating, "esc_cont_current": v_esc_cont_current,
        "esc_max_current": v_esc_max_current, "esc_idle_current": v_esc_idle_current,
        "esc_r": v_esc_r, "esc_weight": v_esc_weight,
        "avionics_voltage_tree": v_avionics_voltage_tree,
        "prop_d": v_prop_d, "prop_pitch": v_prop_pitch, "prop_blades": v_prop_blades,
        "prop_max_rpm": v_prop_max_rpm, "prop_max_thrust": v_prop_max_thrust,
        "prop_table": v_prop_table, "prop_tconst": v_prop_tconst,
        "prop_pconst": v_prop_pconst, "prop_weight": v_prop_weight,
        "mission": v_mission, "alt": v_alt, "temp": v_temp,
        "press": v_press, "wind": v_wind, "wind_dir": v_wind_dir,
        "course_deg": v_course_deg, "climb_rate": v_climb_rate,
        "descent_rate": v_descent_rate, "reserve_percent": v_reserve_percent,
        "rth_reserve_Wh": v_rth_reserve_Wh, "diversion_reserve_Wh": v_div_reserve_Wh,
        "transient_dt_s": v_transient_dt_s, "max_accel_mps2": v_max_accel_mps2,
        "max_decel_mps2": v_max_decel_mps2, "decel_regen_eff": v_decel_regen_eff,
        "inflow_map_enabled": v_inflow_map_enabled, "inflow_mu_bp": v_inflow_mu_bp,
        "inflow_eff_bp": v_inflow_eff_bp,
        "orientation": v_orientation,
        "max_speed_plot": v_max_speed_plot,
    }

    # ================================================================== #
    #  RIGHT PANEL: output notebooks                                      #
    # ================================================================== #
    right = ttk.Frame(main)
    right.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
    right.columnconfigure(0, weight=1)
    right.rowconfigure(0, weight=2)
    right.rowconfigure(1, weight=1)

    display_nb = ttk.Notebook(right)
    display_nb.grid(row=0, column=0, sticky="nsew")

    tab_plot_out         = ttk.Frame(display_nb, padding=0)
    tab_status_out       = ttk.Frame(display_nb, padding=0)
    tab_metrics_out      = ttk.Frame(display_nb, padding=0)
    tab_mission_plots_out= ttk.Frame(display_nb, padding=0)
    for t in (tab_plot_out, tab_status_out, tab_metrics_out, tab_mission_plots_out):
        t.columnconfigure(0, weight=1); t.rowconfigure(0, weight=1)
    display_nb.add(tab_plot_out,          text="Plots")
    display_nb.add(tab_status_out,        text="Status")
    display_nb.add(tab_metrics_out,       text="Metrics")
    display_nb.add(tab_mission_plots_out, text="Mission Plots")
    tab_weight_budget_out = ttk.Frame(display_nb, padding=0)
    tab_weight_budget_out.columnconfigure(0, weight=1)
    tab_weight_budget_out.rowconfigure(0, weight=1)
    display_nb.add(tab_weight_budget_out, text="Weight Budget")

    # ---- Airframe Diagram tab ----------------------------------------
    # A plan-view sketch drawn from the entered dimensions. Its job is to make
    # propeller overlap and tip clearance obvious at a glance, which numbers
    # in a table do not.
    tab_airframe_diagram_out = ttk.Frame(display_nb, padding=0)
    tab_airframe_diagram_out.columnconfigure(0, weight=1)
    tab_airframe_diagram_out.rowconfigure(0, weight=1)
    display_nb.add(tab_airframe_diagram_out, text="Airframe Diagram")

    ad_frame = ttk.LabelFrame(tab_airframe_diagram_out,
                              text="Plan View (to scale)", padding=4)
    ad_frame.grid(row=0, column=0, sticky="nsew")
    ad_frame.columnconfigure(0, weight=1)
    ad_frame.rowconfigure(0, weight=1)
    ad_holder = ttk.Frame(ad_frame)
    ad_holder.grid(row=0, column=0, sticky="nsew")
    ad_holder.columnconfigure(0, weight=1)
    ad_holder.rowconfigure(0, weight=1)
    ad_placeholder = ttk.Label(
        ad_holder,
        text="Press Run Single-Point to draw the airframe.",
        foreground="#888888")
    ad_placeholder.grid(row=0, column=0)
    _ad_canvas = {"widget": None}

    # ================================================================
    # SENSITIVITY  and  CONFIG COMPARISON tabs
    # ================================================================
    # Both work off the configuration from the most recent run, so they never
    # re-read the input fields and cannot disagree with the numbers on screen.

    tab_sensitivity_out = ttk.Frame(display_nb, padding=0)
    tab_sensitivity_out.columnconfigure(0, weight=1)
    tab_sensitivity_out.rowconfigure(1, weight=1)
    display_nb.add(tab_sensitivity_out, text="Sensitivity")

    sens_bar = ttk.Frame(tab_sensitivity_out, padding=(6, 6, 6, 0))
    sens_bar.grid(row=0, column=0, sticky="ew")
    ttk.Label(sens_bar, text="Output:").pack(side="left")
    v_sens_metric = tk.StringVar(value="Flight time (min)")
    ttk.Combobox(sens_bar, textvariable=v_sens_metric, state="readonly", width=22,
                 values=["Flight time (min)", "Range (km)",
                         "Total power (W)"]).pack(side="left", padx=(4, 10))
    ttk.Label(sens_bar, text="Vary each input by ±10% and ±20%",
              foreground="#666666", font=("TkDefaultFont", 8)).pack(side="left")

    sens_body = ttk.Frame(tab_sensitivity_out, padding=4)
    sens_body.grid(row=1, column=0, sticky="nsew")
    sens_body.columnconfigure(0, weight=1)
    sens_body.rowconfigure(0, weight=1)
    sens_body.rowconfigure(1, weight=1)

    _sens_cols = ("param", "m20", "m10", "base", "p10", "p20", "span")
    sens_tv = ttk.Treeview(sens_body, columns=_sens_cols, show="headings", height=8)
    for col, heading, width in [
            ("param", "Input", 190), ("m20", "-20%", 85), ("m10", "-10%", 85),
            ("base", "baseline", 85), ("p10", "+10%", 85), ("p20", "+20%", 85),
            ("span", "swing", 95)]:
        sens_tv.heading(col, text=heading)
        sens_tv.column(col, width=width,
                       anchor="w" if col == "param" else "center", stretch=True)
    sens_tv.grid(row=0, column=0, sticky="nsew")
    sens_tv.tag_configure("strong", font=("TkDefaultFont", 9, "bold"))

    sens_plot = ttk.Frame(sens_body)
    sens_plot.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
    sens_plot.columnconfigure(0, weight=1)
    sens_plot.rowconfigure(0, weight=1)
    sens_ph = ttk.Label(sens_plot,
                        text="Run a single point, then press Run Sensitivity.",
                        foreground="#888888")
    sens_ph.grid(row=0, column=0)
    _sens_canvas = {"widget": None}

    def _sens_levers():
        """
        Inputs worth testing, with a mutator that scales each on a copy.

        Only quantities a designer can actually trade are listed — swapping a
        battery or a prop is a real decision; the air density is not.
        """
        def _scale_capacity(cfg, f):
            cfg.battery.capacity_mAh *= f
            cfg.battery.capacity_Ah *= f          # Wh is derived from this

        return [
            ("All-up weight",      lambda c, f: setattr(c, "drone_weight_g", c.drone_weight_g * f)),
            ("Battery capacity",   _scale_capacity),
            ("Motor Kv",           lambda c, f: setattr(c.motor, "kv", (c.motor.kv or 0) * f)),
            ("Motor resistance",   lambda c, f: setattr(c.motor, "resistance", c.motor.resistance * f)),
            ("Prop diameter",      lambda c, f: setattr(c.propeller, "diameter_in", c.propeller.diameter_in * f)),
            ("Parasite area",      lambda c, f: setattr(c, "parasite_area", (c.parasite_area or 0.0) * f)),
            ("Cruise speed",       lambda c, f: setattr(c, "cruise_speed", (c.cruise_speed or 0.0) * f)),
            ("Avionics draw",      lambda c, f: setattr(c, "periph_current", (c.periph_current or 0.0) * f)),
        ]

    def run_sensitivity():
        if not _last_run.get("drone"):
            messagebox.showinfo("Sensitivity", "Run a single point first.")
            return
        base_cfg = _last_run["drone"]
        speed = float(_last_run.get("speed", base_cfg.cruise_speed or 0.0))
        orient = _last_run.get("orientation", "forward")
        choice = v_sens_metric.get()

        def evaluate(cfg):
            cfg = base_cfg if cfg is None else cfg
            spd = float(getattr(cfg, "cruise_speed", speed) or speed)
            try:
                if choice.startswith("Flight time"):
                    return estimate_flight_time_minutes(cfg, spd, orientation=orient)
                if choice.startswith("Range"):
                    minutes = estimate_flight_time_minutes(cfg, spd, orientation=orient)
                    m = compute_operating_metrics(cfg, spd, orient)
                    gs = float(m.get("groundspeed_mps", spd))
                    return minutes * 60.0 * gs / 1000.0
                return float(compute_operating_metrics(cfg, spd, orient)["total_power_W"])
            except Exception:
                return None

        evaluate.base_config = base_cfg
        rows = core.sensitivity_sweep(_sens_levers(), evaluate)

        for iid in sens_tv.get_children():
            sens_tv.delete(iid)
        if not rows:
            messagebox.showinfo("Sensitivity",
                                "The baseline configuration produced no usable result.")
            return

        decimals = 0 if choice.startswith("Total power") else 2
        for i, row in enumerate(rows):
            res = row["results"]

            def cell(f):
                v = res.get(f)
                return "—" if v is None else f"{v:.{decimals}f}"

            sens_tv.insert("", "end", tags=("strong",) if i == 0 else (), values=(
                row["name"], cell(0.8), cell(0.9), f"{row['baseline']:.{decimals}f}",
                cell(1.1), cell(1.2),
                f"{row['span']:.{decimals}f}  ({row['span_pct']:.0f}%)"))

        # Tornado chart: widest swing at the top.
        fig = plt.Figure(figsize=(max(_view["plot_w"] * 0.7, 6.0), 4.2))
        ax = fig.add_subplot(111)
        names = [r["name"] for r in rows][::-1]
        lows = [r["low"] - r["baseline"] for r in rows][::-1]
        highs = [r["high"] - r["baseline"] for r in rows][::-1]
        ypos = range(len(names))
        ax.barh(list(ypos), [h - l for l, h in zip(lows, highs)],
                left=lows, color="#90CAF9", edgecolor="#1565C0")
        ax.axvline(0.0, color="#37474F", linewidth=1.2)
        ax.set_yticks(list(ypos))
        ax.set_yticklabels(names, fontsize=8)
        ax.set_xlabel(f"change in {choice} vs baseline")
        ax.set_title("Sensitivity (±20% on each input)", fontsize=10)
        ax.grid(True, axis="x", linestyle=":", alpha=0.4)
        fig.tight_layout()

        old = _sens_canvas.get("widget")
        if old is not None:
            try:
                old.get_tk_widget().destroy()
            except Exception:
                pass
        sens_ph.grid_remove()
        canvas = FigureCanvasTkAgg(fig, master=sens_plot)
        canvas.draw()
        canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        _sens_canvas["widget"] = canvas

    ttk.Button(sens_bar, text="⚡  Run Sensitivity",
               command=run_sensitivity).pack(side="right")

    # ---------------- Config comparison ----------------
    tab_compare_out = ttk.Frame(display_nb, padding=0)
    tab_compare_out.columnconfigure(0, weight=1)
    tab_compare_out.rowconfigure(1, weight=1)
    display_nb.add(tab_compare_out, text="Compare")

    cmp_bar = ttk.Frame(tab_compare_out, padding=(6, 6, 6, 0))
    cmp_bar.grid(row=0, column=0, sticky="ew")
    v_cmp_baseline = tk.StringVar(value="No baseline pinned")
    ttk.Label(cmp_bar, text="Baseline:").pack(side="left")
    ttk.Label(cmp_bar, textvariable=v_cmp_baseline, foreground="#0B6BCB",
              font=("TkDefaultFont", 9, "bold")).pack(side="left", padx=(4, 12))

    _cmp_cols = ("metric", "base", "curr", "delta", "pct")
    cmp_tv = ttk.Treeview(tab_compare_out, columns=_cmp_cols,
                          show="headings", height=18)
    for col, heading, width in [("metric", "Metric", 220), ("base", "Baseline", 110),
                                ("curr", "Current", 110), ("delta", "Change", 110),
                                ("pct", "Change %", 100)]:
        cmp_tv.heading(col, text=heading)
        cmp_tv.column(col, width=width,
                      anchor="w" if col == "metric" else "center", stretch=True)
    cmp_tv.grid(row=1, column=0, sticky="nsew", padx=4, pady=4)
    cmp_tv.tag_configure("better", foreground="#1B5E20")
    cmp_tv.tag_configure("worse", foreground="#B71C1C")
    cmp_tv.tag_configure("flat", foreground="#616161")

    _cmp_state = {"baseline": None, "label": None}

    # (key, label, decimals, higher_is_better)
    _CMP_KEYS = [
        ("flight_time_min",   "Flight time (min)",     2,  1),
        ("flight_range_km",   "Range (km)",            2,  1),
        ("total_power_W",     "Total power (W)",       1, -1),
        ("pack_current_A",    "Pack current (A)",      2, -1),
        ("v_load_V",          "Loaded voltage (V)",    2,  1),
        ("thrust_total_N",    "Total thrust (N)",      2,  1),
        ("hover_efficiency_gW", "Hover efficiency (g/W)", 2, 1),
        ("figure_of_merit",   "Figure of merit",       3,  1),
        ("disk_loading_N_m2", "Disk loading (N/m²)",   1, -1),
        ("motor_temp_est_C",  "Motor temp (°C)",       1, -1),
        ("reserve_margin_Wh", "Reserve margin (Wh)",   2,  1),
    ]

    def _current_comparison_metrics():
        """Flatten the last run into the flat dict the comparison expects."""
        if not _last_run.get("drone"):
            return None
        cfg = _last_run["drone"]
        speed = float(_last_run.get("speed", cfg.cruise_speed or 0.0))
        orient = _last_run.get("orientation", "forward")
        m = dict(_last_run.get("metrics") or {})
        try:
            minutes = estimate_flight_time_minutes(cfg, speed, orientation=orient)
            m["flight_time_min"] = minutes
            m["flight_range_km"] = minutes * 60.0 * float(
                m.get("groundspeed_mps", speed)) / 1000.0
        except Exception:
            pass
        return m

    def pin_comparison_baseline():
        metrics = _current_comparison_metrics()
        if not metrics:
            messagebox.showinfo("Compare", "Run a single point first.")
            return
        _cmp_state["baseline"] = metrics
        label = str(v_loaded_cfg.get())
        _cmp_state["label"] = label
        v_cmp_baseline.set(f"{label}  (pinned)")
        refresh_comparison()

    def clear_comparison_baseline():
        _cmp_state["baseline"] = None
        _cmp_state["label"] = None
        v_cmp_baseline.set("No baseline pinned")
        for iid in cmp_tv.get_children():
            cmp_tv.delete(iid)

    def refresh_comparison():
        """Redraw the delta table. Called after every run."""
        for iid in cmp_tv.get_children():
            cmp_tv.delete(iid)
        base = _cmp_state.get("baseline")
        if not base:
            return
        current = _current_comparison_metrics()
        if not current:
            return
        rows = core.compare_metric_sets(
            base, current, [(k, lbl, d) for k, lbl, d, _ in _CMP_KEYS])
        better_map = {k: s for k, _, _, s in _CMP_KEYS}
        for row in rows:
            if not row["comparable"]:
                cmp_tv.insert("", "end", tags=("flat",), values=(
                    row["label"], "—", "—", "—", "—"))
                continue
            d = row["decimals"]
            sign = better_map.get(row["key"], 1) * row["direction"]
            tag = "better" if sign > 0 else ("worse" if sign < 0 else "flat")
            cmp_tv.insert("", "end", tags=(tag,), values=(
                row["label"],
                f"{row['baseline']:.{d}f}", f"{row['current']:.{d}f}",
                core.format_delta(row["delta"], d),
                core.format_delta(row["delta_pct"], 1) + " %"
                if row["delta_pct"] is not None else "—"))

    ttk.Button(cmp_bar, text="📌  Pin Current as Baseline",
               command=pin_comparison_baseline).pack(side="right")
    ttk.Button(cmp_bar, text="✖  Clear Baseline",
               command=clear_comparison_baseline).pack(side="right", padx=(0, 6))

    def _refresh_airframe_diagram(cfg_obj):
        """Redraw the plan view for the configuration just simulated."""
        try:
            fig = make_airframe_diagram_figure(
                cfg_obj, figsize=(max(_view["plot_w"] * 0.7, 6.0), 6.5))
        except Exception as exc:
            ad_placeholder.configure(
                text=f"Could not draw the airframe: {exc}")
            ad_placeholder.grid(row=0, column=0)
            return
        old = _ad_canvas.get("widget")
        if old is not None:
            try:
                old.get_tk_widget().destroy()
            except Exception:
                pass
        ad_placeholder.grid_remove()
        canvas = FigureCanvasTkAgg(fig, master=ad_holder)
        canvas.draw()
        canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        _ad_canvas["widget"] = canvas


    # ---- Plots panel ----
    plot_frame = ttk.LabelFrame(tab_plot_out, text="Performance Plots", padding=4)
    plot_frame.grid(row=0, column=0, sticky="nsew")
    plot_frame.columnconfigure(0, weight=1)
    plot_frame.rowconfigure(0, weight=1)

    plot_canvas = tk.Canvas(plot_frame, highlightthickness=0)
    plot_canvas.grid(row=0, column=0, sticky="nsew")
    plot_scroll = ttk.Scrollbar(plot_frame, orient="vertical", command=plot_canvas.yview)
    plot_scroll.grid(row=0, column=1, sticky="ns")
    plot_canvas.configure(yscrollcommand=plot_scroll.set)

    plot_inner = ttk.Frame(plot_canvas)
    plot_inner_id = plot_canvas.create_window((0, 0), window=plot_inner, anchor="nw")
    plot_inner.columnconfigure(0, weight=1)

    def _update_plot_scrollregion(event=None):
        plot_canvas.configure(scrollregion=plot_canvas.bbox("all"))

    def _match_plot_inner_width(event):
        plot_canvas.itemconfigure(plot_inner_id, width=event.width)

    plot_inner.bind("<Configure>", lambda event: _update_plot_scrollregion(event))
    plot_canvas.bind("<Configure>", lambda event: _match_plot_inner_width(event))
    plot_inner.bind("<Enter>", lambda _: plot_canvas.bind_all("<MouseWheel>", _on_plot_mousewheel))
    plot_inner.bind("<Leave>", lambda _: plot_canvas.unbind_all("<MouseWheel>"))

    _plot_canvases = []  # Track multiple figure canvases
    _current_plot_figs = []  # Track multiple figures

    def _on_plot_mousewheel(evt):
        plot_canvas.yview_scroll(int(-1 * (evt.delta / 120)), "units")
        return "break"

    plot_canvas.bind("<Enter>", lambda _: plot_canvas.bind_all("<MouseWheel>", _on_plot_mousewheel))
    plot_canvas.bind("<Leave>", lambda _: plot_canvas.unbind_all("<MouseWheel>"))

    def _show_figure(fig_or_figs):
        """Display one or more figures in the scrollable plot area."""
        # Clear previous figures
        for canvas in _plot_canvases:
            try:
                canvas.get_tk_widget().destroy()
            except:
                pass
        for fig in _current_plot_figs:
            try:
                plt.close(fig)
            except:
                pass
        _plot_canvases.clear()
        _current_plot_figs.clear()
        
        # Handle both single figure and list of figures
        figures = fig_or_figs if isinstance(fig_or_figs, list) else [fig_or_figs]
        
        for idx, fig in enumerate(figures):
            if fig is None:
                continue
            fc = FigureCanvasTkAgg(fig, master=plot_inner)
            fc.draw()
            fc.get_tk_widget().grid(row=idx, column=0, sticky="nsew")
            _plot_canvases.append(fc)
            _current_plot_figs.append(fig)
        
        # Update scroll region
        plot_inner.update_idletasks()
        plot_canvas.configure(scrollregion=plot_canvas.bbox("all"))


    # Weight Budget panel
    _wb_canvas_ref = [None]
    def _draw_weight_chart(rows):
        data_rows = [r for r in rows if r[0] != "TOTAL"]
        if not data_rows: return
        labels = [r[0] for r in data_rows]
        totals = [r[3] for r in data_rows]
        grand  = sum(totals)
        COLORS = ["#2E75B6","#ED7D31","#A9D18E","#FFC000","#5B9BD5","#FF7F7F"]
        fig, axes = plt.subplots(1, 2, figsize=(7, max(3, len(labels)*0.6+1)))
        fig.patch.set_facecolor("white")
        ax = axes[0]
        left_ = 0.0
        pcts = [t/grand*100 if grand>0 else 0 for t in totals]
        for i,(lbl,pct) in enumerate(zip(labels,pcts)):
            ax.barh(0, pct, left=left_, color=COLORS[i%len(COLORS)],
                    label=lbl, edgecolor="white", linewidth=0.5)
            if pct > 5:
                ax.text(left_+pct/2, 0, f"{pct:.0f}%",
                        ha="center", va="center", fontsize=7.5, color="white")
            left_ += pct
        ax.set_xlim(0,100); ax.set_yticks([])
        ax.set_xlabel("% of total weight"); ax.set_title("Weight Distribution")
        ax.legend(loc="upper center",bbox_to_anchor=(0.5,-0.18),ncol=2,fontsize=7,frameon=False)
        ax.grid(axis="x",alpha=0.3)
        ax2 = axes[1]
        _,_, ats = ax2.pie(totals,labels=None,autopct="%1.0f%%",
            colors=COLORS[:len(labels)],startangle=90,pctdistance=0.75,
            wedgeprops=dict(edgecolor="white",linewidth=0.8))
        for at in ats: at.set_fontsize(7)
        ax2.set_title(f"Total: {grand:.0f} g")
        ax2.legend(labels,loc="lower center",bbox_to_anchor=(0.5,-0.22),ncol=2,fontsize=7,frameon=False)
        fig.tight_layout()
        if _wb_canvas_ref[0]:
            try: _wb_canvas_ref[0].get_tk_widget().destroy()
            except Exception: pass
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        fc = FigureCanvasTkAgg(fig, master=wb_right)
        fc.draw(); fc.get_tk_widget().grid(row=0,column=0,sticky="nsew")
        _wb_canvas_ref[0] = fc; plt.close(fig)
    # Weight Budget layout
    wb_outer = ttk.Frame(tab_weight_budget_out, padding=4)
    wb_outer.grid(row=0, column=0, sticky="nsew")
    wb_outer.columnconfigure(0, weight=1); wb_outer.columnconfigure(1, weight=3)
    wb_outer.rowconfigure(0, weight=1)
    wb_left = ttk.LabelFrame(wb_outer, text="Component Weights", padding=4)
    wb_left.grid(row=0, column=0, sticky="nsew", padx=(0,6))
    wb_left.columnconfigure(0, weight=1); wb_left.rowconfigure(1, weight=1)
    wb_ph = ttk.Label(wb_left, text="Run a simulation to populate.",
                     foreground="#888888", wraplength=200)
    wb_ph.grid(row=0, column=0, columnspan=2, sticky="w", padx=4, pady=(0,4))
    _wb_tv_cols = ("component","unit_w","count","total_w","pct")
    wb_tv = ttk.Treeview(wb_left, columns=_wb_tv_cols, show="headings", height=10)
    for col, heading, width in [
        ("component","Component",160),("unit_w","Unit (g)",70),
        ("count","Qty",40),("total_w","Total (g)",75),("pct","% of Total",70)]:
        wb_tv.heading(col, text=heading)
        wb_tv.column(col, width=width,
                     anchor="w" if col=="component" else "center", stretch=True)
    wb_tv.grid(row=1, column=0, sticky="nsew")
    wb_sb = ttk.Scrollbar(wb_left, orient="vertical", command=wb_tv.yview)
    wb_sb.grid(row=1, column=1, sticky="ns")
    wb_tv.configure(yscrollcommand=wb_sb.set)
    wb_tv.tag_configure("total", font=("TkDefaultFont", 9, "bold"))
    wb_right = ttk.LabelFrame(wb_outer, text="Weight Distribution", padding=4)
    wb_right.grid(row=0, column=1, sticky="nsew")
    wb_right.columnconfigure(0, weight=1); wb_right.rowconfigure(0, weight=1)
    def update_weight_budget(cfg):
        rows = _extract_weight_budget(cfg)
        for iid in wb_tv.get_children(): wb_tv.delete(iid)
        total_g = rows[-1][3] if rows else 1.0
        for label, unit_w, count, total_w in rows[:-1]:
            pct = f"{total_w/total_g*100:.1f}%" if total_g > 0 else "0%"
            wb_tv.insert("","end",values=(label,f"{unit_w:.1f}",
                         str(count),f"{total_w:.1f}",pct))
        if rows:
            label,unit_w,count,total_w = rows[-1]
            wb_tv.insert("","end",values=(label,"","",
                         f"{total_w:.1f}","100%"),tags=("total",))
        _draw_weight_chart(rows)

    # ---- Status panel ----
    sty = ttk.Style()
    try: sty.theme_use(sty.theme_use())
    except Exception: pass

    def _status_color_tag(value, limit, kind):
        try: v = float(value); L = float(limit)
        except Exception: return "na"
        if kind == "max": return "bad" if v > L else ("warn" if v > 0.9*L else "ok")
        else:             return "bad" if v < L else ("warn" if v < 1.1*L else "ok")

    def _make_status_tv(parent, title):
        lf = ttk.LabelFrame(parent, text=title, padding=4)
        lf.pack(fill="both", expand=True, padx=4, pady=4)
        cols = ("metric","value","limit","note")
        tv = ttk.Treeview(lf, columns=cols, show="headings", height=6)
        for c, w in zip(cols, (200, 120, 120, 220)):
            tv.heading(c, text=c.capitalize())
            tv.column(c, width=w, anchor="w" if c in ("metric","note") else "center")
        tv.pack(fill="both", expand=True)
        for tag, bg in [("ok","#d9f2d9"),("warn","#fff2cc"),("bad","#f8d7da"),("na","#efefef")]:
            tv.tag_configure(tag, background=bg)
        return tv

    status_scroll = ttk.Frame(tab_status_out)
    status_scroll.grid(row=0, column=0, sticky="nsew")
    status_scroll.columnconfigure(0, weight=1)

    batt_table    = _make_status_tv(status_scroll, "Battery Status")
    motor_table   = _make_status_tv(status_scroll, "Motor / ESC Status")
    prop_table_tv = _make_status_tv(status_scroll, "Propeller Status")

    def _clear_status_tables():
        for tv in (batt_table, motor_table, prop_table_tv):
            for iid in tv.get_children(): tv.delete(iid)

    def _insert_status_row(tv, metric, val_str, lim_str, tag, note=""):
        tv.insert("", "end", values=(metric, val_str, lim_str, note), tags=(tag,))

    def update_status_tables_from_metrics(config: DroneConfig, metrics: dict):
        _clear_status_tables()
        batt = config.battery
        nm   = max(int(config.num_motors), 1)
        g0   = 9.80665

        Ipack    = float(metrics.get("pack_current_A", 0.0))
        Vload    = float(metrics.get("v_load_V",        0.0))
        Ptot     = float(metrics.get("total_power_W",   0.0))
        Pmotor   = float(metrics.get("motor_power_W",   0.0))
        Periph   = float(metrics.get("periph_power_W",  0.0))
        P_esc    = float(metrics.get("esc_loss_W",       0.0))
        Iesc     = float(metrics.get("motor_I_per_esc_A",0.0))
        tip_mach = metrics.get("tip_mach", None)

        # ── Battery Status ────────────────────────────────────────────────
        Vmin  = float(getattr(batt, "vmin_pack", 0.0))
        Vmax  = float(getattr(batt, "vmax_pack", 0.0))
        Vdrop = Vmax - Vload
        vsag_pct = Vdrop / max(Vmax, 1e-9) * 100
        tag_v = _status_color_tag(Vload, Vmin, "min") if Vmin > 0 else "na"
        _insert_status_row(batt_table, "Pack voltage (loaded)",
            f"{Vload:.2f} V", f">= {Vmin:.2f} V", tag_v,
            f"Sag: {Vdrop:.2f} V ({vsag_pct:.1f}%)")

        if getattr(config, "max_tilt_deg", None) is not None and metrics.get("tilt_required_deg") is not None:
            tilt_req = float(metrics["tilt_required_deg"])
            tilt_lim = float(config.max_tilt_deg)
            _insert_status_row(batt_table, "Tilt req vs max",
                f"{tilt_req:.1f}° / {tilt_lim:.1f}°", f"<= {tilt_lim:.1f}°",
                _status_color_tag(tilt_req, tilt_lim, "max"))

        Icont  = float(getattr(batt, "discharge_cont_A", float("inf")))
        Imax_b = float(getattr(batt, "discharge_max_A",  float("inf")))
        if math.isfinite(Icont):
            _insert_status_row(batt_table, "Pack current vs cont rating",
                f"{Ipack:.2f} A", f"<= {Icont:.2f} A",
                _status_color_tag(Ipack, Icont, "max"))
        else:
            _insert_status_row(batt_table, "Pack current", f"{Ipack:.2f} A", "cont: n/a", "na")
        if math.isfinite(Imax_b):
            _insert_status_row(batt_table, "Pack current vs max rating",
                f"{Ipack:.2f} A", f"<= {Imax_b:.2f} A",
                _status_color_tag(Ipack, Imax_b, "max"))

        cap_Ah = float(getattr(batt, "capacity_Ah", 0.0))
        if cap_Ah > 0:
            c_rate = Ipack / cap_Ah
            c_lim  = Icont / cap_Ah if math.isfinite(Icont) else 999
            _insert_status_row(batt_table, "Discharge C-rate",
                f"{c_rate:.2f} C", f"<= {c_lim:.1f} C (cont)",
                _status_color_tag(c_rate, c_lim, "max"),
                "Continuous C-rate at this speed")

        _insert_status_row(batt_table, "Total electrical power",
            f"{Ptot:.0f} W", "—", "na")

        if Ptot > 0:
            _insert_status_row(batt_table, "Power split (Motor/ESC/Av)",
                f"{Pmotor/Ptot*100:.0f}% / {P_esc/Ptot*100:.0f}% / {Periph/Ptot*100:.0f}%",
                "—", "na", "Motor propulsion / ESC losses / Avionics")

        rsv_margin = float(metrics.get("reserve_margin_Wh", float("nan")))
        if math.isfinite(rsv_margin):
            rsv_tag = "bad" if metrics.get("reserve_breached", False) else (
                      "warn" if rsv_margin < 10.0 else "ok")
            _insert_status_row(batt_table, "Energy reserve margin",
                f"{rsv_margin:+.1f} Wh", ">= 0 Wh", rsv_tag,
                "VIOLATION" if metrics.get("reserve_breached", False) else "OK")

        # ── Motor / ESC Status ────────────────────────────────────────────
        Pmotor_pm = Pmotor / nm
        if getattr(config.motor, "max_power", None) is not None:
            pmax_pm = float(config.motor.max_power)
            _insert_status_row(motor_table, "Motor power / motor",
                f"{Pmotor_pm:.0f} W", f"<= {pmax_pm:.0f} W",
                _status_color_tag(Pmotor_pm, pmax_pm, "max"))
        else:
            _insert_status_row(motor_table, "Motor power / motor",
                f"{Pmotor_pm:.0f} W", "—", "na")

        if getattr(config.motor, "max_current", None) is not None:
            imax_m = float(config.motor.max_current)
            _insert_status_row(motor_table, "Motor current / motor (est)",
                f"{Iesc:.2f} A", f"<= {imax_m:.2f} A",
                _status_color_tag(Iesc, imax_m, "max"))

        weight_N = config.drone_weight_g * g0 / 1000.0
        hover_T  = float(metrics.get("hover_thrust_total_N", weight_N))
        twr_hover = hover_T / max(weight_N, 1e-9)
        _insert_status_row(motor_table, "Hover TWR (total thrust / weight)",
            f"{twr_hover:.2f}:1", ">= 1.5:1",
            "ok" if twr_hover >= 2.0 else ("warn" if twr_hover >= 1.5 else "bad"),
            "< 1.5 is dangerously marginal; > 2.0 is recommended")

        from_physics = float(metrics.get("thrust_total_N", weight_N))
        twr_fwd = from_physics / max(weight_N, 1e-9)
        _insert_status_row(motor_table, "Thrust-to-weight (this speed)",
            f"{twr_fwd:.2f}:1", ">= 1.0:1",
            "ok" if twr_fwd >= 1.0 else "bad",
            "Drops in forward flight due to drag")

        max_pay_g = (hover_T / g0 - config.drone_weight_g / 1000.0) * 1000.0
        _insert_status_row(motor_table, "Max additional payload",
            f"{max_pay_g:.0f} g", ">= 0 g",
            "ok" if max_pay_g >= 0 else "bad",
            "Remaining thrust capacity above current AUW")

        T_motor = float(metrics.get("motor_temp_est_C", float("nan")))
        T_head  = float(metrics.get("motor_thermal_headroom_C", float("nan")))
        if math.isfinite(T_motor):
            _insert_status_row(motor_table, "Motor temperature (est)",
                f"{T_motor:.1f} °C", "<= 100 °C",
                "ok" if T_motor < 80 else ("warn" if T_motor < 100 else "bad"),
                f"Headroom: {T_head:.1f} °C  |  {metrics.get('thermal_status','n/a')}")
        T_esc  = float(metrics.get("esc_temp_est_C", float("nan")))
        T_batt = float(metrics.get("battery_temp_est_C", float("nan")))
        if math.isfinite(T_esc):
            _insert_status_row(motor_table, "ESC temperature (est)",
                f"{T_esc:.1f} °C", "<= 90 °C",
                "ok" if T_esc < 70 else ("warn" if T_esc < 90 else "bad"))
        if math.isfinite(T_batt):
            _insert_status_row(motor_table, "Battery temperature (est)",
                f"{T_batt:.1f} °C", "<= 55 °C",
                "ok" if T_batt < 40 else ("warn" if T_batt < 55 else "bad"))

        esc = getattr(config, "esc", None)
        if esc is not None:
            # ESCConfig.voltage_rating is a CELL COUNT (an "S" rating), not volts.
            # Compare against the pack's series-cell count, never a voltage.
            esc_S  = int(esc.voltage_rating)
            pack_S = int(batt.series_cells)
            _insert_status_row(motor_table, "ESC voltage rating",
                f"{esc_S}S  ({esc_S * batt.operating_voltage_max:.1f} V max)",
                f">= {pack_S}S  ({batt.vmax_pack:.1f} V)",
                "ok" if esc_S >= pack_S else "bad",
                "ESC S-rating must meet or exceed the pack series count")
            _insert_status_row(motor_table, "ESC current vs cont",
                f"{Iesc:.2f} A", f"<= {float(esc.continuous_rating_A):.2f} A",
                _status_color_tag(Iesc, float(esc.continuous_rating_A), "max"))
            _insert_status_row(motor_table, "ESC current vs max",
                f"{Iesc:.2f} A", f"<= {float(esc.max_current_A):.2f} A",
                _status_color_tag(Iesc, float(esc.max_current_A), "max"))
            _insert_status_row(motor_table, "ESC total loss",
                f"{P_esc:.1f} W", "—", "na")

        # ── Propeller / Hover Status ──────────────────────────────────────
        thrust_pm_N = float(metrics.get("thrust_per_motor_N", 0.0))
        thrust_pm_g = thrust_pm_N * 1000.0 / g0
        max_thr_g   = float(getattr(config.propeller, "max_thrust_g", 0.0))
        if max_thr_g > 0:
            _insert_status_row(prop_table_tv, "Thrust / motor",
                f"{thrust_pm_g:.0f} g", f"<= {max_thr_g:.0f} g",
                _status_color_tag(thrust_pm_g, max_thr_g, "max"),
                f"Margin: {max_thr_g-thrust_pm_g:.0f} g")
        else:
            _insert_status_row(prop_table_tv, "Thrust / motor",
                f"{thrust_pm_g:.0f} g", "—", "na")

        rpm = metrics.get("prop_rpm", None)
        mr  = getattr(config.propeller, "max_rpm", None)
        if rpm is not None and float(rpm) > 0:
            rpm_f = float(rpm)
            if mr is not None and float(mr) > 0:
                _insert_status_row(prop_table_tv, "Prop RPM (est)",
                    f"{rpm_f:.0f} rpm", f"<= {float(mr):.0f} rpm",
                    _status_color_tag(rpm_f, float(mr), "max"))
            else:
                _insert_status_row(prop_table_tv, "Prop RPM (est)",
                    f"{rpm_f:.0f} rpm", "—", "na")

        if tip_mach is not None:
            tm = float(tip_mach)
            _insert_status_row(prop_table_tv, "Tip Mach",
                f"{tm:.3f}", "<= 0.60",
                _status_color_tag(tm, 0.60, "max"),
                "Above 0.6 → significant aeroacoustic noise & efficiency loss")

        dl = float(metrics.get("disk_loading_N_m2", float("nan")))
        if math.isfinite(dl):
            _insert_status_row(prop_table_tv, "Disk loading",
                f"{dl:.1f} N/m²", "<= 200 N/m²",
                "ok" if dl <= 100 else ("warn" if dl <= 200 else "bad"),
                "High DL → less efficient hover; < 100 N/m² is good")

        he = float(metrics.get("hover_efficiency_gW", float("nan")))
        if math.isfinite(he):
            _insert_status_row(prop_table_tv, "Hover efficiency",
                f"{he:.2f} g/W", ">= 5 g/W",
                "ok" if he >= 7 else ("warn" if he >= 4 else "bad"),
                "Higher = more lift per watt; typical 5–12 g/W")

        fm = float(metrics.get("figure_of_merit", float("nan")))
        if math.isfinite(fm):
            _insert_status_row(prop_table_tv, "Figure of merit",
                f"{fm:.3f}", ">= 0.65",
                "ok" if fm >= 0.70 else ("warn" if fm >= 0.55 else "bad"),
                "Ideal = 1.0; well-designed prop 0.7–0.8")

        v_wind_max = float(metrics.get("hover_wind_resistance_mps", float("nan")))
        if math.isfinite(v_wind_max):
            _insert_status_row(prop_table_tv, "Max hover wind resistance",
                f"{v_wind_max:.1f} m/s  ({v_wind_max*1.944:.1f} kt)",
                ">= 5 m/s",
                "ok" if v_wind_max >= 8 else ("warn" if v_wind_max >= 4 else "bad"),
                "Maximum wind the vehicle can hold position against")

        sol = float(metrics.get("prop_solidity_sigma", float("nan")))
        if math.isfinite(sol):
            _insert_status_row(prop_table_tv, "Prop solidity σ",
                f"{sol:.3f}", "0.05 – 0.15",
                "ok" if 0.05 <= sol <= 0.15 else "warn",
                "< 0.05 → lightly loaded blades; > 0.15 → high profile losses")

    # ---- Metrics panel ----
    metrics_container = ttk.Frame(tab_metrics_out, padding=4)
    metrics_container.grid(row=0, column=0, sticky="nsew")
    metrics_container.columnconfigure(0, weight=1)
    metrics_container.rowconfigure(0, weight=1)

    metrics_tv = ttk.Treeview(metrics_container, columns=("metric","value"),
                               show="headings", height=28)
    metrics_tv.heading("metric", text="Metric")
    metrics_tv.heading("value",  text="Value")
    metrics_tv.column("metric", width=280, anchor="w")
    metrics_tv.column("value",  width=260, anchor="w")
    metrics_sb = ttk.Scrollbar(metrics_container, orient="vertical", command=metrics_tv.yview)
    metrics_tv.configure(yscrollcommand=metrics_sb.set)
    metrics_tv.grid(row=0, column=0, sticky="nsew")
    metrics_sb.grid(row=0, column=1, sticky="ns")
    try:
        ttk.Style().configure("Metrics.Treeview", rowheight=24)
        metrics_tv.configure(style="Metrics.Treeview")
        metrics_tv.tag_configure("section", font=("TkDefaultFont", 10, "bold"))
    except Exception:
        pass

    def _metrics_clear():
        for iid in metrics_tv.get_children(): metrics_tv.delete(iid)
        _metrics_section_stack["current"] = ""

    # ---- Collapsible metrics sections --------------------------------
    # Sections are real parent nodes, so Treeview gives expand/collapse for
    # free. With 100+ rows a flat list is hard to scan; the sections a user
    # cares about stay open and the rest fold away.
    _metrics_section_stack = {"current": ""}
    _metrics_open_state = {}          # section title -> open/closed, remembered

    def _metrics_add_section(title: str):
        """Start a new collapsible section and make later rows its children."""
        node = metrics_tv.insert(
            "", "end", text=str(title), values=(f"── {title} ──", ""),
            tags=("section",), open=_metrics_open_state.get(str(title), True))
        _metrics_section_stack["current"] = node
        return node

    def _metrics_row(metric: str, value: str):
        """Add a row under the current section (or at top level if none)."""
        metrics_tv.insert(_metrics_section_stack["current"], "end",
                          values=(metric, value))

    def _metrics_remember_open_state(_event=None):
        """Persist which sections the user left open across re-runs."""
        for node in metrics_tv.get_children(""):
            title = str(metrics_tv.item(node, "text"))
            if title:
                _metrics_open_state[title] = bool(metrics_tv.item(node, "open"))

    metrics_tv.bind("<<TreeviewOpen>>", _metrics_remember_open_state, add="+")
    metrics_tv.bind("<<TreeviewClose>>", _metrics_remember_open_state, add="+")

    def _metrics_expand_all():
        for node in metrics_tv.get_children(""):
            metrics_tv.item(node, open=True)
        _metrics_remember_open_state()

    def _metrics_collapse_all():
        for node in metrics_tv.get_children(""):
            metrics_tv.item(node, open=False)
        _metrics_remember_open_state()

    def _metrics_add(metric: str, value: str):
        _metrics_row(metric, value)


    def update_metrics_tab(drone: DroneConfig, metrics: dict, speed_mps: float, orientation: str):
        _metrics_clear()
        def fmt(x, nd=2):
            try: return f"{float(x):.{nd}f}"
            except Exception: return "n/a"
        def fmf(x, nd=2):
            """Format float, returning '—' for NaN/None."""
            try:
                v = float(x)
                return "—" if v != v else f"{v:.{nd}f}"
            except Exception: return "—"

        batt = drone.battery
        nm   = max(int(drone.num_motors), 1)
        g0   = 9.80665

        v_load        = float(metrics.get("v_load_V",          float("nan")))
        I_pack        = float(metrics.get("pack_current_A",    float("nan")))
        P_total       = float(metrics.get("total_power_W",     float("nan")))
        P_motor_total = float(metrics.get("motor_power_W",     0.0))
        P_periph      = float(metrics.get("periph_power_W",    0.0))
        P_esc_loss    = float(metrics.get("esc_loss_W",        0.0))

        cap_mAh     = float(batt.capacity_mAh)
        cap_Ah      = cap_mAh / 1000.0 if cap_mAh else 0.0
        usable_frac = max(0.0, min(1.0, float(getattr(batt, "discharge_percent", 100.0)) / 100.0))
        usable_mAh  = cap_mAh * usable_frac
        usable_Wh   = float(batt.capacity_Wh) * usable_frac
        load_C      = (I_pack / cap_Ah) if cap_Ah > 0 else float("nan")

        t_min          = estimate_flight_time_minutes(drone, speed_mps, orientation=orientation)
        groundspeed_mps= float(metrics.get("groundspeed_mps", speed_mps))
        range_m        = (t_min * 60.0) * groundspeed_mps
        range_km       = range_m / 1000.0

        I_motor = float(metrics.get("motor_I_per_esc_A", float("nan")))
        rpm     = metrics.get("prop_rpm", None)
        rpm     = float(rpm) if rpm is not None else float("nan")
        kv      = getattr(drone.motor, "kv", None)
        Rm      = float(getattr(drone.motor, "resistance", 0.0))
        I0      = float(getattr(drone.motor, "idle_current", 0.0))
        Kt      = 60.0 / (2.0 * math.pi * float(kv)) if kv and float(kv) > 0 else float("nan")

        V_emf = float("nan")
        if kv and kv > 0 and rpm == rpm:
            V_emf = rpm / float(kv)
        V_motor = v_load
        if V_emf == V_emf and I_motor == I_motor:
            V_motor = V_emf + I_motor * Rm

        throttle_linear = (V_motor / v_load) if (v_load == v_load and v_load > 1e-6 and V_motor == V_motor) else float("nan")
        if throttle_linear == throttle_linear:
            throttle_linear = max(0.0, min(1.2, throttle_linear))
        throttle_log = float("nan")
        if throttle_linear == throttle_linear:
            throttle_log = math.log10(1 + 9 * max(0.0, min(1.0, throttle_linear)))

        P_elec_motor = (I_motor * v_load) if (I_motor == I_motor and v_load == v_load) else float("nan")
        P_mech_motor = float("nan")
        if V_emf == V_emf and I_motor == I_motor:
            P_mech_motor = V_emf * max(0.0, I_motor - I0)
        motor_eff = (P_mech_motor / P_elec_motor) if (P_mech_motor==P_mech_motor and P_elec_motor==P_elec_motor and P_elec_motor > 0) else float("nan")

        T_est = float(metrics.get("motor_temp_est_C", float("nan")))

        thrust_total_N = float(metrics.get("thrust_total_N",    float("nan")))
        thrust_pm_N    = float(metrics.get("thrust_per_motor_N",float("nan")))
        weight_kg      = float(drone.drone_weight_g) / 1000.0
        twr            = (thrust_total_N / (weight_kg * g0)) if weight_kg > 0 else float("nan")
        thrust_pm_g    = (thrust_pm_N / g0) * 1000.0 if thrust_pm_N == thrust_pm_N else float("nan")
        spec_thrust    = (thrust_pm_g / P_elec_motor) if (P_elec_motor == P_elec_motor and P_elec_motor > 0 and thrust_pm_g == thrust_pm_g) else float("nan")

        drive_g = 0.0
        if getattr(drone.motor, "weight_g", None): drive_g += float(drone.motor.weight_g) * nm
        if getattr(drone, "esc", None) and getattr(drone.esc,"weight_g",None): drive_g += float(drone.esc.weight_g) * nm
        if getattr(drone.propeller, "weight_g", None): drive_g += float(drone.propeller.weight_g) * nm

        D_m     = float(getattr(drone.propeller,"diameter_in",0.0)) * 0.0254
        A_disk  = math.pi * (D_m/2)**2          # per rotor
        A_total = A_disk * nm                    # total disk area

        p2w_Wkg = (P_total / weight_kg) if weight_kg > 0 else float("nan")
        P_out   = max(0.0, P_total - P_esc_loss - P_periph)
        eff_tot = (P_out / P_total) if P_total > 0 else float("nan")
        tilt    = float(metrics.get("tilt_required_deg", float("nan")))

        # ── Battery ───────────────────────────────────────────────────────
        _metrics_add_section("Battery")
        Vmax  = float(getattr(batt,"vmax_pack", 0.0))
        Vmin  = float(getattr(batt,"vmin_pack", 0.0))
        Vsag  = Vmax - v_load
        _metrics_add("Chemistry",              f"{getattr(batt,'chemistry','—') or '—'}")
        _metrics_add("Configuration",          f"{batt.series_cells}S × {batt.parallel_cells}P  ({batt.total_cells} cells total)")
        _metrics_add("Pack Voltage (no load)", f"{fmt(Vmax,2)} V")
        _metrics_add("Pack Voltage (cutoff)",  f"{fmt(Vmin,2)} V")
        _metrics_add("Pack Voltage (loaded)",  f"{fmt(v_load,2)} V  (sag: {fmt(Vsag,2)} V / {fmt(Vsag/max(Vmax,1e-9)*100,1)}%)")
        _metrics_add("Pack Resistance",        f"{fmt(getattr(batt,'pack_resistance',0)*1000,1)} mΩ")
        _metrics_add("Discharge C-rate",       f"{fmt(load_C,2)} C")
        if cap_Ah > 0 and math.isfinite(float(getattr(batt,'discharge_cont_A',float('inf')))):
            c_cont = float(batt.discharge_cont_A) / cap_Ah
            _metrics_add("Cont C-rate limit",  f"{fmt(c_cont,1)} C  ({fmt(float(batt.discharge_cont_A),1)} A)")
        _metrics_add("Total Capacity",         f"{fmt(cap_mAh,0)} mAh  ({fmt(cap_Ah,3)} Ah)")
        _metrics_add("Usable Capacity",        f"{fmt(usable_mAh,0)} mAh  ({fmt(usable_frac*100,0)}% of pack)")
        _metrics_add("Energy (total)",         f"{fmt(float(batt.capacity_Wh),2)} Wh")
        _metrics_add("Energy (usable)",        f"{fmt(usable_Wh,2)} Wh")
        _metrics_add("Energy Density",         f"{fmt(getattr(batt,'energy_density_Wh_per_kg',0),0)} Wh/kg")
        _metrics_add("Battery Weight",         f"{fmt(batt.weight_g,0)} g  ({fmt(batt.weight_g/max(weight_kg*1000,1e-9)*100,1)}% of AUW)")
        _metrics_add("SoC",                    f"{fmt(metrics.get('soc_percent',100.0),1)} %  [{metrics.get('soc_model_source','linear')}]")
        _metrics_add("Pack Current",           f"{fmt(I_pack,2)} A")
        _metrics_add("Flight Time",            f"{fmt(t_min,2)} min  ({fmt(t_min/60,3)} h)")

        # ── Motor @ Operating Point ───────────────────────────────────────
        _metrics_add_section("Motor @ Operating Point")
        _metrics_add("Kv constant",            f"{fmt(kv,0) if kv else '—'} rpm/V")
        _metrics_add("Kt constant",            f"{fmf(Kt,4)} Nm/A  (torque per amp)")
        _metrics_add("No-load current I₀",     f"{fmt(I0,2)} A")
        _metrics_add("Resistance Rm",          f"{fmt(Rm*1000,1)} mΩ")
        if getattr(drone.motor,'pole_count',None):
            _metrics_add("Pole count",         f"{drone.motor.pole_count}")
        _metrics_add("Max current (rated)",    f"{fmf(getattr(drone.motor,'max_current',float('nan')),0)} A")
        _metrics_add("Max power (rated)",      f"{fmf(getattr(drone.motor,'max_power',float('nan')),0)} W")
        _metrics_add("Back-EMF (est)",         f"{fmf(V_emf,2)} V  ({fmt(rpm,0)} rpm / {fmt(kv,0) if kv else '—'} KV)")
        _metrics_add("Terminal Voltage",       f"{fmt(V_motor,2)} V")
        _metrics_add("Current / motor",        f"{fmt(I_motor,2)} A")
        _metrics_add("Electric Power / motor", f"{fmt(P_elec_motor,1)} W")
        _metrics_add("Mechanical Power / motor",f"{fmt(P_mech_motor,1)} W")
        _metrics_add("Motor efficiency",       f"{fmt(motor_eff*100,1)} %  (P_mech / P_elec)")
        _metrics_add("Copper loss / motor",    f"{fmt(I_motor**2 * Rm if I_motor==I_motor else float('nan'),2)} W  (I²Rm)")
        _metrics_add("Throttle (linear)",      f"{fmt(throttle_linear*100,0)} %")
        _metrics_add("Throttle (log scale)",   f"{fmt(throttle_log*100,0)} %")
        _metrics_add("RPM",                    f"{fmt(rpm,0)} rpm")
        _metrics_add("Thrust / motor",         f"{fmt(thrust_pm_g,0)} g  ({fmt(thrust_pm_N,3)} N)")
        _metrics_add("Thrust total",           f"{fmt((thrust_total_N/g0)*1000,0)} g  ({fmt(thrust_total_N,2)} N)")
        _metrics_add("Specific Thrust",        f"{fmt(spec_thrust,2)} g/W")
        _metrics_add("Hover Efficiency",       f"{fmt(metrics.get('hover_efficiency_gW',float('nan')),2)} g/W")
        _metrics_add("Figure of Merit",        f"{fmt(metrics.get('figure_of_merit',float('nan')),3)}")
        _metrics_add("Ideal Hover Power",      f"{fmt(metrics.get('hover_ideal_power_W',float('nan')),1)} W")
        _metrics_add("Actual Induced Power",   f"{fmt(metrics.get('actual_induced_power_W',float('nan')),1)} W")
        _metrics_add("Est. Temperature", f"{T_est:.1f} °C")
        _metrics_add("Thermal Headroom",       f"{fmt(metrics.get('motor_thermal_headroom_C',float('nan')),1)} °C  (to 120 °C)")
        _metrics_add("Thermal Status",         f"{metrics.get('thermal_status','n/a')}")

        # ── Total Drive & Power ───────────────────────────────────────────
        _metrics_add_section("Total Drive & Power")
        _metrics_add("Drive Weight",           f"{fmt(drive_g,0)} g  ({nm} motors + ESC + props)")
        _metrics_add("Thrust-Weight Ratio",    f"{fmt(twr,3)} : 1")
        hover_TWR = float(metrics.get("hover_thrust_total_N", thrust_total_N)) / max(weight_kg * g0, 1e-9)
        _metrics_add("Hover TWR",              f"{fmt(hover_TWR,3)} : 1  (static, no drag)")
        max_pay_g = (thrust_total_N/g0 - weight_kg)*1000 if thrust_total_N==thrust_total_N else float("nan")
        _metrics_add("Max Extra Payload",      f"{fmt(max_pay_g,0)} g")
        _metrics_add("Total Current",          f"{fmt(I_pack,2)} A")
        _metrics_add("Motor Power (total)",    f"{fmt(P_motor_total,1)} W  ({nm} motors)")
        _metrics_add("ESC Losses (total)",     f"{fmt(P_esc_loss,1)} W")
        _metrics_add("Avionics Power",         f"{fmt(P_periph,1)} W")
        _metrics_add("Total Power In P(in)",   f"{fmt(P_total,1)} W")
        _metrics_add("Propulsion P(out)",      f"{fmt(P_out,1)} W")
        _metrics_add("System Efficiency",      f"{fmt(eff_tot*100,1)} %  (P_out/P_in)")
        _metrics_add("Power / Weight",         f"{fmt(p2w_Wkg,1)} W/kg")
        if P_total > 0:
            _metrics_add("Power split",        f"Motor {P_motor_total/P_total*100:.0f}%  ESC {P_esc_loss/P_total*100:.0f}%  Avionics {P_periph/P_total*100:.0f}%")

        # ── Propeller & Rotor ─────────────────────────────────────────────
        _metrics_add_section("Propeller & Rotor")
        D_in    = float(getattr(drone.propeller,"diameter_in",0.0))
        P_in_p  = float(getattr(drone.propeller,"pitch_in",0.0))
        blades  = int(getattr(drone.propeller,"blades",2))
        pd_ratio= (P_in_p * 0.0254) / max(D_m, 1e-9)
        # tip_mach and tip_speed_mps can be None when RPM is unavailable;
        # .get() returns None (not the default) when key exists with value None.
        _tm = metrics.get("tip_mach");     tip_mach = float(_tm) if _tm is not None else float("nan")
        _ts = metrics.get("tip_speed_mps"); tip_spd  = float(_ts) if _ts is not None else float("nan")
        _metrics_add("Diameter",               f"{D_in:.1f} in  ({D_m*100:.1f} cm)")
        _metrics_add("Pitch",                  f"{P_in_p:.1f} in  ({P_in_p*2.54:.1f} cm)")
        _metrics_add("Blades",                 f"{blades}")
        _metrics_add("Pitch / Diameter ratio", f"{pd_ratio:.3f}")
        _metrics_add("Disk Area / rotor",      f"{A_disk*1e4:.1f} cm²  ({A_disk*1550:.1f} in²)")
        _metrics_add("Total Disk Area",        f"{A_total*1e4:.1f} cm²  (all {nm} rotors)")
        _metrics_add("Disk Loading",           f"{fmt(metrics.get('disk_loading_N_m2',float('nan')),1)} N/m²")
        _metrics_add("RPM",                    f"{fmt(rpm,0)} rpm")
        _metrics_add("Tip Speed", f"{tip_spd:.2f} m/s  ({tip_spd*3.6:.1f} km/h  /  {tip_spd*1.944:.1f} kt)")
        _metrics_add("Tip Mach",               f"{fmf(tip_mach,4)}{'  ⚠ significant noise' if tip_mach==tip_mach and tip_mach>0.6 else ''}")
        _metrics_add("Advance Ratio μ",        f"{fmt(metrics.get('advance_ratio_mu',float('nan')),3)}")
        _metrics_add("Prop Solidity σ",        f"{fmt(metrics.get('prop_solidity_sigma',float('nan')),3)}")
        _metrics_add("Blade Chord (est)",      f"{fmt(metrics.get('blade_chord_est_m',float('nan'))*1000,1)} mm")
        _metrics_add("Inflow Efficiency η",    f"{fmt(metrics.get('inflow_efficiency',float('nan')),3)}")
        _metrics_add("Inflow Power Multiplier",f"{fmt(metrics.get('inflow_power_multiplier',float('nan')),3)}")

        # ── Flight Performance ────────────────────────────────────────────
        _metrics_add_section("Flight Performance")
        _metrics_add("Orientation",            f"{orientation}")
        _metrics_add("Airspeed", f"{speed_mps:.2f} m/s  ({speed_mps*3.6:.1f} km/h  /  {speed_mps*1.944:.1f} kt)")
        _metrics_add("Ground Speed", f"{groundspeed_mps:.2f} m/s  ({groundspeed_mps*3.6:.1f} km/h  /  {groundspeed_mps*1.944:.1f} kt)")
        _metrics_add("Head / Cross Wind",      f"{fmf(metrics.get('wind_head_mps',float('nan')),2)} / {fmf(metrics.get('wind_cross_mps',float('nan')),2)} m/s")
        _metrics_add("Tilt Angle",             f"{fmf(tilt,1)} °")
        if getattr(drone,'max_tilt_deg',None) is not None:
            _metrics_add("Tilt Limit",         f"{float(drone.max_tilt_deg):.1f} °")
        _metrics_add("Estimated Range", f"{range_km:.2f} km  ({range_km*0.6214:.2f} mi  /  {range_km*0.5400:.2f} nm)")
        _metrics_add("Flight Time",            f"{fmt(t_min,2)} min  ({fmt(t_min/60,3)} h)")
        # Hover endurance (always useful to know)
        try:
            t_hover = estimate_flight_time_minutes(drone, 0.0, orientation="hover")
            _metrics_add("Hover Endurance",    f"{fmt(t_hover,2)} min")
        except Exception:
            pass
        # Specific range and endurance
        if P_total > 0 and groundspeed_mps > 0:
            SR = groundspeed_mps / P_total * 3600.0  # m/Wh
            SE = 60.0 / P_total                       # min/Wh
            _metrics_add("Specific Range",     f"{SR/1000:.3f} km/Wh  ({SR:.0f} m/Wh)")
            _metrics_add("Specific Endurance", f"{SE:.3f} min/Wh")
        _metrics_add("Acceleration",           f"{fmf(metrics.get('accel_mps2',float('nan')),2)} m/s²")
        _metrics_add("Kinetic Power Term",     f"{fmf(metrics.get('kinetic_power_W',float('nan')),1)} W")
        _metrics_add("Reserve Target",         f"{fmf(metrics.get('reserve_target_Wh',float('nan')),1)} Wh")
        _metrics_add("Reserve Margin",         f"{fmf(metrics.get('reserve_margin_Wh',float('nan')),1)} Wh  ({'VIOLATION' if metrics.get('reserve_breached',False) else 'OK'})")
        _metrics_add("Transient Segment",      f"{metrics.get('segment_type','steady')}")
        _metrics_add("Hover Wind Resistance",  f"{fmt(metrics.get('hover_wind_resistance_mps',float('nan')),2)} m/s  ({fmt(metrics.get('hover_wind_resistance_mps',0)*1.944,1)} kt)")

        # ── Thermal ───────────────────────────────────────────────────────
        _metrics_add_section("Thermal Estimates")
        _metrics_add("Motor temperature",      f"{fmf(T_est,1)} °C  [limit: 120 °C]")
        _metrics_add("Motor headroom",         f"{fmf(metrics.get('motor_thermal_headroom_C',float('nan')),1)} °C")
        _metrics_add("ESC temperature",        f"{fmf(metrics.get('esc_temp_est_C',float('nan')),1)} °C")
        _metrics_add("Battery temperature",    f"{fmf(metrics.get('battery_temp_est_C',float('nan')),1)} °C")
        _metrics_add("Thermal status",         f"{metrics.get('thermal_status','n/a')}")
        _metrics_add("Motor copper loss",      f"{fmf(metrics.get('motor_copper_loss_W_per_motor',float('nan')),2)} W/motor")
        _metrics_add("Battery I²R loss",       f"{fmf(metrics.get('battery_loss_W',float('nan')),3)} W")

        # ── Environment & Design ──────────────────────────────────────────
        _metrics_add_section("Environment & Design")
        rho = float(getattr(drone,'air_density',1.225))
        rho_drop = (1.225 - rho) / 1.225 * 100
        _metrics_add("Air Density",            f"{rho:.4f} kg/m³  ({rho_drop:+.1f}% vs ISA SL)")
        try:
            T0_k, L_k = 288.15, 0.0065
            da_m = T0_k / L_k * (1.0 - (rho/1.225)**(1.0/(1.0 - L_k*287.05/g0)))
            _metrics_add("Density Altitude",   f"{da_m:.0f} m  ({da_m*3.281:.0f} ft)")
        except Exception:
            pass
        _metrics_add("Vehicle AUW",            f"{fmt(weight_kg*1000,0)} g  ({fmt(weight_kg,3)} kg)")
        payload_g = float(getattr(drone,'payload_mass_g',0.0) or 0.0)
        if payload_g > 0:
            _metrics_add("Payload",            f"{fmt(payload_g,0)} g  ({payload_g/max(weight_kg*1000,1e-9)*100:.1f}% of AUW)")
        _metrics_add("Battery mass fraction",  f"{batt.weight_g/max(weight_kg*1000,1e-9)*100:.1f}%")
        _metrics_add("Drive mass fraction",    f"{drive_g/max(weight_kg*1000,1e-9)*100:.1f}%")
        _metrics_add("Motor configuration",    f"{getattr(drone,'motor_configuration','flat').title()}")
        _metrics_add("Energy density (batt)",  f"{fmt(getattr(batt,'energy_density_Wh_per_kg',0),0)} Wh/kg")

    # ---- Mission Plots panel ----
    mission_container = ttk.Frame(tab_mission_plots_out, padding=4)
    mission_container.grid(row=0, column=0, sticky="nsew")
    mission_container.columnconfigure(0, weight=0)
    mission_container.columnconfigure(1, weight=1)
    mission_container.rowconfigure(0, weight=1)

    mission_controls  = ttk.LabelFrame(mission_container, text="Y-axis variables", padding=4)
    mission_controls.grid(row=0, column=0, sticky="ns", padx=(0,8))
    mission_plot_frame= ttk.LabelFrame(mission_container, text="Mission plot", padding=4)
    mission_plot_frame.grid(row=0, column=1, sticky="nsew")
    mission_plot_frame.columnconfigure(0, weight=1)
    mission_plot_frame.rowconfigure(0, weight=1)

    ttk.Label(mission_controls, text="Select variables to plot vs mission time.").grid(
        row=0, column=0, sticky="w")
    mission_var_list = tk.Listbox(mission_controls, selectmode="extended", height=16, exportselection=False)
    mission_var_list.grid(row=1, column=0, sticky="nsew", pady=(4,4))
    mission_controls.rowconfigure(1, weight=1)
    mission_controls.columnconfigure(0, weight=1)
    ml_sb = ttk.Scrollbar(mission_controls, orient="vertical", command=mission_var_list.yview)
    ml_sb.grid(row=1, column=1, sticky="ns", pady=(4,4))
    mission_var_list.configure(yscrollcommand=ml_sb.set)

    mission_btns = ttk.Frame(mission_controls)
    mission_btns.grid(row=2, column=0, columnspan=2, sticky="ew")
    ttk.Button(mission_btns, text="Plot selected", command=lambda: _update_mission_plot()).grid(row=0, column=0, padx=(0,4))
    ttk.Button(mission_btns, text="Clear", command=lambda: _clear_mission_plot()).grid(row=0, column=1)

    mission_canvas_ref = [None]
    last_mission_series = [None]

    MISSION_VARS = [
        ("segment_type",        "Segment type",                  "—"),
        ("airspeed_mps",        "Vehicle airspeed",              "m/s"),
        ("commanded_airspeed_mps","Commanded airspeed",          "m/s"),
        ("accel_mps2",          "Acceleration",                  "m/s²"),
        ("groundspeed_mps",     "Ground speed",                  "m/s"),
        ("headwind_mps",        "Headwind",                      "m/s"),
        ("crosswind_mps",       "Crosswind",                     "m/s"),
        ("distance_km",         "Distance traveled",             "km"),
        ("altitude_m",          "Altitude",                      "m"),
        ("tilt_deg",            "Tilt angle",                    "deg"),
        ("climb_rate_cmd_mps",  "Climb rate command",            "m/s"),
        ("descent_rate_cmd_mps","Descent rate command",          "m/s"),
        ("climb_power_add_W",   "Climb/descent power",           "W"),
        ("kinetic_power_W",     "Kinetic power term",            "W"),
        ("battery_voltage_V",   "Battery voltage (loaded)",      "V"),
        ("battery_current_A",   "Battery current",               "A"),
        ("battery_energy_Wh",   "Battery energy remaining",      "Wh"),
        ("battery_soc_frac",    "Battery SoC fraction",          "—"),
        ("battery_soc_percent", "Battery SoC",                   "%"),
        ("reserve_target_Wh",   "Reserve target",                "Wh"),
        ("reserve_margin_Wh",   "Reserve margin",                "Wh"),
        ("reserve_breach",      "Reserve breach flag",           "bool"),
        ("battery_capacity_mAh","Battery capacity remaining",    "mAh"),
        ("total_power_W",       "Total power",                   "W"),
        ("motor_power_W",       "Motor power (total)",           "W"),
        ("motor_power_per_motor_W","Motor power (per motor)",    "W"),
        ("motor_current_A",     "Motor/ESC current (per ESC)",   "A"),
        ("motor_rpm",           "Motor RPM",                     "rpm"),
        ("tip_mach",            "Tip Mach",                      "—"),
        ("advance_ratio_mu",    "Advance ratio (μ)",             "—"),
        ("inflow_efficiency",   "Inflow efficiency (η)",         "—"),
        ("inflow_power_multiplier", "Inflow power multiplier",   "—"),
        ("hover_efficiency_gW", "Hover efficiency",              "g/W"),
        ("figure_of_merit",     "Figure of merit",               "—"),
        ("disk_loading_N_m2",   "Disk loading",                  "N/m²"),
        ("motor_temp_est_C",    "Motor temperature (est)",       "°C"),
        ("esc_temp_est_C",      "ESC temperature (est)",         "°C"),
        ("battery_temp_est_C",  "Battery temperature (est)",     "°C"),
        ("thermal_status",      "Thermal status",                "—"),
        ("hover_wind_resistance_mps", "Hover wind resistance",   "m/s"),
        ("prop_solidity_sigma", "Propeller solidity",            "—"),
        ("motor_thrust_N",      "Motor thrust (per motor)",      "N"),
        ("thrust_total_N",      "Total thrust",                  "N"),
        ("periph_power_W",      "Avionics/peripherals power",    "W"),
        ("esc_loss_W",          "ESC loss power",                "W"),
    ]
    _mission_items = []
    for k, lbl, unit in MISSION_VARS:
        mission_var_list.insert(tk.END, f"{lbl} ({unit})")
        _mission_items.append((k, lbl, unit))

    def _clear_mission_plot():
        for w in mission_plot_frame.winfo_children(): w.destroy()
        mission_canvas_ref[0] = None

    def _update_mission_plot():
        ms = last_mission_series[0]
        if ms is None:
            messagebox.showinfo("Mission plot", "Run a mission first."); return
        sel = list(mission_var_list.curselection())
        if not sel:
            messagebox.showinfo("Mission plot", "Select at least one variable."); return

        t_min = [x/60.0 for x in ms.get('t_s', [])]
        if not t_min: return

        fig = plt.Figure(figsize=(7.5, 4.5), dpi=100)
        ax0 = fig.add_subplot(111)
        selected = [_mission_items[i] for i in sel]
        by_unit = {}
        for k, lbl, unit in selected:
            by_unit.setdefault(unit, []).append((k, lbl))

        axes = [(ax0, list(by_unit.keys())[0])]
        ax0.set_ylabel(list(by_unit.keys())[0])
        for ui, unit in enumerate(list(by_unit.keys())[1:], 1):
            axn = ax0.twinx()
            axn.spines['right'].set_position(('outward', 55*(ui-0)))
            axn.set_ylabel(unit)
            axes.append((axn, unit))

        lines, labels = [], []
        for ax, unit in axes:
            for key, lbl in by_unit.get(unit, []):
                y = ms.get(key, [])
                if y is None: continue
                yy = []
                for v in y:
                    try: yy.append(float("nan") if isinstance(v,float) and v!=v else float(v))
                    except Exception: yy.append(float("nan"))
                ln, = ax.plot(t_min, yy, label=lbl)
                lines.append(ln); labels.append(lbl)

        ax0.set_xlabel("Mission time (min)")
        ax0.grid(True)
        fig.suptitle("Mission variables vs time")
        if lines: ax0.legend(lines, labels, loc="best")

        for w in mission_plot_frame.winfo_children(): w.destroy()
        mc = FigureCanvasTkAgg(fig, master=mission_plot_frame)
        mc.draw()
        mc.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        mission_canvas_ref[0] = mc

    def _on_list_wheel(evt):
        if evt.delta: mission_var_list.yview_scroll(int(-1*(evt.delta/120)), "units")
        return "break"
    mission_var_list.bind("<MouseWheel>", _on_list_wheel)

    # ---- Output text ----
    out_frame = ttk.LabelFrame(right, text="Output", padding=4)
    out_frame.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
    out_frame.columnconfigure(0, weight=1)
    out_frame.rowconfigure(0, weight=1)

    out_text = tk.Text(out_frame, height=8, wrap="word", state="disabled")
    out_text.grid(row=0, column=0, sticky="nsew")
    out_sb = ttk.Scrollbar(out_frame, orient="vertical", command=out_text.yview)
    out_sb.grid(row=0, column=1, sticky="ns")
    out_text.configure(yscrollcommand=out_sb.set)

    def out_print(msg: str):
        out_text.configure(state="normal")
        out_text.delete("1.0", "end")
        out_text.insert("end", msg)
        out_text.configure(state="disabled")

    def out_append(msg: str):
        out_text.configure(state="normal")
        out_text.insert("end", msg + "\n")
        out_text.see("end")
        out_text.configure(state="disabled")

    def clear_log():
        out_text.configure(state="normal")
        out_text.delete("1.0", "end")
        out_text.configure(state="disabled")

    # ================================================================== #
    #  BUILD CONFIG FROM GUI                                              #
    # ================================================================== #
    def build_config_from_gui() -> DroneConfig:
        batt = BatteryConfig(
            chemistry              = v_batt_chem.get().strip() or None,
            operating_voltage_min  = parse_float("Vmin/cell", v_batt_vmin.get()),
            operating_voltage_nominal = parse_float("Vnom/cell", v_batt_vnom.get()),
            operating_voltage_max  = parse_float("Vmax/cell", v_batt_vmax.get()),
            cell_capacity_mAh      = parse_float_opt("Cell capacity", v_batt_cell_capacity.get()),
            pack_capacity_mAh      = parse_float_opt("Pack capacity", v_batt_pack_capacity.get()),
            cell_weight_g          = parse_float_opt("Cell weight", v_batt_cell_weight.get()),
            pack_weight_g          = parse_float_opt("Pack weight", v_batt_pack_weight.get()),
            unit_energy_density    = parse_float_opt("Energy density", v_batt_energy_density.get()),
            charge_current_max     = parse_float_opt("Max charge current", v_batt_chg.get()),
            discharge_cont_A       = parse_float_opt("Cont discharge", v_batt_a_cont.get()),
            discharge_max_A        = parse_float_opt("Max discharge", v_batt_a_max.get()),
            discharge_c_cont       = parse_float_opt("Cont C-rate", v_batt_c_cont.get()),
            discharge_c_max        = parse_float_opt("Max C-rate", v_batt_c_max.get()),
            discharge_percent      = parse_float("Discharge %", v_batt_dischg_pct.get()),
            resistance_cell_mOhm   = parse_float("Rcell", v_batt_r.get()),
            unit_mode              = v_batt_unit_mode.get().strip().lower(),
            series_units           = parse_int("Series units", v_batt_series.get()),
            parallel_units         = parse_int("Parallel units", v_batt_parallel.get()),
            cells_series_per_unit  = parse_int("Cells series/unit", v_batt_cells_series.get()),
            cells_parallel_per_unit= parse_int("Cells parallel/unit", v_batt_cells_parallel.get()),
            soc_model              = v_batt_soc_model.get().strip() or "auto",
            soc_curve_csv          = v_batt_soc_curve_csv.get().strip() or None,
            soc_bp                 = parse_float_list(v_batt_soc_bp.get().strip()),
            ocv_cell_bp            = parse_float_list(v_batt_ocv_cell_bp.get().strip()),
            r_scale_bp             = parse_float_list(v_batt_r_scale_bp.get().strip()),
        )
        motor = MotorConfig(
            kv            = parse_float("Kv", v_motor_kv.get()),
            idle_current  = parse_float_opt("Idle current", v_motor_i0.get(), 0.0),
            idle_voltage  = parse_float_opt("Idle voltage", v_motor_v0.get(), 10.0),
            rated_voltage = parse_int_opt("Rated voltage", v_motor_rated_v.get(), 0),
            resistance    = parse_float("Motor resistance", v_motor_r.get()),
            max_current   = parse_float_opt("Motor max current", v_motor_imax.get()),
            max_power     = parse_float_opt("Motor max power", v_motor_pmax.get()),
            pole_count    = parse_int_opt("Pole count", v_motor_pole_count.get(), 14),
            weight_g      = parse_float_opt("Motor weight", v_motor_weight.get()),
            size_mm       = v_motor_size.get().strip() or None,
        )
        esc = None
        _ef = [v_esc_voltage_rating.get().strip(), v_esc_cont_current.get().strip(),
               v_esc_max_current.get().strip(), v_esc_idle_current.get().strip(),
               v_esc_r.get().strip(), v_esc_weight.get().strip()]
        if any(_ef):
            esc = ESCConfig(
                voltage_rating       = parse_int_opt("ESC voltage rating", v_esc_voltage_rating.get(), 0),
                continuous_current_A = parse_float_opt("ESC cont current", v_esc_cont_current.get(), 0.0),
                max_current_A        = parse_float_opt("ESC max current", v_esc_max_current.get(), 0.0),
                idle_current_A       = parse_float_opt("ESC idle current", v_esc_idle_current.get(), 0.0),
                resistance           = parse_float_opt("ESC resistance", v_esc_r.get(), 0.0),
                weight_g             = parse_float_opt("ESC weight", v_esc_weight.get()),
            )
        avionics = AvionicsConfig(voltage_tree=_get_voltage_tree_from_table())

        prop_table_path = v_prop_table.get().strip() or None
        tc = v_prop_tconst.get().strip(); pc = v_prop_pconst.get().strip()
        inflow_mu_bp = parse_float_list(v_inflow_mu_bp.get().strip())
        inflow_eff_bp = parse_float_list(v_inflow_eff_bp.get().strip())
        inflow_enabled_raw = parse_optional_bool(v_inflow_map_enabled.get().strip())
        inflow_map_enabled = True if inflow_enabled_raw is None else bool(inflow_enabled_raw)
        drag_model_raw = v_drag_model_mode.get().strip().lower()
        if drag_model_raw not in ("auto", "manual", "geometry"):
            drag_model_raw = "auto"
        prop = PropellerConfig(
            diameter_in  = parse_float("Prop diameter", v_prop_d.get()),
            pitch_in     = parse_float("Prop pitch", v_prop_pitch.get()),
            max_rpm      = parse_float_opt("Prop max RPM", v_prop_max_rpm.get(), 0.0),
            max_thrust_g = parse_float_opt("Prop max thrust", v_prop_max_thrust.get(), 0.0),
            blades       = parse_int_opt("Prop blades", v_prop_blades.get(), 2),
            table_csv    = prop_table_path,
            TConst       = float(tc) if tc else None,
            PConst       = float(pc) if pc else None,
            weight_g     = parse_float_opt("Prop weight", v_prop_weight.get()),
        )
        base_weight_g = parse_float("Weight", v_weight.get())
        payload_mass_g = max(parse_float_opt("Payload mass", v_payload_mass.get(), 0.0), 0.0)
        drone = DroneConfig(
            num_motors               = parse_int("Num motors", v_num_motors.get()),
            battery                  = batt,
            motor                    = motor,
            propeller                = prop,
            drone_weight_g           = base_weight_g + payload_mass_g,
            profile_drag_coefficient = (parse_float("Profile Cd", v_profile_drag.get()) if v_profile_drag.get().strip() else 0.0),
            profile_area             = (parse_float("Profile area", v_profile_area.get()) if v_profile_area.get().strip() else 0.0),
            parasite_drag_coefficient= (parse_float("Parasite Cd", v_parasite_drag.get()) if v_parasite_drag.get().strip() else 0.0),
            parasite_area            = (parse_float("Parasite area", v_parasite_area.get()) if v_parasite_area.get().strip() else 0.0),
            frontal_area             = (parse_float("Frontal area", v_area.get()) if v_area.get().strip() else 0.0),
            cruise_speed             = parse_float("Cruise speed", v_speed.get()),
            periph_current           = parse_float("Peripheral current", v_periph_current.get()),
            esc                      = esc,
            avionics                 = avionics,
            air_density              = AIR_DENSITY,
            body_length_m            = (parse_float("Body length", v_body_length_m.get()) if v_body_length_m.get().strip() else None),
            body_width_m             = (parse_float("Body width", v_body_width_m.get()) if v_body_width_m.get().strip() else None),
            body_height_m            = (parse_float("Body height", v_body_height_m.get()) if v_body_height_m.get().strip() else None),
            arm_length_m             = (parse_float("Arm length", v_arm_length_m.get()) if v_arm_length_m.get().strip() else None),
            arm_width_m              = (parse_float("Arm width", v_arm_width_m.get()) if v_arm_width_m.get().strip() else None),
            coaxial_spacing_m        = (parse_float("Coaxial spacing", v_coaxial_spacing_m.get()) if v_coaxial_spacing_m.get().strip() else None),
            max_tilt_deg             = (parse_float("Max tilt", v_max_tilt_deg.get()) if v_max_tilt_deg.get().strip() else None),
            motor_configuration      = (v_motor_configuration.get().strip().lower() or "flat"),
            drag_model               = drag_model_raw,
            transient_dt_s           = parse_float("Transient dt", v_transient_dt_s.get()),
            max_accel_mps2           = parse_float("Max accel", v_max_accel_mps2.get()),
            max_decel_mps2           = parse_float("Max decel", v_max_decel_mps2.get()),
            decel_regen_eff          = parse_float("Decel regen efficiency", v_decel_regen_eff.get()),
            inflow_map_enabled       = inflow_map_enabled,
            inflow_mu_bp             = inflow_mu_bp,
            inflow_eff_bp            = inflow_eff_bp,
        )
        drone.payload_mass_g = payload_mass_g
        alt  = parse_float("Altitude", v_alt.get())
        temp = v_temp.get().strip()
        pres = v_press.get().strip()
        drone.air_density = compute_air_density(
            altitude_m    = alt,
            temperature_C = float(temp) if temp else None,
            pressure_Pa   = float(pres) if pres else None,
        )
        return drone

    # ================================================================== #
    #  RUN CALLBACKS                                                      #
    # ================================================================== #
    def run_single_point():
        clear_log()
        try:
            drone       = build_config_from_gui()
            orientation = v_orientation.get().strip().lower()
            if orientation not in ("hover","forward"):
                raise ValueError("Orientation must be 'hover' or 'forward'.")
            speed   = parse_float("Speed", v_speed.get())
            wind = parse_float("Wind speed", v_wind.get())
            wind_dir = parse_float("Wind direction", v_wind_dir.get())
            course_deg = parse_float("Course heading", v_course_deg.get())
            climb_rate = max(parse_float("Climb rate", v_climb_rate.get()), 0.0)
            descent_rate = max(parse_float("Descent rate", v_descent_rate.get()), 0.0)
            if climb_rate > 0 and descent_rate > 0:
                descent_rate = 0.0
            be_v, be_min, br_v, br_km = find_optimal_speeds(drone)
            metrics = compute_operating_metrics(
                drone,
                speed_mps=speed,
                orientation=orientation,
                wind_mps=wind,
                wind_direction_deg=wind_dir,
                course_deg=course_deg,
                ambient_temp_C=(float(v_temp.get()) if v_temp.get().strip() else 25.0),
            )
            potential_power_w = (drone.drone_weight_g * 9.81 / 1000.0) * (climb_rate - descent_rate)
            base_total_w = float(metrics.get("total_power_W", 0.0))
            adj_total_w = max(base_total_w + potential_power_w, 0.0)
            if base_total_w > 0:
                scale = adj_total_w / base_total_w
                metrics["motor_power_W"] = float(metrics.get("motor_power_W", 0.0)) * scale
                metrics["esc_loss_W"] = float(metrics.get("esc_loss_W", 0.0)) * scale
            metrics["climb_rate_cmd_mps"] = climb_rate
            metrics["descent_rate_cmd_mps"] = descent_rate
            metrics["potential_power_W"] = potential_power_w
            metrics["climb_power_add_W"] = potential_power_w
            metrics["total_power_W"] = adj_total_w
            metrics["pack_current_A"] = adj_total_w / max(float(drone.battery.vnom_pack), 1.0)
            metrics["v_load_V"] = battery_voltage_under_load(drone.battery, metrics["pack_current_A"])
            reserve_target_wh = max(
                drone.battery.usable_Wh * (parse_float("Reserve percent", v_reserve_percent.get()) / 100.0),
                parse_float("RTH reserve", v_rth_reserve_Wh.get()) + parse_float("Diversion reserve", v_div_reserve_Wh.get()),
            )
            metrics["reserve_target_Wh"] = reserve_target_wh
            metrics["reserve_margin_Wh"] = drone.battery.usable_Wh - reserve_target_wh
            metrics["reserve_breached"] = bool(metrics["reserve_margin_Wh"] < 0)
            if adj_total_w > 0:
                t_min = drone.battery.usable_Wh / adj_total_w * 60.0
            else:
                t_min = 0.0
            d_km = float(metrics.get("groundspeed_mps", 0.0)) * (t_min * 60.0) / 1000.0
            update_status_tables_from_metrics(drone, metrics)
            update_metrics_tab(drone, metrics, speed_mps=speed, orientation=orientation)

            max_spd = parse_float("Max speed plot", v_max_speed_plot.get())
            _last_run["drone"]   = drone
            _last_run["max_spd"] = max_spd
            # Speed and orientation are needed by the Sensitivity and Compare
            # tabs, which re-evaluate the same operating point.
            _last_run["speed"] = speed
            _last_run["orientation"] = orientation
            # Capture sweep data for CSV/Excel export
            _mc_v = [0.5 + (max_spd-0.5)*i/200 for i in range(201)]
            _last_run_sweep.clear()
            _last_run_sweep.update({
                "Speed (m/s)": _mc_v,
                "Flight Time (min)": [estimate_flight_time_minutes(drone,v,"forward") for v in _mc_v],
                "Range (km)": [estimate_flight_distance_km(drone,v,"forward") for v in _mc_v],
                "Power Fwd (W)": [power_required(drone,v,"forward") for v in _mc_v],
                "Power Hov (W)": [power_required(drone,v,"hover") for v in _mc_v],
                "Thrust Fwd (N)": [thrust_required(drone,v,"forward") for v in _mc_v],
                "Thrust Hov (N)": [thrust_required(drone,v,"hover") for v in _mc_v],
                "Disk Loading (N/m²)": [disk_loading_N_m2(drone)] * len(_mc_v),
                "Hover Wind Resistance (m/s)": [hover_wind_resistance_mps(drone)] * len(_mc_v),
                "Prop Solidity (σ)": [propeller_solidity(drone.propeller.diameter_in, drone.propeller.blades)] * len(_mc_v),
            })
            _last_run_cfg[0] = drone
            update_weight_budget(drone)
            _refresh_airframe_diagram(drone)
            refresh_comparison()
            fig = make_performance_figure(
                drone, max_speed=max_spd,
                figsize=(_view["plot_w"], _view["plot_h"]))
            # Store metrics for regeneration during plot scale changes
            _last_run["metrics"] = metrics
            # Generate motor operating point figure if propeller table is available
            motor_fig = None
            if drone.propeller.table is not None:
                try:
                    motor_fig = make_motor_operating_point_figure(drone, metrics, figsize=(_view["plot_w"], 6))
                except Exception:
                    pass
            # Display both figures
            if motor_fig:
                _show_figure([fig, motor_fig])
            else:
                _show_figure(fig)
            display_nb.select(tab_plot_out)

            def _fmt_out(x, nd=2):
                try:
                    return f"{float(x):.{nd}f}"
                except Exception:
                    return "n/a"

            out_print(
                f"=== Single-Point Run @ {speed:.1f} m/s ({orientation}) ===\n"
                f"Air density     : {drone.air_density:.3f} kg/m³\n"
                f"Flight time     : {t_min:.2f} min\n"
                f"Flight distance : {d_km:.2f} km\n"
                f"SoC             : {_fmt_out(metrics.get('soc_percent', float('nan')),1)} % ({metrics.get('soc_model_source', 'linear-fallback')})\n"
                f"Ground speed    : {_fmt_out(metrics.get('groundspeed_mps', float('nan')),2)} m/s\n"
                f"Head/Cross wind : {_fmt_out(metrics.get('wind_head_mps', float('nan')),2)} / {_fmt_out(metrics.get('wind_cross_mps', float('nan')),2)} m/s\n"
                f"Best endurance  : {be_v:.2f} m/s → {be_min:.2f} min\n"
                f"Best range      : {br_v:.2f} m/s → {br_km:.2f} km\n"
                f"Hover eff.      : {_fmt_out(metrics.get('hover_efficiency_gW', float('nan')),2)} g/W\n"
                f"Figure of merit : {_fmt_out(metrics.get('figure_of_merit', float('nan')),3)}\n"
                f"Disk loading    : {_fmt_out(metrics.get('disk_loading_N_m2', float('nan')),1)} N/m²\n"
                f"Tip Mach        : {_fmt_out(metrics.get('tip_mach', float('nan')),3)}\n"
                f"Potential power : {_fmt_out(metrics.get('potential_power_W', float('nan')),1)} W\n"
                f"Reserve target  : {_fmt_out(metrics.get('reserve_target_Wh', float('nan')),1)} Wh\n"
                f"Reserve margin  : {_fmt_out(metrics.get('reserve_margin_Wh', float('nan')),1)} Wh\n"
                f"Thermal status  : {metrics.get('thermal_status', 'n/a')} ({_fmt_out(metrics.get('motor_temp_est_C', float('nan')),1)} °C)\n"
                f"Thermal M/ESC/B : {_fmt_out(metrics.get('motor_temp_est_C', float('nan')),1)} / {_fmt_out(metrics.get('esc_temp_est_C', float('nan')),1)} / {_fmt_out(metrics.get('battery_temp_est_C', float('nan')),1)} °C\n"
                f"Hover wind max  : {_fmt_out(metrics.get('hover_wind_resistance_mps', float('nan')),2)} m/s\n"
                f"Prop solidity σ : {_fmt_out(metrics.get('prop_solidity_sigma', float('nan')),3)}\n"
            )
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def run_mission():
        clear_log()
        try:
            mission_path = v_mission.get().strip()
            if not mission_path:
                raise ValueError("Select a mission JSON file first (Mission/Env tab).")
            drone       = build_config_from_gui()
            orientation = v_orientation.get().strip().lower()
            if orientation not in ("hover","forward"):
                raise ValueError("Orientation must be 'hover' or 'forward'.")
            temp = v_temp.get().strip()
            pres = v_press.get().strip()
            wind = parse_float("Wind speed", v_wind.get())
            wind_dir = parse_float("Wind direction", v_wind_dir.get())
            course_deg = parse_float("Course heading", v_course_deg.get())
            climb_rate = max(parse_float("Climb rate", v_climb_rate.get()), 0.0)
            descent_rate = max(parse_float("Descent rate", v_descent_rate.get()), 0.0)
            if climb_rate > 0 and descent_rate > 0:
                descent_rate = 0.0

            mission = MissionProfile.from_json(mission_path)
            mission.reserve_percent = parse_float("Reserve percent", v_reserve_percent.get())
            mission.rth_reserve_Wh = parse_float("RTH reserve", v_rth_reserve_Wh.get())
            mission.diversion_reserve_Wh = parse_float("Diversion reserve", v_div_reserve_Wh.get())
            mission.wind_direction_deg = wind_dir
            mission.transient_dt_s = parse_float("Transient step dt", v_transient_dt_s.get())
            mission.max_accel_mps2 = parse_float("Max accel", v_max_accel_mps2.get())
            mission.max_decel_mps2 = parse_float("Max decel", v_max_decel_mps2.get())
            mission.decel_regen_eff = parse_float("Decel regen efficiency", v_decel_regen_eff.get())
            for _p in mission.phases:
                _p.course_deg = float(_p.course_deg if _p.course_deg else course_deg)
                if _p.climb_rate_mps is None:
                    _p.climb_rate_mps = climb_rate
                if _p.descent_rate_mps is None:
                    _p.descent_rate_mps = descent_rate
            results, worst_metrics, mission_series = simulate_mission(
                drone, mission, orientation=orientation,
                temperature_C = float(temp) if temp else None,
                pressure_Pa   = float(pres) if pres else None,
                wind_mps      = wind,
            )
            last_mission_series[0] = mission_series

            lines = [f"=== Mission: {os.path.basename(mission_path)} ===",
                     f"Orientation: {orientation}  |  Wind: {wind:.2f} m/s @ {wind_dir:.1f}°", ""]
            total_t = total_d = 0.0
            for name, t_min, d_km, status in results:
                total_t += t_min; total_d += d_km
                lines.append(f"  {name}: {t_min:.2f} min, {d_km:.2f} km  [{status}]")
            lines += ["", f"TOTAL: {total_t:.2f} min, {total_d:.2f} km"]
            if worst_metrics is not None:
                lines += [
                    "",
                    f"Reserve target/margin: {float(worst_metrics.get('reserve_target_Wh',0.0)):.1f} / {float(worst_metrics.get('reserve_margin_Wh',0.0)):+.1f} Wh",
                    f"SoC (min/model): {float(worst_metrics.get('soc_percent',100.0)):.1f}% / {str(worst_metrics.get('soc_model_source','linear-fallback'))}",
                    f"Transient (dt/a+/a-): {float(mission.transient_dt_s):.2f}s / {float(mission.max_accel_mps2):.2f} / {float(mission.max_decel_mps2):.2f} m/s²",
                    f"Inflow μ/η/mult: {float(worst_metrics.get('advance_ratio_mu',0.0)):.3f} / {float(worst_metrics.get('inflow_efficiency',1.0)):.3f} / {float(worst_metrics.get('inflow_power_multiplier',1.0)):.3f}",
                    f"Thermal M/ESC/B: {float(worst_metrics.get('motor_temp_est_C',0.0)):.1f} / {float(worst_metrics.get('esc_temp_est_C',0.0)):.1f} / {float(worst_metrics.get('battery_temp_est_C',0.0)):.1f} °C [{worst_metrics.get('thermal_status','OK')}]",
                ]
            out_print("\n".join(lines))

            if worst_metrics is not None:
                update_status_tables_from_metrics(drone, worst_metrics)

            max_spd = parse_float("Max speed plot", v_max_speed_plot.get())
            _last_run["drone"]   = drone
            _last_run["max_spd"] = max_spd
            _last_run_cfg[0] = drone
            update_weight_budget(drone)
            fig = make_performance_figure(
                drone, max_speed=max_spd,
                figsize=(_view["plot_w"], _view["plot_h"]))
            _show_figure(fig)
            display_nb.select(tab_plot_out)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ================================================================== #
    #  SAVE / LOAD CONFIG                                                 #
    # ================================================================== #
    def _extract_avionics_rails() -> list:
        rails = []
        for iid in avionics_tree.get_children():
            v_s, i_s, e_s = avionics_tree.item(iid, "values")
            try:
                rails.append({"voltage": float(v_s), "current": float(i_s), "eff": float(e_s)})
            except Exception: continue
        rails.sort(key=lambda r: r["voltage"])
        return rails

    def save_config_to_file(path: str) -> None:
        data = {
            "schema": "multicopter_power_sim_gui_config",
            "version": 1,
            "vars": {k: v.get() for k, v in config_vars.items()},
            "avionics_rails": _extract_avionics_rails(),
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

    def load_config_from_file(path: str) -> None:
        # Keep the mode-bar label in step with whatever was just loaded.
        try:
            v_loaded_cfg.set(os.path.basename(str(path)))
        except Exception:
            pass
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for k, val in data.get("vars", {}).items():
            if k in config_vars:
                try: config_vars[k].set("" if val is None else str(val))
                except Exception: pass
        rails = data.get("avionics_rails", None)
        if isinstance(rails, list) and rails:
            _av_load_rows(rails)
        else:
            try:
                d = parse_voltage_tree(v_avionics_voltage_tree.get().strip())
                _av_load_rows([{"voltage": v, "current": ci[0], "eff": ci[1]}
                               for v, ci in d.items()])
            except Exception: pass
        on_unit_mode_change()
        # Re-apply visibility: a loaded config may switch cell/pack mode,
        # and the newly-restored values must respect the current UI mode.
        _apply_field_mode()




    # ------------------------------------------------------------------
    # SIMPLE / ADVANCED MODE
    # ------------------------------------------------------------------
    def _apply_field_mode(*_a):
        """
        Show or hide input rows according to the Simple/Advanced setting.

        Simple mode grids only the rows whose key is in MC_SIMPLE_FIELDS.
        Advanced mode grids everything.  Rows are removed with grid_remove(),
        which preserves their grid options so re-showing is a plain grid().
        Hidden fields keep their values, so switching modes never changes a
        result — it only changes what you can see.
        """
        simple = (v_ui_mode.get() == "Simple")
        for row in _field_rows:
            show = (not simple) or (row["key"] in MC_SIMPLE_FIELDS)
            for w in row["widgets"]:
                try:
                    if show: w.grid()
                    else:    w.grid_remove()
                except Exception:
                    pass
        # A section heading is hidden when every field under it is hidden.
        for sec in _section_rows:
            show = (not simple) or (not sec["keys"]) or bool(sec["keys"] & MC_SIMPLE_FIELDS)
            for w in sec["widgets"]:
                try:
                    if show: w.grid()
                    else:    w.grid_remove()
                except Exception:
                    pass


    v_ui_mode.trace_add("write", _apply_field_mode)
    _apply_field_mode()          # apply the default (Simple) at startup


    # Startup banner: states the version and points at the two new UI features,
    # so a stale copy of the script is immediately obvious.

    def prompt_save_config():
        path = filedialog.asksaveasfilename(
            title="Save configuration", defaultextension=".json",
            filetypes=[("JSON files","*.json"),("All files","*.*")])
        if not path: return
        try:
            save_config_to_file(path)
            messagebox.showinfo("Saved", f"Saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error saving config", str(e))

    def prompt_load_config():
        path = filedialog.askopenfilename(
            title="Load configuration",
            filetypes=[("JSON files","*.json"),("All files","*.*")])
        if not path: return
        try:
            load_config_from_file(path)
            messagebox.showinfo("Loaded", f"Loaded from:\n{path}")
        except Exception as e:
            messagebox.showerror("Error loading config", str(e))

    file_menu.add_command(label="Load Config…", command=prompt_load_config)
    file_menu.add_command(label="Save Config…", command=prompt_save_config)

    _status_tv_pairs = [
        (batt_table, "Battery Status"),
        (motor_table, "Motor / ESC Status"),
        (prop_table_tv, "Propeller Status"),
    ]
    _report_title = "Multicopter Power Simulator — Performance Analysis"

    # ================================================================== #
    # EXPORT & REPORT FUNCTIONS
    # ================================================================== #
    _last_run_sweep: dict = {}    # populated after each run
    _last_run_cfg  = [None]       # populated after each run

    def _get_metrics_rows() -> list:
        """Read all rows from the metrics Treeview."""
        rows = []
        for iid in metrics_tv.get_children():
            vals = metrics_tv.item(iid, "values")
            if vals and len(vals) >= 2:
                rows.append((str(vals[0]), str(vals[1])))
        return rows

    def _get_status_sections() -> list:
        """Read all status Treeview tables as (title, [(metric,val,lim,note,tag),...])."""
        result = []
        for tv, title in _status_tv_pairs:
            sec_rows = []
            for iid in tv.get_children():
                vals = tv.item(iid, "values")
                tags = tv.item(iid, "tags")
                tag  = tags[0] if tags else "na"
                if vals and len(vals) >= 4:
                    sec_rows.append((str(vals[0]), str(vals[1]),
                                     str(vals[2]), str(vals[3]), str(tag)))
                elif vals and len(vals) == 3:
                    sec_rows.append((str(vals[0]), str(vals[1]),
                                     str(vals[2]), "", "na"))
            if sec_rows:
                result.append((title, sec_rows))
        return result

    def _get_log_text() -> str:
        try:
            out_text.configure(state="normal")
            t = out_text.get("1.0", "end")
            out_text.configure(state="disabled")
            return t
        except Exception:
            return ""

    def _get_inputs_rows() -> list:
        """Return all GUI input variables as (label, value) pairs."""
        rows = []
        for k, var in config_vars.items():
            try:
                val = var.get()
                if val not in ("", None):
                    rows.append((k.replace("_", " ").title(), str(val)))
            except Exception:
                pass
        return rows

    def _do_export_csv():
        if not _last_run_sweep:
            messagebox.showinfo("No data", "Run a simulation first to generate sweep data.")
            return
        path = filedialog.asksaveasfilename(
            title="Export CSV", defaultextension=".csv",
            filetypes=[("CSV files","*.csv"),("All files","*.*")])
        if not path: return
        try:
            _export_csv_file(path, _last_run_sweep, _get_metrics_rows())
            messagebox.showinfo("Exported", f"CSV saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))

    def _do_export_excel():
        if not _last_run_sweep:
            messagebox.showinfo("No data", "Run a simulation first to generate sweep data.")
            return
        path = filedialog.asksaveasfilename(
            title="Export Excel", defaultextension=".xlsx",
            filetypes=[("Excel files","*.xlsx"),("All files","*.*")])
        if not path: return
        try:
            cfg = _last_run_cfg[0]
            wb  = _extract_weight_budget(cfg) if cfg else []
            _export_excel_file(path, _last_run_sweep, _get_metrics_rows(), wb)
            messagebox.showinfo("Exported", f"Excel saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))

    def _do_generate_report():
        if not _last_run_sweep and not _get_metrics_rows():
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        path = filedialog.asksaveasfilename(
            title="Save PDF Report", defaultextension=".pdf",
            filetypes=[("PDF files","*.pdf"),("All files","*.*")])
        if not path: return
        try:
            cfg   = _last_run_cfg[0]
            figs  = []
            # Add all current plot figures from the scrollable area
            figs.extend(_current_plot_figs)
            if mission_canvas_ref[0] is not None:
                figs.append(mission_canvas_ref[0].figure)
            for num in plt.get_fignums():
                fig = plt.figure(num)
                if fig not in figs:
                    figs.append(fig)
            wb    = _extract_weight_budget(cfg) if cfg else []
            _generate_pdf_report(
                path         = path,
                report_title = _report_title,
                inputs_rows  = _get_inputs_rows(),
                metrics_rows = _get_metrics_rows(),
                status_sections = _get_status_sections(),
                log_text     = _get_log_text(),
                figures      = figs,
                weight_budget = wb,
            )
            messagebox.showinfo("Report generated", f"PDF report saved to:\n{path}")
        except Exception as e:
            import traceback
            messagebox.showerror("Report error", traceback.format_exc())

    # Wire export menu items
    file_menu.add_separator()
    file_menu.add_command(label="Export CSV…",       command=_do_export_csv)
    file_menu.add_command(label="Export Excel…",     command=_do_export_excel)
    file_menu.add_command(label="Generate PDF Report…", command=_do_generate_report)

    file_menu.add_separator()
    file_menu.add_command(label="Exit", command=exit_app)

    # ================================================================== #
    #  BUTTON ROW                                                         #
    # ================================================================== #
    btn_frame = ttk.Frame(main)
    btn_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 0))
    btn_frame.columnconfigure(1, weight=1)

    ttk.Button(btn_frame, text="▶  Run Single-Point",
               command=run_single_point).grid(row=0, column=0, padx=(0, 6), pady=4)
    ttk.Button(btn_frame, text="📋  Run Mission (JSON)",
               command=run_mission).grid(row=0, column=1, padx=(0, 6), pady=4, sticky="w")
    ttk.Button(btn_frame, text="💾  Save Config",
               command=prompt_save_config).grid(row=0, column=2, padx=4, pady=4)
    ttk.Button(btn_frame, text="📂  Load Config",
               command=prompt_load_config).grid(row=0, column=3, padx=4, pady=4)
    ttk.Button(btn_frame, text="📊  Export CSV",
               command=_do_export_csv).grid(row=0, column=4, padx=4, pady=4)
    ttk.Button(btn_frame, text="📗  Export Excel",
               command=_do_export_excel).grid(row=0, column=5, padx=4, pady=4)
    ttk.Button(btn_frame, text="📄  Generate Report",
               command=_do_generate_report).grid(row=0, column=6, padx=4, pady=4)

    root.protocol("WM_DELETE_WINDOW", exit_app)


    _banner_lines = [
        f"Multicopter Power Simulator  v{SIM_VERSION}",
        "=" * 46,
        "Input detail is set to Simple (selector above the input tabs).",
        "Hover the blue ? beside any field for an explanation and a",
        "typical value. Switch to Advanced to reveal every input.",
        "Metrics sections collapse - click a section heading to fold it.",
        "",
    ]
    _banner_lines.append("Load Config -> examples/configs/ for ready-made aircraft.")
    _banner_lines.append("Run Mission (JSON) -> examples/missions/ for flight profiles.")
    out_print("\n".join(_banner_lines) + "\n")

    root.mainloop()

# -------------------------------
# CLI
# -------------------------------
def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Multicopter Flight Simulator (CLI + optional GUI)")

    parser.add_argument("--gui", action="store_true", help="Launch the GUI instead of CLI run.")

    # Drone
    parser.add_argument("--num_motors", type=int, required=False)
    parser.add_argument("--weight", type=float, required=False, help="Base drone weight excluding payload (g)")
    parser.add_argument("--payload_mass_g", type=float, default=0.0, help="Payload mass added to base drone weight (g)")
    parser.add_argument("--profile_drag", type=float, default=0.0, help="Profile drag coefficient")
    parser.add_argument("--profile_area", type=float, default=0.0, help="Rotor/arms profile reference area (m^2)")
    parser.add_argument("--parasite_drag", type=float, default=0.0, help="Parasite drag coefficient (fuselage/arms)")
    parser.add_argument("--parasite_area", type=float, default=0.0, help="Parasite reference area (m^2)")
    parser.add_argument(
        "--drag_model_mode",
        type=str,
        default="auto",
        choices=["auto", "manual"],
        help="Drag model source: auto uses geometry fallback when drag inputs are missing; manual always uses provided profile/parasite values.",
    )
    parser.add_argument("--area", type=float, required=False, help="Frontal area (m^2)")
    parser.add_argument("--body_length_m", type=float, default=None, help="Body length (m) for geometry-based drag fallback")
    parser.add_argument("--body_width_m", type=float, default=None, help="Body width (m) for geometry-based drag fallback")
    parser.add_argument("--body_height_m", type=float, default=None, help="Body height (m) for geometry-based drag fallback")
    parser.add_argument("--arm_length_m", type=float, default=None, help="Arm length (m) (center to motor) for geometry-based drag fallback")
    parser.add_argument("--arm_width_m", type=float, default=None, help="Arm width (m) for square-tube arm drag fallback")
    parser.add_argument("--max_tilt_deg", type=float, default=None, help="Maximum tilt angle (deg) for forward flight")
    parser.add_argument("--motor_configuration", type=str, default="flat", choices=["flat","coaxial"], help="Motor layout: flat or coaxial")
    parser.add_argument("--coaxial_spacing_m", type=float, default=None, help="Vertical spacing between coaxial rotors (m). If omitted, assumes ~0.2D")
    parser.add_argument("--speed", type=float, default=10.0, help="Speed (m/s) for single-point run")
    parser.add_argument("--periph_current", type=float, default=0.0, help="Peripheral current draw (A)")

    # Battery
    parser.add_argument("--battery_operating_voltage_min", type=float, required=False)
    parser.add_argument("--battery_operating_voltage_nominal", type=float, required=False)
    parser.add_argument("--battery_operating_voltage_max", type=float, required=False)
    parser.add_argument("--battery_energy_density", type=float, required=False)
    parser.add_argument("--battery_charge_current_max", type=float, required=False)
    parser.add_argument("--battery_discharge_cont_A", type=float, required=False)
    parser.add_argument("--battery_discharge_max_A", type=float, required=False)
    parser.add_argument("--battery_discharge_c_cont", type=float, required=False,
                        help="Continuous discharge C-rate (e.g., 15 for 15C). Used if --battery_discharge_cont not provided.")
    parser.add_argument("--battery_discharge_c_max", type=float, required=False,
                        help="Max/burst discharge C-rate (e.g., 25 for 25C). Defaults to continuous if omitted.")
    parser.add_argument("--battery_discharge_percent", type=float, default=100.0,
                        help="Percent of pack capacity to use (e.g., 80 means stop at 20 percent remaining).")
    parser.add_argument("--battery_chemistry", type=str, default=None)

    parser.add_argument("--battery_unit_mode", choices=["cell", "pack"], default="cell") # cells or pack
    parser.add_argument("--battery_series_units", type=int, required=False)    
    parser.add_argument("--battery_parallel_units", type=int, required=False, default=1)

    parser.add_argument("--battery_cells_series_per_unit", type=int, default=1)
    parser.add_argument("--battery_cells_parallel_per_unit", type=int, default=1)

    parser.add_argument("--battery_cell_capacity", type=float, required=False)  # mAh per cell
    parser.add_argument("--battery_resistance_cell", type=float, required=False) # mΩ per cell
    parser.add_argument("--battery_pack_capacity", type=float, required=False)  # mAh per pack
    parser.add_argument("--battery_pack_weight_g", type=float, required=False)
    parser.add_argument("--battery_cell_weight_g", type=float, required=False)
    parser.add_argument("--battery_soc_model", type=str, default="auto",
                        help="Battery SoC model: auto/linear/lipo/liion/lifepo4")
    parser.add_argument("--battery_soc_curve_csv", type=str, default=None,
                        help="CSV path with SoC curve columns: soc, ocv_cell, r_scale")
    parser.add_argument("--battery_soc_bp", type=str, default=None,
                        help="Comma-separated SoC breakpoints in [0,1]")
    parser.add_argument("--battery_ocv_cell_bp", type=str, default=None,
                        help="Comma-separated OCV-per-cell breakpoints (V)")
    parser.add_argument("--battery_r_scale_bp", type=str, default=None,
                        help="Comma-separated resistance scale breakpoints")
    # Motor
    parser.add_argument("--motor_kv", type=float, required=False)
    parser.add_argument("--motor_idle_current", type=float, required=False)
    parser.add_argument("--motor_idle_voltage", type=float, required=False)
    parser.add_argument("--motor_rated_voltage", type=float, required=False)
    parser.add_argument("--motor_resistance", type=float, required=False)
    parser.add_argument("--motor_max_current", type=float, required=False)
    parser.add_argument("--motor_max_power", type=float, required=False)
    parser.add_argument("--motor_pole_count", type=int, required=False)
    parser.add_argument("--motor_weight", type=float, required=False)
    parser.add_argument("--motor_size", type=str, default=None, help="Motor size/form factor (e.g., 28x28mm)")

    # ESC
    parser.add_argument("--esc_voltage_rating", type=float, required=False)
    parser.add_argument("--esc_cont_current", type=float, required=False)
    parser.add_argument("--esc_max_current", type=float, required=False)
    parser.add_argument("--esc_idle_current", type=float, required=False)
    parser.add_argument("--esc_resistance", type=float, required=False)
    parser.add_argument("--esc_weight", type=float, required=False)

    # Avionics
    parser.add_argument("--avionics_voltage_tree", type=str, default=None, help="Voltage tree for avionics power draw, e.g., '5.0:(2,0.9), 12.0:(1.5,0.85)' means 2A at 5V with 90 percent efficiency, and 1.5A at 12V with 85 percent efficiency")

    # Propeller
    parser.add_argument("--prop_diameter", type=float, required=False)
    parser.add_argument("--prop_pitch", type=float, required=False)
    parser.add_argument("--prop_blades", type=int, default=2)
    parser.add_argument("--prop_weight", type=float, required=False, help="Propeller weight per motor (g)")
    parser.add_argument("--prop_weight_g", dest="prop_weight", type=float, required=False, help=argparse.SUPPRESS)
    # These two were read via getattr() in build_drone_from_args() but never
    # defined as arguments, so they always fell back to 0 and their status
    # checks were silently skipped in every CLI and batch run.
    parser.add_argument("--prop_max_rpm", type=float, default=None,
                        help="Manufacturer prop RPM limit, used for a status check.")
    parser.add_argument("--prop_max_thrust", type=float, default=None,
                        help="Max thrust of one motor+prop (g), used for a status check.")
    parser.add_argument("--prop_table", type=str, default=None)
    parser.add_argument("--prop_tconst", type=float, default=None, help="Prop thrust coefficient (C_T-like)")
    parser.add_argument("--prop_pconst", type=float, default=None, help="Prop power coefficient (C_P-like)")

    # Mission
    parser.add_argument("--mission", type=str, default=None, help="Path to mission profile JSON; if omitted, do single-point run.")

    # Environment
    parser.add_argument("--altitude", type=float, default=0.0, help="Altitude above sea level (m)")
    parser.add_argument("--temperature", type=float, default=None, help="Ambient temperature (°C)")
    parser.add_argument("--pressure", type=float, default=None, help="Ambient pressure (Pa)")
    parser.add_argument("--wind", type=float, default=0.0, help="Wind speed (m/s)")
    parser.add_argument("--wind_direction_deg", type=float, default=0.0,
                        help="Wind direction FROM (deg), meteorological convention")
    parser.add_argument("--course_deg", type=float, default=0.0,
                        help="Vehicle course heading (deg)")
    parser.add_argument("--climb_rate_mps", type=float, default=0.0,
                        help="Commanded climb rate (m/s)")
    parser.add_argument("--descent_rate_mps", type=float, default=0.0,
                        help="Commanded descent rate (m/s)")
    parser.add_argument("--reserve_percent", type=float, default=20.0,
                        help="Mission reserve percent of usable energy")
    parser.add_argument("--rth_reserve_Wh", type=float, default=0.0,
                        help="Return-to-home reserve (Wh)")
    parser.add_argument("--diversion_reserve_Wh", type=float, default=0.0,
                        help="Diversion reserve (Wh)")
    parser.add_argument("--transient_dt_s", type=float, default=0.5,
                        help="Mission simulation timestep for transient dynamics (s)")
    parser.add_argument("--max_accel_mps2", type=float, default=2.0,
                        help="Maximum forward acceleration for mission transient model (m/s²)")
    parser.add_argument("--max_decel_mps2", type=float, default=2.5,
                        help="Maximum deceleration for mission transient model (m/s²)")
    parser.add_argument("--decel_regen_eff", type=float, default=0.0,
                        help="Fraction of deceleration kinetic power recovered (0..1)")
    parser.add_argument(
        "--inflow_map_enabled",
        type=str,
        default=None,
        help="Enable/disable rotor inflow map (true/false, 1/0, yes/no). Default: enabled.",
    )
    parser.add_argument("--inflow_mu_bp", type=str, default=None,
                        help="Comma-separated advance-ratio mu breakpoints, e.g. '0,0.1,0.2,0.3'")
    parser.add_argument("--inflow_eff_bp", type=str, default=None,
                        help="Comma-separated inflow-efficiency breakpoints, same length as mu")

    # Options
    parser.add_argument("--orientation", type=str, default="forward", choices=["hover", "forward"])
    parser.add_argument("--plot", action="store_true", help="Show matplotlib window with performance curves (CLI only)")

    return parser


def validate_required_cli_args(args):
    """
    Check that CLI mode has the arguments the physics genuinely needs.

    Only truly load-bearing inputs are required.  Anything that is either
    (a) an alternative to another input, or (b) used solely for a status
    check, is optional — demanding all of them made batch runs and the
    bundled example configs fail for no physical reason.

    Mutually exclusive alternatives, checked separately below:
      * cell_capacity / cell_weight   vs  pack_capacity / pack_weight
        (whichever matches --battery_unit_mode)
      * discharge_cont_A              vs  discharge_c_cont
    """
    required = [
        "num_motors", "weight",
        "battery_operating_voltage_min", "battery_operating_voltage_nominal",
        "battery_operating_voltage_max",
        "battery_resistance_cell", "battery_series_units",
        "motor_kv", "motor_resistance",
        "prop_diameter", "prop_pitch",
    ]
    missing = [k for k in required if getattr(args, k) is None]

    # Capacity: need whichever pair matches the unit mode.
    unit_mode = str(getattr(args, "battery_unit_mode", "cell") or "cell").strip().lower()
    if unit_mode == "pack":
        if getattr(args, "battery_pack_capacity", None) is None:
            missing.append("battery_pack_capacity")
    else:
        if getattr(args, "battery_cell_capacity", None) is None:
            missing.append("battery_cell_capacity")

    # Current limit: either an absolute amp figure or a C-rate is enough.
    if (getattr(args, "battery_discharge_cont_A", None) is None
            and getattr(args, "battery_discharge_c_cont", None) is None):
        missing.append("battery_discharge_cont_A (or --battery_discharge_c_cont)")
    if missing:
        raise SystemExit(f"Missing required CLI args: {', '.join('--' + m for m in missing)}\n"
                         f"Tip: run with --gui to use the graphical interface.")



def _load_mission_or_exit(path: str) -> "MissionProfile":
    """
    Load a mission JSON for a CLI run, failing with a readable message.

    A mistyped path or a malformed file is ordinary user error, so it should
    produce a one-line explanation rather than a Python traceback.
    """
    import os
    if not os.path.exists(path):
        raise SystemExit(
            f"Mission file not found: {path}\n"
            "Check the path, or try one of the bundled examples in "
            "examples/missions/."
        )
    try:
        return MissionProfile.from_json(path)
    except json.JSONDecodeError as e:
        raise SystemExit(f"Mission file {path} is not valid JSON: {e}")
    except (KeyError, TypeError, ValueError) as e:
        raise SystemExit(
            f"Mission file {path} could not be read: {e}\n"
            "Each entry in \"phases\" needs a name and either a duration or a "
            "distance. See examples/missions/ for working files."
        )


def main():
    parser = build_arg_parser()
    args = parser.parse_args()

    if args.gui:
        launch_gui()
        return

    validate_required_cli_args(args)
    drone = build_drone_from_args(args)
    orientation = args.orientation.strip().lower()
    if orientation not in ("hover", "forward"):
        raise ValueError("orientation must be 'hover' or 'forward'")

    print(f"Using air density = {drone.air_density:.3f} kg/m^3 at {args.altitude:.1f} m altitude")

    if args.mission:
        mission = _load_mission_or_exit(args.mission)
        mission.reserve_percent = float(args.reserve_percent)
        mission.rth_reserve_Wh = float(args.rth_reserve_Wh)
        mission.diversion_reserve_Wh = float(args.diversion_reserve_Wh)
        mission.wind_direction_deg = float(args.wind_direction_deg)
        mission.transient_dt_s = float(args.transient_dt_s)
        mission.max_accel_mps2 = float(args.max_accel_mps2)
        mission.max_decel_mps2 = float(args.max_decel_mps2)
        mission.decel_regen_eff = float(args.decel_regen_eff)
        for _p in mission.phases:
            _p.course_deg = float(_p.course_deg if _p.course_deg else args.course_deg)
            if _p.climb_rate_mps is None:
                _p.climb_rate_mps = float(args.climb_rate_mps)
            if _p.descent_rate_mps is None:
                _p.descent_rate_mps = float(args.descent_rate_mps)
        results, _worst_metrics, _mission_series = simulate_mission(
            drone,
            mission,
            orientation=orientation,
            temperature_C=args.temperature,
            pressure_Pa=args.pressure,
            wind_mps=args.wind,
        )
        for name, time_min, dist_km, status in results:
            print(f"{name}: {time_min:.1f} min, {dist_km:.2f} km, {status}")
    else:
        be_v, be_min, br_v, br_km = find_optimal_speeds(drone)
        metrics = _compute_operating_metrics_core(
            drone,
            speed_mps=args.speed,
            orientation=orientation,
            wind_mps=args.wind,
            wind_direction_deg=args.wind_direction_deg,
            course_deg=args.course_deg,
            ambient_temp_C=(float(args.temperature) if args.temperature is not None else 25.0),
        )
        climb_rate = max(float(args.climb_rate_mps), 0.0)
        descent_rate = max(float(args.descent_rate_mps), 0.0)
        if climb_rate > 0 and descent_rate > 0:
            descent_rate = 0.0
        potential_power_w = (drone.drone_weight_g * 9.81 / 1000.0) * (climb_rate - descent_rate)
        base_total_w = float(metrics.get("total_power_W", 0.0))
        adj_total_w = max(base_total_w + potential_power_w, 0.0)
        if base_total_w > 0:
            scale = adj_total_w / base_total_w
            metrics["motor_power_W"] = float(metrics.get("motor_power_W", 0.0)) * scale
            metrics["esc_loss_W"] = float(metrics.get("esc_loss_W", 0.0)) * scale
        metrics["total_power_W"] = adj_total_w
        metrics["pack_current_A"] = adj_total_w / max(float(drone.battery.vnom_pack), 1.0)
        metrics["v_load_V"] = battery_voltage_under_load(drone.battery, metrics["pack_current_A"])
        metrics["climb_rate_cmd_mps"] = climb_rate
        metrics["descent_rate_cmd_mps"] = descent_rate
        metrics["potential_power_W"] = potential_power_w
        metrics["reserve_target_Wh"] = max(
            float(drone.battery.usable_Wh) * (float(args.reserve_percent) / 100.0),
            float(args.rth_reserve_Wh) + float(args.diversion_reserve_Wh),
        )
        metrics["reserve_margin_Wh"] = float(drone.battery.usable_Wh) - float(metrics["reserve_target_Wh"])
        metrics["reserve_breached"] = bool(metrics["reserve_margin_Wh"] < 0.0)
        if adj_total_w > 0:
            t_min = float(drone.battery.usable_Wh) / adj_total_w * 60.0
        else:
            t_min = 0.0
        d_km = float(metrics.get("groundspeed_mps", 0.0)) * (t_min * 60.0) / 1000.0

        print(f"Estimated flight time at {args.speed:.2f} m/s ({orientation}): {t_min:.1f} min")
        print(f"Estimated flight distance at {args.speed:.2f} m/s ({orientation}): {d_km:.2f} km")
        print(f"SoC / model source    : {metrics.get('soc_percent', 100.0):.1f}% / {metrics.get('soc_model_source', 'linear-fallback')}")
        print(f"Ground speed          : {metrics.get('groundspeed_mps', 0.0):.2f} m/s")
        print(f"Head / Cross wind     : {metrics.get('wind_head_mps', 0.0):+.2f} / {metrics.get('wind_cross_mps', 0.0):+.2f} m/s")
        print(f"Best endurance speed (forward): {be_v:.1f} m/s -> {be_min:.1f} min")
        print(f"Best range speed (forward): {br_v:.1f} m/s -> {br_km:.2f} km")
        print(f"Hover Efficiency      : {metrics.get('hover_efficiency_gW', 0.0):.2f} g/W")
        print(f"Figure of Merit (FM)  : {metrics.get('figure_of_merit', 0.0):.3f}")
        print(f"Disk Loading          : {metrics.get('disk_loading_N_m2', 0.0):.1f} N/m²")
        tip_mach = metrics.get("tip_mach", None)
        if tip_mach is not None:
            tip_mach_f = float(tip_mach)
            noise_note = " (significant aeroacoustic noise likely)" if tip_mach_f > 0.6 else ""
            print(f"Tip Mach              : {tip_mach_f:.3f}{noise_note}")
        print(f"Advance Ratio μ       : {metrics.get('advance_ratio_mu', 0.0):.3f}")
        print(f"Inflow Efficiency η   : {metrics.get('inflow_efficiency', 1.0):.3f}")
        print(f"Inflow Power Mult.    : {metrics.get('inflow_power_multiplier', 1.0):.3f}")
        print(f"Commanded Airspeed    : {metrics.get('commanded_airspeed_mps', args.speed):.2f} m/s")
        print(f"Acceleration          : {metrics.get('accel_mps2', 0.0):+.2f} m/s²")
        print(f"Kinetic Power Term    : {metrics.get('kinetic_power_W', 0.0):+.1f} W")
        print(f"Potential Power Term  : {metrics.get('potential_power_W', 0.0):+.1f} W")
        print(f"Reserve Target/Margin : {metrics.get('reserve_target_Wh', 0.0):.1f} / {metrics.get('reserve_margin_Wh', 0.0):+.1f} Wh")
        print(f"Reserve Status        : {'VIOLATION' if metrics.get('reserve_breached', False) else 'OK'}")
        print(f"Motor Thermal Status  : {metrics.get('thermal_status', 'n/a')} @ {metrics.get('motor_temp_est_C', 0.0):.1f} °C")
        print(f"Thermal M/ESC/Batt    : {metrics.get('motor_temp_est_C', 0.0):.1f} / {metrics.get('esc_temp_est_C', 0.0):.1f} / {metrics.get('battery_temp_est_C', 0.0):.1f} °C")
        print(f"Hover Wind Resistance : {metrics.get('hover_wind_resistance_mps', 0.0):.2f} m/s")
        print(f"Prop Solidity σ       : {metrics.get('prop_solidity_sigma', 0.0):.3f}")

    if args.plot:
        plot_performance(drone)


if __name__ == "__main__":
    main()
